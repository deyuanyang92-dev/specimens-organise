from __future__ import annotations

import os
import py_compile
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from PIL import Image

from specimen_app.app_settings import DEFAULT_PHOTO_FILENAME_FILL_SHORTCUT, load_settings, save_settings, settings_path
from specimen_app.classification_fields import (
    CLASSIFICATION_COLUMNS,
    REQUIRED_CLASSIFICATION_COLUMNS,
    classification_values_from_family_match,
    classification_values_from_species_match,
)
from specimen_app.excel_store import ExcelStore
from specimen_app.image_cache import ThumbnailCache
from specimen_app.image_search import (
    _get_or_build_search_index,
    append_images_to_index,
    clear_image_index,
    default_image_query,
    extract_core_identifier,
    image_file_filter,
    image_index_exists,
    image_search_results,
    is_supported_image,
    iter_workspace_images,
    suffixes_for_image_type,
)
from specimen_app.models import CLASSIFICATION_HEADERS, ImportConflictError, WorkspaceNotInitializedError
from specimen_app.parsing import (
    derive_specimen_fields_from_tube_number,
    extract_bottle_label,
    extract_collection_date,
    extract_location_code,
    extract_photo_date,
    extract_photo_seq,
    extract_photo_seq_from_filename,
    extract_save_method_from_filename,
    extract_save_method_from_tube_number,
    extract_specimen_tube_from_filename,
    extract_tube_from_filename,
)
from specimen_app.release_manager import list_releases
from specimen_app.species import FamilyMatch, SpeciesMatch, SpeciesMatcher
from specimen_app.ui import (
    classification_column_value_from_taxonomy_match,
    default_photo_filename_fill_fields,
    format_taxonomy_candidate_label,
    grid_shape,
    photo_filename_source_for_specimen_fill,
    specimen_updates_from_photo_filename,
)
from specimen_app.workspace import has_workspace_data, initialize_workspace, is_generated_workspace_path, is_workspace


class CoreTests(unittest.TestCase):
    def setUp(self) -> None:
        clear_image_index()
        self.tmp = Path(tempfile.mkdtemp())

    def tearDown(self) -> None:
        shutil.rmtree(self.tmp)
        clear_image_index()

    def _write_species_preset(self) -> Path:
        path = self.tmp / "species_preset.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["物种中文名", "物种拉丁名", "科中文名", "科拉丁名"])
        ws.append(["珠江川纽虫", "Amniclineus zhujiangensis", "纵沟纽虫科", "Lineidae"])
        ws.append(["青纵沟纽虫", "Lineus fuscoviridis", "纵沟纽虫科", "Lineidae"])
        ws.append(["习见脑纽虫", "Cerebratulina communis", "纵沟纽虫科", "Lineidae"])
        ws.append(["戴氏脑纽虫", "Cerebratulina darvelli", "纵沟纽虫科", "Lineidae"])
        ws.append(["珠角裸沙蚕", "Nicon moniloceras", "沙蚕科", "Nereididae"])
        wb.save(path)
        wb.close()
        return path

    def test_create_vouchers_increment(self) -> None:
        store = ExcelStore(self.tmp)
        self.assertEqual(store.create_specimen(), "YZZ000001")
        self.assertEqual(store.create_specimen(), "YZZ000002")
        self.assertEqual(store.list_vouchers(), ["YZZ000001", "YZZ000002"])

    def test_build_release_script_compiles(self) -> None:
        project_root = Path(__file__).resolve().parents[1]
        py_compile.compile(str(project_root / "build_release.py"), doraise=True)

    def test_old_classification_config_names_are_not_used(self) -> None:
        project_root = Path(__file__).resolve().parents[1]
        old_names = [
            "classification_config",
            "CLASSIFICATION_FIELDS",
            "CLASSIFICATION_REQUIRED_FIELDS",
            "CLASSIFICATION_DISPLAY_FIELDS",
            "SPECIES_COMPLETER_FIELDS",
            "FAMILY_COMPLETER_FIELDS",
            "CLASSIFICATION_COMPLETER_FIELDS",
            "species_autofill_updates",
            "family_autofill_updates",
        ]
        for path in (project_root / "specimen_app").glob("*.py"):
            if path.name == "classification_fields.py":
                continue
            text = path.read_text(encoding="utf-8")
            for name in old_names:
                self.assertNotIn(name, text, f"{name} still appears in {path.name}")

    def test_classification_schema_includes_optional_higher_taxonomy_fields(self) -> None:
        for field in ["属名", "目", "纲", "门", "备注"]:
            self.assertIn(field, CLASSIFICATION_COLUMNS)
            self.assertNotIn(field, REQUIRED_CLASSIFICATION_COLUMNS)

    def test_species_matcher_searches_species_and_family_fields(self) -> None:
        matcher = SpeciesMatcher(self._write_species_preset())

        self.assertEqual(matcher.species_matches("珠江")[0].latin_name, "Amniclineus zhujiangensis")
        self.assertEqual(matcher.species_matches("Amnic")[0].chinese_name, "珠江川纽虫")

        family_names = [match.family_name for match in matcher.family_matches("纵沟")]
        self.assertEqual(family_names, ["纵沟纽虫科"])
        self.assertEqual(matcher.family_matches("Line")[0].family_name, "纵沟纽虫科")
        self.assertEqual(matcher.species_matches("纵沟")[0].chinese_name, "青纵沟纽虫")

        self.assertEqual(matcher.resolve_unique_species("珠江").chinese_name, "珠江川纽虫")
        self.assertEqual(
            matcher.resolve_unique_species("amniclineus zhujiangensis").chinese_name,
            "珠江川纽虫",
        )
        self.assertIsNone(matcher.resolve_unique_species("珠"))
        self.assertIsNone(matcher.resolve_unique_species("纵沟"))
        self.assertEqual(matcher.resolve_unique_family("line").family_latin, "Lineidae")

    def test_classification_autofill_update_maps(self) -> None:
        species = SpeciesMatch(
            chinese_name="珠江川纽虫",
            latin_name="Amniclineus zhujiangensis",
            family_name="纵沟纽虫科",
            family_latin="Lineidae",
        )
        family = FamilyMatch(family_name="纵沟纽虫科", family_latin="Lineidae")

        self.assertEqual(
            classification_values_from_species_match(species),
            {
                "种名*": "珠江川纽虫",
                "种拉丁": "Amniclineus zhujiangensis",
                "属名": "Amniclineus",
                "科*": "纵沟纽虫科",
                "科拉丁": "Lineidae",
            },
        )
        self.assertEqual(
            classification_values_from_family_match(family),
            {
                "科*": "纵沟纽虫科",
                "科拉丁": "Lineidae",
            },
        )

    def test_taxonomy_candidate_display_is_not_inserted_into_species_field(self) -> None:
        species = SpeciesMatch(
            chinese_name="青纵沟纽虫",
            latin_name="Lineus fuscoviridis",
            family_name="纵沟纽虫科",
            family_latin="Lineidae",
        )

        self.assertEqual(
            format_taxonomy_candidate_label("种名*", "species", species),
            "青纵沟纽虫  Lineus fuscoviridis  纵沟纽虫科  Lineidae",
        )
        self.assertEqual(
            classification_column_value_from_taxonomy_match("种名*", "species", species),
            "青纵沟纽虫",
        )
        self.assertEqual(
            classification_column_value_from_taxonomy_match("种拉丁", "species", species),
            "Lineus fuscoviridis",
        )
        self.assertEqual(
            classification_column_value_from_taxonomy_match("属名", "species", species),
            "Lineus",
        )

        self.assertEqual(
            classification_values_from_species_match(species),
            {
                "种名*": "青纵沟纽虫",
                "种拉丁": "Lineus fuscoviridis",
                "属名": "Lineus",
                "科*": "纵沟纽虫科",
                "科拉丁": "Lineidae",
            },
        )

    def test_existing_classification_workbook_gets_new_optional_columns(self) -> None:
        workspace = self.tmp / "old_classification_schema"
        workspace.mkdir()
        ExcelStore(workspace)
        old_headers = ["入库编号*", "种名*", "种拉丁", "科*", "科拉丁"]
        path = workspace / "数据" / "分类信息.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(old_headers)
        ws.append(["YZZ000001", "旧种", "Oldus species", "旧科", "Oldidae"])
        wb.save(path)
        wb.close()

        reopened = ExcelStore(workspace)
        row = reopened.get_classification("YZZ000001")
        self.assertEqual(row["种名*"], "旧种")
        self.assertEqual(row["种拉丁"], "Oldus species")
        self.assertEqual(row["科*"], "旧科")
        self.assertEqual(row["属名"], "")
        self.assertEqual(row["备注"], "")

        upgraded = load_workbook(path, read_only=True, data_only=True)
        try:
            headers = [str(cell.value or "") for cell in next(upgraded.active.iter_rows(max_row=1))]
        finally:
            upgraded.close()
        for field in CLASSIFICATION_HEADERS:
            self.assertIn(field, headers)

        reopened.set_fields("classification", "YZZ000001", {"属名": "Oldus", "备注": "只能鉴定到属"})
        updated = reopened.get_classification("YZZ000001")
        self.assertEqual(updated["种名*"], "旧种")
        self.assertEqual(updated["属名"], "Oldus")
        self.assertEqual(updated["备注"], "只能鉴定到属")

    def test_tube_number_parsing(self) -> None:
        self.assertEqual(extract_location_code("QD-LSD-SC001-1-R-250923"), "QD-LSD")
        self.assertEqual(extract_bottle_label("QD-LSD-SC001-1-R-250923"), "QD-LSD-SC001")
        self.assertEqual(extract_bottle_label("QD-CK-SC008-260827"), "QD-CK-SC008")
        self.assertEqual(extract_collection_date("QD-LSD-SC001-1-R-250923"), "2025-09-23")
        self.assertEqual(extract_collection_date("QD-LSD-SC001-20250923"), "2025-09-23")

    def test_tube_number_parsing_extracts_photo_sequence(self) -> None:
        examples = [
            ("QD-CK-SC008-20240315", "QD-CK-SC008", 1, "2024-03-15"),
            ("QD-CK-SC008-2-20240315", "QD-CK-SC008", 2, "2024-03-15"),
            ("QD-CK-SC008-3-20240315", "QD-CK-SC008", 3, "2024-03-15"),
            ("QD-CK-SC008--2-20240315-xxx", "QD-CK-SC008", 2, "2024-03-15"),
        ]
        for tube, bottle_label, photo_seq, collection_date in examples:
            with self.subTest(tube=tube):
                self.assertEqual(extract_bottle_label(tube), bottle_label)
                self.assertEqual(extract_photo_seq(tube), photo_seq)
                self.assertEqual(extract_collection_date(tube), collection_date)

    def test_photo_filename_parsing_extracts_tube_date_and_sequence(self) -> None:
        self.assertEqual(extract_tube_from_filename("QD-CK-SC008-2-20240315-a.jpg"), "QD-CK-SC008")
        # 兼容 GXRG-A-BZC001.tif：原规则要求第二段地点码至少 2 位，导致无法填充 GXRG-A。
        self.assertEqual(extract_tube_from_filename("GXRG-A-BZC001.tif"), "GXRG-A-BZC001")
        self.assertEqual(extract_photo_seq_from_filename("QD-CK-SC008-2-20240315-a.jpg"), 2)
        self.assertEqual(extract_photo_seq_from_filename("QD-CK-SC008--3-20240315-a.jpg"), 3)
        self.assertEqual(extract_photo_seq_from_filename("QD-CK-SC008-20240315-a.jpg"), 1)
        self.assertEqual(extract_photo_date("QD-CK-SC008--3-20240315-a.jpg"), "2024-03-15")

    def test_photo_filename_parsing_extracts_specimen_fill_fields(self) -> None:
        self.assertEqual(
            extract_specimen_tube_from_filename("QD-CK-SC008-2-20240315-a.jpg"),
            "QD-CK-SC008-2-20240315",
        )
        self.assertEqual(
            extract_specimen_tube_from_filename("QD_CK_SC008_3_20250923.tif"),
            "QD-CK-SC008-3-20250923",
        )
        self.assertEqual(
            extract_specimen_tube_from_filename("QD-LSD-SC001-1-R-250923.jpg"),
            "QD-LSD-SC001-1-R-250923",
        )
        # 没有日期或保存方式也应保留核心编号，后续可填充采集地点缩写 GXRG-A。
        self.assertEqual(extract_specimen_tube_from_filename("GXRG-A-BZC001.tif"), "GXRG-A-BZC001")
        self.assertEqual(extract_specimen_tube_from_filename("QD-CK-SC008-1.tif"), "QD-CK-SC008")
        self.assertEqual(extract_save_method_from_filename("QD-LSD-SC001-1-R-250923.jpg"), "RE")
        self.assertEqual(extract_save_method_from_filename("XM-ABC-SC001-FE-250924.jpg"), "FE")

    def test_tube_number_derives_specimen_fill_fields(self) -> None:
        # 原测试只覆盖日期和地点；旧版本管内编号也能派生保存方式，这里防止再次丢失。
        self.assertEqual(extract_save_method_from_tube_number("QD-LSD-SC001-1-R-250923"), "RE")
        self.assertEqual(extract_save_method_from_tube_number("QD-LSD-SC001-250923-R"), "RE")
        self.assertEqual(extract_save_method_from_tube_number("XM-ABC-SC001-FE-250924"), "FE")
        self.assertEqual(extract_save_method_from_tube_number("XM-ABC-SC001-9E-250924"), "9E")
        self.assertEqual(extract_save_method_from_tube_number("XM-ABC-SC001-7E-250924"), "7E")
        self.assertEqual(extract_save_method_from_tube_number("XM-ABC-SC001-79-250924"), "79")
        self.assertEqual(
            derive_specimen_fields_from_tube_number("QD-LSD-SC001-1-R-250923"),
            {"采集地点缩写*": "QD-LSD", "采集日期": "2025-09-23", "保存方式": "RE"},
        )
        self.assertEqual(
            derive_specimen_fields_from_tube_number("GXRG-A-BZC001"),
            {"采集地点缩写*": "GXRG-A"},
        )

    def test_photo_filename_fill_helpers_are_conservative(self) -> None:
        row = {"文件名": "archived_2.jpg", "原始文件名": "QD-LSD-SC001-1-R-250923.jpg"}
        self.assertEqual(photo_filename_source_for_specimen_fill(row), "QD-LSD-SC001-1-R-250923.jpg")
        updates = specimen_updates_from_photo_filename(photo_filename_source_for_specimen_fill(row))
        self.assertEqual(
            updates,
            {
                "管内编号*": "QD-LSD-SC001-1-R-250923",
                "采集地点缩写*": "QD-LSD",
                "采集日期": "2025-09-23",
                "保存方式": "RE",
            },
        )
        defaults = default_photo_filename_fill_fields(
            updates,
            {"管内编号*": "", "采集地点缩写*": "OLD", "采集日期": "", "保存方式": "FE"},
        )
        self.assertEqual(defaults, ["管内编号*", "采集日期"])

        # 兼容无日期/保存方式的文件名：至少能从核心编号填充管内编号和采集地点。
        gxrg_updates = specimen_updates_from_photo_filename("GXRG-A-BZC001.tif")
        self.assertEqual(gxrg_updates, {"管内编号*": "GXRG-A-BZC001", "采集地点缩写*": "GXRG-A"})

    def test_set_field_autofills_tube_derived_fields(self) -> None:
        store = ExcelStore(self.tmp)
        voucher = store.create_specimen()
        store.set_fields("specimen", voucher, {"管内编号*": "QD-LSD-SC001-1-R-250923"})
        row = store.get_specimen(voucher)
        self.assertEqual(row["采集日期"], "2025-09-23")
        self.assertEqual(row["采集地点缩写*"], "QD-LSD")
        self.assertEqual(row["保存方式"], "RE")
        store.set_fields("specimen", voucher, {"管内编号*": "XM-ABC-SC001-1-R-250924"})
        row = store.get_specimen(voucher)
        self.assertEqual(row["采集日期"], "2025-09-24")
        self.assertEqual(row["采集地点缩写*"], "XM-ABC")
        self.assertEqual(row["保存方式"], "RE")
        store.set_fields("specimen", voucher, {"管内编号*": "XM-ABC-SC001-FE-250924"})
        row = store.get_specimen(voucher)
        self.assertEqual(row["保存方式"], "FE")

    def test_photo_filename_fill_can_disable_hidden_derived_overwrite(self) -> None:
        store = ExcelStore(self.tmp)
        voucher = store.create_specimen()
        store.set_fields("specimen", voucher, {"采集日期": "2024-01-01", "采集地点缩写*": "OLD"})
        store.set_fields(
            "specimen",
            voucher,
            {"管内编号*": "QD-LSD-SC001-1-R-250923"},
            auto_derive_specimen_fields=False,
        )
        row = store.get_specimen(voucher)
        self.assertEqual(row["管内编号*"], "QD-LSD-SC001-1-R-250923")
        self.assertEqual(row["采集日期"], "2024-01-01")
        self.assertEqual(row["采集地点缩写*"], "OLD")

    def test_undo_redo_field_update(self) -> None:
        store = ExcelStore(self.tmp)
        voucher = store.create_specimen()
        store.set_fields("specimen", voucher, {"管内编号*": "QD-LSD-SC001-1-R-250923"})
        self.assertEqual(store.get_specimen(voucher)["管内编号*"], "QD-LSD-SC001-1-R-250923")
        store.undo_last()
        self.assertEqual(store.get_specimen(voucher)["管内编号*"], "")
        store.redo_last()
        self.assertEqual(store.get_specimen(voucher)["管内编号*"], "QD-LSD-SC001-1-R-250923")

    def test_redo_order_after_multiple_undo(self) -> None:
        store = ExcelStore(self.tmp)
        voucher = store.create_specimen()
        store.set_fields("specimen", voucher, {"管内编号*": "QD-LSD-SC001-1-R-250923"})
        # 原测试手动再写 RE；现在管内编号会自动派生 RE，因此改用 FE 保留“两步修改”的测试意图。
        store.set_fields("specimen", voucher, {"保存方式": "FE"})
        store.undo_last()
        store.undo_last()
        self.assertEqual(store.get_specimen(voucher)["管内编号*"], "")
        self.assertEqual(store.get_specimen(voucher)["保存方式"], "")
        store.redo_last()
        self.assertEqual(store.get_specimen(voucher)["管内编号*"], "QD-LSD-SC001-1-R-250923")
        self.assertEqual(store.get_specimen(voucher)["保存方式"], "RE")
        store.redo_last()
        self.assertEqual(store.get_specimen(voucher)["保存方式"], "FE")

    def test_import_conflict_blocks_write(self) -> None:
        target = self.tmp / "target"
        source = self.tmp / "source"
        target.mkdir()
        source.mkdir()
        target_store = ExcelStore(target)
        voucher = target_store.create_specimen()
        target_store.set_fields("specimen", voucher, {"管内编号*": "QD-LSD-SC001-1-R-250923"})

        source_store = ExcelStore(source)
        source_voucher = source_store.create_specimen()
        self.assertEqual(source_voucher, voucher)
        source_store.set_fields("specimen", source_voucher, {"管内编号*": "QD-CK-SC008-1-R-250923"})

        with self.assertRaises(ImportConflictError) as ctx:
            target_store.import_workspace(source)
        self.assertIsNotNone(ctx.exception.report_path)
        self.assertEqual(target_store.list_vouchers(), [voucher])

    def test_import_same_record_skips(self) -> None:
        target = self.tmp / "target_same"
        source = self.tmp / "source_same"
        target.mkdir()
        source.mkdir()
        target_store = ExcelStore(target)
        voucher = target_store.create_specimen()
        target_store.set_fields("specimen", voucher, {"管内编号*": "QD-LSD-SC001-1-R-250923"})

        source_store = ExcelStore(source)
        source_voucher = source_store.create_specimen()
        source_store.set_fields("specimen", source_voucher, {"管内编号*": "QD-LSD-SC001-1-R-250923"})

        result = target_store.import_workspace(source)
        self.assertEqual(result.imported, 0)
        self.assertEqual(result.skipped, 1)
        self.assertEqual(target_store.list_vouchers(), [voucher])

    def test_import_from_file_success(self) -> None:
        target = self.tmp / "target_file"
        target.mkdir()
        source = self.tmp / "source.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["入库编号*", "管内编号*", "保存方式", "采集日期", "采集地点缩写*", "入库日期", "标本存放位置", "信息录入人员", "核对人员", "备注"])
        ws.append(["YZZ000003", "QD-LSD-SC001-1-R-250923", "", "", "", "", "", "", "", ""])
        wb.save(source)

        store = ExcelStore(target)
        result = store.import_from_file(source)
        self.assertEqual(result.imported, 1)
        self.assertEqual(store.list_vouchers(), ["YZZ000003"])
        self.assertEqual(store.get_specimen("YZZ000003")["管内编号*"], "QD-LSD-SC001-1-R-250923")

    def test_import_from_file_blocks_duplicate_source_ids(self) -> None:
        target = self.tmp / "target_file_duplicate"
        target.mkdir()
        source = self.tmp / "source_duplicate.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["入库编号*", "管内编号*", "保存方式", "采集日期", "采集地点缩写*", "入库日期", "标本存放位置", "信息录入人员", "核对人员", "备注"])
        ws.append(["YZZ000003", "QD-LSD-SC001-1-R-250923", "", "", "", "", "", "", "", ""])
        ws.append(["YZZ000003", "QD-CK-SC008-1-R-250923", "", "", "", "", "", "", "", ""])
        wb.save(source)

        store = ExcelStore(target)
        with self.assertRaises(ImportConflictError):
            store.import_from_file(source)

    def test_duplicate_detection(self) -> None:
        store = ExcelStore(self.tmp)
        store.create_specimen()
        path = self.tmp / "数据" / "标本信息.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["入库编号*", "管内编号*", "保存方式", "采集日期", "采集地点缩写*", "入库日期", "标本存放位置", "备注"])
        ws.append(["YZZ000001", "A", "", "", "", "", "", ""])
        ws.append(["YZZ000001", "B", "", "", "", "", "", ""])
        wb.save(path)
        with self.assertRaises(ImportConflictError):
            store.assert_unique_vouchers()

    def test_multiple_photo_records_are_preserved(self) -> None:
        store = ExcelStore(self.tmp)
        voucher = store.create_specimen()
        photo_dir = self.tmp / "照片"
        photo_dir.mkdir()
        first = photo_dir / "photo1.tif"
        second = photo_dir / "photo2.jpg"
        first.write_bytes(b"fake-tiff")
        second.write_bytes(b"fake-jpg")
        store.add_photo(voucher, first)
        store.add_photo(voucher, second)
        photos = store.get_photos(voucher)
        self.assertEqual(len(photos), 2)
        self.assertEqual(photos[0]["文件名"], "photo1.tif")
        self.assertEqual(photos[1]["文件名"], "photo2.jpg")
        self.assertEqual(photos[0]["相对路径"], "./照片/photo1.tif")
        self.assertEqual(photos[1]["相对路径"], "./照片/photo2.jpg")
        self.assertEqual(photos[0]["原始文件名"], "photo1.tif")
        self.assertEqual(photos[1]["原始文件名"], "photo2.jpg")
        self.assertEqual(photos[0]["归档状态"], "已归档")
        self.assertTrue(store.resolve_photo_path(photos[0]).exists())
        self.assertTrue(store.resolve_photo_path(photos[1]).exists())

    def test_same_name_different_photo_uses_numbered_archive_name(self) -> None:
        workspace = self.tmp / "workspace_same_name"
        workspace.mkdir()
        first_dir = self.tmp / "first"
        second_dir = self.tmp / "second"
        first_dir.mkdir()
        second_dir.mkdir()
        first = first_dir / "same.jpg"
        second = second_dir / "same.jpg"
        first.write_bytes(b"one")
        second.write_bytes(b"two")
        store = ExcelStore(workspace)
        voucher = store.create_specimen()

        first_row = store.add_photo(voucher, first, allow_outside=True)
        second_row = store.add_photo(voucher, second, allow_outside=True)

        self.assertEqual(first_row["文件名"], "same.jpg")
        self.assertEqual(second_row["文件名"], "same_2.jpg")
        self.assertEqual(first_row["相对路径"], "./照片/same.jpg")
        self.assertEqual(second_row["相对路径"], "./照片/same_2.jpg")
        self.assertEqual(first_row["原始文件名"], "same.jpg")
        self.assertEqual(second_row["原始文件名"], "same.jpg")
        self.assertTrue((workspace / "照片" / "same.jpg").exists())
        self.assertTrue((workspace / "照片" / "same_2.jpg").exists())

    def test_photo_description_update_does_not_crash(self) -> None:
        store = ExcelStore(self.tmp)
        voucher = store.create_specimen()
        photo_dir = self.tmp / "照片"
        photo_dir.mkdir()
        first = photo_dir / "photo1.jpg"
        first.write_bytes(b"fake-jpg")
        store.add_photo(voucher, first)
        self.assertTrue(store.set_photo_description(voucher, 0, "背面照片"))
        self.assertEqual(store.get_photos(voucher)[0]["描述"], "背面照片")

    def test_photo_filename_update_preserves_archive_metadata(self) -> None:
        store = ExcelStore(self.tmp)
        voucher = store.create_specimen()
        photo_dir = self.tmp / "照片"
        photo_dir.mkdir()
        first = photo_dir / "raw_name.jpg"
        first.write_bytes(b"fake-jpg")
        store.add_photo(voucher, first)
        old_row = store.get_photos(voucher)[0]

        new_name = "QD-CK-SC008-2-20240315.jpg"
        self.assertTrue(store.set_photo_filename(voucher, 0, new_name))
        updated = store.get_photos(voucher)[0]
        self.assertEqual(updated["文件名"], new_name)
        self.assertEqual(updated["原始文件名"], "raw_name.jpg")
        self.assertEqual(updated["相对路径"], f"./照片/{new_name}")
        self.assertTrue(store.resolve_photo_path(updated).exists())
        self.assertFalse(store.resolve_photo_path(old_row).exists())

        self.assertTrue(store.undo_last())
        self.assertEqual(store.get_photos(voucher)[0]["文件名"], "raw_name.jpg")
        self.assertTrue((photo_dir / "raw_name.jpg").exists())
        self.assertFalse((photo_dir / new_name).exists())
        self.assertTrue(store.redo_last())
        self.assertEqual(store.get_photos(voucher)[0]["文件名"], new_name)
        self.assertFalse((photo_dir / "raw_name.jpg").exists())
        self.assertTrue((photo_dir / new_name).exists())

    def test_external_photo_is_copied_into_workspace_archive(self) -> None:
        workspace = self.tmp / "workspace"
        workspace.mkdir()
        outside_dir = self.tmp / "external"
        outside_dir.mkdir()
        outside = outside_dir / "outside.jpg"
        outside.write_bytes(b"fake-jpg")
        store = ExcelStore(workspace)
        voucher = store.create_specimen()
        row = store.add_photo(voucher, outside, allow_outside=True)
        resolved = store.resolve_photo_path(row)
        self.assertEqual(row["文件名"], "outside.jpg")
        self.assertEqual(row["原始文件名"], "outside.jpg")
        self.assertEqual(Path(row["原始路径"]), outside)
        self.assertEqual(row["来源工作区根路径"], "")
        self.assertEqual(row["相对路径"], "./照片/outside.jpg")
        self.assertEqual(resolved.parent, workspace / "照片")
        self.assertTrue(resolved.exists())
        self.assertTrue(outside.exists())

    def test_delete_photo_removes_unreferenced_archive_file(self) -> None:
        workspace = self.tmp / "workspace_delete_photo"
        workspace.mkdir()
        outside = self.tmp / "delete_me.jpg"
        outside.write_bytes(b"fake-jpg")
        store = ExcelStore(workspace)
        voucher = store.create_specimen()
        row = store.add_photo(voucher, outside, allow_outside=True)
        archived = store.resolve_photo_path(row)
        self.assertTrue(archived.exists())

        self.assertTrue(store.delete_photo(voucher, 0))

        self.assertEqual(store.get_photos(voucher), [])
        self.assertFalse(archived.exists())
        self.assertTrue(outside.exists())

    def test_hash_prefixed_archive_paths_are_migrated_to_original_names(self) -> None:
        workspace = self.tmp / "workspace_hash_migrate"
        data_dir = workspace / "数据"
        archive_dir = workspace / "照片"
        data_dir.mkdir(parents=True)
        archive_dir.mkdir()
        hashed = archive_dir / "abcdef123456__legacy.jpg"
        hashed.write_bytes(b"legacy")
        store = ExcelStore(workspace)
        voucher = store.create_specimen()
        rows = store.read_rows("photo")
        rows.append(
            {
                "入库编号*": voucher,
                "文件名": "legacy.jpg",
                "相对路径": "./照片/abcdef123456__legacy.jpg",
                "描述": "",
                "来源工作区根路径": "",
                "原始文件名": "legacy.jpg",
                "原始路径": str(self.tmp / "legacy.jpg"),
                "文件SHA256": store._file_sha256(hashed),
                "文件大小": str(hashed.stat().st_size),
                "归档时间": "",
                "归档状态": "已归档",
            }
        )
        store._write_rows("photo", rows)
        store.config["data_schema_version"] = "1.1.0"
        store._save_config()

        migrated = ExcelStore(workspace)
        photo = migrated.get_photos(voucher)[0]
        self.assertEqual(photo["文件名"], "legacy.jpg")
        self.assertEqual(photo["相对路径"], "./照片/legacy.jpg")
        self.assertTrue((archive_dir / "legacy.jpg").exists())
        self.assertFalse(hashed.exists())

    def test_photo_conflicts_use_archived_file_hash(self) -> None:
        workspace = self.tmp / "workspace_hash_conflict"
        workspace.mkdir()
        outside = self.tmp / "same_photo.jpg"
        outside.write_bytes(b"same-content")
        store = ExcelStore(workspace)
        first = store.create_specimen()
        second = store.create_specimen()
        store.add_photo(first, outside, allow_outside=True)

        conflicts = store.find_photo_conflicts([outside], second)
        self.assertEqual(conflicts, {str(outside.resolve()): first})

    def test_resolve_photo_path_blocks_parent_traversal_without_source_root(self) -> None:
        workspace = self.tmp / "workspace_traversal"
        workspace.mkdir()
        outside = self.tmp / "outside.jpg"
        outside.write_bytes(b"fake-jpg")
        store = ExcelStore(workspace)
        resolved = store.resolve_photo_path({"相对路径": "../outside.jpg", "来源工作区根路径": ""})
        self.assertNotEqual(resolved, outside)
        self.assertFalse(resolved.exists())

    def test_replace_photo_failure_preserves_old_record(self) -> None:
        workspace = self.tmp / "workspace_replace"
        workspace.mkdir()
        photo_dir = workspace / "照片"
        photo_dir.mkdir()
        old = photo_dir / "old.jpg"
        old.write_bytes(b"old")
        missing = self.tmp / "missing_replace.jpg"
        store = ExcelStore(workspace)
        voucher = store.create_specimen()
        old_row = store.add_photo(voucher, old)
        with self.assertRaises(FileNotFoundError):
            store.replace_photo(voucher, 0, missing, allow_outside=False)
        self.assertEqual(store.get_photos(voucher)[0], old_row)

    def test_replace_photo_archives_external_copy(self) -> None:
        workspace = self.tmp / "workspace_replace_external"
        workspace.mkdir()
        photo_dir = workspace / "照片"
        photo_dir.mkdir()
        old = photo_dir / "old.jpg"
        old.write_bytes(b"old")
        outside = self.tmp / "outside_replace.jpg"
        outside.write_bytes(b"new")
        store = ExcelStore(workspace)
        voucher = store.create_specimen()
        old_row = store.add_photo(voucher, old)
        old_archived = store.resolve_photo_path(old_row)

        new_row = store.replace_photo(voucher, 0, outside, allow_outside=False)
        self.assertIsNotNone(new_row)
        photos = store.get_photos(voucher)
        self.assertEqual(photos[0]["文件名"], "outside_replace.jpg")
        self.assertEqual(Path(photos[0]["原始路径"]), outside)
        self.assertEqual(store.resolve_photo_path(photos[0]).parent, workspace / "照片")
        self.assertFalse(old_archived.exists())
        self.assertTrue(outside.exists())

    def test_import_workspace_archives_found_photos_and_reports_missing(self) -> None:
        source = self.tmp / "source_archive_import"
        target = self.tmp / "target_archive_import"
        source.mkdir()
        target.mkdir()
        source_store = ExcelStore(source)
        found_voucher = source_store.create_specimen()
        missing_voucher = source_store.create_specimen()
        found_original = self.tmp / "found_original.jpg"
        missing_original = self.tmp / "missing_original.jpg"
        found_original.write_bytes(b"found")
        missing_original.write_bytes(b"missing")
        source_store.add_photo(found_voucher, found_original, allow_outside=True)
        missing_row = source_store.add_photo(missing_voucher, missing_original, allow_outside=True)
        source_store.resolve_photo_path(missing_row).unlink()
        missing_original.unlink()

        target_store = ExcelStore(target)
        result = target_store.import_workspace(source)

        self.assertEqual(result.imported, 2)
        self.assertEqual(result.photos_imported, 1)
        self.assertIsNotNone(result.report_path)
        self.assertTrue(result.report_path.exists())
        found_photos = target_store.get_photos(found_voucher)
        self.assertEqual(len(found_photos), 1)
        self.assertEqual(target_store.resolve_photo_path(found_photos[0]).parent, target / "照片")
        self.assertEqual(target_store.get_photos(missing_voucher), [])

    def test_move_photos_is_atomic_and_undoable(self) -> None:
        store = ExcelStore(self.tmp)
        source = store.create_specimen()
        target = store.create_specimen()
        photo_dir = self.tmp / "照片"
        photo_dir.mkdir()
        first = photo_dir / "move1.jpg"
        second = photo_dir / "move2.jpg"
        first.write_bytes(b"one")
        second.write_bytes(b"two")
        store.add_photo(source, first)
        store.add_photo(source, second)

        self.assertEqual(store.move_photos(source, target, [0]), 1)
        self.assertEqual([row["文件名"] for row in store.get_photos(source)], ["move2.jpg"])
        self.assertEqual([row["文件名"] for row in store.get_photos(target)], ["move1.jpg"])
        self.assertEqual(store.undo_last(), "move_photos")
        self.assertEqual([row["文件名"] for row in store.get_photos(source)], ["move2.jpg", "move1.jpg"])
        self.assertEqual(store.get_photos(target), [])
        self.assertEqual(store.redo_last(), "move_photos")
        self.assertEqual([row["文件名"] for row in store.get_photos(target)], ["move1.jpg"])

    def test_grid_filename_setting_defaults_and_roundtrips(self) -> None:
        old_appdata = os.environ.get("APPDATA")
        os.environ["APPDATA"] = str(self.tmp / "config")
        try:
            path = settings_path()
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text('{"preview_quality":"standard"}', encoding="utf-8")
            settings = load_settings()
            self.assertTrue(settings.show_grid_filenames)
            self.assertEqual(settings.photo_filename_fill_shortcut, DEFAULT_PHOTO_FILENAME_FILL_SHORTCUT)

            settings.show_grid_filenames = False
            settings.photo_filename_fill_shortcut = "Ctrl+Shift+F"
            save_settings(settings)
            reloaded = load_settings()
            self.assertFalse(reloaded.show_grid_filenames)
            self.assertEqual(reloaded.photo_filename_fill_shortcut, "Ctrl+Shift+F")
        finally:
            if old_appdata is None:
                os.environ.pop("APPDATA", None)
            else:
                os.environ["APPDATA"] = old_appdata

    def test_batch_photo_records_undo_and_redo_as_one_action(self) -> None:
        store = ExcelStore(self.tmp)
        voucher = store.create_specimen()
        photo_dir = self.tmp / "照片"
        photo_dir.mkdir()
        paths = []
        for name in ["batch1.jpg", "batch2.jpg", "batch3.jpg"]:
            path = photo_dir / name
            path.write_bytes(b"fake-jpg")
            paths.append(path)

        added = store.add_photos(voucher, paths)
        self.assertEqual(len(added), 3)
        self.assertEqual(len(store.get_photos(voucher)), 3)
        self.assertEqual(store.undo_last(), "add_photos")
        self.assertEqual(store.get_photos(voucher), [])
        self.assertEqual(store.redo_last(), "add_photos")
        self.assertEqual([row["文件名"] for row in store.get_photos(voucher)], ["batch1.jpg", "batch2.jpg", "batch3.jpg"])

    def test_image_scan_excludes_generated_and_version_dirs(self) -> None:
        kept = self.tmp / "采集照片" / "YZZ000001_QD-LSD.jpg"
        kept.parent.mkdir()
        kept.write_bytes(b"jpg")
        for relative in [
            "build/tmp.jpg",
            "dist/app.jpg",
            "releases/v0.2.2/old.jpg",
            "数据/数据版本/snapshot.jpg",
            "数据/缩略图缓存/cache.jpg",
        ]:
            path = self.tmp / relative
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(b"jpg")

        scanned = {path.relative_to(self.tmp).as_posix() for path in iter_workspace_images(self.tmp)}
        self.assertEqual(scanned, {"采集照片/YZZ000001_QD-LSD.jpg"})

    def test_image_search_requires_core_identifier_at_file_name_start(self) -> None:
        photo_dir = self.tmp / "照片"
        photo_dir.mkdir()
        first = photo_dir / "QD-CK-WenSC004-2-20250923-青岛沧口吻沙蚕.tif"
        second = photo_dir / "QD_CK_WenSC004_3_20250923.tif"
        wrong_number = photo_dir / "QD-CK-WenSC0042-20250923.tif"
        wrong_core = photo_dir / "QD-CK-WenSC005-20250923.tif"
        camera_name = photo_dir / "PB060001.tif"
        jpg_match = photo_dir / "QD-CK-WenSC004-4-20250923.jpg"
        for path in [first, second, wrong_number, wrong_core, camera_name, jpg_match]:
            path.write_bytes(b"image")

        results = image_search_results(
            self.tmp,
            "YZZ000001",
            {},
            {},
            [second],
            query="QD-CK-WenSC004",
        )
        self.assertEqual([result.file_name for result in results], [first.name, second.name])
        self.assertEqual(results[0].matched_keywords, ("QD-CK-WenSC004",))
        self.assertTrue(results[1].is_linked)

    def test_image_search_can_switch_to_jpg_or_tif_jpg(self) -> None:
        photo_dir = self.tmp / "照片"
        photo_dir.mkdir()
        tif = photo_dir / "QD-CK-WenSC004-1.tif"
        jpg = photo_dir / "QD-CK-WenSC004-2.jpg"
        jpeg = photo_dir / "QD-CK-WenSC004-3.jpeg"
        for path in [tif, jpg, jpeg]:
            path.write_bytes(b"image")

        default_results = image_search_results(self.tmp, "YZZ000001", {}, {}, [], query="QD-CK-WenSC004")
        self.assertEqual([result.file_name for result in default_results], [tif.name])

        jpg_results = image_search_results(
            self.tmp,
            "YZZ000001",
            {},
            {},
            [],
            query="QD-CK-WenSC004",
            suffixes=suffixes_for_image_type("jpg"),
        )
        self.assertEqual([result.file_name for result in jpg_results], [jpg.name, jpeg.name])

        combined_results = image_search_results(
            self.tmp,
            "YZZ000001",
            {},
            {},
            [],
            query="QD-CK-WenSC004",
            suffixes=suffixes_for_image_type("tif_jpg"),
        )
        self.assertEqual([result.file_name for result in combined_results], [tif.name, jpg.name, jpeg.name])

    def test_image_search_supports_common_image_suffixes(self) -> None:
        photo_dir = self.tmp / "照片"
        photo_dir.mkdir()
        names = [
            "P001-a.webp",
            "P001-b.gif",
            "P001-c.jfif",
            "P001-d.jpe",
            "P001-e.jp2",
            "P001-f.j2k",
        ]
        for name in names:
            (photo_dir / name).write_bytes(b"image")

        results = image_search_results(
            self.tmp,
            "YZZ000001",
            {},
            {},
            [],
            query="P001",
            suffixes=suffixes_for_image_type("all"),
        )
        self.assertEqual([result.file_name for result in results], names)
        self.assertIn(".webp", suffixes_for_image_type("all"))
        self.assertIn(".jfif", suffixes_for_image_type("jpg"))
        self.assertTrue(is_supported_image(photo_dir / "P001-a.webp"))
        self.assertFalse(is_supported_image(photo_dir / "P001.txt"))
        self.assertIn("*.webp", image_file_filter())

    def test_image_search_falls_back_to_filename_contains_match(self) -> None:
        photo_dir = self.tmp / "照片"
        photo_dir.mkdir()
        middle_code = photo_dir / "sampleP001middle.webp"
        hyphen_code = photo_dir / "图版-A-111-背面.jpg"
        for path in [middle_code, hyphen_code]:
            path.write_bytes(b"image")

        p001_results = image_search_results(
            self.tmp,
            "YZZ000001",
            {},
            {},
            [],
            query="P001",
            suffixes=suffixes_for_image_type("all"),
        )
        self.assertEqual([result.file_name for result in p001_results], [middle_code.name])
        self.assertEqual(p001_results[0].score, 60)

        hyphen_results = image_search_results(
            self.tmp,
            "YZZ000001",
            {},
            {},
            [],
            query="A-111",
            suffixes=suffixes_for_image_type("all"),
        )
        self.assertEqual([result.file_name for result in hyphen_results], [hyphen_code.name])

    def test_image_search_matches_single_letter_hyphen_prefix(self) -> None:
        project = self.tmp / "广西海洋大学图谱项目"
        first = project / "钩齿短脊虫（6张）" / "A-钩齿短脊虫.tif"
        second = project / "扁蛰虫（3张）" / "图200-A-扁蛰虫体前部背面观.tif"
        for path in [first, second]:
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(b"image")

        results = image_search_results(
            self.tmp,
            "YZZ000001",
            {},
            {},
            [],
            query="A-",
            extra_roots=[project],
            suffixes=suffixes_for_image_type("tif_jpg"),
        )
        self.assertEqual([result.file_name for result in results], [first.name, second.name])

    def test_image_search_ignores_cached_index_from_different_scope(self) -> None:
        photo_dir = self.tmp / "照片"
        photo_dir.mkdir()
        default_photo = photo_dir / "QD-CK-SC008-1.tif"
        default_photo.write_bytes(b"image")
        default_index = _get_or_build_search_index([photo_dir], cache_root=self.tmp)

        project = self.tmp / "广西海洋大学图谱项目"
        project_photo = project / "钩齿短脊虫（6张）" / "A-钩齿短脊虫.tif"
        project_photo.parent.mkdir(parents=True)
        project_photo.write_bytes(b"image")

        results = image_search_results(
            self.tmp,
            "YZZ000001",
            {},
            {},
            [],
            query="A-",
            extra_roots=[project],
            search_index=default_index,
        )
        self.assertEqual([result.file_name for result in results], [project_photo.name])

    def test_image_search_empty_query_does_not_scan_or_match(self) -> None:
        photo_dir = self.tmp / "照片"
        photo_dir.mkdir()
        path = photo_dir / "QD-CK-WenSC004-1.tif"
        path.write_bytes(b"image")
        self.assertEqual(image_search_results(self.tmp, "YZZ000001", {}, {}, [], query=""), [])

    def test_image_search_reuses_index_and_force_rebuilds(self) -> None:
        photo_dir = self.tmp / "照片"
        photo_dir.mkdir()
        first = photo_dir / "QD-CK-WenSC004-1.tif"
        second = photo_dir / "QD-CK-WenSC004-2.tif"
        first.write_bytes(b"image")
        initial = image_search_results(self.tmp, "YZZ000001", {}, {}, [], query="QD-CK-WenSC004")
        self.assertEqual([result.file_name for result in initial], [first.name])

        second.write_bytes(b"image")
        cached = image_search_results(self.tmp, "YZZ000001", {}, {}, [], query="QD-CK-WenSC004")
        self.assertEqual([result.file_name for result in cached], [first.name])
        rebuilt = image_search_results(
            self.tmp,
            "YZZ000001",
            {},
            {},
            [],
            query="QD-CK-WenSC004",
            force_rebuild=True,
        )
        self.assertEqual([result.file_name for result in rebuilt], [first.name, second.name])

    def test_image_search_index_appends_new_photos_without_rescan(self) -> None:
        photo_dir = self.tmp / "照片"
        photo_dir.mkdir()
        first = photo_dir / "QD-CK-WenSC004-1.tif"
        second = photo_dir / "QD-CK-WenSC004-2.tif"
        third = photo_dir / "QD-CK-WenSC004-3.jpg"
        first.write_bytes(b"image")

        initial = image_search_results(self.tmp, "YZZ000001", {}, {}, [], query="QD-CK-WenSC004")
        self.assertEqual([result.file_name for result in initial], [first.name])
        self.assertTrue(image_index_exists(self.tmp))

        second.write_bytes(b"image")
        self.assertEqual(append_images_to_index(self.tmp, [second]), 1)
        appended = image_search_results(self.tmp, "YZZ000001", {}, {}, [], query="QD-CK-WenSC004")
        self.assertEqual([result.file_name for result in appended], [first.name, second.name])
        self.assertEqual(append_images_to_index(self.tmp, [second]), 0)

        third.write_bytes(b"image")
        clear_image_index()
        self.assertEqual(append_images_to_index(self.tmp, [third]), 1)
        combined = image_search_results(
            self.tmp,
            "YZZ000001",
            {},
            {},
            [],
            query="QD-CK-WenSC004",
            suffixes=suffixes_for_image_type("tif_jpg"),
        )
        self.assertEqual([result.file_name for result in combined], [first.name, second.name, third.name])

    def test_image_search_uses_core_identifier_from_tube_number(self) -> None:
        photo_dir = self.tmp / "照片"
        photo_dir.mkdir()
        first = photo_dir / "QD-CK-SC008-1-20250923-青岛沧口沙蚕.tif"
        second = photo_dir / "QD-CK-SC008-2-20250923-青岛沧口沙蚕.tif"
        unrelated = photo_dir / "QD-CK-WSC003-20250923-青岛沧口围沙蚕.tif"
        for path in [first, second, unrelated]:
            path.write_bytes(b"image")

        specimen = {"管内编号*": "QD-CK-SC008-260827", "采集地点缩写*": "QD-CK"}
        self.assertEqual(default_image_query(specimen), "QD-CK-SC008")
        self.assertEqual(extract_core_identifier("QD_CK_SC008_260827"), "QD-CK-SC008")
        results = image_search_results(self.tmp, "YZZ000003", specimen, {}, [], query="QD-CK-SC008-260827")
        self.assertEqual([result.file_name for result in results], [first.name, second.name])

    def test_thumbnail_cache_reuses_and_invalidates_by_source_metadata(self) -> None:
        source = self.tmp / "照片"
        source.mkdir()
        image_path = source / "large.jpg"
        Image.new("RGB", (120, 80), "red").save(image_path)
        cache = ThumbnailCache(self.tmp)
        first = cache.thumbnail(image_path, (40, 40))
        cached_files = sorted((self.tmp / "数据" / "缩略图缓存").glob("*.jpg"))
        self.assertEqual(len(cached_files), 1)
        second = cache.thumbnail(image_path, (40, 40))
        self.assertEqual(second.size, first.size)
        self.assertEqual(sorted((self.tmp / "数据" / "缩略图缓存").glob("*.jpg")), cached_files)
        Image.new("RGB", (130, 90), "blue").save(image_path)
        third = cache.thumbnail(image_path, (40, 40))
        self.assertLessEqual(third.width, 40)
        self.assertEqual(len(list((self.tmp / "数据" / "缩略图缓存").glob("*.jpg"))), 2)

    def test_grid_shape_options(self) -> None:
        self.assertEqual(grid_shape(2), (2, 1))
        self.assertEqual(grid_shape(4), (2, 2))
        self.assertEqual(grid_shape(6), (3, 2))
        self.assertEqual(grid_shape(8), (4, 2))

    def test_snapshot_and_restore(self) -> None:
        store = ExcelStore(self.tmp)
        voucher = store.create_specimen()
        store.set_fields("specimen", voucher, {"管内编号*": "QD-LSD-SC001-1-R-250923"})
        snapshot = store.create_data_snapshot("测试快照", "保存初始管内编号")
        store.set_fields("specimen", voucher, {"管内编号*": "XM-ABC-SC001-1-R-250924"})
        self.assertEqual(store.get_specimen(voucher)["管内编号*"], "XM-ABC-SC001-1-R-250924")
        store.restore_data_snapshot(snapshot)
        self.assertEqual(store.get_specimen(voucher)["管内编号*"], "QD-LSD-SC001-1-R-250923")
        versions = store.list_data_versions()
        self.assertTrue(any(row["操作类型"] == "回退数据版本" for row in versions))

    def test_workspace_initialization_and_detection(self) -> None:
        source = self.tmp / "source_workspace"
        target = self.tmp / "new_workspace"
        (source / "字段模版").mkdir(parents=True)
        (source / "字段模版" / "表格信息预设字段.xlsx").write_text("placeholder", encoding="utf-8")
        initialize_workspace(target, source)
        self.assertTrue(is_workspace(target))
        self.assertTrue((target / "数据").exists())
        self.assertTrue((target / "字段模版" / "表格信息预设字段.xlsx").exists())
        self.assertTrue(has_workspace_data(ExcelStore(target).root))

    def test_excel_store_does_not_create_uninitialized_workspace_without_permission(self) -> None:
        target = self.tmp / "empty_workspace"
        target.mkdir()
        with self.assertRaises(WorkspaceNotInitializedError):
            ExcelStore(target, create_if_missing=False)
        self.assertFalse((target / "数据").exists())

    def test_generated_release_dirs_are_not_valid_workspaces(self) -> None:
        release_dir = self.tmp / "releases" / "v0.2.3"
        (release_dir / "数据").mkdir(parents=True)
        (release_dir / "数据" / "工作区配置.json").write_text("{}", encoding="utf-8")
        self.assertTrue(is_generated_workspace_path(release_dir))
        self.assertFalse(is_workspace(release_dir))

    def test_release_listing(self) -> None:
        release_dir = self.tmp / "releases" / "v0.2.0"
        release_dir.mkdir(parents=True)
        if sys.platform == "win32":
            exe = release_dir / "标本入库管理_v0.2.0.exe"
            exe.write_bytes(b"exe")
        else:
            exe = release_dir / "标本入库管理_v0.2.0"
            exe.write_text("#!/bin/sh", encoding="utf-8")
            os.chmod(exe, 0o755)
        (release_dir / "release_notes.md").write_text("# notes", encoding="utf-8")
        releases = list_releases(self.tmp)
        self.assertEqual(releases[0].version, "v0.2.0")
        self.assertEqual(releases[0].exe_path, exe)


if __name__ == "__main__":
    unittest.main()
