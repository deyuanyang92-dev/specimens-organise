"""通过 WoRMS REST API 构建 bootstrap 缓存。

DwC-A 全库 zip 返回 403。改用 `AphiaRecordsByMatchNames` 端点：
- 支持模糊匹配（TAXAMATCH），自动纠正拼写错误
- 批量查询（一次 ~50 个名字），大幅减少请求数
- 用户预设列表里的轻微拼写错误也能找到正确条目

输入：specimen_app/字段模版/表格信息预设字段.xlsx（物种拉丁名列）
输出：specimen_app/assets/worms_cache_bootstrap.sqlite.gz

使用：
    python tools/build_worms_cache_via_api.py
    # 选项: --limit N 只跑前 N 个（测试用），--batch 调批大小
"""

from __future__ import annotations

import argparse
import json
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import load_workbook  # noqa: E402

from specimen_app.worms_client import (  # noqa: E402
    WoRMSRecord,
    cache_stats,
    export_cache_gz,
    write_to_cache,
)


PRESET_PATH = PROJECT_ROOT / "specimen_app" / "字段模版" / "表格信息预设字段.xlsx"
FALLBACK_PRESET_PATH = PROJECT_ROOT / "字段模版" / "表格信息预设字段.xlsx"
OUTPUT_PATH = PROJECT_ROOT / "specimen_app" / "assets" / "worms_cache_bootstrap.sqlite.gz"

# 批量模糊匹配端点：一次请求多个名字，自动纠正拼写。
_MATCH_URL = "https://www.marinespecies.org/rest/AphiaRecordsByMatchNames"
_USER_AGENT = "specimen-inventory-build/1.0"


def _resolve_preset() -> Path:
    if PRESET_PATH.is_file():
        return PRESET_PATH
    if FALLBACK_PRESET_PATH.is_file():
        return FALLBACK_PRESET_PATH
    raise SystemExit(f"找不到物种预设：{PRESET_PATH} 或 {FALLBACK_PRESET_PATH}")


def _read_latin_names(xlsx_path: Path) -> list[str]:
    """读取「物种拉丁名」列，去重去空。"""
    wb = load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active
    headers = [(c.value or "").strip() if c.value else "" for c in next(ws.iter_rows())]
    try:
        latin_idx = headers.index("物种拉丁名")
    except ValueError:
        raise SystemExit(f"找不到「物种拉丁名」列：{headers}")
    names: list[str] = []
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[latin_idx]
        if not val:
            continue
        name = str(val).strip()
        if not name or name in seen:
            continue
        seen.add(name)
        names.append(name)
    wb.close()
    return names


def _match_batch(names: list[str], timeout: int = 30) -> list[list[dict]]:
    """批量模糊匹配。返回与输入等长的列表，每项是该名字的匹配数组。"""
    qs = "&".join("scientificnames%5B%5D=" + urllib.parse.quote(n) for n in names)
    url = f"{_MATCH_URL}?{qs}"
    req = urllib.request.Request(url, headers={
        "User-Agent": _USER_AGENT,
        "Accept": "application/json",
    })
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        raw = resp.read()
        status = getattr(resp, "status", None) or resp.getcode()
    if status == 204 or not raw or not raw.strip():
        return [[] for _ in names]
    data = json.loads(raw)
    # 服务端按输入顺序返回，但有时长度不匹配，做容错。
    if not isinstance(data, list):
        return [[] for _ in names]
    while len(data) < len(names):
        data.append([])
    return data


def _record_from_dict(d: dict) -> WoRMSRecord:
    return WoRMSRecord(
        aphia_id=int(d.get("AphiaID") or 0),
        status=str(d.get("status", "")),
        valid_name=str(d.get("valid_name") or d.get("scientificname", "")),
        valid_aphia_id=int(d.get("valid_AphiaID") or 0),
        phylum=str(d.get("phylum") or ""),
        class_=str(d.get("class") or ""),
        order=str(d.get("order") or ""),
        family=str(d.get("family") or ""),
        genus=str(d.get("genus") or ""),
        authority=str(d.get("authority") or ""),
        rank=str(d.get("rank") or ""),
    )


def build(limit: int | None, batch_size: int, delay: float) -> None:
    preset = _resolve_preset()
    print(f"读取物种预设：{preset}")
    names = _read_latin_names(preset)
    if limit:
        names = names[:limit]
    print(f"待查询 {len(names)} 个物种，每批 {batch_size}，请求间隔 {delay}s")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    hit = miss = err = 0
    start = time.time()
    total_batches = (len(names) + batch_size - 1) // batch_size
    for batch_i in range(total_batches):
        batch = names[batch_i * batch_size:(batch_i + 1) * batch_size]
        try:
            results = _match_batch(batch)
        except (urllib.error.URLError, json.JSONDecodeError) as e:
            err += len(batch)
            print(f"  批 {batch_i + 1}/{total_batches} 失败: {e}")
            if delay > 0:
                time.sleep(delay)
            continue
        for name, matches in zip(batch, results):
            if not matches:
                miss += 1
                continue
            # 优先选 status=accepted；否则取首个。
            accepted = next((m for m in matches if m.get("status") == "accepted"), None)
            pick = accepted or matches[0]
            try:
                write_to_cache(_record_from_dict(pick))
                hit += 1
            except Exception as e:
                err += 1
                print(f"  写入失败 {name}: {e}")
        done = min((batch_i + 1) * batch_size, len(names))
        elapsed = time.time() - start
        rate = done / elapsed if elapsed > 0 else 0
        eta = (len(names) - done) / rate if rate > 0 else 0
        print(f"  [{done}/{len(names)}] 命中 {hit} 未匹配 {miss} 错 {err} | "
              f"{rate:.1f}/s | ETA {eta:.0f}s")
        if delay > 0 and batch_i + 1 < total_batches:
            time.sleep(delay)

    stats = cache_stats()
    print(f"\n缓存统计：共 {stats.get('count', 0)} 条记录")

    print(f"导出 gz：{OUTPUT_PATH}")
    n = export_cache_gz(OUTPUT_PATH)
    size_mb = OUTPUT_PATH.stat().st_size / 1024 / 1024
    print(f"完成：{n} 条记录，{size_mb:.2f} MB")


def main() -> None:
    ap = argparse.ArgumentParser(description="通过 REST API 构建 WoRMS bootstrap 缓存")
    ap.add_argument("--limit", type=int, default=None, help="只跑前 N 个（调试用）")
    ap.add_argument("--batch", type=int, default=50,
                    help="每批查询数（WoRMS 上限约 50，默认 50）")
    ap.add_argument("--delay", type=float, default=1.0,
                    help="批之间休眠秒数（默认 1.0）")
    args = ap.parse_args()
    build(args.limit, args.batch, args.delay)


if __name__ == "__main__":
    main()
