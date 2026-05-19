from __future__ import annotations

import copy
import hashlib
import json
import os
import re
import shutil
import sys
import uuid
from collections import Counter, OrderedDict
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterator

# 规范化软件设计 2026-05 P1 优化:openpyxl 改 lazy import。
# 旧:模块顶层 import,加载 lxml + XML 字符串表 ~10-15MB,模块加载即占。
# 现:首次 _ensure_openpyxl() 才加载;启动 splash 出现前不必占。
# 注 (Grill G1): lazy 后 numpy 阻塞从顶层时刻迁移到首次调用时刻;首次调用时
#   numpy 可能已被 tifffile/PIL 间接 import,该阻塞条件可能不再命中。
#   openpyxl 纯 Python 路径不强依赖 numpy 阻塞,语义可接受。
# 注:Workbook / load_workbook 仍在模块级作为名字;每个用点先调 _ensure_openpyxl()。
Workbook = None  # type: ignore[assignment]
load_workbook = None  # type: ignore[assignment]


def _ensure_openpyxl() -> None:
    """首次调用时把 openpyxl 的 Workbook / load_workbook 注入到本模块 globals。"""
    global Workbook, load_workbook
    if Workbook is not None:
        return
    _numpy_module = sys.modules.get("numpy")
    _blocked_numpy_for_openpyxl = "numpy" not in sys.modules
    if _blocked_numpy_for_openpyxl:
        sys.modules["numpy"] = None
    try:
        from openpyxl import Workbook as _Wb, load_workbook as _lwb
    finally:
        if _blocked_numpy_for_openpyxl:
            sys.modules.pop("numpy", None)
        elif _numpy_module is not None:
            sys.modules["numpy"] = _numpy_module
    Workbook = _Wb
    load_workbook = _lwb

from . import __version__
from .models import (
    ACTION_LOG_FILE,
    ACTION_LOG_HEADERS,
    ALLOC_LOG_FILE,
    ALLOC_LOG_HEADERS,
    CATEGORY_FILES,
    CATEGORY_HEADERS,
    CHANGE_LOG_FILE,
    CHANGE_LOG_HEADERS,
    CHANGE_SUMMARY_HEADERS,
    CLASSIFICATION_FILE,
    CLASSIFICATION_HEADERS,
    CLASSIFICATION_REQUIRED,
    CLASSIFICATION_SUMMARY_FIELDS,
    CURRENT_DATA_SCHEMA_VERSION,
    DATA_VERSION_DIR,
    DATA_VERSION_LOG_FILE,
    DATA_VERSION_LOG_HEADERS,
    DISPLAY_CATEGORY_NAMES,
    DuplicateVoucherError,
    INDEX_FILE,
    INDEX_HEADERS,
    ImportConflictError,
    ImportResult,
    PHOTO_COUNT_COLUMN,
    PHOTO_DESC_COLUMN,
    PHOTO_FILE,
    PHOTO_FILENAME_COLUMN,
    PHOTO_HEADERS,
    PHOTO_PATH_COLUMN,
    SPECIMEN_FILE,
    SPECIMEN_HEADERS,
    SPECIMEN_REQUIRED,
    SUMMARY_COLUMNS,
    SUMMARY_COLUMN_SOURCE,
    WORKSPACE_CONFIG_FILE,
    WorkspaceLockedError,
    WorkspaceNotInitializedError,
    Row,
    StatusFlags,
)
from .app_settings import PHOTO_MANAGEMENT_OPTIONS
from .accession_series import AccessionSeries, format_series_number, series_prefix_of
from .parsing import derive_specimen_fields_from_tube_number, format_voucher, parse_voucher_serial
from .startup_diag import mark as _startup_mark


def _voucher_sort_key(value: str) -> tuple[int, int, str]:
    """YZZ 编号按流水号排在最前；非 YZZ 编号追加在后按字母排序。"""
    serial = parse_voucher_serial(value)
    if serial is not None:
        return (0, serial, "")
    return (1, 0, str(value))


DEFAULT_CONFIG = {
    "workspace_id": "",
    "prefix": "YZZ",
    "next_serial": 1,
    "undo_depth": 200,
    "data_schema_version": CURRENT_DATA_SCHEMA_VERSION,
    # 多系列入库编号支持（旧工作区缺失这两个键时用此默认值，行为与升级前完全一致）
    "active_series_name": "YZZ",
    "accession_series": [],
}


class ExcelStore:
    def __init__(self, workspace_root: Path | str, lock: bool = False,
                 create_if_missing: bool = True, read_only: bool = False):
        """初始化 ExcelStore。

        规范化软件设计 2026-05 多窗口支持:read_only=True 时跳过 lock_workspace,
        允同工作区多只读副本窗口同时存在;所有写操作 raise PermissionError。
        """
        self.root = Path(workspace_root).resolve()
        self.data_dir = self.root / "数据"
        self.lock_file = self.data_dir / ".workspace.lock"
        self._locked = False
        self._read_only = bool(read_only)
        self._create_if_missing = create_if_missing
        # 规范化软件设计 2026-05 P1 审查修复:_row_cache 加 LRU 上限。
        # 2026-05 内存档位扩展:maxsize 由 memory_profile 驱动 (3/4/6/12/20)。
        # settings 不可用 fallback 到 8 (老默认)。
        self._row_cache: OrderedDict[str, list[Row]] = OrderedDict()
        self._file_mtimes: dict[str, float] = {}
        try:
            from .app_settings import load_settings
            from .env_detect import memory_profile_params
            profile = load_settings().memory_profile
            self._row_cache_maxsize = memory_profile_params(profile)["row_cache_maxsize"]
        except Exception:
            self._row_cache_maxsize = 8
        if not self.data_dir.exists():
            if not create_if_missing:
                raise WorkspaceNotInitializedError(f"该工作目录尚未初始化，缺少数据目录：{self.data_dir}")
            self.data_dir.mkdir(exist_ok=True)
        if not self.data_dir.is_dir():
            raise WorkspaceNotInitializedError(f"数据路径不是目录：{self.data_dir}")
        if not create_if_missing and not self._has_workspace_seed_files():
            raise WorkspaceNotInitializedError(f"该工作目录尚未初始化，缺少数据文件：{self.data_dir}")
        self.config = self._load_or_create_config()
        # 只读模式跳锁:同工作区多只读副本可共存,主写窗口仍持锁
        if lock and not self._read_only:
            self.acquire_lock()
            import atexit
            # C2: 限时执行 release_lock，防 SMB/NAS 网络锁挂死导致进程无法退出。
            # 用守护线程 + Event 超时：3 秒内释放即正常；卡住则放弃（10min stale 自愈兜底）。
            atexit.register(self._release_lock_with_timeout, 3.0)

        # 只读模式守卫:覆盖所有写入 API
        if self._read_only:
            self._install_readonly_guards()
        # 启动诊断埋点：逐子步骤打点，定位"启动死机"卡在哪一步。
        self.ensure_files()
        _startup_mark("ExcelStore.ensure_files")
        self._upgrade_workspace_schema()
        _startup_mark("ExcelStore._upgrade_workspace_schema")
        self._assert_supported_data_schema()
        self.ensure_index()
        _startup_mark("ExcelStore.ensure_index")
        self._sync_next_serial()
        _startup_mark("ExcelStore._sync_next_serial")

    def close(self) -> None:
        """释放工作区锁文件。退出应用前应调用，避免遗留过期锁。

        `__init__` 已注册 atexit 钩子，但显式调用更可靠。
        """
        self.release_lock()

    def _install_readonly_guards(self) -> None:
        """规范化软件设计 2026-05 多窗口:只读模式覆盖所有写方法。

        覆盖 write API → raise PermissionError;只读副本窗口禁所有数据变更。
        读 API (read_rows / get_specimen / read_alloc_log / workspace_overview 等) 不动。
        """
        def _ro(name):
            def _denied(*args, **kwargs):
                raise PermissionError(
                    f"只读模式禁止写入 ({name})。请在主窗口操作或关闭只读副本。"
                )
            return _denied
        write_methods = [
            "create_specimen", "set_fields", "import_workspace",
            "create_data_snapshot", "restore_data_snapshot",
            "undo_last", "redo_last", "set_undo_depth",
            "downgrade_schema_version", "batch_reserve_vouchers",
            "log_alloc_event", "set_active_series",
            "ensure_assignee_series", "upgrade_to_multi_user_protocol",
        ]
        for name in write_methods:
            if hasattr(self, name):
                setattr(self, name, _ro(name))

    def acquire_lock(self) -> None:
        if self._locked:
            return
        payload = {
            "pid": os.getpid(),
            "time": datetime.now().isoformat(timespec="seconds"),
            "workspace": str(self.root),
        }
        try:
            fd = os.open(self.lock_file, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        except FileExistsError:
            if not self._lock_is_stale():
                content = ""
                try:
                    content = self.lock_file.read_text(encoding="utf-8")
                except OSError:
                    pass
                raise WorkspaceLockedError(f"工作区已被占用：{content}")
            try:
                self.lock_file.unlink()
            except OSError:
                pass
            try:
                fd = os.open(self.lock_file, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            except FileExistsError as exc:
                raise WorkspaceLockedError("工作区已被占用") from exc
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)
        self._locked = True

    def _lock_is_stale(self) -> bool:
        try:
            content = self.lock_file.read_text(encoding="utf-8")
            info = json.loads(content)
            pid = int(info.get("pid", 0))
            lock_time = info.get("time", "")
        except (OSError, json.JSONDecodeError, ValueError, TypeError):
            return True
        if pid <= 0 or pid == os.getpid():
            return True
        if lock_time:
            try:
                locked_at = datetime.fromisoformat(lock_time)
                if (datetime.now() - locked_at).total_seconds() > 600:
                    return True
            except (ValueError, TypeError):
                pass
        try:
            os.kill(pid, 0)
        except ProcessLookupError:
            return True
        except PermissionError:
            return False
        return False

    def release_lock(self) -> None:
        if not self._locked:
            return
        try:
            self.lock_file.unlink(missing_ok=True)
        finally:
            self._locked = False

    def _release_lock_with_timeout(self, timeout_seconds: float) -> None:
        """C2: 限时执行 release_lock。

        SMB / NAS / OneDrive 等网络文件系统在断网或同步抢占时，`unlink()` 可能阻塞
        数十秒甚至挂死，会让应用退出卡住。本方法用后台线程做 release，主线程超时即
        返回 — 留下的锁文件由现有 10min stale 检测机制兜底自愈。

        atexit 调用方收到的"返回"≠ 实际"释放完成"，仅表示"主进程不再等"。
        """
        if not self._locked:
            return
        import threading

        done = threading.Event()

        def _worker() -> None:
            try:
                self.release_lock()
            except Exception:
                pass
            finally:
                done.set()

        t = threading.Thread(target=_worker, name="release_lock", daemon=True)
        t.start()
        done.wait(timeout=timeout_seconds)
        # 不 join — 让 daemon 线程在主进程退出时自动死

    def ensure_files(self) -> None:
        # 工作区锁保证单进程访问，启动时遗留的 .tmp 文件都是上次崩溃留下的，安全删除。
        for _stale in self.data_dir.glob("*.tmp"):
            try:
                _stale.unlink()
            except OSError:
                pass
        self._ensure_workbook(self.data_dir / SPECIMEN_FILE, SPECIMEN_HEADERS)
        self._ensure_workbook(self.data_dir / PHOTO_FILE, PHOTO_HEADERS)
        self._ensure_workbook(self.data_dir / CLASSIFICATION_FILE, CLASSIFICATION_HEADERS)
        self._ensure_workbook(self.data_dir / INDEX_FILE, INDEX_HEADERS)
        self._ensure_change_log()
        self._ensure_workbook(self.data_dir / ACTION_LOG_FILE, ACTION_LOG_HEADERS)
        self._ensure_workbook(self.data_dir / DATA_VERSION_LOG_FILE, DATA_VERSION_LOG_HEADERS)
        self._ensure_alloc_log()

    def _has_workspace_seed_files(self) -> bool:
        return any(
            (self.data_dir / file_name).exists()
            for file_name in [WORKSPACE_CONFIG_FILE, SPECIMEN_FILE, PHOTO_FILE, CLASSIFICATION_FILE, INDEX_FILE]
        )

    def list_vouchers(self, series_filter: str | None = None) -> list[str]:
        """返回工作区内全部入库编号，按编号流水号升序排序。

        series_filter: None=全部；"YZZ"=仅 YZZ；其他字符串=按前缀匹配非 YZZ 系列。
        """
        rows = self.read_rows("specimen")
        vouchers = [self._value(row, "入库编号*") for row in rows if self._value(row, "入库编号*")]
        if series_filter is not None:
            if series_filter == "YZZ":
                vouchers = [v for v in vouchers if parse_voucher_serial(v) is not None]
            else:
                vouchers = [v for v in vouchers if series_prefix_of(v) == series_filter]
        return sorted(vouchers, key=_voucher_sort_key)

    def workspace_overview(self) -> dict[str, Any]:
        """汇总主界面凭证列表所需的概览数据。

        规范化软件设计 2026-05 启动卡死优化:
        - 旧: 三次 read_rows("specimen"/"classification"/"photo") 读全表 + 缓存到 _row_cache,
          中型工作区 (5000 凭证) 瞬时 RSS +95MB,2GB 机器触发 swap 卡死。
        - 现: 流式 _stream_columns() 只读必要列,不缓存全表,峰值降 50%。
          下游 get_specimen / get_classification / get_photos 等仍走 read_rows 加 _row_cache,
          首次访问时才触发完整读取(按需)。

        返回 dict 键(不变,API 兼容):
        - ``vouchers``: list[str],全部入库编号(按流水号排序)
        - ``flags``: dict[voucher -> StatusFlags]
        - ``photo_counts``: dict[voucher -> int]
        - ``tube_numbers``: dict[voucher -> str]
        - ``photo_filenames``: dict[voucher -> list[str]]
        """
        # 只读必要列(具体字段集随 StatusFlags 必需字段变化)。
        spec_cols = set(SPECIMEN_REQUIRED) | {"入库编号*", "管内编号*"}
        class_cols = set(CLASSIFICATION_REQUIRED) | {"入库编号*"}
        photo_cols = {"入库编号*", "文件名"}

        # photo 表聚合: 计数 + 文件名 list
        photo_counts: dict[str, int] = {}
        photo_filenames: dict[str, list[str]] = {}
        photo_path = self.data_dir / CATEGORY_FILES["photo"]
        for row in self._stream_columns(photo_path, photo_cols):
            voucher = row.get("入库编号*", "")
            if not voucher:
                continue
            photo_counts[voucher] = photo_counts.get(voucher, 0) + 1
            file_name = row.get("文件名", "")
            if file_name:
                photo_filenames.setdefault(voucher, []).append(file_name)

        # classification 表聚合: voucher -> required 字段 dict
        class_by_voucher: dict[str, dict[str, str]] = {}
        class_path = self.data_dir / CATEGORY_FILES["classification"]
        for row in self._stream_columns(class_path, class_cols):
            voucher = row.get("入库编号*", "")
            if voucher:
                class_by_voucher[voucher] = row

        # specimen 表聚合: 凭证列表 + tube + flags
        vouchers: list[str] = []
        flags: dict[str, StatusFlags] = {}
        tube_numbers: dict[str, str] = {}
        spec_path = self.data_dir / CATEGORY_FILES["specimen"]
        for row in self._stream_columns(spec_path, spec_cols):
            voucher = row.get("入库编号*", "")
            if not voucher:
                continue
            vouchers.append(voucher)
            tube = row.get("管内编号*", "")
            if tube:
                tube_numbers[voucher] = tube
            class_row = class_by_voucher.get(voucher, {})
            flags[voucher] = StatusFlags(
                specimen_complete=all(row.get(field, "") for field in SPECIMEN_REQUIRED),
                has_photo=photo_counts.get(voucher, 0) > 0,
                classification_complete=bool(class_row) and all(class_row.get(field, "") for field in CLASSIFICATION_REQUIRED),
            )
        vouchers.sort(key=_voucher_sort_key)
        return {
            "vouchers": vouchers,
            "flags": flags,
            "photo_counts": photo_counts,
            "tube_numbers": tube_numbers,
            "photo_filenames": photo_filenames,
        }

    def _stream_columns(self, path: Path, wanted_columns: set[str]) -> "Iterator[dict[str, str]]":
        """流式读 Excel,只 yield 包含 wanted_columns 字段的 sparse dict。

        规范化软件设计 2026-05 新增,供 workspace_overview 用,避免 read_rows 全列读 + 缓存。
        - 不进 _row_cache,本方法只服务 overview 的轻量聚合。
        - 流式 iter_rows,不 list() 物化。
        - 不在 wanted_columns 内的列直接跳,sparse dict 进一步省内存。
        - 文件不存在 / 表头为空 -> yield 0 行。
        """
        if not path.exists():
            return
        _ensure_openpyxl()
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                return
            headers = [self._string(v) for v in header_row]
            # 预计算 wanted 列在 raw 中的 (idx, header) 列表,避免每行重判定。
            wanted_idx: list[tuple[int, str]] = [
                (i, h) for i, h in enumerate(headers) if h in wanted_columns
            ]
            for raw in rows_iter:
                row: dict[str, str] = {}
                for idx, header in wanted_idx:
                    if idx < len(raw):
                        value = self._string(raw[idx])
                        if value != "":
                            row[header] = value
                if row:
                    yield row
        finally:
            wb.close()

    def summary_records(self) -> list[dict[str, Any]]:
        """把分散在多个 Excel 的字段汇总成一张宽表（纯内存视图，不改任何文件结构）。

        每条记录是扁平 dict，键为 SUMMARY_COLUMNS：标本全字段 + 分类全字段（分类"备注"
        用"分类备注"消歧）+ 照片数 + 照片聚合列（照片文件名 / 照片绝对路径 / 照片描述，
        均为 list，按入库编号聚合该编号下所有照片对应值）。
        左连接：缺分类信息的入库编号也会出现，分类列留空。
        """
        specimens = self.read_rows("specimen")
        classifications = self.read_rows("classification")
        photos = self.read_rows("photo")
        class_by_voucher = {
            self._value(row, "入库编号*"): row
            for row in classifications
            if self._value(row, "入库编号*")
        }
        photo_counts: dict[str, int] = {}
        photo_filenames: dict[str, list[str]] = {}
        # 同模式再聚合绝对路径 / 描述，供入库汇总的照片聚合列使用。
        photo_abs_paths: dict[str, list[str]] = {}
        photo_descs: dict[str, list[str]] = {}
        for row in photos:
            voucher = self._value(row, "入库编号*")
            if not voucher:
                continue
            photo_counts[voucher] = photo_counts.get(voucher, 0) + 1
            file_name = self._value(row, "文件名")
            if file_name:
                photo_filenames.setdefault(voucher, []).append(file_name)
            abs_path = self._value(row, "绝对路径")
            if abs_path:
                photo_abs_paths.setdefault(voucher, []).append(abs_path)
            desc = self._value(row, "描述")
            if desc:
                photo_descs.setdefault(voucher, []).append(desc)
        records: list[dict[str, Any]] = []
        for row in specimens:
            voucher = self._value(row, "入库编号*")
            if not voucher:
                continue
            class_row = class_by_voucher.get(voucher, {})
            record: dict[str, Any] = {}
            for col in SUMMARY_COLUMNS:
                category, excel_field = SUMMARY_COLUMN_SOURCE[col]
                if col == PHOTO_COUNT_COLUMN:
                    record[col] = photo_counts.get(voucher, 0)
                elif col == PHOTO_FILENAME_COLUMN:
                    record[col] = photo_filenames.get(voucher, [])
                elif col == PHOTO_PATH_COLUMN:
                    record[col] = photo_abs_paths.get(voucher, [])
                elif col == PHOTO_DESC_COLUMN:
                    record[col] = photo_descs.get(voucher, [])
                elif category == "classification":
                    record[col] = self._value(class_row, excel_field)
                else:  # specimen 列与主键"入库编号*"都取自标本行
                    record[col] = self._value(row, excel_field)
            # 旧逻辑：循环外单独 record["照片文件名"] = ...；现"照片文件名"已是 SUMMARY_COLUMN，
            # 由上面循环统一设置（键名不变，_apply_filters 仍按 record["照片文件名"] 取）。
            records.append(record)
        records.sort(key=lambda r: _voucher_sort_key(r["入库编号*"]))
        return records

    def voucher_photo_counts(self) -> dict[str, int]:
        """Return {voucher: photo_count} for all vouchers that have photos."""
        counts: dict[str, int] = {}
        for row in self.read_rows("photo"):
            v = self._value(row, "入库编号*")
            if v:
                counts[v] = counts.get(v, 0) + 1
        return counts

    def status_for(self, voucher: str) -> StatusFlags:
        specimen = self.get_specimen(voucher) or {}
        classification = self.get_classification(voucher) or {}
        photos = self.get_photos(voucher)
        return StatusFlags(
            specimen_complete=all(self._value(specimen, field) for field in SPECIMEN_REQUIRED),
            has_photo=bool(photos),
            classification_complete=all(self._value(classification, field) for field in CLASSIFICATION_REQUIRED),
        )

    def all_status_flags(self) -> dict[str, StatusFlags]:
        specimens = self.read_rows("specimen")
        classifications = self.read_rows("classification")
        photos = self.read_rows("photo")
        class_by_voucher: dict[str, Row] = {}
        for row in classifications:
            v = self._value(row, "入库编号*")
            if v:
                class_by_voucher[v] = row
        photo_vouchers = {self._value(row, "入库编号*") for row in photos if self._value(row, "入库编号*")}
        result: dict[str, StatusFlags] = {}
        for row in specimens:
            v = self._value(row, "入库编号*")
            if not v:
                continue
            class_row = class_by_voucher.get(v, {})
            result[v] = StatusFlags(
                specimen_complete=all(self._value(row, f) for f in SPECIMEN_REQUIRED),
                has_photo=v in photo_vouchers,
                classification_complete=bool(class_row) and all(self._value(class_row, f) for f in CLASSIFICATION_REQUIRED),
            )
        return result

    def get_specimen(self, voucher: str) -> Row | None:
        """返回该入库编号的标本信息行（dict）；不存在返回 None。"""
        return self._find_one("specimen", voucher)

    def get_classification(self, voucher: str) -> Row | None:
        """返回该入库编号的分类信息行（dict）；不存在返回 None。"""
        return self._find_one("classification", voucher)

    def get_photos(self, voucher: str) -> list[Row]:
        """返回该入库编号关联的全部照片信息行（一对多，可能为空 list）。"""
        return [row for row in self.read_rows("photo") if self._value(row, "入库编号*") == voucher]

    def get_all_photo_voucher_map(self) -> dict[str, list[str]]:
        """Return mapping from resolved photo path to list of voucher numbers.

        Used by the image search dialog to show which voucher(s) an
        already-linked photo belongs to.
        """
        result: dict[str, list[str]] = {}
        for row in self.read_rows("photo"):
            voucher = self._value(row, "入库编号*")
            if not voucher:
                continue
            resolved = str(self.resolve_photo_path(row))
            result.setdefault(resolved, []).append(voucher)
        return result

    def create_specimen(self) -> str:
        voucher = self.next_voucher()
        now = self._now()
        row = {header: "" for header in SPECIMEN_HEADERS}
        row["入库编号*"] = voucher
        row["入库日期"] = datetime.now().date().isoformat()
        self._append_row("specimen", row)
        self._append_index(voucher, now, "", "", self.record_fingerprint(voucher, specimen_override=row))
        self._ensure_summary_row(voucher, created_at=now)
        self._record_action("create_specimen", voucher, "specimen", "", {}, row)
        active = self.config.get("active_series_name", "YZZ")
        if active == "YZZ":
            # 原逻辑：按 parse_voucher_serial 推进 next_serial
            self.config["next_serial"] = max(int(self.config.get("next_serial", 1)), (parse_voucher_serial(voucher) or 0) + 1)
        else:
            # 非 YZZ 系列：推进该系列的 next_counter
            self._advance_series_counter(active)
        self._save_config()
        return voucher

    def create_specimen_with_voucher(self, voucher: str) -> str:
        """规范化软件设计 2026-05 Phase 5:手动指定 voucher 创建 specimen。

        跳过 next_serial 自增,直接用 voucher 字串。校验:
        - 不能为空 / 全空格
        - 不能与已存在 voucher 重复 (检 编号索引)

        与 create_specimen 一致流程:写 specimen / index / summary / action_log。
        不更新 next_serial (手动添加视为外部预留,不参与自增体系)。

        返回 voucher 字串。重复时 raise DuplicateVoucherError。
        """
        voucher = (voucher or "").strip()
        if not voucher:
            raise ValueError("voucher 不能为空")
        # 重复检测:走编号索引 (跟现有 next_voucher 重复保护一致)。注意 INDEX 表用 "入库编号" 列名(无 *)。
        index_rows = self._read_plain_rows(self.data_dir / INDEX_FILE, INDEX_HEADERS)
        for row in index_rows:
            if self._value(row, "入库编号") == voucher:
                raise DuplicateVoucherError(f"入库编号 {voucher} 已存在")
        now = self._now()
        row = {header: "" for header in SPECIMEN_HEADERS}
        row["入库编号*"] = voucher
        row["入库日期"] = datetime.now().date().isoformat()
        self._append_row("specimen", row)
        self._append_index(voucher, now, "", "", self.record_fingerprint(voucher, specimen_override=row))
        self._ensure_summary_row(voucher, created_at=now)
        self._record_action("create_specimen_manual", voucher, "specimen", "", {}, row)
        # 不推 next_serial — 用户既然手填,自增体系由他自管
        return voucher

    def delete_specimen(self, voucher: str) -> None:
        specimen = self.get_specimen(voucher)
        if not specimen:
            return
        old = {
            "specimen": specimen,
            "classification": self.get_classification(voucher),
            "photos": self.get_photos(voucher),
            "index": self._find_index(voucher),
        }
        remaining_photos = [row for row in self.read_rows("photo") if self._value(row, "入库编号*") != voucher]
        self._delete_rows("specimen", voucher)
        self._delete_rows("classification", voucher)
        self._write_rows("photo", remaining_photos)
        for photo in old["photos"]:
            self._delete_unreferenced_photo_file(photo, remaining_photos)
        self._delete_index(voucher)
        self._record_action("delete_specimen", voucher, "specimen", "", old, {})

    def clear_photos(self, voucher: str) -> int:
        photos = self.get_photos(voucher)
        if not photos:
            return 0
        self._record_action("clear_photos", voucher, "photo", "", {"photos": photos}, {})
        remaining_photos = [row for row in self.read_rows("photo") if self._value(row, "入库编号*") != voucher]
        self._write_rows("photo", remaining_photos)
        for photo in photos:
            self._delete_unreferenced_photo_file(photo, remaining_photos)
        return len(photos)

    def set_fields(
        self,
        category: str,
        voucher: str,
        updates: dict[str, Any],
        action_type: str = "update_fields",
        auto_derive_specimen_fields: bool = True,
    ) -> bool:
        """更新某入库编号在 specimen / classification 表中的若干字段。

        Args:
            category: ``"specimen"`` 或 ``"classification"``。
            voucher: 入库编号；该编号在目标表中不存在时会新建一行。
            updates: ``{字段名: 新值}``；不在该表表头里的键会被忽略。
            action_type: 写入操作日志的类型标签（用于撤销/重做）。
            auto_derive_specimen_fields: 为 True 且更新了 ``管内编号*`` 时，
                自动联动推导 ``采集日期`` / ``采集地点缩写*`` / ``保存方式``。

        Returns:
            是否有字段真正发生变化（无变化返回 False，不写日志）。

        变更会自动写入修改记录并追加可撤销的操作日志条目。
        """
        if category not in ("specimen", "classification"):
            raise ValueError(f"Unsupported category: {category}")
        headers = CATEGORY_HEADERS[category]
        updates = {field: self._string(value) for field, value in updates.items() if field in headers}
        if not updates:
            return False

        rows = self.read_rows(category)
        index = self._row_index(rows, voucher)
        if index is None:
            new_row = {header: "" for header in headers}
            new_row["入库编号*"] = voucher
            rows.append(new_row)
            index = len(rows) - 1
        old_row = rows[index].copy()
        changed = {field: value for field, value in updates.items() if self._value(old_row, field) != value}
        if not changed:
            return False

        rows[index].update(changed)
        if auto_derive_specimen_fields and category == "specimen" and "管内编号*" in changed:
            tube = rows[index].get("管内编号*", "")
            # 原代码只自动派生“采集日期”和“采集地点缩写*”；旧版本还支持保存方式。
            # 现在统一走管内编号派生函数，恢复保存方式，同时保留原字段兼容。
            auto_updates = derive_specimen_fields_from_tube_number(tube)
            for field, value in auto_updates.items():
                if value and rows[index].get(field) != value:
                    rows[index][field] = value
                    changed[field] = value

        self._write_rows(category, rows)
        new_row = rows[index].copy()
        self._write_changes_and_summary(voucher, category, old_row, new_row, action_type)
        self._update_index_fingerprint(voucher)
        self._record_action(action_type, voucher, category, "", old_row, new_row)
        return True

    def add_photo(
        self,
        voucher: str,
        photo_path: Path | str,
        allow_outside: bool = False,
        photo_management_mode: str = "copy_with_absolute",
        photo_library_path: Path | str | None = None,
    ) -> Row:
        row = self._photo_row(
            voucher,
            photo_path,
            allow_outside=allow_outside,
            photo_management_mode=photo_management_mode,
            photo_library_path=photo_library_path,
        )
        self._append_row("photo", row)
        self._update_summary_modified(voucher)
        self._record_action("add_photo", voucher, "photo", "", {}, row)
        return row

    def add_photos(
        self,
        voucher: str,
        photo_paths: list[Path | str],
        allow_outside: bool = False,
        photo_management_mode: str = "copy_with_absolute",
        photo_library_path: Path | str | None = None,
    ) -> list[Row]:
        rows_to_add = [
            self._photo_row(
                voucher,
                path,
                allow_outside=allow_outside,
                photo_management_mode=photo_management_mode,
                photo_library_path=photo_library_path,
            )
            for path in photo_paths
        ]
        if not rows_to_add:
            return []
        rows = self.read_rows("photo")
        rows.extend(rows_to_add)
        self._write_rows("photo", rows)
        self._update_summary_modified(voucher)
        self._record_action("add_photos", voucher, "photo", "", {}, rows_to_add)
        return rows_to_add

    def find_photo_conflicts(self, photo_paths: list[Path | str], target_voucher: str) -> dict[str, str]:
        resolved_inputs = {Path(p).resolve() for p in photo_paths}
        if not resolved_inputs:
            return {}
        input_hashes: dict[str, str] = {}
        for path in resolved_inputs:
            try:
                input_hashes[str(path)] = self._file_sha256(path)
            except OSError:
                continue
        conflicts: dict[str, str] = {}
        for row in self.read_rows("photo"):
            voucher = self._value(row, "入库编号*")
            if voucher == target_voucher:
                continue
            row_hash = self._value(row, "文件SHA256")
            if row_hash:
                for input_path, input_hash in input_hashes.items():
                    if input_hash == row_hash:
                        conflicts[input_path] = voucher
                continue
            row_path = self.resolve_photo_path(row)
            if row_path:
                resolved = Path(row_path).resolve()
                if resolved in resolved_inputs:
                    conflicts[str(resolved)] = voucher
        return conflicts

    def find_archive_name_conflicts(self, photo_paths: list[Path | str]) -> dict[str, str]:
        """Return input photos whose original filename collides with a different archived file."""
        archive_dir = self._photo_archive_dir()
        conflicts: dict[str, str] = {}
        for raw_path in photo_paths:
            path = Path(raw_path).resolve()
            if not path.is_file():
                continue
            target = archive_dir / self._safe_photo_filename(path.name)
            if not target.exists():
                continue
            try:
                if target.resolve() == path:
                    continue
                if self._file_sha256(target) == self._file_sha256(path):
                    continue
            except OSError:
                continue
            conflicts[str(path)] = str(target.resolve())
        return conflicts

    def export_all_data(self, target: Path) -> int:
        _ensure_openpyxl()
        wb = Workbook()
        wb.remove(wb.active)
        count = 0
        for category in CATEGORY_FILES:
            rows = self.read_rows(category)
            if not rows:
                continue
            ws = wb.create_sheet(title=category)
            headers = CATEGORY_HEADERS.get(category, [])
            ws.append(headers)
            for row in rows:
                ws.append([row.get(h, "") for h in headers])
            count += len(rows)
        wb.save(str(target))
        return count

    def import_from_file(self, source: Path) -> ImportResult:
        source_rows = self._read_external_rows(source, SPECIMEN_HEADERS)
        if not source_rows:
            return ImportResult(imported=0, skipped=0, photos_imported=0)
        source_ids = [self._value(row, "入库编号*") for row in source_rows if self._value(row, "入库编号*")]
        duplicate_source = [voucher for voucher, count in Counter(source_ids).items() if count > 1]
        if duplicate_source:
            report = self._write_conflict_report(
                [{"入库编号": voucher, "冲突类型": "导入文件内部重复", "源记录摘要": "", "目标记录摘要": ""} for voucher in duplicate_source]
            )
            raise ImportConflictError("导入文件存在重复入库编号，导入已阻止。", report)
        existing = set(self.list_vouchers())
        imported_ids: list[str] = []
        skipped = 0
        for row in source_rows:
            voucher = self._value(row, "入库编号*")
            if not voucher:
                continue
            if voucher in existing:
                skipped += 1
                continue
            imported_ids.append(voucher)
        if not imported_ids:
            return ImportResult(imported=0, skipped=skipped, photos_imported=0)
        self.create_data_snapshot("导入前快照", f"导入数据文件前自动快照：{source}")
        target_specimens = self.read_rows("specimen")
        target_classes = self.read_rows("classification")
        target_photos = self.read_rows("photo")
        target_index = self._read_plain_rows(self.data_dir / INDEX_FILE)
        now = self._now()
        import_set = set(imported_ids)
        for row in source_rows:
            voucher = self._value(row, "入库编号*")
            if voucher in import_set:
                target_specimens.append(self._fit_headers(row, SPECIMEN_HEADERS))
                target_index.append(
                    {
                        "入库编号": voucher,
                        "record_id": str(uuid.uuid4()),
                        "创建时间": now,
                        "来源工作区": str(source),
                        "来源记录ID": "",
                        # 原代码：self._fingerprint_from_rows(row)
                        # _fingerprint_from_rows 需要同时接收标本和分类两部分；单文件导入没有分类表时传 None。
                        "记录指纹": self._fingerprint_from_rows(row, None),
                    }
                )
                self._ensure_summary_row(voucher, created_at=now)
        self._write_rows("specimen", target_specimens)
        self._write_rows("classification", target_classes)
        self._write_rows("photo", target_photos)
        self._write_plain_rows(self.data_dir / INDEX_FILE, INDEX_HEADERS, target_index)
        self._sync_next_serial()
        self._record_action("import_file", "", "workspace", "", {}, {"source": str(source), "imported": imported_ids})
        self._record_data_version("导入数据文件", f"来源：{source}；导入 {len(imported_ids)} 个标本")
        return ImportResult(imported=len(imported_ids), skipped=skipped, photos_imported=0)

    def _photo_row(
        self,
        voucher: str,
        photo_path: Path | str,
        allow_outside: bool = False,
        source_row: Row | None = None,
        photo_management_mode: str = "copy_with_absolute",
        photo_library_path: Path | str | None = None,
    ) -> Row:
        original_name = self._value(source_row, "原始文件名") or self._value(source_row, "文件名") or Path(photo_path).name
        mode = photo_management_mode if photo_management_mode in PHOTO_MANAGEMENT_OPTIONS else "copy_with_absolute"
        if source_row is not None:
            # 导入其他工作区时沿用旧逻辑：复制为当前工作区副本，保证目标工作区可独立使用。
            mode = "copy_with_absolute"
        if mode == "absolute_only":
            return self._absolute_photo_row(voucher, Path(photo_path), original_name, source_row)
        archive_dir = Path(photo_library_path).expanduser() if mode == "copy_to_custom_library" and photo_library_path else None
        archived = self._archive_photo_file(Path(photo_path), original_name=original_name, archive_dir=archive_dir)
        return {
            "入库编号*": voucher,
            "文件名": archived["file_name"],
            "相对路径": archived["relative_path"],
            "绝对路径": archived["path"],
            "描述": self._value(source_row, "描述"),
            "来源工作区根路径": "",
            "原始文件名": archived["original_name"],
            "原始路径": archived["source_path"],
            "文件SHA256": archived["sha256"],
            "文件大小": archived["size"],
            "归档时间": archived["archived_at"],
            "归档状态": "已归档",
        }

    def _absolute_photo_row(self, voucher: str, source: Path, original_name: str, source_row: Row | None = None) -> Row:
        source = source.resolve()
        if not source.is_file():
            raise FileNotFoundError(f"照片文件不存在：{source}")
        return {
            "入库编号*": voucher,
            "文件名": source.name,
            "相对路径": "",
            "绝对路径": str(source),
            "描述": self._value(source_row, "描述"),
            "来源工作区根路径": "",
            "原始文件名": Path(original_name or source.name).name,
            "原始路径": str(source),
            "文件SHA256": self._file_sha256(source),
            "文件大小": str(source.stat().st_size),
            "归档时间": "",
            "归档状态": "仅记录",
        }

    def delete_photo(self, voucher: str, photo_index: int) -> bool:
        rows = self.read_rows("photo")
        matching_positions = [i for i, row in enumerate(rows) if self._value(row, "入库编号*") == voucher]
        if photo_index < 0 or photo_index >= len(matching_positions):
            return False
        position = matching_positions[photo_index]
        old_row = rows.pop(position)
        self._write_rows("photo", rows)
        self._delete_unreferenced_photo_file(old_row, rows)
        self._update_summary_modified(voucher)
        self._record_action("delete_photo", voucher, "photo", "", old_row, {})
        return True

    def set_photo_filename(self, voucher: str, photo_index: int, filename: str) -> bool:
        return self._rename_photo_file(voucher, photo_index, filename)

    def set_photo_description(self, voucher: str, photo_index: int, description: str) -> bool:
        return self._set_photo_text_field(voucher, photo_index, "描述", description)

    def _set_photo_text_field(self, voucher: str, photo_index: int, field: str, value: str) -> bool:
        if field not in {"文件名", "描述"}:
            raise ValueError(f"不支持修改照片字段：{field}")
        rows = self.read_rows("photo")
        matching_positions = [i for i, row in enumerate(rows) if self._value(row, "入库编号*") == voucher]
        if photo_index < 0 or photo_index >= len(matching_positions):
            return False
        position = matching_positions[photo_index]
        old_row = rows[position].copy()
        rows[position][field] = self._string(value)
        if old_row == rows[position]:
            return False
        self._write_rows("photo", rows)
        # 原代码只有“描述”可保存；现在“文件名”和“描述”统一走修改明细 + 汇总写入。
        self._write_changes_and_summary(voucher, "photo", old_row, rows[position], "update_photo")
        self._record_action("update_photo", voucher, "photo", field, old_row, rows[position].copy())
        return True

    def _rename_photo_file(self, voucher: str, photo_index: int, filename: str) -> bool:
        rows = self.read_rows("photo")
        matching_positions = [i for i, row in enumerate(rows) if self._value(row, "入库编号*") == voucher]
        if photo_index < 0 or photo_index >= len(matching_positions):
            return False
        position = matching_positions[photo_index]
        old_row = rows[position].copy()
        if self._value(old_row, "归档状态") == "仅记录":
            return self._set_photo_text_field(voucher, photo_index, "文件名", filename)
        old_path = self.resolve_photo_path(old_row)
        if not old_path.exists():
            raise FileNotFoundError(f"照片文件不存在：{old_path}")
        archive_dir = old_path.parent if not self._is_workspace_archive_path(old_path) else None
        target = self._move_archive_file_to_name(old_path, filename, archive_dir=archive_dir)
        new_row = old_row.copy()
        new_row["文件名"] = target.name
        new_row["相对路径"] = self._archive_relative_path(target) if self._is_under_root(target, self.root) else ""
        new_row["绝对路径"] = str(target.resolve())
        new_row["来源工作区根路径"] = ""
        if old_row == new_row:
            return False
        rows[position] = self._fit_headers(new_row, PHOTO_HEADERS)
        self._write_rows("photo", rows)
        self._delete_unreferenced_photo_file(old_row, rows)
        self._write_changes_and_summary(voucher, "photo", old_row, rows[position], "update_photo")
        self._record_action("update_photo", voucher, "photo", "文件名", old_row, rows[position].copy())
        return True

    def replace_photo(
        self,
        voucher: str,
        photo_index: int,
        photo_path: Path | str,
        allow_outside: bool = False,
        photo_management_mode: str = "copy_with_absolute",
        photo_library_path: Path | str | None = None,
    ) -> Row | None:
        rows = self.read_rows("photo")
        matching_positions = [i for i, row in enumerate(rows) if self._value(row, "入库编号*") == voucher]
        if photo_index < 0 or photo_index >= len(matching_positions):
            return None
        position = matching_positions[photo_index]
        new_row = self._photo_row(
            voucher,
            photo_path,
            allow_outside=allow_outside,
            photo_management_mode=photo_management_mode,
            photo_library_path=photo_library_path,
        )
        old_row = rows[position].copy()
        if old_row == new_row:
            return new_row
        rows[position] = new_row
        self._write_rows("photo", rows)
        self._delete_unreferenced_photo_file(old_row, rows)
        self._write_changes_and_summary(voucher, "photo", old_row, new_row, "update_photo")
        # 原代码：self._update_summary_modified(voucher)
        # _write_changes_and_summary 已更新修改汇总，避免重复计数。
        self._record_action("update_photo", voucher, "photo", "", old_row, new_row.copy())
        return new_row

    def move_photos(self, source_voucher: str, target_voucher: str, photo_indices: list[int] | set[int] | tuple[int, ...]) -> int:
        if not source_voucher or not target_voucher or source_voucher == target_voucher:
            return 0
        requested_set: set[int] = set()
        for index in photo_indices:
            try:
                parsed = int(index)
            except (TypeError, ValueError):
                continue
            if parsed >= 0:
                requested_set.add(parsed)
        requested = sorted(requested_set)
        if not requested:
            return 0
        rows = self.read_rows("photo")
        source_positions = [i for i, row in enumerate(rows) if self._value(row, "入库编号*") == source_voucher]
        selected_positions = [source_positions[index] for index in requested if index < len(source_positions)]
        if not selected_positions:
            return 0
        old_rows = [rows[position].copy() for position in selected_positions]
        moved_rows: list[Row] = []
        for row in old_rows:
            moved = row.copy()
            moved["入库编号*"] = target_voucher
            moved_rows.append(self._fit_headers(moved, PHOTO_HEADERS))
        selected_set = set(selected_positions)
        new_rows = [row for index, row in enumerate(rows) if index not in selected_set]
        new_rows.extend(moved_rows)
        self._write_rows("photo", new_rows)
        self._update_summary_modified(source_voucher)
        self._update_summary_modified(target_voucher)
        # 原代码在 UI 中逐张 add_photo/delete_photo；任何一步失败都会留下半完成状态。
        self._record_action(
            "move_photos",
            "",
            "photo",
            "",
            {"source": source_voucher, "photos": old_rows},
            {"target": target_voucher, "photos": moved_rows},
        )
        return len(moved_rows)

    def next_voucher(self) -> str:
        self.assert_unique_vouchers()
        active = self.config.get("active_series_name", "YZZ")
        if active == "YZZ":
            # 旧：max(existing+1, 1)，不考虑批量预留
            # 新：若曾批量预留，reserved_through_serial 记录上次预留的最末编号；
            # 下一个创建的编号必须在预留段之后，避免与已打印标签冲突。
            reserved = int(self.config.get("reserved_through_serial", 0))
            return format_voucher(max(self._max_existing_serial() + 1, reserved + 1))
        series = self._get_series_config(active)
        if series is None:
            return format_voucher(max(self._max_existing_serial() + 1, 1))
        return format_series_number(series)

    def assert_unique_vouchers(self) -> None:
        duplicate_messages: list[str] = []
        for category in ("specimen", "classification"):
            ids = [self._value(row, "入库编号*") for row in self.read_rows(category) if self._value(row, "入库编号*")]
            duplicates = [voucher for voucher, count in Counter(ids).items() if count > 1]
            if duplicates:
                duplicate_messages.append(f"{DISPLAY_CATEGORY_NAMES[category]} 重复: {', '.join(duplicates)}")

        index_ids = [self._value(row, "入库编号") for row in self._read_plain_rows(self.data_dir / INDEX_FILE) if self._value(row, "入库编号")]
        duplicates = [voucher for voucher, count in Counter(index_ids).items() if count > 1]
        if duplicates:
            duplicate_messages.append(f"编号索引重复: {', '.join(duplicates)}")

        if duplicate_messages:
            raise ImportConflictError("发现重复入库编号，已阻止继续写入。\n" + "\n".join(duplicate_messages))

    def import_workspace(
        self,
        source_root: Path | str,
        photo_duplicate_policy: str = "import",
    ) -> ImportResult:
        """合并源工作区到当前工作区。

        photo_duplicate_policy: M4 跨 voucher 同 SHA256 照片审核策略
        - "import"（默认，向后兼容）：源 photo 行原样写入，即使中心已有同 SHA256
          但属于不同 voucher 的记录（物理文件层 SHA256 去重仍生效，1 份文件被多条记录引用）。
          手动「导入工作区」菜单走此路径。
        - "skip"：源 photo 行被静默 skip，不写入；ImportResult.duplicate_candidates 留空。
        - "report"：与 "skip" 一致地不写入，但把疑似重复入库的照片记录到 duplicate_candidates，
          调用方可写报告供主管审核。aggregate_incoming 默认走此路径。
        """
        source = Path(source_root).resolve()
        source_data = source / "数据"
        if not source_data.exists():
            source_data = source
        source_specimens = self._read_external_rows(source_data / SPECIMEN_FILE, SPECIMEN_HEADERS)
        source_classes = self._read_external_rows(source_data / CLASSIFICATION_FILE, CLASSIFICATION_HEADERS)
        source_photos = self._read_external_rows(source_data / PHOTO_FILE, PHOTO_HEADERS)

        source_ids = [self._value(row, "入库编号*") for row in source_specimens if self._value(row, "入库编号*")]
        duplicate_source = [voucher for voucher, count in Counter(source_ids).items() if count > 1]
        if duplicate_source:
            report = self._write_conflict_report(
                [{"入库编号": voucher, "冲突类型": "源工作区内部重复", "源记录摘要": "", "目标记录摘要": ""} for voucher in duplicate_source]
            )
            raise ImportConflictError("源工作区存在重复入库编号，导入已阻止。", report)

        target_fingerprints = {voucher: self.record_fingerprint(voucher) for voucher in self.list_vouchers()}
        source_classes_by_id = {self._value(row, "入库编号*"): row for row in source_classes if self._value(row, "入库编号*")}
        conflicts: list[dict[str, str]] = []
        skipped = 0
        import_ids: list[str] = []

        for row in source_specimens:
            voucher = self._value(row, "入库编号*")
            if not voucher:
                continue
            source_fp = self._fingerprint_from_rows(row, source_classes_by_id.get(voucher))
            if voucher in target_fingerprints:
                if source_fp == target_fingerprints[voucher]:
                    skipped += 1
                    continue
                conflicts.append(
                    {
                        "入库编号": voucher,
                        "冲突类型": "目标工作区已有不同标本",
                        "源记录摘要": self._record_summary(row, source_classes_by_id.get(voucher)),
                        "目标记录摘要": self._record_summary(self.get_specimen(voucher), self.get_classification(voucher)),
                    }
                )
            else:
                import_ids.append(voucher)

        if conflicts:
            report = self._write_conflict_report(conflicts)
            raise ImportConflictError("发现入库编号冲突，导入已阻止。", report)

        if not import_ids:
            return ImportResult(imported=0, skipped=skipped, photos_imported=0)

        self.create_data_snapshot("导入前快照", f"导入工作区前自动快照：{source}")
        target_specimens = self.read_rows("specimen")
        target_classes = self.read_rows("classification")
        target_photos = self.read_rows("photo")
        target_index = self._read_plain_rows(self.data_dir / INDEX_FILE)
        source_classes_by_id = {self._value(row, "入库编号*"): row for row in source_classes if self._value(row, "入库编号*")}
        import_id_set = set(import_ids)
        now = self._now()

        for row in source_specimens:
            voucher = self._value(row, "入库编号*")
            if voucher in import_id_set:
                target_specimens.append(self._fit_headers(row, SPECIMEN_HEADERS))
                class_row = source_classes_by_id.get(voucher)
                if class_row:
                    target_classes.append(self._fit_headers(class_row, CLASSIFICATION_HEADERS))
                target_index.append(
                    {
                        "入库编号": voucher,
                        "record_id": str(uuid.uuid4()),
                        "创建时间": now,
                        "来源工作区": str(source),
                        "来源记录ID": "",
                        "记录指纹": self._fingerprint_from_rows(row, class_row),
                    }
                )
                self._ensure_summary_row(voucher, created_at=now)

        # M4：扫中心机已有照片 SHA256 → voucher 映射，用于跨 voucher 同 SHA256 检测
        # 仅在 photo_duplicate_policy != "import" 时才构建（性能优化）
        existing_sha_to_voucher: dict[str, str] = {}
        if photo_duplicate_policy != "import":
            for existing_photo in target_photos:
                sha = str(self._value(existing_photo, "文件SHA256") or "").lower()
                vch = str(self._value(existing_photo, "入库编号*") or "")
                if sha and vch and sha not in existing_sha_to_voucher:
                    existing_sha_to_voucher[sha] = vch

        photos_imported = 0
        missing_photos: list[dict[str, str]] = []
        duplicate_candidates: list[dict] = []
        for photo in source_photos:
            voucher = self._value(photo, "入库编号*")
            if voucher in import_id_set:
                source_path = self._resolve_import_photo_path(photo, source)
                if not source_path.exists():
                    missing_photos.append(
                        {
                            "入库编号": voucher,
                            "文件名": self._value(photo, "文件名"),
                            "相对路径": self._value(photo, "相对路径"),
                            "来源工作区根路径": self._value(photo, "来源工作区根路径"),
                            "解析路径": str(source_path),
                        }
                    )
                    continue
                # M4：跨 voucher 同 SHA256 检测
                if photo_duplicate_policy != "import":
                    photo_sha = str(self._value(photo, "文件SHA256") or "").lower()
                    if not photo_sha:
                        # 源 photo 行没存 SHA256（旧版数据 / 外部导入），即时算
                        try:
                            photo_sha = self._file_sha256(source_path).lower()
                        except OSError:
                            photo_sha = ""
                    if photo_sha and photo_sha in existing_sha_to_voucher:
                        existing_voucher = existing_sha_to_voucher[photo_sha]
                        if existing_voucher != voucher:
                            # 命中"潜在重复入库"：同照片已被关联到其它 voucher
                            if photo_duplicate_policy == "report":
                                duplicate_candidates.append(
                                    {
                                        "入库编号": voucher,
                                        "已有voucher": existing_voucher,
                                        "文件SHA256": photo_sha,
                                        "源相对路径": self._value(photo, "相对路径"),
                                        "源原始路径": self._value(photo, "原始路径"),
                                        "源解析路径": str(source_path),
                                    }
                                )
                            # skip / report 都不写入新 photo 行
                            continue
                fitted = self._photo_row(voucher, source_path, allow_outside=True, source_row=photo)
                target_photos.append(fitted)
                photos_imported += 1
                # 把刚导入的照片也加入查表，避免同源工作区内的二次重复
                if photo_duplicate_policy != "import":
                    new_sha = str(fitted.get("文件SHA256") or "").lower()
                    if new_sha and new_sha not in existing_sha_to_voucher:
                        existing_sha_to_voucher[new_sha] = voucher

        self._write_rows("specimen", target_specimens)
        self._write_rows("classification", target_classes)
        self._write_rows("photo", target_photos)
        self._write_plain_rows(self.data_dir / INDEX_FILE, INDEX_HEADERS, target_index)
        self._sync_next_serial()
        report_path = self._write_photo_missing_report(missing_photos) if missing_photos else None
        self._record_action("import_workspace", "", "workspace", "", {}, {"source": str(source), "imported": import_ids})
        summary = f"来源：{source}；导入 {len(import_ids)} 个标本，照片 {photos_imported} 张"
        if report_path:
            summary += f"；缺失照片 {len(missing_photos)} 张，报告：{report_path}"
        if duplicate_candidates:
            summary += f"；潜在重复入库照片 {len(duplicate_candidates)} 张（已记录待审核）"
        self._record_data_version("导入工作区", summary)
        return ImportResult(
            imported=len(import_ids),
            skipped=skipped,
            photos_imported=photos_imported,
            report_path=report_path,
            duplicate_candidates=duplicate_candidates,
        )

    def create_data_snapshot(self, operation_type: str = "手动快照", summary: str = "") -> Path:
        version_id = datetime.now().strftime("v%Y%m%d_%H%M%S")
        snapshot_dir = self.data_dir / DATA_VERSION_DIR / version_id
        suffix = 1
        while snapshot_dir.exists():
            snapshot_dir = self.data_dir / DATA_VERSION_DIR / f"{version_id}_{suffix}"
            suffix += 1
        snapshot_dir.mkdir(parents=True)
        for path in self.data_dir.iterdir():
            if path.name == DATA_VERSION_DIR or path.name == ".workspace.lock":
                continue
            if path.is_file() and path.suffix.lower() in {".xlsx", ".json"}:
                shutil.copy2(path, snapshot_dir / path.name)
        manifest = {
            "version_id": snapshot_dir.name,
            "created_at": self._now(),
            "operation_type": operation_type,
            "software_version": __version__,
            "data_schema_version": self.config.get("data_schema_version", CURRENT_DATA_SCHEMA_VERSION),
            "summary": summary,
            "workspace": str(self.root),
        }
        with (snapshot_dir / "snapshot_manifest.json").open("w", encoding="utf-8") as handle:
            json.dump(manifest, handle, ensure_ascii=False, indent=2)
        self._record_data_version(operation_type, summary or operation_type, snapshot_dir)
        return snapshot_dir

    def list_data_versions(self) -> list[Row]:
        rows = self._read_plain_rows(self.data_dir / DATA_VERSION_LOG_FILE, DATA_VERSION_LOG_HEADERS)
        return [row for row in rows if self._value(row, "快照路径")]

    def restore_data_snapshot(self, snapshot_path: Path | str) -> None:
        snapshot = Path(snapshot_path).resolve()
        expected_parent = (self.data_dir / DATA_VERSION_DIR).resolve()
        try:
            snapshot.relative_to(expected_parent)
        except ValueError:
            # 原代码：if not str(snapshot).startswith(str(expected_parent))
            # 避免 /path/data_versions_old 这类字符串前缀误判。
            raise ValueError(f"快照路径不在数据版本目录内：{snapshot}")
        if not snapshot.exists() or not snapshot.is_dir():
            raise FileNotFoundError(f"数据版本不存在：{snapshot}")
        self.create_data_snapshot("回退前快照", f"回退到 {snapshot.name} 前自动保存当前状态")
        for path in snapshot.iterdir():
            if not path.is_file():
                continue
            if path.name in {DATA_VERSION_LOG_FILE, "snapshot_manifest.json", ".workspace.lock"}:
                continue
            if path.suffix.lower() in {".xlsx", ".json"}:
                shutil.copy2(path, self.data_dir / path.name)
        self.config = self._load_or_create_config()
        self._record_data_version("回退数据版本", f"已恢复：{snapshot.name}", snapshot)
        self.ensure_files()
        self.ensure_index()
        self._sync_next_serial()

    def undo_last(self) -> str | None:
        # 规范化软件设计 2026-05 P1 审查修复:_apply_action 失败时不能把 action 标"已撤销",
        # 否则下次 undo 跳过它造成"幽灵 action"。包 try/except,异常时清晰传播给上层,
        # 不修改 action 状态。Excel 文件多步写入仍无真事务(已知限制),日后专项重构。
        rows = self._read_plain_rows(self.data_dir / ACTION_LOG_FILE)
        depth = int(self.config.get("undo_depth", 200))
        candidates = [row for row in rows[-depth:] if self._value(row, "是否撤销") != "是"]
        if not candidates:
            return None
        action = candidates[-1]
        try:
            self._apply_action(action, undo=True)
        except Exception:
            # apply 失败 → 不标记 + 重抛,让上层弹错误对话框。
            # 注意:数据可能部分被 undo(_apply_action 内多个写入步骤)。
            raise
        action["是否撤销"] = "是"
        self._write_plain_rows(self.data_dir / ACTION_LOG_FILE, ACTION_LOG_HEADERS, rows)
        return self._value(action, "操作类型")

    def redo_last(self) -> str | None:
        # 同 undo_last 的 try/except 保护。
        rows = self._read_plain_rows(self.data_dir / ACTION_LOG_FILE)
        if not any(self._value(row, "是否撤销") == "是" for row in rows):
            return None
        start = len(rows) - 1
        while start >= 0 and self._value(rows[start], "是否撤销") == "是":
            start -= 1
        action_index = start + 1
        if action_index >= len(rows):
            return None
        action = rows[action_index]
        try:
            self._apply_action(action, undo=False)
        except Exception:
            raise
        action["是否撤销"] = ""
        self._write_plain_rows(self.data_dir / ACTION_LOG_FILE, ACTION_LOG_HEADERS, rows)
        return self._value(action, "操作类型")

    def set_undo_depth(self, depth: int) -> None:
        self.config["undo_depth"] = max(1, min(int(depth), 1000))
        self._save_config()

    def resolve_photo_path(self, photo_row: Row) -> Path:
        relative = self._value(photo_row, "相对路径")
        candidates: list[Path] = []
        if relative:
            candidates.append(self._resolve_relative(self.root, relative))
        source_root = self._value(photo_row, "来源工作区根路径")
        if relative and source_root:
            src = Path(source_root)
            if src.is_dir():
                candidates.append(self._resolve_relative(src, relative))
        absolute = self._value(photo_row, "绝对路径")
        if absolute:
            candidates.append(Path(absolute).expanduser().resolve())
        original = self._value(photo_row, "原始路径")
        if original:
            candidates.append(Path(original).expanduser().resolve())
        for path in candidates:
            if path.exists():
                return path
        if candidates:
            return candidates[0]
        return self.root / self._value(photo_row, "文件名")

    def _resolve_import_photo_path(self, photo_row: Row, source_root: Path) -> Path:
        relative = self._value(photo_row, "相对路径")
        candidates = []
        if relative:
            candidates.append(self._resolve_relative(source_root, relative))
        source_photo_root = self._value(photo_row, "来源工作区根路径")
        if relative and source_photo_root:
            src = Path(source_photo_root)
            if src.is_dir():
                candidates.append(self._resolve_relative(src, relative))
        absolute_path = self._value(photo_row, "绝对路径")
        if absolute_path:
            candidates.append(Path(absolute_path).resolve())
        original_path = self._value(photo_row, "原始路径")
        if original_path:
            candidates.append(Path(original_path).resolve())
        for path in candidates:
            if path.exists():
                return path
        if candidates:
            return candidates[0]
        return source_root / self._value(photo_row, "文件名")

    def relative_photo_path(self, path: Path, allow_outside: bool = False) -> str:
        path = path.resolve()
        try:
            relative = path.relative_to(self.root)
            return "./" + relative.as_posix()
        except ValueError:
            if not allow_outside:
                raise ValueError("照片不在当前工作区内，无法生成稳定的工作区相对路径")
            return Path(os.path.relpath(path, self.root)).as_posix()

    def _photo_location(self, path: Path, allow_outside: bool = False) -> tuple[str, str]:
        try:
            relative = path.relative_to(self.root)
            return "./" + relative.as_posix(), ""
        except ValueError:
            if not allow_outside:
                raise ValueError("照片不在当前工作区内，无法生成稳定的工作区相对路径")
            # 原代码会把外部照片保存成 ../xxx；现在用来源根路径 + 文件名避免路径穿越。
            return "./" + path.name, str(path.parent)

    def _archive_photo_file(
        self,
        source: Path,
        original_name: str | None = None,
        archive_dir: Path | None = None,
    ) -> dict[str, str]:
        source = source.resolve()
        if not source.is_file():
            raise FileNotFoundError(f"照片文件不存在：{source}")
        digest = self._file_sha256(source)
        archive_dir = (archive_dir or self._photo_archive_dir()).resolve()
        archive_dir.mkdir(parents=True, exist_ok=True)
        clean_name = self._safe_photo_filename(original_name or source.name)
        target = self._archive_target_path(archive_dir, digest, clean_name)
        if not target.exists():
            tmp = archive_dir / f".{uuid.uuid4().hex}.tmp{source.suffix.lower()}"
            try:
                shutil.copy2(source, tmp)
                copied_digest = self._file_sha256(tmp)
                if copied_digest != digest:
                    raise OSError(f"照片复制校验失败：{source}")
                tmp.replace(target)
            finally:
                try:
                    tmp.unlink(missing_ok=True)
                except OSError:
                    pass
        return {
            "path": str(target),
            "file_name": target.name,
            "relative_path": self._archive_relative_path(target) if self._is_under_root(target, self.root) else "",
            "original_name": Path(original_name or source.name).name,
            "source_path": str(source),
            "sha256": digest,
            "size": str(source.stat().st_size),
            "archived_at": self._now(),
        }

    def _archive_target_path(self, archive_dir: Path, digest: str, clean_name: str) -> Path:
        return self._available_archive_target(archive_dir, clean_name, digest)

    def _photo_archive_dir(self) -> Path:
        return self.root / "照片"

    def _archive_relative_path(self, path: Path) -> str:
        return "./" + path.resolve().relative_to(self.root).as_posix()

    def _is_under_root(self, path: Path, root: Path) -> bool:
        try:
            path.resolve().relative_to(root.resolve())
            return True
        except ValueError:
            return False

    def _available_archive_target(self, archive_dir: Path, clean_name: str, digest: str | None = None) -> Path:
        target = archive_dir / clean_name
        if self._target_available_for_digest(target, digest):
            return target
        path = Path(clean_name)
        stem = path.stem or "photo"
        suffix = path.suffix
        # 规范化软件设计 2026-05 P1 审查修复:counter 加 100000 上限防 O(n) 性能悬崖。
        # 100000 个同名碰撞已是天文数字,触顶时回退到 uuid 后缀保证可继续。
        counter = 2
        _MAX_COUNTER = 100000
        while counter <= _MAX_COUNTER:
            candidate = archive_dir / f"{stem}_{counter}{suffix}"
            if self._target_available_for_digest(candidate, digest):
                return candidate
            counter += 1
        # 达上限 → uuid 兜底
        import uuid as _uuid
        return archive_dir / f"{stem}_{_uuid.uuid4().hex[:12]}{suffix}"

    def _target_available_for_digest(self, target: Path, digest: str | None) -> bool:
        if not target.exists():
            return True
        if digest:
            try:
                return self._file_sha256(target) == digest
            except OSError:
                return False
        return False

    def _move_archive_file_to_name(self, source: Path, filename: str, archive_dir: Path | None = None) -> Path:
        source = source.resolve()
        if not source.is_file():
            raise FileNotFoundError(f"照片文件不存在：{source}")
        digest = self._file_sha256(source)
        archive_dir = (archive_dir or self._photo_archive_dir()).resolve()
        archive_dir.mkdir(parents=True, exist_ok=True)
        default_suffix = source.suffix if source.suffix else ""
        clean_name = self._safe_photo_filename(filename, default_suffix=default_suffix)
        target = self._available_archive_target(archive_dir, clean_name, digest)
        if source == target.resolve():
            return target
        if target.exists():
            # 同内容同名文件已存在时复用目标，避免留下重复副本。
            return target
        if self._is_workspace_archive_path(source):
            source.replace(target)
        else:
            tmp = archive_dir / f".{uuid.uuid4().hex}.tmp{source.suffix.lower()}"
            try:
                shutil.copy2(source, tmp)
                if self._file_sha256(tmp) != digest:
                    raise OSError(f"照片复制校验失败：{source}")
                tmp.replace(target)
            finally:
                try:
                    tmp.unlink(missing_ok=True)
                except OSError:
                    pass
        return target

    def _delete_unreferenced_photo_file(self, photo_row: Row, remaining_rows: list[Row] | None = None) -> bool:
        try:
            path = self.resolve_photo_path(photo_row).resolve()
        except Exception:
            return False
        if not self._is_managed_photo_path(photo_row, path) or not path.exists():
            return False
        rows = remaining_rows if remaining_rows is not None else self.read_rows("photo")
        for row in rows:
            try:
                other = self.resolve_photo_path(row).resolve()
            except Exception:
                continue
            if other == path:
                return False
        return self._delete_archive_file_if_safe(path)

    def _delete_archive_file_if_safe(self, path: Path) -> bool:
        path = path.resolve()
        if not path.exists():
            return False
        try:
            path.unlink()
            return True
        except OSError:
            return False

    def _is_workspace_archive_path(self, path: Path) -> bool:
        try:
            path.resolve().relative_to(self._photo_archive_dir().resolve())
            return True
        except ValueError:
            return False

    def _is_managed_photo_path(self, photo_row: Row, path: Path) -> bool:
        if self._is_workspace_archive_path(path):
            return True
        if self._value(photo_row, "归档状态") != "已归档":
            return False
        original = self._value(photo_row, "原始路径")
        if original:
            try:
                if Path(original).resolve() == path.resolve():
                    return False
            except OSError:
                return False
        absolute = self._value(photo_row, "绝对路径")
        return bool(absolute) and Path(absolute).expanduser().resolve() == path.resolve()

    def _safe_photo_filename(self, filename: str, default_suffix: str = "") -> str:
        name = Path(filename or "photo").name.strip() or "photo"
        name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name)
        path = Path(name)
        suffix = path.suffix or default_suffix
        stem = path.stem or "photo"
        if len(stem) > 140:
            stem = stem[:140].rstrip(" ._") or "photo"
        return f"{stem}{suffix}"

    def _file_sha256(self, path: Path) -> str:
        h = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                h.update(chunk)
        return h.hexdigest()

    def _cached_rows(self, file_key: str, loader: Callable[[], list[Row]]) -> list[Row]:
        file_path = self.data_dir / file_key
        try:
            current_mtime = file_path.stat().st_mtime
        except OSError:
            current_mtime = 0.0
        cached_mtime = self._file_mtimes.get(file_key, -1.0)
        if file_key in self._row_cache and cached_mtime == current_mtime:
            # LRU: move_to_end 让命中项标为最近使用
            self._row_cache.move_to_end(file_key)
            return [row.copy() for row in self._row_cache[file_key]]
        rows = loader()
        self._row_cache[file_key] = rows
        self._row_cache.move_to_end(file_key)
        self._file_mtimes[file_key] = current_mtime
        # LRU 驱逐:超 maxsize 时弹最旧项(popitem(last=False))
        while len(self._row_cache) > self._row_cache_maxsize:
            evicted_key, _ = self._row_cache.popitem(last=False)
            self._file_mtimes.pop(evicted_key, None)
        return [row.copy() for row in rows]

    def _invalidate_cache(self, *file_keys: str) -> None:
        for key in file_keys:
            self._row_cache.pop(key, None)
            self._file_mtimes.pop(key, None)

    def _enforce_row_cache_size(self) -> None:
        """规范化软件设计 2026-05 内存档位:用户改小档位后立即驱逐多余项,缩内存到位。

        正常情况下 _cached_rows 内置 while 循环就会驱逐;但用户调小 maxsize 后
        到下次 _cached_rows 触发前内存不会立刻降,本 helper 供 SettingsDialog 保存路径
        手动 enforce 一次。
        """
        while len(self._row_cache) > self._row_cache_maxsize:
            evicted_key, _ = self._row_cache.popitem(last=False)
            self._file_mtimes.pop(evicted_key, None)

    def read_rows(self, category: str) -> list[Row]:
        """读取分类下的全部行（dense dict —— 缺失字段补 ""）。

        内部 _row_cache 持有 sparse dict（_read_plain_rows 只存非空字段），
        本方法出口处按 CATEGORY_HEADERS 把每行补成 dense，保证下游 `row["字段"]`
        直接索引不会 KeyError（向后兼容 v0.5.0 及以前的 dense 契约）。
        """
        file_key = CATEGORY_FILES[category]
        headers = CATEGORY_HEADERS[category]
        sparse_rows = self._cached_rows(
            file_key,
            lambda: self._read_plain_rows(self.data_dir / file_key, headers),
        )
        # _cached_rows 已 [row.copy()]，这里返回的 dense 是临时局部表，调用方用完即回收。
        return [{h: row.get(h, "") for h in headers} for row in sparse_rows]

    def record_fingerprint(
        self,
        voucher: str,
        specimen_override: Row | None = None,
        classification_override: Row | None = None,
    ) -> str:
        specimen = specimen_override if specimen_override is not None else self.get_specimen(voucher)
        classification = classification_override if classification_override is not None else self.get_classification(voucher)
        return self._fingerprint_from_rows(specimen, classification)

    def ensure_index(self) -> None:
        index_rows = self._read_plain_rows(self.data_dir / INDEX_FILE)
        indexed = {self._value(row, "入库编号") for row in index_rows if self._value(row, "入库编号")}
        specimens = self.read_rows("specimen")
        classifications = self.read_rows("classification")
        class_by_voucher = {
            self._value(row, "入库编号*"): row
            for row in classifications
            if self._value(row, "入库编号*")
        }
        now = self._now()
        changed = False
        for row in specimens:
            voucher = self._value(row, "入库编号*")
            if not voucher:
                continue
            if voucher not in indexed:
                index_rows.append(
                    {
                        "入库编号": voucher,
                        "record_id": str(uuid.uuid4()),
                        "创建时间": now,
                        "来源工作区": "",
                        "来源记录ID": "",
                        "记录指纹": self._fingerprint_from_rows(row, class_by_voucher.get(voucher)),
                    }
                )
                changed = True
        if changed:
            self._write_plain_rows(self.data_dir / INDEX_FILE, INDEX_HEADERS, index_rows)

    def _load_or_create_config(self) -> dict[str, Any]:
        path = self.data_dir / WORKSPACE_CONFIG_FILE
        # 旧实现：merged = {**DEFAULT_CONFIG, **data} / DEFAULT_CONFIG.copy()
        # 都是浅拷贝 → DEFAULT_CONFIG["accession_series"]（list）等可变值会被多个实例共享，
        # 一个实例 add_series 后，下次 ExcelStore() 启动看到的"默认"已被污染。
        # 改用 deepcopy 杜绝跨实例 mutable 共享。
        if path.exists():
            with path.open("r", encoding="utf-8") as handle:
                data = json.load(handle)
            merged = copy.deepcopy(DEFAULT_CONFIG)
            merged.update(data)
        else:
            if not self._create_if_missing:
                raise WorkspaceNotInitializedError(f"该工作目录尚未初始化，缺少配置文件：{path}")
            merged = copy.deepcopy(DEFAULT_CONFIG)
        if not merged.get("workspace_id"):
            merged["workspace_id"] = str(uuid.uuid4())
        if not merged.get("data_schema_version"):
            merged["data_schema_version"] = CURRENT_DATA_SCHEMA_VERSION
        self.config = merged
        self._save_config()
        return merged

    def _assert_supported_data_schema(self) -> None:
        current = str(self.config.get("data_schema_version", CURRENT_DATA_SCHEMA_VERSION))
        if _version_tuple(current) > _version_tuple(CURRENT_DATA_SCHEMA_VERSION):
            raise ImportConflictError(
                f"该工作区数据版本为 {current}，高于当前软件支持的 {CURRENT_DATA_SCHEMA_VERSION}，已禁止写入。\n\n"
                "请升级软件到最新版本后再打开；或先用新版软件的「工具 → 降低工作区兼容版本」"
                "将工作区版本降至 1.0.0，旧版软件即可重新打开。"
            )

    def downgrade_schema_version(self, target: str = "1.0.0") -> None:
        """将工作区兼容版本降至 target，以便旧版软件可以打开。

        旧：无此方法，用户用新版打开工作区后数据版本升至 1.1.x，旧软件因版本检查锁死。
        只修改 工作区配置.json 里的 data_schema_version，不回滚任何数据内容。
        下次用新版软件打开时，_upgrade_workspace_schema 会自动重新升级。
        """
        self.config["data_schema_version"] = target
        self._save_config()

    def _upgrade_workspace_schema(self) -> None:
        current = str(self.config.get("data_schema_version", "1.0.0"))
        if _version_tuple(current) < _version_tuple("1.1.1"):
            self._migrate_hash_prefixed_photos()
        if _version_tuple(current) < _version_tuple(CURRENT_DATA_SCHEMA_VERSION):
            self.config["data_schema_version"] = CURRENT_DATA_SCHEMA_VERSION
            self._save_config()

    def _migrate_hash_prefixed_photos(self) -> None:
        rows = self.read_rows("photo")
        if not rows:
            return
        changed = False
        moved_paths: dict[Path, Path] = {}
        old_paths: set[Path] = set()
        pattern = re.compile(r"^[0-9a-fA-F]{12}(?:_[0-9a-fA-F]{8})?__(.+)$")
        archive_dir = self._photo_archive_dir().resolve()
        for idx, row in enumerate(rows):
            relative = self._value(row, "相对路径")
            old_path = self._resolve_relative(self.root, relative).resolve()
            if old_path in moved_paths:
                target = moved_paths[old_path]
            else:
                try:
                    old_path.relative_to(archive_dir)
                except ValueError:
                    continue
                match = pattern.match(old_path.name)
                if not match:
                    continue
                if not old_path.exists():
                    continue
                desired = self._value(row, "原始文件名") or self._value(row, "文件名") or match.group(1)
                target = self._move_archive_file_to_name(old_path, desired)
                moved_paths[old_path] = target
                old_paths.add(old_path)
            new_row = row.copy()
            new_row["文件名"] = target.name
            new_row["相对路径"] = self._archive_relative_path(target)
            new_row["来源工作区根路径"] = ""
            fitted = self._fit_headers(new_row, PHOTO_HEADERS)
            if self._fit_headers(row, PHOTO_HEADERS) != fitted:
                rows[idx] = fitted
                changed = True
        if changed:
            self._write_rows("photo", rows)
        for old_path in old_paths:
            self._delete_archive_file_if_safe(old_path)

    def _save_config(self) -> None:
        # 原：直接写 path，崩溃/断电会留下截断的 JSON，下次打开工作区失败。
        # 现：写临时文件再原子替换，确保要么新版本完整、要么旧版本保留。
        path = self.data_dir / WORKSPACE_CONFIG_FILE
        tmp = path.with_suffix(f".{os.getpid()}.tmp")
        try:
            with tmp.open("w", encoding="utf-8") as handle:
                json.dump(self.config, handle, ensure_ascii=False, indent=2)
            tmp.replace(path)
        except Exception:
            try:
                tmp.unlink(missing_ok=True)
            except OSError:
                pass
            raise

    def _ensure_workbook(self, path: Path, headers: list[str]) -> None:
        if not path.exists():
            _ensure_openpyxl()
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(headers)
            # 原：wb.save(path) 直接写，崩溃留下残缺文件导致下次 path.exists() 为 True
            # 但内容损坏。现：用原子替换。
            tmp = path.with_suffix(f".{os.getpid()}.tmp")
            wb.save(tmp)
            tmp.replace(path)
            return
        rows = self._read_plain_rows(path)
        existing_headers = self._headers(path)
        missing = [header for header in headers if header not in existing_headers]
        if missing:
            self._write_plain_rows(path, existing_headers + missing, rows)

    def _ensure_change_log(self) -> None:
        path = self.data_dir / CHANGE_LOG_FILE
        if path.exists():
            return
        _ensure_openpyxl()
        wb = Workbook()
        ws = wb.active
        ws.title = "修改明细"
        ws.append(CHANGE_LOG_HEADERS)
        summary = wb.create_sheet("修改汇总")
        summary.append(CHANGE_SUMMARY_HEADERS)
        # 原：直接写，改用原子替换，与 _ensure_workbook 保持一致。
        tmp = path.with_suffix(f".{os.getpid()}.tmp")
        wb.save(tmp)
        tmp.replace(path)

    def _ensure_alloc_log(self) -> None:
        self._ensure_workbook(self.data_dir / ALLOC_LOG_FILE, ALLOC_LOG_HEADERS)

    # ── 编号分发日志 ──────────────────────────────────────────────────────────

    def batch_reserve_vouchers(self, n: int, series_name: str | None = None) -> list[str]:
        """预留 n 个连续编号（不创建标本记录），返回编号列表，并推进计数器。"""
        active = series_name or self.config.get("active_series_name", "YZZ")
        if active == "YZZ":
            reserved = int(self.config.get("reserved_through_serial", 0))
            start = max(self._max_existing_serial() + 1, reserved + 1)
            numbers = [format_voucher(start + i) for i in range(n)]
            self.config["reserved_through_serial"] = start + n - 1
        else:
            series = self._get_series_config(active)
            if series is None:
                raise ValueError(f"系列 {active!r} 未找到")
            numbers = []
            counter = series.next_counter
            for _ in range(n):
                numbers.append(format_series_number(series, counter))
                counter += series.step
            for item in self.config.get("accession_series", []):
                if item.get("name") == active:
                    item["next_counter"] = counter
        self._save_config()
        return numbers

    def log_alloc_event(self, record: dict) -> None:
        """追加一条分发记录（批量领取 / 任务开始 / 任务结束）。原子重写。"""
        rows = self._read_plain_rows(self.data_dir / ALLOC_LOG_FILE, ALLOC_LOG_HEADERS)
        row = {h: str(record.get(h, "")) for h in ALLOC_LOG_HEADERS}
        rows.append(row)
        self._write_plain_rows(self.data_dir / ALLOC_LOG_FILE, ALLOC_LOG_HEADERS, rows)

    def read_alloc_log(self) -> list[Row]:
        """读取全部分发记录。"""
        return self._read_plain_rows(self.data_dir / ALLOC_LOG_FILE, ALLOC_LOG_HEADERS)

    def _record_data_version(self, operation_type: str, summary: str, snapshot_path: Path | None = None) -> None:
        rows = self._read_plain_rows(self.data_dir / DATA_VERSION_LOG_FILE, DATA_VERSION_LOG_HEADERS)
        rows.append(
            {
                "版本ID": snapshot_path.name if snapshot_path else datetime.now().strftime("v%Y%m%d_%H%M%S"),
                "时间": self._now(),
                "操作类型": operation_type,
                "软件版本": __version__,
                "数据结构版本": self.config.get("data_schema_version", CURRENT_DATA_SCHEMA_VERSION),
                "操作者": os.environ.get("USERNAME") or os.environ.get("USER") or "",
                "摘要": summary,
                "快照路径": str(snapshot_path.resolve()) if snapshot_path else "",
            }
        )
        self._write_plain_rows(self.data_dir / DATA_VERSION_LOG_FILE, DATA_VERSION_LOG_HEADERS, rows)

    def _write_changes_and_summary(self, voucher: str, category: str, old_row: Row, new_row: Row, action_type: str) -> None:
        """Append field changes and update summary in a single file write."""
        now = self._now()
        detail_rows = self._read_change_detail_rows()
        summary_rows = self._read_summary_rows()
        for field in CATEGORY_HEADERS[category]:
            old = self._value(old_row, field)
            new = self._value(new_row, field)
            if old != new:
                detail_rows.append(
                    {
                        "入库编号": voucher,
                        "信息类别": DISPLAY_CATEGORY_NAMES[category],
                        "字段名": field,
                        "旧值": old,
                        "新值": new,
                        "修改时间": now,
                        "操作类型": action_type,
                    }
                )
        if not any(self._value(row, "入库编号") == voucher for row in summary_rows):
            summary_rows.append(
                {
                    "入库编号": voucher,
                    "创建时间": now,
                    "第一次修改时间": "",
                    "第二次修改时间": "",
                    "最近修改时间": "",
                    "修改次数": 0,
                }
            )
        for row in summary_rows:
            if self._value(row, "入库编号") == voucher:
                count = int(row.get("修改次数") or 0) + 1
                row["修改次数"] = count
                if count == 1:
                    row["第一次修改时间"] = now
                elif count == 2:
                    row["第二次修改时间"] = now
                row["最近修改时间"] = now
                break
        path = self.data_dir / CHANGE_LOG_FILE
        self._ensure_change_log()
        with self._open_workbook(path) as wb:
            if "修改明细" not in wb.sheetnames:
                wb.create_sheet("修改明细")
            self._replace_sheet(wb["修改明细"], CHANGE_LOG_HEADERS, detail_rows)
            if "修改汇总" not in wb.sheetnames:
                wb.create_sheet("修改汇总")
            self._replace_sheet(wb["修改汇总"], CHANGE_SUMMARY_HEADERS, summary_rows)
            tmp = path.with_suffix(f".{os.getpid()}.tmp")
            wb.save(tmp)
            tmp.replace(path)

    def _read_change_detail_rows(self) -> list[Row]:
        path = self.data_dir / CHANGE_LOG_FILE
        return self._read_sheet_rows(path, "修改明细", CHANGE_LOG_HEADERS)

    def _write_change_detail_rows(self, rows: list[Row]) -> None:
        path = self.data_dir / CHANGE_LOG_FILE
        with self._open_workbook(path) as wb:
            if "修改明细" not in wb.sheetnames:
                wb.create_sheet("修改明细")
            ws = wb["修改明细"]
            self._replace_sheet(ws, CHANGE_LOG_HEADERS, rows)
            tmp = path.with_suffix(f".{os.getpid()}.tmp")
            wb.save(tmp)
            tmp.replace(path)

    def _read_summary_rows(self) -> list[Row]:
        return self._read_sheet_rows(self.data_dir / CHANGE_LOG_FILE, "修改汇总", CHANGE_SUMMARY_HEADERS)

    def _write_summary_rows(self, rows: list[Row]) -> None:
        path = self.data_dir / CHANGE_LOG_FILE
        with self._open_workbook(path) as wb:
            if "修改汇总" not in wb.sheetnames:
                wb.create_sheet("修改汇总")
            ws = wb["修改汇总"]
            self._replace_sheet(ws, CHANGE_SUMMARY_HEADERS, rows)
            tmp = path.with_suffix(f".{os.getpid()}.tmp")
            wb.save(tmp)
            tmp.replace(path)

    def _ensure_summary_row(self, voucher: str, created_at: str | None = None) -> None:
        rows = self._read_summary_rows()
        if any(self._value(row, "入库编号") == voucher for row in rows):
            return
        rows.append(
            {
                "入库编号": voucher,
                "创建时间": created_at or self._now(),
                "第一次修改时间": "",
                "第二次修改时间": "",
                "最近修改时间": "",
                "修改次数": 0,
            }
        )
        self._write_summary_rows(rows)

    def _update_summary_modified(self, voucher: str) -> None:
        rows = self._read_summary_rows()
        if not any(self._value(row, "入库编号") == voucher for row in rows):
            self._ensure_summary_row(voucher)
            rows = self._read_summary_rows()
        now = self._now()
        for row in rows:
            if self._value(row, "入库编号") == voucher:
                count = int(row.get("修改次数") or 0) + 1
                row["修改次数"] = count
                if count == 1:
                    row["第一次修改时间"] = now
                elif count == 2:
                    row["第二次修改时间"] = now
                row["最近修改时间"] = now
                break
        self._write_summary_rows(rows)

    def _record_action(
        self,
        action_type: str,
        voucher: str,
        category: str,
        field: str,
        old_value: Any,
        new_value: Any,
    ) -> None:
        rows = self._read_plain_rows(self.data_dir / ACTION_LOG_FILE, ACTION_LOG_HEADERS)
        rows.append(
            {
                "操作ID": str(uuid.uuid4()),
                "时间": self._now(),
                "操作类型": action_type,
                "入库编号": voucher,
                "信息类别": category,
                "字段名": field,
                "旧值JSON": json.dumps(old_value, ensure_ascii=False, default=str),
                "新值JSON": json.dumps(new_value, ensure_ascii=False, default=str),
                "是否撤销": "",
            }
        )
        self._write_plain_rows(self.data_dir / ACTION_LOG_FILE, ACTION_LOG_HEADERS, rows)

    def _apply_action(self, action: Row, undo: bool) -> None:
        action_type = self._value(action, "操作类型")
        voucher = self._value(action, "入库编号")
        category = self._value(action, "信息类别")
        field = self._value(action, "字段名")
        old_value = self._json(action.get("旧值JSON"))
        new_value = self._json(action.get("新值JSON"))
        value = old_value if undo else new_value

        if action_type in ("update_fields", "classification_autofill"):
            rows = self.read_rows(category)
            index = self._row_index(rows, voucher)
            if index is None and value:
                rows.append(value)
            elif index is not None:
                rows[index] = self._fit_headers(value, CATEGORY_HEADERS[category])
            self._write_rows(category, rows)
            self._update_index_fingerprint(voucher)
        elif action_type == "update_photo":
            rows = self.read_rows("photo")
            target = old_value if undo else new_value
            opposite = new_value if undo else old_value
            idx = self._find_photo_row_index(rows, opposite)
            if idx is not None:
                if field == "文件名":
                    source_path = self.resolve_photo_path(opposite)
                    if source_path.exists():
                        target_path = self._move_archive_file_to_name(source_path, self._value(target, "文件名"))
                        target = dict(target)
                        target["文件名"] = target_path.name
                        target["相对路径"] = self._archive_relative_path(target_path)
                        target["来源工作区根路径"] = ""
                rows[idx] = self._fit_headers(target, PHOTO_HEADERS)
                self._write_rows("photo", rows)
        elif action_type == "add_photo":
            if undo:
                self._remove_photo_row(old_value=new_value)
            else:
                self._append_row("photo", new_value)
        elif action_type == "add_photos":
            photos = new_value if isinstance(new_value, list) else []
            if undo:
                for photo in photos:
                    self._remove_photo_row(old_value=photo)
            else:
                rows = self.read_rows("photo")
                rows.extend(self._fit_headers(photo, PHOTO_HEADERS) for photo in photos)
                self._write_rows("photo", rows)
        elif action_type == "delete_photo":
            if undo:
                self._append_row("photo", old_value)
            else:
                self._remove_photo_row(old_value=old_value)
        elif action_type == "create_specimen":
            if undo:
                self._delete_rows("specimen", voucher)
                self._delete_index(voucher)
            else:
                self._append_row("specimen", new_value)
                self._append_index(voucher, self._now(), "", "", self.record_fingerprint(voucher, specimen_override=new_value))
        elif action_type == "delete_specimen":
            if undo:
                specimen = old_value.get("specimen")
                classification = old_value.get("classification")
                photos = old_value.get("photos") or []
                index = old_value.get("index")
                if specimen:
                    self._append_row("specimen", specimen)
                if classification:
                    self._append_row("classification", classification)
                for photo in photos:
                    self._append_row("photo", photo)
                if index:
                    self._append_index_row(index)
            else:
                self._delete_rows("specimen", voucher)
                self._delete_rows("classification", voucher)
                self._delete_rows("photo", voucher)
                self._delete_index(voucher)
        elif action_type == "clear_photos":
            if undo:
                for photo in old_value.get("photos") or []:
                    self._append_row("photo", photo)
                self._invalidate_cache(PHOTO_FILE)
            else:
                self._delete_rows("photo", voucher)
                self._invalidate_cache(PHOTO_FILE)
        elif action_type == "move_photos":
            old_photos = old_value.get("photos") or []
            new_photos = new_value.get("photos") or []
            rows = self.read_rows("photo")
            if undo:
                for photo in new_photos:
                    self._remove_photo_from_rows(rows, photo)
                rows.extend(self._fit_headers(photo, PHOTO_HEADERS) for photo in old_photos)
            else:
                for photo in old_photos:
                    self._remove_photo_from_rows(rows, photo)
                rows.extend(self._fit_headers(photo, PHOTO_HEADERS) for photo in new_photos)
            self._write_rows("photo", rows)

    def _remove_photo_row(self, old_value: Row) -> None:
        rows = self.read_rows("photo")
        idx = self._find_photo_row_index(rows, old_value)
        if idx is not None:
            removed = rows.pop(idx)
            self._write_rows("photo", rows)
            self._delete_unreferenced_photo_file(removed, rows)

    def _remove_photo_from_rows(self, rows: list[Row], target: Row) -> bool:
        idx = self._find_photo_row_index(rows, target)
        if idx is None:
            return False
        rows.pop(idx)
        return True

    def _find_photo_row_index(self, rows: list[Row], target: Row) -> int | None:
        fitted = self._fit_headers(target, PHOTO_HEADERS)
        for idx, row in enumerate(rows):
            if self._fit_headers(row, PHOTO_HEADERS) == fitted:
                return idx
        return None

    def _append_row(self, category: str, row: Row) -> None:
        rows = self.read_rows(category)
        rows.append(self._fit_headers(row, CATEGORY_HEADERS[category]))
        self._write_rows(category, rows)

    def _write_rows(self, category: str, rows: list[Row]) -> None:
        self._write_plain_rows(self.data_dir / CATEGORY_FILES[category], CATEGORY_HEADERS[category], rows)
        self._invalidate_cache(CATEGORY_FILES[category])

    def _delete_rows(self, category: str, voucher: str) -> None:
        rows = [row for row in self.read_rows(category) if self._value(row, "入库编号*") != voucher]
        self._write_rows(category, rows)

    def _find_one(self, category: str, voucher: str) -> Row | None:
        for row in self.read_rows(category):
            if self._value(row, "入库编号*") == voucher:
                return row
        return None

    def _row_index(self, rows: list[Row], voucher: str) -> int | None:
        for idx, row in enumerate(rows):
            if self._value(row, "入库编号*") == voucher:
                return idx
        return None

    def _append_index(self, voucher: str, created_at: str, source_workspace: str, source_record_id: str, fingerprint: str) -> None:
        self._append_index_row(
            {
                "入库编号": voucher,
                "record_id": str(uuid.uuid4()),
                "创建时间": created_at,
                "来源工作区": source_workspace,
                "来源记录ID": source_record_id,
                "记录指纹": fingerprint,
            }
        )

    def _append_index_row(self, row: Row) -> None:
        rows = self._read_plain_rows(self.data_dir / INDEX_FILE, INDEX_HEADERS)
        voucher = self._value(row, "入库编号")
        if any(self._value(existing, "入库编号") == voucher for existing in rows):
            return
        rows.append(self._fit_headers(row, INDEX_HEADERS))
        self._write_plain_rows(self.data_dir / INDEX_FILE, INDEX_HEADERS, rows)

    def _find_index(self, voucher: str) -> Row | None:
        for row in self._read_plain_rows(self.data_dir / INDEX_FILE, INDEX_HEADERS):
            if self._value(row, "入库编号") == voucher:
                return row
        return None

    def _delete_index(self, voucher: str) -> None:
        rows = [row for row in self._read_plain_rows(self.data_dir / INDEX_FILE, INDEX_HEADERS) if self._value(row, "入库编号") != voucher]
        self._write_plain_rows(self.data_dir / INDEX_FILE, INDEX_HEADERS, rows)

    def _update_index_fingerprint(self, voucher: str) -> None:
        rows = self._read_plain_rows(self.data_dir / INDEX_FILE, INDEX_HEADERS)
        changed = False
        for row in rows:
            if self._value(row, "入库编号") == voucher:
                row["记录指纹"] = self.record_fingerprint(voucher)
                changed = True
        if changed:
            self._write_plain_rows(self.data_dir / INDEX_FILE, INDEX_HEADERS, rows)

    def _max_existing_serial(self) -> int:
        serials: list[int] = []
        for category in ("specimen", "classification", "photo"):
            header = "入库编号*"
            serials.extend(
                serial
                for row in self.read_rows(category)
                for serial in [parse_voucher_serial(self._value(row, header))]
                if serial is not None
            )
        serials.extend(
            serial
            for row in self._read_plain_rows(self.data_dir / INDEX_FILE, INDEX_HEADERS)
            for serial in [parse_voucher_serial(self._value(row, "入库编号"))]
            if serial is not None
        )
        return max(serials, default=0)

    def _sync_next_serial(self) -> None:
        self.config["next_serial"] = self._max_existing_serial() + 1
        self._save_config()

    # ── 多系列编号辅助方法 ─────────────────────────────────────────────────

    def _get_series_config(self, name: str) -> AccessionSeries | None:
        """按名称查找非 YZZ 系列配置，未找到返回 None。"""
        for item in self.config.get("accession_series", []):
            if item.get("name") == name:
                return AccessionSeries.from_dict(item)
        return None

    def _advance_series_counter(self, name: str) -> None:
        """将指定系列的 next_counter 按 step 推进一步（写回 config，调用方负责 _save_config）。"""
        for item in self.config.get("accession_series", []):
            if item.get("name") == name:
                item["next_counter"] = item.get("next_counter", 1) + item.get("step", 1)
                return

    def get_all_series_names(self) -> list[str]:
        """返回全部系列名称列表，YZZ 始终排第一。"""
        others = [s.get("name", "") for s in self.config.get("accession_series", [])]
        return ["YZZ"] + [n for n in others if n]

    def get_active_series_name(self) -> str:
        return self.config.get("active_series_name", "YZZ")

    def set_active_series(self, name: str) -> None:
        self.config["active_series_name"] = name
        self._save_config()

    def add_series(self, series: AccessionSeries) -> None:
        """新增一个非 YZZ 系列配置。若同名已存在则覆盖。"""
        series_list: list[dict] = self.config.setdefault("accession_series", [])
        for i, item in enumerate(series_list):
            if item.get("name") == series.name:
                series_list[i] = series.to_dict()
                self._save_config()
                return
        series_list.append(series.to_dict())
        self._save_config()

    def remove_series(self, name: str) -> None:
        """删除指定系列配置（不影响已录入的编号数据）。"""
        self.config["accession_series"] = [
            s for s in self.config.get("accession_series", []) if s.get("name") != name
        ]
        if self.config.get("active_series_name") == name:
            self.config["active_series_name"] = "YZZ"
        self._save_config()

    def count_vouchers_by_series(self, series_name: str) -> int:
        """返回已分发给指定系列的编号数量（从标本数据直接计数）。"""
        from .accession_series import series_prefix_of
        from .parsing import parse_voucher_serial
        rows = self.read_rows("specimen")
        if series_name == "YZZ":
            return sum(1 for r in rows if parse_voucher_serial(self._value(r, "入库编号*")) is not None)
        series = self._get_series_config(series_name)
        prefix = series.prefix if series else series_name
        return sum(
            1 for r in rows
            if series_prefix_of(str(self._value(r, "入库编号*") or "")) == prefix
        )

    def update_series_counter(self, name: str, new_counter: int) -> None:
        """手动设置系列的 next_counter（用于跳过已用编号）。"""
        for item in self.config.get("accession_series", []):
            if item.get("name") == name:
                item["next_counter"] = new_counter
                self._save_config()
                return

    # ── M5 多人协作：旧版工作区识别 + 升级到多人协作协议 ────────────────────

    def detect_legacy_workspace(self) -> bool:
        """识别旧版工作区（已录过数据 + 还没贴多人协作协议标记）。

        返回 True 时主管 UI 建议用户走「升级到多人协作格式」让 M5 自动归档历史段。
        空工作区不算 legacy（没东西可"丢失"），返回 False。
        已升级（含 `multi_user_protocol_version` 键）也返回 False。
        """
        has_marker = bool(self.config.get("multi_user_protocol_version"))
        if has_marker:
            return False
        # 用 specimen 行数（read_rows 已缓存，开销低）判定是否含历史数据
        return any(self._value(r, "入库编号*") for r in self.read_rows("specimen"))

    def upgrade_to_multi_user_protocol(self) -> dict:
        """把当前工作区升级到多人协作格式。

        升级**只动 `工作区配置.json`**：加 `multi_user_protocol_version` 键 +
        `legacy_yzz_segment`（记录历史 YZZ 段范围，方便事后追溯"哪些是升级前录入的"）。
        所有 Excel / 照片文件原样保留；schema 版本不 bump（保持跨版本兼容）。
        升级前**强制**创建快照（已有 `create_data_snapshot` 机制）。

        返回升级摘要：
        - already_upgraded: bool（True 表示无需升级）
        - snapshot_path:    Path（成功升级时返回）
        - legacy_yzz_segment: [start, end]（成功升级时返回）
        """
        if self.config.get("multi_user_protocol_version"):
            return {"already_upgraded": True}
        snapshot_path = self.create_data_snapshot(
            "升级到多人协作格式前快照",
            "在升级工作区到多人协作格式之前自动创建快照，方便回退。",
        )
        legacy_segment = [
            1,
            int(self.config.get("reserved_through_serial", 0)) or self._max_existing_serial(),
        ]
        self.config["multi_user_protocol_version"] = "1.0"
        self.config["legacy_yzz_segment"] = legacy_segment
        self._save_config()
        self._record_action(
            "upgrade_to_multi_user_protocol",
            "",
            "workspace",
            "",
            {},
            {
                "snapshot": str(snapshot_path),
                "multi_user_protocol_version": "1.0",
                "legacy_yzz_segment": legacy_segment,
            },
        )
        return {
            "already_upgraded": False,
            "snapshot_path": snapshot_path,
            "multi_user_protocol_version": "1.0",
            "legacy_yzz_segment": legacy_segment,
        }

    # ── M2 多人协作：录入员独立系列（前缀分人，避免离线撞号） ─────────────────

    def ensure_assignee_series(
        self,
        assignee: str,
        prefix: str,
        digits: int = 6,
        separator: str = "",
        year_pos: str = "none",
    ) -> str:
        """给指定录入员"按需"创建或复用独立编号系列；返回系列名。

        语义：同名 (assignee) 的系列已存在则直接复用（不动 next_counter）；
        不存在则新建一个，prefix/digits/separator/year_pos 立刻写入工作区配置。
        系列名固定为 `{assignee}_系列`，便于回溯。

        与 `add_series` 区别：add_series 是"按需新增/覆盖"；本方法是"按需创建（不
        覆盖已有计数器）"，更适合多次给同一录入员发号的场景。
        """
        assignee = (assignee or "").strip()
        if not assignee:
            raise ValueError("录入员名称不能为空")
        clean_prefix = (prefix or "").strip()
        if not clean_prefix:
            raise ValueError("录入员前缀不能为空")
        if not re.fullmatch(r"[A-Za-z0-9\-_]+", clean_prefix):
            raise ValueError(
                f"录入员前缀只支持 ASCII 字母/数字/横线/下划线（避免跨平台与 Excel 字符问题）：{clean_prefix!r}"
            )
        name = f"{assignee}_系列"
        existing = self._get_series_config(name)
        if existing is not None:
            return name
        # 也检查 prefix 不与其它系列重复（避免两个录入员用同一前缀，破坏分人语义）
        for item in self.config.get("accession_series", []):
            if item.get("prefix") == clean_prefix and item.get("name") != name:
                raise ValueError(
                    f"前缀 {clean_prefix!r} 已被系列 {item.get('name')!r} 占用，请换一个前缀。"
                )
        series = AccessionSeries(
            name=name,
            prefix=clean_prefix,
            digits=digits,
            separator=separator,
            year_pos=year_pos,
            next_counter=1,
            step=1,
        )
        self.add_series(series)
        return name

    def _write_conflict_report(self, conflicts: list[dict[str, str]]) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = self.data_dir / f"导入冲突报告_{timestamp}.xlsx"
        headers = ["入库编号", "冲突类型", "源记录摘要", "目标记录摘要"]
        self._write_plain_rows(path, headers, conflicts)
        return path

    def _write_photo_missing_report(self, rows: list[dict[str, str]]) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = self.data_dir / f"照片导入缺失报告_{timestamp}.xlsx"
        headers = ["入库编号", "文件名", "相对路径", "来源工作区根路径", "解析路径"]
        self._write_plain_rows(path, headers, rows)
        return path

    def _record_summary(self, specimen: Row | None, classification: Row | None) -> str:
        if not specimen and not classification:
            return ""
        specimen = specimen or {}
        classification = classification or {}
        parts = [
            f"管内编号={self._value(specimen, '管内编号*')}",
            f"地点={self._value(specimen, '采集地点缩写*')}",
            f"日期={self._value(specimen, '采集日期')}",
        ]
        parts.extend(
            f"{label}={self._value(classification, field)}"
            for label, field in CLASSIFICATION_SUMMARY_FIELDS
        )
        return "; ".join(parts)

    def _fingerprint_from_rows(self, specimen: Row | None, classification: Row | None) -> str:
        payload = {
            "specimen": self._fit_headers(specimen or {}, SPECIMEN_HEADERS),
            "classification": self._fit_headers(classification or {}, CLASSIFICATION_HEADERS),
        }
        encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
        return hashlib.sha256(encoded.encode("utf-8")).hexdigest()

    def _read_external_rows(self, path: Path, required_headers: list[str]) -> list[Row]:
        if not path.exists():
            return []
        return [self._fit_headers(row, required_headers) for row in self._read_plain_rows(path, required_headers)]

    def _headers(self, path: Path) -> list[str]:
        _ensure_openpyxl()
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            return [self._string(cell.value) for cell in next(ws.iter_rows(max_row=1))]
        finally:
            wb.close()

    def _read_plain_rows(self, path: Path, fallback_headers: list[str] | None = None) -> list[Row]:
        if not path.exists():
            return []
        _ensure_openpyxl()
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            # 规范化软件设计 2026-05 启动卡死优化:
            # 旧 `rows = list(ws.iter_rows(values_only=True))` 一次性物化整张表 -> 瞬时 RSS +30MB
            # (2GB 机器立刻触发 swap 卡死)。改流式 iter -> 解析一行处理一行,峰值减半。
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                return []
            headers = [self._string(value) for value in header_row]
            if fallback_headers:
                headers = headers or fallback_headers
            # sparse row dict:只保留非空字段。调用方走 `_value(row, field)` 或 `row.get(field, "")`,
            # 空字段返 ""。`read_rows` 出口处补 dense 保 API 契约。
            data: list[Row] = []
            for raw in rows_iter:
                row: Row = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    if idx >= len(raw):
                        continue
                    value = self._string(raw[idx])
                    if value != "":
                        row[header] = value
                if row:  # 非空行才进数据
                    data.append(row)
            return data
        finally:
            wb.close()

    def _read_sheet_rows(self, path: Path, sheet_name: str, fallback_headers: list[str]) -> list[Row]:
        if not path.exists():
            return []
        _ensure_openpyxl()
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return []
            ws = wb[sheet_name]
            # 同 _read_plain_rows: 流式 iter,不 list() 物化(避免 2GB 机内存峰值)。
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                return []
            headers = [self._string(value) for value in header_row] or fallback_headers
            data: list[Row] = []
            for raw in rows_iter:
                row: Row = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    if idx >= len(raw):
                        continue
                    value = self._string(raw[idx])
                    if value != "":
                        row[header] = value
                if row:
                    data.append(row)
            return data
        finally:
            wb.close()

    def _write_plain_rows(self, path: Path, headers: list[str], rows: list[Row]) -> None:
        _ensure_openpyxl()
        wb = Workbook()
        # 规范化软件设计 2026-05 P1 审查修复:Workbook 用 try/finally close,防 save/replace 异常时文件句柄泄漏。
        try:
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(headers)
            for row in rows:
                fitted = self._fit_headers(row, headers)
                ws.append([fitted.get(header, "") for header in headers])
            tmp = path.with_suffix(f".{os.getpid()}.tmp")
            wb.save(tmp)
            tmp.replace(path)
        finally:
            try:
                wb.close()
            except Exception:
                pass
        self._invalidate_cache(path.name)

    def _replace_sheet(self, ws: Any, headers: list[str], rows: list[Row]) -> None:
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)
        for row in rows:
            fitted = self._fit_headers(row, headers)
            ws.append([fitted.get(header, "") for header in headers])

    @contextmanager
    def _open_workbook(self, path: Path) -> Iterator[Any]:
        _ensure_openpyxl()
        wb = load_workbook(path)
        try:
            yield wb
        finally:
            wb.close()

    def _fit_headers(self, row: Row, headers: list[str]) -> Row:
        return {header: self._string((row or {}).get(header, "")) for header in headers}

    def _value(self, row: Row | None, field: str) -> str:
        if not row:
            return ""
        return self._string(row.get(field, ""))

    def _string(self, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat(sep=" ", timespec="seconds")
        s = str(value).strip()
        # 规范化软件设计 2026-05 P1 优化: sys.intern() 短字符串共享池。
        # Excel 高重复字段(凭证号 9 字符 / 管内编号 22 字符 / 地点缩写 ≤10 字符 /
        # 录入人员 / 保存方式 RE/FE 等)5000+ 行有大量重复,intern 后同值共享同一对象。
        # 阈值 ≤64 字符:覆盖全部短字段;长备注 / 描述不 intern(避免长跑泄漏)。
        # 估省 1-3MB on 中型工作区。
        if 0 < len(s) <= 64:
            return sys.intern(s)
        return s

    def _json(self, value: object) -> Any:
        if value in (None, ""):
            return {}
        try:
            return json.loads(str(value))
        except json.JSONDecodeError:
            return {}

    def _resolve_relative(self, root: Path, relative: str) -> Path:
        text = str(relative or "").strip()
        if text.startswith("./"):
            text = text[2:]
        base = root.resolve()
        raw = Path(text)
        resolved = raw.resolve() if raw.is_absolute() else (base / raw).resolve()
        try:
            resolved.relative_to(base)
        except ValueError:
            # 原代码：if not str(resolved).startswith(str(root.resolve())): return root / text
            # 使用 Path.relative_to 做严格边界检查；越界路径返回一个确定不存在的占位路径。
            return base / "__invalid_photo_path__" / raw.name
        return resolved

    def _now(self) -> str:
        return datetime.now().isoformat(sep=" ", timespec="seconds")


def _version_tuple(value: str) -> tuple[int, int, int]:
    parts = []
    for raw in str(value).split(".")[:3]:
        try:
            parts.append(int(raw))
        except ValueError:
            parts.append(0)
    while len(parts) < 3:
        parts.append(0)
    return tuple(parts[:3])
