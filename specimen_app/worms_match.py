"""WoRMS 分类匹配 — 非模态独立窗口（替代原 worms_dialog.py 的模态对话框）。

两个 Tab：
  Tab 1「匹配更新」— 与原 WormsMatchDialog 功能相同，改用本地缓存优先查询。
  Tab 2「分类浏览」— 输入任意分类名称，展示完整分类路径及同义名。

本地缓存：lookup → write 自动累积；用户可通过「管理本地数据库」手动导入 DwC-A zip。

追加规则（与用户约定，同原 worms_dialog.py）：
- 默认「只填空白」：已有内容的字段跳过。
- 勾选「覆盖已有内容」：全量写入。
- 备注字段特殊：无论覆盖模式，始终追加"WoRMS:AphiaID"；
  已含"WoRMS:"前缀则跳过（避免重复）。
"""

from __future__ import annotations

import urllib.request
import urllib.error
import urllib.parse
from pathlib import Path
from typing import TYPE_CHECKING

from PyQt5.QtCore import QThread, Qt, QUrl, pyqtSignal
from PyQt5.QtGui import QColor, QDesktopServices
from PyQt5.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QDialog,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMenu,
    QMessageBox,
    QPlainTextEdit,
    QProgressBar,
    QPushButton,
    QSizePolicy,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from . import __version__
from .app_settings import app_config_dir
from .worms_client import (
    WormsError,
    WoRMSRecord,
    WORMS_DWCA_URL,
    cache_stats,
    clear_cache,
    crawl_full_rest,
    export_cache_gz,
    import_dwca,
    install_cache_gz,
    query_worms_with_cache,
)

_USER_AGENT = f"specimen-inventory/{__version__}"

if TYPE_CHECKING:
    from .excel_store import ExcelStore


# ---------------------------------------------------------------------------
# Result-table column layout (shared between _MatchTab and _BrowseTab header)
# ---------------------------------------------------------------------------

_RESULT_COLS = [
    "输入名称", "WoRMS 接受名", "AphiaID", "门", "纲", "目", "科", "属", "状态", "影响标本数",
]
_COL_INPUT  = 0
_COL_VALID  = 1
_COL_APHIA  = 2
_COL_PHYLUM = 3
_COL_CLASS  = 4
_COL_ORDER  = 5
_COL_FAMILY = 6
_COL_GENUS  = 7
_COL_STATUS = 8
_COL_COUNT  = 9

_STATUS_OK   = "✓ 已接受"
_STATUS_NONE = "✗ 未找到"
_STATUS_ERR  = "⚠ 查询失败"

_COLOR_OK   = QColor("#d4edda")
_COLOR_NONE = QColor("#fff3cd")
_COLOR_ERR  = QColor("#f8d7da")

_WORMS_DETAIL_URL = "https://www.marinespecies.org/aphia.php?p=taxdetails&id={}"


# ---------------------------------------------------------------------------
# Background workers
# ---------------------------------------------------------------------------

class _WormsQueryWorker(QThread):
    """Queries WoRMS (cache-first) for each name in *names*, one at a time."""

    progress    = pyqtSignal(int, int)         # current, total
    result_ready = pyqtSignal(str, object)     # name, WoRMSRecord | None | WormsError
    finished_all = pyqtSignal()

    def __init__(self, names: list[str], cache: dict, parent=None) -> None:
        super().__init__(parent)
        self._names  = names
        self._cache  = cache
        self._abort  = False

    def abort(self) -> None:
        self._abort = True

    def run(self) -> None:
        total = len(self._names)
        for i, name in enumerate(self._names):
            if self._abort:
                break
            if name in self._cache:
                rec = self._cache[name]
            else:
                try:
                    rec = query_worms_with_cache(name)
                except WormsError as exc:
                    rec = exc
                self._cache[name] = rec
            self.result_ready.emit(name, rec)
            self.progress.emit(i + 1, total)
        self.finished_all.emit()


class _DwcaImportWorker(QThread):
    """Calls import_dwca() in a background thread, emitting progress."""

    progress = pyqtSignal(int, int)    # rows_processed, total
    finished = pyqtSignal(int, str)    # count_imported, error_msg (empty = success)

    def __init__(self, zip_path: str, parent=None) -> None:
        super().__init__(parent)
        self._zip_path = zip_path

    def run(self) -> None:
        try:
            count = import_dwca(
                self._zip_path,
                progress_cb=lambda n, t: self.progress.emit(n, t),
            )
            self.finished.emit(count, "")
        except Exception as exc:
            self.finished.emit(0, str(exc))


class _SynonymsFetchWorker(QThread):
    """Fetch synonyms list from WoRMS REST for a given AphiaID."""

    finished = pyqtSignal(list, str)   # synonyms_list, error_msg

    def __init__(self, aphia_id: int, parent=None) -> None:
        super().__init__(parent)
        self._aphia_id = aphia_id

    def run(self) -> None:
        import json as _json
        url = (
            f"https://www.marinespecies.org/rest/AphiaSynonymsByAphiaID"
            f"/{self._aphia_id}"
        )
        try:
            req = urllib.request.Request(
                url,
                headers={"Accept": "application/json"},
            )
            # 规范化软件设计 2026-05 P1 审查修复:
            # 旧:`urlopen(..., timeout=10)` 仅 connect 超时,resp.read() 可能因服务器停顿无限挂。
            # 现:用 socket 级 default timeout 控制读阶段(socket.setdefaulttimeout 全局,
            # 此处局部 set + restore)。读上限 10MB 防 OOM。
            import socket
            _old_to = socket.getdefaulttimeout()
            socket.setdefaulttimeout(10)
            try:
                with urllib.request.urlopen(req, timeout=10) as resp:
                    raw = resp.read(10 * 1024 * 1024 + 1)
                    if len(raw) > 10 * 1024 * 1024:
                        raise RuntimeError("Synonyms 响应过大(>10MB)")
                    data = _json.loads(raw)
            finally:
                socket.setdefaulttimeout(_old_to)
            if isinstance(data, list):
                syns = [
                    str(r.get("scientificname", "") or r.get("valid_name", ""))
                    for r in data
                    if isinstance(r, dict)
                ]
                self.finished.emit(syns, "")
            else:
                self.finished.emit([], "")
        except Exception as exc:
            self.finished.emit([], str(exc))


# 旧：`_WormsDwcaDownloadWorker` 直接拉 WORMS_DWCA_URL（zip 全库）。
# 2026-05 起 WoRMS 将 DwC-A zip 限制为 GBIF IP 白名单，客户端固定 403。
# 改用 REST 递归抓取（worms_client.crawl_full_rest）。
# 旧 worker 类已删除；如需仍走 DwC-A zip，请用「导入本地 DwC-A zip」按钮
# 配合 `_DwcaImportWorker`（仅本地 zip 输入，无网络下载）。

class _WormsRestCrawlWorker(QThread):
    """Recursively crawl WoRMS via REST and import into local cache.

    Replaces the deprecated `_WormsDwcaDownloadWorker` (DwC-A zip is now
    IP-restricted to GBIF, returns 403 to general clients).

    Signals:
      progress(imported_count, current_taxon_name)  — emitted every ~500 records.
      finished(imported_count, error_msg)           — error_msg empty on success.
    """

    progress = pyqtSignal(int, str)
    finished = pyqtSignal(int, str)

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self._stop_requested = False

    def request_stop(self) -> None:
        self._stop_requested = True

    def run(self) -> None:
        try:
            count = crawl_full_rest(
                progress_cb=lambda n, name: self.progress.emit(n, name),
                resume_state_path=app_config_dir() / "worms_crawl_state.json",
                should_stop=lambda: self._stop_requested,
            )
            self.finished.emit(count, "")
        except InterruptedError:
            # 用户取消，已保存断点
            self.finished.emit(0, "已取消（进度已保存，下次点击会自动续传）")
        except WormsError as exc:
            self.finished.emit(0, str(exc))
        except Exception as exc:
            self.finished.emit(0, f"{type(exc).__name__}: {exc}")


# 模块级可复制错误对话框：替代 `_status_lbl.setText(f"失败：{err}")` 单行文本。
# 用户报告的 HTTP 403 等长错误需要能复制反馈给开发者，QLabel 不支持选中复制。
class _CopyableErrorDialog(QDialog):
    """Error dialog with selectable + copyable QPlainTextEdit and a "Copy all" button."""

    def __init__(self, title: str, message: str, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(680, 380)
        layout = QVBoxLayout(self)
        self._edit = QPlainTextEdit(message)
        self._edit.setReadOnly(True)
        layout.addWidget(self._edit)
        btn_row = QHBoxLayout()
        copy_btn = QPushButton("复制全部")
        copy_btn.clicked.connect(lambda: QApplication.clipboard().setText(message))
        ok_btn = QPushButton("关闭")
        ok_btn.clicked.connect(self.accept)
        btn_row.addWidget(copy_btn)
        btn_row.addStretch(1)
        btn_row.addWidget(ok_btn)
        layout.addLayout(btn_row)


class _GithubCacheDownloadWorker(QThread):
    """Downloads worms_cache*.sqlite.gz from the latest GitHub Release and installs it."""

    dl_progress = pyqtSignal(int, int)   # downloaded_bytes, total_bytes
    finished    = pyqtSignal(int, str)   # record_count, error_msg

    def run(self) -> None:
        import json, ssl, tempfile
        _GH_HOSTS = ("github.com", "githubusercontent.com")
        tmp_path = None
        try:
            from .updater import GITHUB_REPO
            api_url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
            api_req = urllib.request.Request(
                api_url,
                headers={"User-Agent": _USER_AGENT, "Accept": "application/vnd.github+json"},
            )
            with urllib.request.urlopen(api_req, timeout=15, context=ssl.create_default_context()) as r:
                data = json.loads(r.read())

            asset_url = None
            for asset in data.get("assets", []):
                name = str(asset.get("name", ""))
                if name.startswith("worms_cache") and name.endswith(".sqlite.gz"):
                    asset_url = str(asset.get("browser_download_url", ""))
                    break
            if not asset_url:
                self.finished.emit(
                    0,
                    "此 GitHub Release 中未找到 worms_cache*.sqlite.gz 资产。\n"
                    "请改用「从 WoRMS 官网下载」或手动导入 DwC-A zip。",
                )
                return

            host = (urllib.parse.urlparse(asset_url).hostname or "").lower()
            if not any(host == h or host.endswith("." + h) for h in _GH_HOSTS):
                self.finished.emit(0, f"拒绝从非 GitHub 域名下载：{host}")
                return

            with tempfile.NamedTemporaryFile(suffix=".sqlite.gz", delete=False) as tmp:
                tmp_path = Path(tmp.name)
            dl_req = urllib.request.Request(asset_url, headers={"User-Agent": _USER_AGENT})
            # 规范化软件设计 2026-05 P1 审查修复:下载加总大小硬上限 + Content-Length 校验,防恶意/异常 header 导致 OOM/磁盘满。
            # WoRMS bootstrap sqlite.gz 通常 < 50MB,设 500MB 上限远超合理值兜底。
            _MAX_DL_BYTES = 500 * 1024 * 1024
            with urllib.request.urlopen(dl_req, timeout=120, context=ssl.create_default_context()) as resp:
                total = int(resp.headers.get("Content-Length", 0) or 0)
                if total > _MAX_DL_BYTES:
                    raise RuntimeError(f"WoRMS 缓存包过大(Content-Length={total},上限 {_MAX_DL_BYTES})")
                done = 0
                with open(tmp_path, "wb") as f:
                    while True:
                        chunk = resp.read(256 * 1024)
                        if not chunk:
                            break
                        f.write(chunk)
                        done += len(chunk)
                        if done > _MAX_DL_BYTES:
                            raise RuntimeError(f"WoRMS 缓存包超过 {_MAX_DL_BYTES // 1024 // 1024}MB,已中止")
                        self.dl_progress.emit(done, total)

            count = install_cache_gz(tmp_path)
            self.finished.emit(count, "")
        except Exception as exc:
            self.finished.emit(0, str(exc))
        finally:
            if tmp_path and tmp_path.exists():
                tmp_path.unlink(missing_ok=True)


def _make_separator() -> QFrame:
    sep = QFrame()
    sep.setFrameShape(QFrame.HLine)
    sep.setFrameShadow(QFrame.Sunken)
    return sep


def _disconnect_worker_signals(worker) -> None:
    """规范化软件设计 2026-05 P1 审查修复:重赋 self._worker 前断开旧 worker 全部 signal。

    防 防 旧 worker 已 finished 但 slot 仍连着,新 worker 触发同名 slot 时旧 worker 残留信号
    也跟着触发(双重处理 / use-after-free 风险)。
    若 worker 已无 signal connected 也安全(disconnect 失败 swallowed)。
    """
    if worker is None:
        return
    for sig_name in ("dl_progress", "progress", "finished", "result_ready"):
        sig = getattr(worker, sig_name, None)
        if sig is None:
            continue
        try:
            sig.disconnect()
        except (TypeError, RuntimeError):
            pass  # 没连接过 / 已断开


# ---------------------------------------------------------------------------
# DB manager dialog
# ---------------------------------------------------------------------------

class _DbManagerDialog(QDialog):
    """Shows cache stats, online download (GitHub / WoRMS official), manual import, clear."""

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("管理本地 WoRMS 数据库")
        self.setMinimumWidth(480)
        self._worker: QThread | None = None
        self._build_ui()
        self._refresh_stats()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)

        self._stats_lbl = QLabel()
        layout.addWidget(self._stats_lbl)

        attr_lbl = QLabel(
            '数据来源：<a href="https://www.marinespecies.org">WoRMS</a>'
            "，CC BY 4.0，WoRMS Editorial Board"
        )
        attr_lbl.setOpenExternalLinks(True)
        layout.addWidget(attr_lbl)

        layout.addWidget(_make_separator())
        layout.addWidget(QLabel("在线下载："))

        self._gh_btn = QPushButton("从 GitHub Release 下载缓存数据库（推荐，~15 MB）")
        self._gh_btn.setToolTip(
            "下载随软件发布的预置 WoRMS SQLite 缓存（gzip 压缩）。\n"
            "文件较小（~15 MB），不依赖 WoRMS 网站直连速度。\n"
            "需要 GitHub Release 中上传了 worms_cache*.sqlite.gz 资产。"
        )
        self._gh_btn.clicked.connect(self._start_gh_download)
        layout.addWidget(self._gh_btn)

        # 旧按钮文案：「从 WoRMS 官网下载全量 DwC-A（~30–50 MB）」
        # 改用 REST 递归抓取（DwC-A zip 已 IP 限制为 GBIF）。文案 + tooltip 全部重写。
        self._worms_btn = QPushButton("通过 WoRMS REST 全量抓取（约 30-60 分钟，约 50 万条）")
        self._worms_btn.setToolTip(
            "WoRMS 官方 DwC-A zip 已限制为 GBIF IP 专用，直链返回 403，无法下载。\n"
            "本按钮改用公开 REST API（AphiaChildrenByAphiaID）递归抓取整库。\n"
            "速率约 3 请求/秒，整库约需 30-60 分钟，支持断点续传。\n"
            "完成后自动导出 worms_cache_YYYY-QN.sqlite.gz 到桌面，可上传 GitHub Release\n"
            "供其他用户秒级下载安装。"
        )
        self._worms_btn.clicked.connect(self._start_worms_download)
        layout.addWidget(self._worms_btn)

        # W1: 后台独立进程下载选项（即使关闭应用也继续）
        self._bg_chk = QCheckBox(
            "在独立后台进程下载（即使关闭应用也继续，断点续传）"
        )
        self._bg_chk.setToolTip(
            "勾选后，「通过 WoRMS REST 全量抓取」会启动独立子进程跑下载。\n"
            "关闭应用 / 关闭本对话框后下载继续，直到完成或手动停止。\n"
            "现有 state 文件（worms_crawl_state.json）已支持断点续传 — 中途\n"
            "kill / 断电后再次启动会自动续传。\n"
            "不勾选 = 走 UI 线程（QThread），关闭应用立刻停（state 已保存，下次续传）。"
        )
        layout.addWidget(self._bg_chk)

        # W1: 后台 daemon 管理按钮
        self._bg_status_btn = QPushButton("检查后台下载状态")
        self._bg_status_btn.clicked.connect(self._check_bg_daemon_status)
        layout.addWidget(self._bg_status_btn)

        self._bg_stop_btn = QPushButton("停止后台下载")
        self._bg_stop_btn.clicked.connect(self._stop_bg_daemon)
        layout.addWidget(self._bg_stop_btn)

        layout.addWidget(_make_separator())

        self._import_btn = QPushButton("导入本地 DwC-A zip 文件…")
        self._import_btn.setToolTip(
            "已在本地下载好 WoRMS DwC-A zip？直接选择该文件导入。"
        )
        self._import_btn.clicked.connect(self._start_import)
        layout.addWidget(self._import_btn)

        self._progress = QProgressBar()
        self._progress.setVisible(False)
        layout.addWidget(self._progress)

        self._status_lbl = QLabel("")
        self._status_lbl.setWordWrap(True)
        layout.addWidget(self._status_lbl)

        layout.addWidget(_make_separator())

        clear_btn = QPushButton("清空缓存…")
        clear_btn.clicked.connect(self._clear_cache)
        layout.addWidget(clear_btn)

        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)

    def _refresh_stats(self) -> None:
        stats = cache_stats()
        count = stats.get("count", 0)
        last = stats.get("last_import") or "（从未导入）"
        self._stats_lbl.setText(f"本地缓存：{count:,} 条记录\n最后更新：{last}")

    def _set_busy(self, busy: bool) -> None:
        self._gh_btn.setEnabled(not busy)
        # _worms_btn 在 busy 期间仍要响应「取消」点击，单独管理
        self._import_btn.setEnabled(not busy)
        self._progress.setVisible(busy)
        if not busy:
            self._progress.setMaximum(100)
            self._progress.setValue(0)
            # 恢复 REST 抓取按钮的初始文案（_start_worms_download 已重命名）
            self._worms_btn.setEnabled(True)
            self._worms_btn.setText("通过 WoRMS REST 全量抓取（约 30-60 分钟，约 50 万条）")
        else:
            self._worms_btn.setEnabled(True)  # 保持可点（用作取消）

    def _start_gh_download(self) -> None:
        if self._worker and self._worker.isRunning():
            return
        _disconnect_worker_signals(self._worker)  # P1 审查修复:断开旧 worker signal
        self._set_busy(True)
        self._status_lbl.setText("正在连接 GitHub…")
        self._worker = _GithubCacheDownloadWorker(parent=self)
        self._worker.dl_progress.connect(self._on_dl_progress)
        self._worker.finished.connect(self._on_worker_finished)
        self._worker.start()

    def _start_worms_download(self) -> None:
        # W1: 后台模式 — 启 detached 子进程后立即返回
        if self._bg_chk.isChecked():
            self._launch_bg_daemon()
            return

        # 抓取中 → 第二次点击 = 请求取消
        if self._worker and self._worker.isRunning():
            if isinstance(self._worker, _WormsRestCrawlWorker):
                self._worker.request_stop()
                self._status_lbl.setText("已请求取消，等待当前请求结束…")
                self._worms_btn.setEnabled(False)
            return
        _disconnect_worker_signals(self._worker)  # P1 审查修复:断开旧 worker signal
        self._set_busy(True)
        # 抓取启动后改按钮为「取消抓取」（一次性 affordance）
        self._worms_btn.setEnabled(True)
        self._worms_btn.setText("取消抓取")
        self._status_lbl.setText("正在通过 REST 递归抓取 WoRMS(请保持联网,预计 30-60 分钟)…")
        self._progress.setMaximum(0)  # 总数未知 → 不确定模式
        self._worker = _WormsRestCrawlWorker(parent=self)
        self._worker.progress.connect(self._on_crawl_progress)
        self._worker.finished.connect(self._on_worker_finished)
        self._worker.start()

    def _on_crawl_progress(self, imported: int, current_taxon: str) -> None:
        """REST 抓取进度回调。imported = 累计写入条数，current_taxon = 当前节点 valid_name。"""
        if current_taxon:
            self._status_lbl.setText(
                f"抓取中… 已写入 {imported:,} 条（当前: {current_taxon}）"
            )
        else:
            self._status_lbl.setText(f"抓取中… 已写入 {imported:,} 条")

    def _start_import(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 WoRMS DwC-A zip", "", "ZIP 文件 (*.zip)"
        )
        if not path:
            return
        if self._worker and self._worker.isRunning():
            return
        _disconnect_worker_signals(self._worker)  # P1 审查修复
        self._set_busy(True)
        self._status_lbl.setText("正在导入…")
        self._worker = _DwcaImportWorker(path, parent=self)
        self._worker.progress.connect(self._on_imp_progress)
        self._worker.finished.connect(self._on_worker_finished)
        self._worker.start()

    def _on_dl_progress(self, done: int, total: int) -> None:
        mb_done = done / 1_048_576
        if total > 0:
            self._progress.setMaximum(total)
            self._progress.setValue(done)
            mb_total = total / 1_048_576
            self._status_lbl.setText(f"下载中… {mb_done:.1f} MB / {mb_total:.1f} MB")
        else:
            self._progress.setMaximum(0)
            self._status_lbl.setText(f"下载中… {mb_done:.1f} MB")

    def _on_imp_progress(self, n: int, total: int) -> None:
        if total > 0:
            self._progress.setMaximum(total)
            self._progress.setValue(n)
        self._status_lbl.setText(f"导入中… {n:,} / {total:,} 行")

    def _on_worker_finished(self, count: int, err: str) -> None:
        was_rest_crawl = isinstance(self._worker, _WormsRestCrawlWorker)
        self._set_busy(False)
        if err:
            # 旧：QLabel 单行展示 → 用户复制 HTTP 403 等长报错困难。
            # 新：弹 `_CopyableErrorDialog`（QPlainTextEdit + 一键复制）。
            self._status_lbl.setText(f"失败：{err.splitlines()[0][:80]}（详见对话框，可复制）")
            _CopyableErrorDialog("WoRMS 操作失败", err, parent=self).exec_()
        else:
            self._status_lbl.setText(f"完成！已安装 {count:,} 条记录。")
            if was_rest_crawl and count > 0:
                # REST 抓取成功 → 自动导出 .sqlite.gz 到桌面，便于上传 GitHub Release
                self._auto_export_to_desktop()
        self._refresh_stats()

    def _auto_export_to_desktop(self) -> None:
        """REST 抓取完成后自动导出本地缓存为 .sqlite.gz 到 ~/Desktop（fallback: ~/）。"""
        try:
            from datetime import date
            today = date.today()
            fname = f"worms_cache_{today.year}-Q{(today.month - 1) // 3 + 1}.sqlite.gz"
            desktop = Path.home() / "Desktop"
            if desktop.is_dir():
                out = desktop / fname
            else:
                # WSL / 无桌面环境：写到 ~/
                out = Path.home() / fname
            exported = export_cache_gz(out)
            size_mb = out.stat().st_size / 1_048_576
            QMessageBox.information(
                self,
                "已导出可分发缓存",
                f"已生成压缩缓存文件供分发：\n\n"
                f"  路径: {out}\n"
                f"  大小: {size_mb:.1f} MB\n"
                f"  条数: {exported:,}\n\n"
                f"如需让其他用户秒级安装，请将此文件作为资产上传到本项目的\n"
                f"GitHub Release（文件名匹配 worms_cache*.sqlite.gz）。",
            )
        except Exception as exc:
            # 导出失败不影响主流程（缓存已写入 SQLite）
            _CopyableErrorDialog(
                "自动导出 .sqlite.gz 失败",
                f"抓取成功，但导出可分发文件时出错：\n\n{type(exc).__name__}: {exc}\n\n"
                f"本地缓存已写入，不影响 app 使用；如需分发可在 CLI 手动跑\n"
                f"`python tools/build_worms_cache.py --export`",
                parent=self,
            ).exec_()

    def _clear_cache(self) -> None:
        reply = QMessageBox.question(
            self, "确认清空",
            "将删除本地缓存中的所有 WoRMS 记录，确认继续？",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            clear_cache()
            self._refresh_stats()
            self._status_lbl.setText("缓存已清空")

    def closeEvent(self, event) -> None:
        # W2 + E3：优雅请求停止 + 限时等待 + terminate fallback。
        # worker 已有 request_stop() (worms 抓取 worker) 或 requestInterruption()。
        self._stop_worker(wait_ms=3000)
        # 从主窗口的 WindowManager 注销自己（如果已注册）
        try:
            manager = getattr(self.parent(), "manager", None)
            if manager is not None and hasattr(manager, "unregister_dialog_stopper"):
                manager.unregister_dialog_stopper(self)
        except Exception:
            pass
        super().closeEvent(event)

    # ──────────────────── W1: 后台 detached daemon 管理 ─────────────────

    def _bg_paths(self) -> tuple["Path", "Path"]:
        """返回 (state_path, pid_path) — 与 daemon 共享同一 state 文件即可续传。"""
        from .app_settings import app_config_dir
        cfg = app_config_dir()
        cfg.mkdir(parents=True, exist_ok=True)
        return (
            cfg / "worms_crawl_state.json",
            cfg / "worms_crawler.pid",
        )

    def _launch_bg_daemon(self) -> None:
        """W1: subprocess.Popen 启动 detached daemon 子进程。"""
        from .worms_crawler_daemon import daemon_running, spawn_detached
        state_path, pid_path = self._bg_paths()
        existing = daemon_running(pid_path)
        if existing is not None:
            QMessageBox.information(
                self, "已有后台下载",
                f"后台已有 daemon 在跑（PID={existing}）。\n"
                "请先点击「停止后台下载」结束它，或等待完成。",
            )
            return
        try:
            pid = spawn_detached(state_path, pid_path)
        except Exception as exc:
            QMessageBox.critical(
                self, "启动失败",
                f"无法启动后台 daemon：{exc}",
            )
            return
        self._status_lbl.setText(
            f"已启动后台下载（PID={pid}）。\n"
            f"关闭本应用后下载继续进行。点击「检查后台下载状态」查看进度。"
        )

    def _check_bg_daemon_status(self) -> None:
        """W1: 显示后台 daemon 状态 + 进度（从 state file 读 imported 数）。"""
        from .worms_crawler_daemon import daemon_running
        import json
        state_path, pid_path = self._bg_paths()
        pid = daemon_running(pid_path)
        if pid is None:
            # 检查 state 文件是否暗示有未完成抓取
            if state_path.exists():
                try:
                    data = json.loads(state_path.read_text(encoding="utf-8"))
                    imported = data.get("imported_total", "?")
                    queue_remaining = len(data.get("queue", []))
                    QMessageBox.information(
                        self, "状态",
                        f"后台 daemon 当前未运行。\n"
                        f"上次抓取的 state 文件存在：\n"
                        f"  已导入：{imported} 条\n"
                        f"  剩余队列：{queue_remaining} 个节点\n"
                        f"勾选「在独立后台进程下载」后点抓取按钮可续传。",
                    )
                    return
                except Exception:
                    pass
            QMessageBox.information(self, "状态", "后台 daemon 未运行，无续传状态。")
            return
        # daemon 在跑 — 读 state 显示进度
        progress_text = "（state 文件尚未生成）"
        try:
            if state_path.exists():
                data = json.loads(state_path.read_text(encoding="utf-8"))
                imported = data.get("imported_total", 0)
                queue_remaining = len(data.get("queue", []))
                visited = len(data.get("visited", []))
                progress_text = (
                    f"已导入：{imported} 条\n"
                    f"已访问节点：{visited}\n"
                    f"剩余队列：{queue_remaining}"
                )
        except Exception:
            pass
        QMessageBox.information(
            self, "后台下载状态",
            f"后台 daemon 运行中（PID={pid}）。\n\n{progress_text}",
        )
        # 刷新缓存统计（daemon 写的是同一个 sqlite）
        self._refresh_stats()

    def _stop_bg_daemon(self) -> None:
        """W1: 给后台 daemon 发 SIGTERM，让它优雅停 + 保存 state。"""
        from .worms_crawler_daemon import daemon_running, request_stop_daemon
        _, pid_path = self._bg_paths()
        pid = daemon_running(pid_path)
        if pid is None:
            QMessageBox.information(self, "停止", "后台 daemon 未运行。")
            return
        reply = QMessageBox.question(
            self, "停止后台下载",
            f"将给 PID={pid} 的后台 daemon 发停止信号。\n"
            "state 已自动保存，下次启动会自动续传。\n\n是否继续？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )
        if reply != QMessageBox.Yes:
            return
        ok = request_stop_daemon(pid_path)
        if ok:
            self._status_lbl.setText(
                f"已请求停止后台 daemon（PID={pid}）。\n"
                "几秒后会优雅退出，state 已保存可续传。"
            )
        else:
            QMessageBox.warning(self, "停止失败", "无法发送停止信号（进程可能已退出）。")

    def _stop_worker(self, wait_ms: int = 3000) -> None:
        """W2 + E3：优雅停 worker → 限时等 → 强杀兜底。供 closeEvent 和
        WindowManager.stop_all_dialog_workers 复用（C1）。"""
        worker = self._worker
        if worker is None or not worker.isRunning():
            return
        # 优先调 request_stop()（WormsRestCrawlWorker 已实现）；
        # 退化到 requestInterruption()（QThread 标准接口）。
        try:
            if hasattr(worker, "request_stop"):
                worker.request_stop()
            else:
                worker.requestInterruption()
        except Exception:
            pass
        if not worker.wait(wait_ms) and worker.isRunning():
            try:
                worker.terminate()
                worker.wait(500)
            except Exception:
                pass


# ---------------------------------------------------------------------------
# Tab 1: Match tab (same functionality as original worms_dialog.py)
# ---------------------------------------------------------------------------

class _MatchTab(QWidget):
    """WoRMS 匹配更新 Tab — cache-first queries, right-click context menu."""

    def __init__(self, store: "ExcelStore", session_cache: dict, parent=None) -> None:
        super().__init__(parent)
        self._store = store
        self._cache = session_cache  # shared with _BrowseTab
        self._vouchers: list[str] | None = None  # None = tools-menu mode

        self._row_names:   list[str]    = []
        self._row_records: list[object] = []  # WoRMSRecord | None | WormsError
        self._worker: _WormsQueryWorker | None = None

        self._build_ui()

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setSpacing(8)

        # --- input area ---
        input_row = QHBoxLayout()
        self._name_edit = QPlainTextEdit()
        self._name_edit.setPlaceholderText(
            "每行输入一个拉丁学名，例如：\nGadus morhua\nThunnus thynnus"
        )
        self._name_edit.setFixedHeight(90)
        input_row.addWidget(self._name_edit)

        btn_col = QVBoxLayout()
        import_btn = QPushButton("导入 Excel 第一列")
        import_btn.setToolTip("选择 Excel 文件，读取第一列（跳过表头）作为物种名列表")
        import_btn.clicked.connect(self._import_excel)
        btn_col.addWidget(import_btn)

        db_btn = QPushButton("管理本地数据库…")
        db_btn.setToolTip("查看缓存统计、导入 DwC-A、清空缓存")
        db_btn.clicked.connect(self._open_db_manager)
        btn_col.addWidget(db_btn)
        btn_col.addStretch()
        input_row.addLayout(btn_col)
        root.addLayout(input_row)

        # --- control row ---
        ctrl = QHBoxLayout()
        self._query_btn = QPushButton("查询 WoRMS")
        self._query_btn.clicked.connect(self._start_query)
        ctrl.addWidget(self._query_btn)

        self._progress_lbl = QLabel("")
        ctrl.addWidget(self._progress_lbl)
        ctrl.addStretch()

        self._overwrite_cb = QCheckBox("覆盖已有内容")
        self._overwrite_cb.setToolTip(
            "默认只填空白字段。勾选后将用 WoRMS 值覆盖已有内容。\n"
            "备注字段始终是追加（不覆盖），无论此选项状态。"
        )
        ctrl.addWidget(self._overwrite_cb)
        root.addLayout(ctrl)

        # --- results table ---
        self._table = QTableWidget(0, len(_RESULT_COLS))
        self._table.setHorizontalHeaderLabels(_RESULT_COLS)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.horizontalHeader().setSectionResizeMode(
            _COL_INPUT, QHeaderView.ResizeToContents
        )
        self._table.horizontalHeader().setSectionResizeMode(
            _COL_VALID, QHeaderView.Stretch
        )
        self._table.horizontalHeader().setSectionResizeMode(
            _COL_STATUS, QHeaderView.ResizeToContents
        )
        self._table.horizontalHeader().setSectionResizeMode(
            _COL_COUNT, QHeaderView.ResizeToContents
        )
        self._table.itemChanged.connect(self._on_check_changed)
        # Right-click context menu for opening WoRMS webpage.
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._show_context_menu)
        root.addWidget(self._table)

        # --- summary + action buttons ---
        bottom = QHBoxLayout()
        self._summary_lbl = QLabel("查询完成后可应用到工作区")
        self._summary_lbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        bottom.addWidget(self._summary_lbl)

        self._apply_btn = QPushButton("应用到工作区")
        self._apply_btn.setEnabled(False)
        self._apply_btn.clicked.connect(self._apply_to_store)
        bottom.addWidget(self._apply_btn)
        root.addLayout(bottom)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def prefill_from_vouchers(self, vouchers: list[str]) -> None:
        """Pre-populate input text from vouchers and lock to those vouchers."""
        self._vouchers = vouchers
        seen: set[str] = set()
        names: list[str] = []
        for v in vouchers:
            cls = self._store.get_classification(v)
            if cls:
                name = (cls.get("种拉丁") or "").strip()
                if not name:
                    name = (cls.get("种名*") or "").strip()
            else:
                name = ""
            if name and name not in seen:
                seen.add(name)
                names.append(name)
        self._name_edit.setPlainText("\n".join(names))

    # ------------------------------------------------------------------
    # Input helpers
    # ------------------------------------------------------------------

    def _import_excel(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)"
        )
        if not path:
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            names: list[str] = []
            seen: set[str] = set()
            first = True
            for row in ws.iter_rows(values_only=True):
                if first:
                    first = False
                    val = str(row[0] or "").strip() if row else ""
                    if val.lower() in {
                        "物种名", "种名", "latinname", "scientificname",
                        "scientific name", "species", "name", "物种拉丁名",
                    }:
                        continue
                val = str(row[0] or "").strip() if row else ""
                if val and val not in seen:
                    seen.add(val)
                    names.append(val)
            wb.close()
        except Exception as exc:
            QMessageBox.warning(self, "读取失败", f"无法读取 Excel 文件：\n{exc}")
            return
        existing = self._name_edit.toPlainText().strip()
        if existing:
            self._name_edit.setPlainText(existing + "\n" + "\n".join(names))
        else:
            self._name_edit.setPlainText("\n".join(names))

    def _open_db_manager(self) -> None:
        dlg = _DbManagerDialog(parent=self)
        dlg.exec_()

    # ------------------------------------------------------------------
    # Query
    # ------------------------------------------------------------------

    def _start_query(self) -> None:
        if self._worker and self._worker.isRunning():
            self._worker.abort()
            self._worker.wait()

        raw = self._name_edit.toPlainText()
        names: list[str] = []
        seen: set[str] = set()
        for line in raw.splitlines():
            n = line.strip()
            if n and n not in seen:
                seen.add(n)
                names.append(n)

        if not names:
            QMessageBox.information(self, "无输入", "请先输入物种名（每行一个）。")
            return

        self._table.setRowCount(0)
        self._row_names.clear()
        self._row_records.clear()
        self._apply_btn.setEnabled(False)
        self._progress_lbl.setText(f"正在查询 0/{len(names)}…")
        self._query_btn.setEnabled(False)

        _disconnect_worker_signals(self._worker)  # P1 审查修复
        self._worker = _WormsQueryWorker(names, self._cache, parent=self)
        self._worker.result_ready.connect(self._on_result_ready)
        self._worker.progress.connect(self._on_progress)
        self._worker.finished_all.connect(self._on_query_finished)
        self._worker.start()

    def _on_progress(self, current: int, total: int) -> None:
        self._progress_lbl.setText(f"正在查询 {current}/{total}…")

    def _on_result_ready(self, name: str, rec: object) -> None:
        row = self._table.rowCount()
        self._table.insertRow(row)
        self._row_names.append(name)
        self._row_records.append(rec)

        chk = QTableWidgetItem(name)
        chk.setFlags(chk.flags() | Qt.ItemIsUserCheckable)

        if isinstance(rec, WoRMSRecord):
            chk.setCheckState(Qt.Checked)
            bg = _COLOR_OK
            status_text = _STATUS_OK
            valid  = rec.valid_name
            aphia  = str(rec.aphia_id)
            phylum = rec.phylum
            class_ = rec.class_
            order  = rec.order
            family = rec.family
            genus  = rec.genus
        elif rec is None:
            chk.setCheckState(Qt.Unchecked)
            bg = _COLOR_NONE
            status_text = _STATUS_NONE
            valid = aphia = phylum = class_ = order = family = genus = ""
        else:
            chk.setCheckState(Qt.Unchecked)
            bg = _COLOR_ERR
            status_text = f"{_STATUS_ERR}: {rec}"
            valid = aphia = phylum = class_ = order = family = genus = ""

        cells = [chk, valid, aphia, phylum, class_, order, family, genus, status_text]
        for col, cell in enumerate(cells):
            item = cell if isinstance(cell, QTableWidgetItem) else QTableWidgetItem(str(cell))
            item.setBackground(bg)
            self._table.setItem(row, col, item)

        count_item = QTableWidgetItem(str(self._count_affected(name)))
        count_item.setBackground(bg)
        self._table.setItem(row, _COL_COUNT, count_item)

    def _on_query_finished(self) -> None:
        self._query_btn.setEnabled(True)
        total = self._table.rowCount()
        ok = sum(1 for rec in self._row_records if isinstance(rec, WoRMSRecord))
        self._progress_lbl.setText(
            f"查询完成：{total} 个物种，{ok} 个已接受，{total - ok} 个未找到/失败"
        )
        self._apply_btn.setEnabled(ok > 0)
        self._update_summary_label()

    def _on_check_changed(self, item: QTableWidgetItem) -> None:
        if item.column() == _COL_INPUT:
            self._update_summary_label()

    def _update_summary_label(self) -> None:
        checked_ok = 0
        total_vouchers = 0
        for row in range(self._table.rowCount()):
            chk_item = self._table.item(row, _COL_INPUT)
            if chk_item and chk_item.checkState() == Qt.Checked:
                rec = self._row_records[row] if row < len(self._row_records) else None
                if isinstance(rec, WoRMSRecord):
                    checked_ok += 1
                    name = self._row_names[row]
                    total_vouchers += self._count_affected(name)
        self._summary_lbl.setText(
            f"已勾选 {checked_ok} 个物种，将更新 {total_vouchers} 个标本的分类信息"
        )

    # ------------------------------------------------------------------
    # Right-click context menu
    # ------------------------------------------------------------------

    def _show_context_menu(self, pos) -> None:
        row = self._table.rowAt(pos.y())
        if row < 0 or row >= len(self._row_records):
            return
        rec = self._row_records[row]
        menu = QMenu(self)
        if isinstance(rec, WoRMSRecord) and rec.aphia_id:
            url = _WORMS_DETAIL_URL.format(rec.aphia_id)
            menu.addAction(
                "在 WoRMS 网站查看",
                lambda: QDesktopServices.openUrl(QUrl(url)),
            )
        if menu.actions():
            menu.exec_(self._table.viewport().mapToGlobal(pos))

    # ------------------------------------------------------------------
    # Apply
    # ------------------------------------------------------------------

    def _count_affected(self, name: str) -> int:
        if self._vouchers is not None:
            return sum(
                1 for v in self._vouchers
                if (self._store.get_classification(v) or {}).get("种拉丁", "").strip() == name
            )
        return sum(
            1 for v in self._store.list_vouchers()
            if (self._store.get_classification(v) or {}).get("种拉丁", "").strip() == name
        )

    def _apply_to_store(self) -> None:
        overwrite = self._overwrite_cb.isChecked()
        changed_vouchers = 0
        changed_species = 0

        for row in range(self._table.rowCount()):
            chk_item = self._table.item(row, _COL_INPUT)
            if not chk_item or chk_item.checkState() != Qt.Checked:
                continue
            rec = self._row_records[row] if row < len(self._row_records) else None
            if not isinstance(rec, WoRMSRecord):
                continue

            name = self._row_names[row]

            if self._vouchers is not None:
                targets = [
                    v for v in self._vouchers
                    if (self._store.get_classification(v) or {}).get("种拉丁", "").strip() == name
                ]
            else:
                targets = [
                    v for v in self._store.list_vouchers()
                    if (self._store.get_classification(v) or {}).get("种拉丁", "").strip() == name
                ]

            if not targets:
                continue

            worms_fields = {
                "种拉丁": rec.valid_name,
                "属名":   rec.genus,
                "科*":    rec.family,
                "科拉丁": rec.family,
                "目":     rec.order,
                "纲":     rec.class_,
                "门":     rec.phylum,
            }

            species_updated = False
            for voucher in targets:
                existing = self._store.get_classification(voucher) or {}
                updates: dict[str, str] = {}

                for field, val in worms_fields.items():
                    if not val:
                        continue
                    if overwrite or not existing.get(field, "").strip():
                        updates[field] = val

                # 备注: always append AphiaID tag (never overwrite existing content).
                worms_tag = f"WoRMS:{rec.aphia_id}"
                note = existing.get("备注", "") or ""
                if worms_tag not in note:
                    updates["备注"] = (
                        f"{note}; {worms_tag}".lstrip("; ") if note else worms_tag
                    )

                if updates:
                    self._store.set_fields("classification", voucher, updates)
                    changed_vouchers += 1
                    species_updated = True

            if species_updated:
                changed_species += 1

        QMessageBox.information(
            self,
            "应用完成",
            f"已更新 {changed_species} 个物种、{changed_vouchers} 个标本的分类信息。",
        )
        for row in range(self._table.rowCount()):
            name = self._row_names[row] if row < len(self._row_names) else ""
            count_item = self._table.item(row, _COL_COUNT)
            if count_item and name:
                count_item.setText(str(self._count_affected(name)))

    # ------------------------------------------------------------------
    # Cleanup
    # ------------------------------------------------------------------

    def stop_worker(self) -> None:
        if self._worker and self._worker.isRunning():
            self._worker.abort()
            self._worker.wait()


# ---------------------------------------------------------------------------
# Tab 2: Browse tab
# ---------------------------------------------------------------------------

class _BrowseTab(QWidget):
    """分类浏览 Tab — 输入分类名称，展示完整分类路径及同义名。"""

    def __init__(self, session_cache: dict, parent=None) -> None:
        super().__init__(parent)
        self._cache = session_cache
        self._query_worker: _WormsQueryWorker | None = None
        self._syn_worker:   _SynonymsFetchWorker | None = None
        self._build_ui()

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setSpacing(8)

        # Search bar
        search_row = QHBoxLayout()
        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("输入拉丁学名，例如：Gadidae 或 Gadus morhua")
        self._search_edit.returnPressed.connect(self._start_browse)
        search_row.addWidget(self._search_edit)

        self._search_btn = QPushButton("查询")
        self._search_btn.clicked.connect(self._start_browse)
        search_row.addWidget(self._search_btn)
        root.addLayout(search_row)

        # Status label
        self._status_lbl = QLabel("")
        root.addWidget(self._status_lbl)

        # Classification path label
        self._path_lbl = QLabel("")
        self._path_lbl.setWordWrap(True)
        self._path_lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)
        root.addWidget(self._path_lbl)

        # Record detail label
        self._detail_lbl = QLabel("")
        self._detail_lbl.setWordWrap(True)
        self._detail_lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)
        root.addWidget(self._detail_lbl)

        # WoRMS link button
        self._link_btn = QPushButton("在 WoRMS 网站查看")
        self._link_btn.setVisible(False)
        self._link_btn.clicked.connect(self._open_worms_url)
        root.addWidget(self._link_btn)

        # Synonyms table
        syn_lbl = QLabel("同义名列表：")
        root.addWidget(syn_lbl)

        self._syn_table = QTableWidget(0, 1)
        self._syn_table.setHorizontalHeaderLabels(["同义名"])
        self._syn_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self._syn_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._syn_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        root.addWidget(self._syn_table)

        self._current_aphia_id: int = 0

    def _start_browse(self) -> None:
        name = self._search_edit.text().strip()
        if not name:
            return

        # Stop any previous workers.
        if self._query_worker and self._query_worker.isRunning():
            self._query_worker.abort()
            self._query_worker.wait()

        self._status_lbl.setText("正在查询…")
        self._path_lbl.setText("")
        self._detail_lbl.setText("")
        self._link_btn.setVisible(False)
        self._syn_table.setRowCount(0)
        self._current_aphia_id = 0
        self._search_btn.setEnabled(False)

        _disconnect_worker_signals(getattr(self, "_query_worker", None))  # P1 审查修复
        self._query_worker = _WormsQueryWorker([name], self._cache, parent=self)
        self._query_worker.result_ready.connect(self._on_browse_result)
        self._query_worker.finished_all.connect(lambda: self._search_btn.setEnabled(True))
        self._query_worker.start()

    def _on_browse_result(self, name: str, rec: object) -> None:
        if isinstance(rec, WoRMSRecord):
            self._display_record(rec)
        elif rec is None:
            self._status_lbl.setText(f"未在 WoRMS 找到「{name}」的接受名。")
        else:
            self._status_lbl.setText(f"查询失败：{rec}")

    def _display_record(self, rec: WoRMSRecord) -> None:
        self._current_aphia_id = rec.aphia_id
        self._status_lbl.setText(f"AphiaID: {rec.aphia_id}  |  状态: {rec.status}  |  Rank: {rec.rank}")

        # Build classification path.
        path_parts = []
        for level, value in [
            ("Phylum", rec.phylum),
            ("Class",  rec.class_),
            ("Order",  rec.order),
            ("Family", rec.family),
            ("Genus",  rec.genus),
        ]:
            if value:
                path_parts.append(f"{level}: {value}")
        if rec.valid_name:
            path_parts.append(f"<b>{rec.rank}: {rec.valid_name}</b>")

        self._path_lbl.setText("  ›  ".join(path_parts) if path_parts else "（无分类路径信息）")
        self._detail_lbl.setText(
            f"接受名：{rec.valid_name}　　权威：{rec.authority or '—'}"
        )
        self._link_btn.setVisible(rec.aphia_id > 0)

        # Fetch synonyms in background (fail silently).
        if rec.aphia_id > 0:
            if self._syn_worker and self._syn_worker.isRunning():
                self._syn_worker.wait()
            self._syn_worker = _SynonymsFetchWorker(rec.aphia_id, parent=self)
            self._syn_worker.finished.connect(self._on_synonyms_ready)
            self._syn_worker.start()

    def _on_synonyms_ready(self, synonyms: list, err: str) -> None:
        self._syn_table.setRowCount(0)
        for name in synonyms:
            row = self._syn_table.rowCount()
            self._syn_table.insertRow(row)
            self._syn_table.setItem(row, 0, QTableWidgetItem(name))

    def _open_worms_url(self) -> None:
        if self._current_aphia_id:
            url = _WORMS_DETAIL_URL.format(self._current_aphia_id)
            QDesktopServices.openUrl(QUrl(url))

    def stop_worker(self) -> None:
        if self._query_worker and self._query_worker.isRunning():
            self._query_worker.abort()
            self._query_worker.wait()
        if self._syn_worker and self._syn_worker.isRunning():
            self._syn_worker.wait()


# ---------------------------------------------------------------------------
# Main window
# ---------------------------------------------------------------------------

class WormsMatchWindow(QDialog):
    """非模态单实例 WoRMS 分类匹配窗口（替代原 WormsMatchDialog）。

    由 SpecimenWindow._open_worms_match() 持有引用保证单实例。
    用 show() + raise_() + activateWindow() 而非 exec()。
    """

    def __init__(self, parent, store: "ExcelStore") -> None:
        super().__init__(parent, Qt.Window)
        self._store = store

        # Session memory cache (name → WoRMSRecord | None | WormsError),
        # shared between tabs so a name queried in one tab is visible in the other.
        self._session_cache: dict[str, object] = {}

        self.setWindowTitle("WoRMS 分类匹配")
        self.setMinimumWidth(960)
        self.setMinimumHeight(600)
        self._build_ui()

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)

        tabs = QTabWidget()
        self._match_tab  = _MatchTab(self._store, self._session_cache, parent=self)
        self._browse_tab = _BrowseTab(self._session_cache, parent=self)
        tabs.addTab(self._match_tab,  "匹配更新")
        tabs.addTab(self._browse_tab, "分类浏览")
        root.addWidget(tabs)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def prefill_vouchers(self, vouchers: list[str]) -> None:
        """从入库汇总右键入口预填物种名（切换到匹配 tab）。"""
        # Find the QTabWidget — it's the only direct child.
        for child in self.children():
            if isinstance(child, QTabWidget):
                child.setCurrentIndex(0)
                break
        self._match_tab.prefill_from_vouchers(vouchers)

    # ------------------------------------------------------------------
    # Cleanup
    # ------------------------------------------------------------------

    def closeEvent(self, event) -> None:
        self._match_tab.stop_worker()
        self._browse_tab.stop_worker()
        super().closeEvent(event)
