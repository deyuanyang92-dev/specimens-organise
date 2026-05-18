"""软件自带的「字段模版」参考数据访问层。

`字段模版/` 是**随软件分发**的参考数据（不是工作区数据）：
- `表格信息预设字段.xlsx` —— 分类预设，供 `SpeciesMatcher` 做种名/科名自动匹配。
- `数据录入字段及字段说明.xlsx` —— 每个录入字段的填写示例 / 说明，供面板里字段旁的「?」提示。

本模块负责定位这些自带文件（源码运行 / PyInstaller 打包都能找到），并解析字段说明。
只用 openpyxl / pathlib，import 安全（不依赖 QApplication）。
"""

from __future__ import annotations

import sys
from pathlib import Path

# 规范化软件设计 2026-05 P1 优化:openpyxl 改函数内 lazy import,启动期不触发加载。
# Python import 缓存保证首次真加载,后续是 sys.modules 字典查询(0 开销)。

_TEMPLATE_DIR_NAME = "字段模版"
_FIELD_HELP_FILE = "数据录入字段及字段说明.xlsx"

# 照片面板字段名 -> 字段说明表里的字段名（说明表用「照片1_前缀」）。
_PHOTO_FIELD_ALIASES = {
    "文件名": "照片1_文件名",
    "相对路径": "照片1_相对路径",
    "描述": "照片1_描述",
}

_field_help_cache: dict[str, dict[str, str]] | None = None


def bundled_template_path(filename: str) -> Path | None:
    """定位软件自带 `字段模版/<filename>`；源码运行 / 打包都能找到，找不到返回 None。

    仿 icon._variant_icon_path 的多根解析。打包时由 build_release.py 的
    `--add-data specimen_app/字段模版` 带进 `_internal/specimen_app/字段模版/`。
    """
    rel = Path("specimen_app") / _TEMPLATE_DIR_NAME / filename
    roots: list[Path] = []
    if getattr(sys, "frozen", False):
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            roots.append(Path(meipass))
        roots.append(Path(sys.executable).resolve().parent / "_internal")
    # 源码运行：本文件在 specimen_app/ 下，自带模板就在同目录的 字段模版/。
    roots.append(Path(__file__).resolve().parent.parent)
    roots.append(Path.cwd())
    for root in roots:
        candidate = root / rel
        if candidate.exists():
            return candidate
    # 兜底：本文件同级的 字段模版/（不带 specimen_app/ 前缀的布局）。
    direct = Path(__file__).resolve().parent / _TEMPLATE_DIR_NAME / filename
    if direct.exists():
        return direct
    return None


def load_field_help() -> dict[str, dict[str, str]]:
    """读自带 `数据录入字段及字段说明.xlsx` 的 Sheet2 → {字段名称: {示例,说明,其他要求}}。

    Sheet2 列：大类 / 字段名称 / 填写示例 / 字段说明 / 其他要求。
    模块级缓存；文件缺失 / 解析失败一律返回 {}（容错，不该让缺说明挡住界面）。
    """
    global _field_help_cache
    if _field_help_cache is not None:
        return _field_help_cache
    result: dict[str, dict[str, str]] = {}
    path = bundled_template_path(_FIELD_HELP_FILE)
    if path is None:
        _field_help_cache = result
        return result
    try:
        from openpyxl import load_workbook  # lazy, P1 优化
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            # Sheet2 是字段说明表；找不到就退而用第二个 sheet。
            ws = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb.worksheets[min(1, len(wb.worksheets) - 1)]
            rows = ws.iter_rows(values_only=True)
            next(rows, None)  # 跳过表头
            for row in rows:
                cells = [("" if c is None else str(c)).strip() for c in row]
                cells += [""] * (5 - len(cells))
                _, name, example, desc, extra = cells[:5]
                if not name:
                    continue
                result[name] = {"示例": example, "说明": desc, "其他要求": extra}
        finally:
            wb.close()
    except Exception:
        result = {}
    _field_help_cache = result
    return result


def field_help_for(panel_field: str) -> dict[str, str] | None:
    """按面板字段名查填写说明；照片字段走别名映射。无说明返回 None。"""
    help_map = load_field_help()
    key = _PHOTO_FIELD_ALIASES.get(panel_field, panel_field)
    info = help_map.get(key) or help_map.get(panel_field)
    if not info:
        return None
    # 全空（示例/说明/其他要求都没有）视作无说明。
    if not any(info.get(k) for k in ("示例", "说明", "其他要求")):
        return None
    return info
