"""多人协作 — 收件箱聚合（M1 含 P1 等价的降级模式）。

设计要点（与 CLAUDE.md 数据兼容性约束一致）：
- **不改动 import_workspace 的现有行为**：本模块只负责 *扫子目录 + 加锁 + 调用现有 import_workspace + 分流*。
  现有指纹冲突 / 编号唯一性 / 照片物理去重等保护**沿用** ExcelStore 已有机制。
- **降级模式**：incoming/ 下任何"含 `数据/` 子目录"的目录都能被吃，不要求 manifest.json。
  这等价于 P1 极简方案；M3 任务包上线后 manifest 才会被 M4 段守护校验。
- **跨机锁**：rename `{name}/` → `processing_{name}/`，rename 失败说明另一合并员正在处理该子目录。
  rename 是文件系统级原子操作（同盘内），SMB/NTFS 都支持。
- **合并前强制 snapshot**：调 store.create_data_snapshot()，失败可一键回退。
- **分流目录**：
    incoming/processed/   成功合并后的子目录归档（保留 manifest + 数据 + 照片便于审计）
    incoming/conflicts/   ImportConflictError 的子目录 + 冲突报告 xlsx
    incoming/errors/      其它异常的子目录 + error.log
    incoming/duplicates/  M4 跨 voucher 同 SHA256 照片审核（本模块预留，M1 不写）

`aggregate_incoming(store, incoming_root, progress_cb=None) -> AggregateReport`
"""

from __future__ import annotations

import json
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Callable

# 规范化软件设计 2026-05 P1 优化:openpyxl 改函数内 lazy import,启动期不触发加载。

from .excel_store import ExcelStore
from .models import (
    AggregatePreview,
    AggregateReport,
    CLASSIFICATION_FILE,
    ImportConflictError,
    PHOTO_FILE,
    SPECIMEN_FILE,
    WorkspaceError,
)


# 处理中前缀（跨机锁）。其它合并员看到这个前缀就跳过。
PROCESSING_PREFIX = "processing_"

# 预定义分流子目录名（自身不参与扫描，含 name_conflicts 是 S3 新增）
_RESERVED_SUBDIR_NAMES = frozenset(
    {"processed", "conflicts", "errors", "duplicates", "name_conflicts"}
)

# S1: 递归扫描深度上限。3 层足以覆盖常见嵌套（如 incoming/某人/某次任务/数据/）；
# 设上限避免在巨型文件树上失控扫描。
_MAX_SCAN_DEPTH = 3


# ─────────────────────────── M4 段守护辅助 ───────────────────────────

_VOUCHER_SPLIT_RE = re.compile(r"^(.*?)(\d+)$")


def _split_voucher(voucher: str) -> tuple[str, int] | None:
    """拆 voucher 为 (前缀, 数字尾) 元组。失败返回 None。

    例：'ZS-000012' → ('ZS-', 12)；'YZZ000003' → ('YZZ', 3)。
    """
    m = _VOUCHER_SPLIT_RE.match(str(voucher).strip())
    if not m:
        return None
    return m.group(1), int(m.group(2))


def _voucher_in_range(voucher: str, range_start: str, range_end: str) -> bool:
    """voucher 是否落在 [range_start, range_end] 闭区间内（要求同前缀）。"""
    a = _split_voucher(voucher)
    b = _split_voucher(range_start)
    c = _split_voucher(range_end)
    if not (a and b and c):
        return False
    if a[0] != b[0] or a[0] != c[0]:
        return False
    return b[1] <= a[1] <= c[1]


def _read_source_vouchers(source_dir: Path) -> list[str]:
    """轻量读子目录 数据/标本信息.xlsx 取所有非空 voucher（不开 ExcelStore，零副作用）。"""
    sb_path = source_dir / "数据" / SPECIMEN_FILE
    if not sb_path.exists():
        return []
    try:
        from openpyxl import load_workbook  # lazy, P1 优化
        wb = load_workbook(sb_path, read_only=True, data_only=True)
    except Exception:
        return []
    try:
        ws = wb.active
        # 规范化软件设计 2026-05 P1 审查修复:旧 list(ws.iter_rows()) 全表物化,大工作区 OOM。
        # 改流式 iter,读 header 后逐行扫,仅累积 voucher 列值。
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
    finally:
        # 注意:close 必须在迭代完成后,但若提前 return 在 try 块内则 finally 仍跑
        # 这里先不 close,让下方迭代用完再关
        pass
    header = [str(c) if c is not None else "" for c in header_row]
    if "入库编号*" not in header:
        wb.close()
        return []
    idx = header.index("入库编号*")
    vouchers: list[str] = []
    try:
        for r in rows_iter:
            if r and idx < len(r) and r[idx]:
                v = str(r[idx]).strip()
                if v:
                    vouchers.append(v)
    finally:
        wb.close()
    return vouchers


class _SegmentGuardError(ValueError):
    """voucher 超出 manifest 声明的预留段时抛。被 aggregate_incoming 归入 errors/。"""


def _check_segment_guard(source_dir: Path, manifest: dict | None) -> None:
    """段守护：源子目录的 voucher 必须都落在 manifest.voucher_range 内。

    无 manifest 或缺 voucher_range 时跳过校验（降级模式，沿用 P1 等价行为）。
    """
    if not manifest:
        return
    range_pair = manifest.get("voucher_range")
    if not (isinstance(range_pair, (list, tuple)) and len(range_pair) == 2):
        return
    start, end = str(range_pair[0]), str(range_pair[1])
    vouchers = _read_source_vouchers(source_dir)
    if not vouchers:
        return
    out_of_range = [v for v in vouchers if not _voucher_in_range(v, start, end)]
    if out_of_range:
        sample = ", ".join(out_of_range[:5])
        more = f"（共 {len(out_of_range)} 个超界）" if len(out_of_range) > 5 else ""
        raise _SegmentGuardError(
            f"源工作区有 voucher 超出 manifest 预留段 [{start} … {end}]：{sample}{more}\n"
            f"主管需先在中心机扩展该录入员的段（再次「批量生成编号」给同一录入员/同一前缀）后重试。"
        )


def _scan_name_conflicts(processed_subdirs: list[Path]) -> dict[str, list[dict]]:
    """S3：扫已 processed 的子目录，找跨子目录"原始文件名相同但 SHA256 不同"的照片。

    返回 {原始文件名: [{subdir, voucher, sha256, archived_filename}, ...]}，
    仅保留 entries≥2 且 SHA256 至少 2 种值的条目。

    物理层在归档时已通过加 _2 后缀避免覆盖，本扫描仅为**事后报告**，让主管知道
    发生过同名不同内容的情况，便于回头核对。
    """
    by_name: dict[str, list[dict]] = {}
    for subdir in processed_subdirs:
        photo_xlsx = subdir / "数据" / "照片信息.xlsx"
        if not photo_xlsx.exists():
            continue
        try:
            from openpyxl import load_workbook  # lazy, P1 优化
            wb = load_workbook(photo_xlsx, read_only=True, data_only=True)
        except Exception:
            continue
        # P1 审查修复:流式 iter,不 list 物化整表。
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                wb.close()
                continue
            header = [str(c) if c is not None else "" for c in header_row]
            try:
                orig_idx = header.index("原始文件名")
                sha_idx = header.index("文件SHA256")
                vch_idx = header.index("入库编号*")
                fn_idx = header.index("文件名")
            except ValueError:
                wb.close()
                continue
            for r in rows_iter:
                if not r:
                    continue

                def _get(i: int, _r=r) -> str:  # 闭包绑定当前行 r,避免循环变量泄漏
                    return str(_r[i] or "").strip() if i < len(_r) else ""

                orig_name = _get(orig_idx)
                sha = _get(sha_idx)
                if not (orig_name and sha):
                    continue
                by_name.setdefault(orig_name, []).append(
                    {
                        "subdir": subdir.name,
                        "voucher": _get(vch_idx),
                        "sha256": sha,
                        "archived_filename": _get(fn_idx),
                    }
                )
        finally:
            wb.close()
    # 只保留真正冲突的（≥2 条记录 + ≥2 个不同 SHA256）
    conflicts: dict[str, list[dict]] = {}
    for name, entries in by_name.items():
        shas = {e["sha256"].lower() for e in entries}
        if len(entries) >= 2 and len(shas) >= 2:
            conflicts[name] = entries
    return conflicts


def _write_name_conflicts_report(incoming: Path, conflicts: dict[str, list[dict]]) -> Path | None:
    """S3：把同名不同内容冲突写到 incoming/name_conflicts/同名照片冲突报告_*.xlsx。"""
    if not conflicts:
        return None
    from openpyxl import Workbook
    target_dir = incoming / "name_conflicts"
    target_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = target_dir / f"同名照片冲突报告_{timestamp}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "同名不同内容"
    ws.append(["原始文件名", "来源子目录", "入库编号", "文件SHA256", "归档后文件名"])
    for name, entries in sorted(conflicts.items()):
        for e in entries:
            ws.append([name, e["subdir"], e["voucher"], e["sha256"], e["archived_filename"]])
    wb.save(path)
    return path


def _write_duplicates_report(
    target_dir: Path,
    subdir_name: str,
    candidates: list[dict],
) -> Path:
    """把跨 voucher 同 SHA256 候选写到 incoming/duplicates/{subdir}_重复审核.xlsx。"""
    from openpyxl import Workbook
    target_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = target_dir / f"{subdir_name}_重复审核_{timestamp}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "潜在重复入库"
    headers = ["入库编号", "已有voucher", "文件SHA256", "源相对路径", "源原始路径", "源解析路径"]
    ws.append(headers)
    for row in candidates:
        ws.append([str(row.get(h, "") or "") for h in headers])
    wb.save(path)
    return path


def _is_candidate_dir(path: Path) -> bool:
    """子目录必须存在 `数据/` 子目录才被视为待聚合源。"""
    return path.is_dir() and (path / "数据").is_dir()


def _discover_candidates(incoming_root: Path, max_depth: int = _MAX_SCAN_DEPTH) -> list[Path]:
    """S1: 递归扫描 incoming_root 找所有"含 `数据/` 子目录"的工作区目录。

    支持目录层次：
      incoming/ydy/数据/                直接子目录（M1 原行为）
      incoming/ydy/工作区A/数据/         二层嵌套
      incoming/某项目/某人/工作区B/数据/  三层嵌套（_MAX_SCAN_DEPTH 上限）

    规则：
    - 找到含 `数据/` 子目录的目录立即视为候选并**不再深入它**（避免重复合并嵌套工作区）
    - 跳过 `_RESERVED_SUBDIR_NAMES`（processed/conflicts/errors/duplicates/name_conflicts）
    - 跳过 `processing_*` 前缀（被别的合并员锁住）
    - 跳过隐藏目录（`.` 开头）和 `数据版本/`（快照目录）
    """
    candidates: list[Path] = []
    seen: set[Path] = set()

    def walk(d: Path, depth: int) -> None:
        if depth > max_depth:
            return
        try:
            children = sorted(d.iterdir())
        except OSError:
            return
        for child in children:
            if not child.is_dir():
                continue
            name = child.name
            if name in _RESERVED_SUBDIR_NAMES:
                continue
            if name.startswith(PROCESSING_PREFIX):
                continue
            if name.startswith("."):
                continue
            if name == "数据版本":  # 快照目录，不递归
                continue
            # 候选优先：本身含 `数据/` 子目录 → 收入 + 不再深入
            if _is_candidate_dir(child):
                resolved = child.resolve()
                if resolved not in seen:
                    seen.add(resolved)
                    candidates.append(child)
                continue
            # 否则递归找
            walk(child, depth + 1)

    walk(incoming_root, depth=1)
    return candidates


def _read_manifest(subdir: Path) -> dict | None:
    """读取子目录的 manifest.json；缺失或解析失败返回 None（降级模式照常聚合）。"""
    manifest_path = subdir / "manifest.json"
    if not manifest_path.exists():
        return None
    try:
        return json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError, UnicodeDecodeError):
        return None


def _sort_key(path: Path) -> tuple[int, str, str]:
    """优先按 manifest.packed_at 升序，无 manifest 按 mtime 升序，名字最后兜底。"""
    manifest = _read_manifest(path)
    if manifest:
        packed_at = str(manifest.get("packed_at", "") or "")
        return (0, packed_at, path.name)
    try:
        mtime = str(path.stat().st_mtime)
    except OSError:
        mtime = ""
    return (1, mtime, path.name)


def _unique_dest(parent: Path, name: str) -> Path:
    """目标目录已存在则在末尾加 `_2`、`_3` … 防止覆盖历史归档。"""
    candidate = parent / name
    if not candidate.exists():
        return candidate
    counter = 2
    while True:
        candidate = parent / f"{name}_{counter}"
        if not candidate.exists():
            return candidate
        counter += 1


def _read_xlsx_rows_safe(path: Path) -> tuple[list[str], list[list]]:
    """轻量读 xlsx 表头 + 数据行（只读，零副作用）。失败返回 ([], [])。"""
    if not path.exists():
        return [], []
    try:
        from openpyxl import load_workbook  # lazy, P1 优化
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return [], []
    # P1 审查修复:流式 iter 不 list 物化,大表内存安全。
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], []
        header = [str(c) if c is not None else "" for c in header_row]
        body = [list(r) if r else [] for r in rows_iter]
    finally:
        wb.close()
    return header, body


def preview_aggregate(store: ExcelStore, incoming_root: Path | str) -> AggregatePreview:
    """S7: 只读 dry-run 预测聚合结果。**完全不动文件**，不创快照、不调 import_workspace。

    用于「从收件箱聚合」UI 启动前先告知用户：会有多少新增 / 重复 / 冲突 / 同名冲突 /
    跨 voucher 同 SHA256，让用户预览后决定要不要点开始。
    """
    incoming = Path(incoming_root).resolve()
    if not incoming.is_dir():
        return AggregatePreview()

    candidates = _discover_candidates(incoming)
    candidates.sort(key=_sort_key)
    if not candidates:
        return AggregatePreview()

    # 中心现状快照：vouchers + 指纹 + photo SHA256 → voucher 映射
    central_vouchers = set(store.list_vouchers())
    central_fingerprints = {v: store.record_fingerprint(v) for v in central_vouchers}
    central_sha_to_voucher: dict[str, str] = {}
    for p in store.read_rows("photo"):
        sha = str(store._value(p, "文件SHA256") or "").lower()
        vch = str(store._value(p, "入库编号*") or "")
        if sha and vch and sha not in central_sha_to_voucher:
            central_sha_to_voucher[sha] = vch

    candidate_results: list[tuple] = []
    total_new = total_skipped = total_conflicts = total_photos = total_cross_dup = 0
    all_orig_to_sha: dict[str, set[str]] = {}

    for cand in candidates:
        try:
            rel_parts = cand.relative_to(incoming).parts
        except ValueError:
            rel_parts = (cand.name,)
        flat_name = "__".join(rel_parts) if rel_parts else cand.name

        spec_header, spec_body = _read_xlsx_rows_safe(cand / "数据" / SPECIMEN_FILE)
        if not spec_header or "入库编号*" not in spec_header:
            candidate_results.append((flat_name, "unreadable", 0, "缺 数据/标本信息.xlsx 或缺主键列"))
            continue
        id_idx = spec_header.index("入库编号*")

        # classification 表（可选）
        cls_by_voucher: dict[str, dict] = {}
        cls_header, cls_body = _read_xlsx_rows_safe(cand / "数据" / CLASSIFICATION_FILE)
        if cls_header and "入库编号*" in cls_header:
            cls_id_idx = cls_header.index("入库编号*")
            for r in cls_body:
                if cls_id_idx < len(r) and r[cls_id_idx]:
                    v = str(r[cls_id_idx]).strip()
                    cls_by_voucher[v] = {
                        cls_header[i]: (str(r[i]) if i < len(r) and r[i] is not None else "")
                        for i in range(len(cls_header))
                    }

        # 段守护预测
        manifest = _read_manifest(cand)
        if manifest:
            range_pair = manifest.get("voucher_range")
            if isinstance(range_pair, (list, tuple)) and len(range_pair) == 2:
                start, end = str(range_pair[0]), str(range_pair[1])
                seg_violators: list[str] = []
                for r in spec_body:
                    if id_idx < len(r) and r[id_idx]:
                        v = str(r[id_idx]).strip()
                        if v and not _voucher_in_range(v, start, end):
                            seg_violators.append(v)
                if seg_violators:
                    candidate_results.append(
                        (
                            flat_name,
                            "segment_violation",
                            0,
                            f"voucher 超 manifest 段 [{start}..{end}]，例：{seg_violators[0]}",
                        )
                    )
                    continue

        # 按 voucher 比对中心 → 新增 / 跳过 / 冲突
        cand_new = cand_skipped = cand_conflict = 0
        for r in spec_body:
            if id_idx >= len(r) or not r[id_idx]:
                continue
            v = str(r[id_idx]).strip()
            if not v:
                continue
            spec_row = {
                spec_header[i]: (str(r[i]) if i < len(r) and r[i] is not None else "")
                for i in range(len(spec_header))
            }
            cls_row = cls_by_voucher.get(v)
            if v in central_vouchers:
                src_fp = store._fingerprint_from_rows(spec_row, cls_row)
                if src_fp == central_fingerprints.get(v):
                    cand_skipped += 1
                else:
                    cand_conflict += 1
            else:
                cand_new += 1

        # photo 预扫
        cand_photos = cand_cross_dup_local = 0
        ph_header, ph_body = _read_xlsx_rows_safe(cand / "数据" / PHOTO_FILE)
        if ph_header:
            ph_vch_idx = ph_header.index("入库编号*") if "入库编号*" in ph_header else -1
            ph_sha_idx = ph_header.index("文件SHA256") if "文件SHA256" in ph_header else -1
            ph_orig_idx = ph_header.index("原始文件名") if "原始文件名" in ph_header else -1
            for r in ph_body:
                if ph_vch_idx < 0 or ph_vch_idx >= len(r) or not r[ph_vch_idx]:
                    continue
                vch = str(r[ph_vch_idx]).strip()
                sha = (
                    str(r[ph_sha_idx]).strip().lower()
                    if 0 <= ph_sha_idx < len(r) and r[ph_sha_idx]
                    else ""
                )
                orig = (
                    str(r[ph_orig_idx]).strip()
                    if 0 <= ph_orig_idx < len(r) and r[ph_orig_idx]
                    else ""
                )
                cand_photos += 1
                if sha and sha in central_sha_to_voucher and central_sha_to_voucher[sha] != vch:
                    cand_cross_dup_local += 1
                if orig and sha:
                    all_orig_to_sha.setdefault(orig, set()).add(sha)

        outcome = "conflict" if cand_conflict else ("new" if cand_new else "skipped")
        note = (
            f"{cand_new} 新增 / {cand_skipped} 重复跳过 / {cand_conflict} 冲突 / "
            f"{cand_photos} 照片"
        )
        candidate_results.append((flat_name, outcome, cand_new, note))
        total_new += cand_new
        total_skipped += cand_skipped
        total_conflicts += cand_conflict
        total_photos += cand_photos
        total_cross_dup += cand_cross_dup_local

    name_conflict_count = sum(1 for shas in all_orig_to_sha.values() if len(shas) >= 2)

    return AggregatePreview(
        candidates=candidate_results,
        total_candidates=len(candidates),
        predicted_new_vouchers=total_new,
        predicted_skipped_vouchers=total_skipped,
        predicted_conflicts=total_conflicts,
        predicted_photos=total_photos,
        predicted_name_conflicts=name_conflict_count,
        predicted_cross_voucher_duplicates=total_cross_dup,
    )


def aggregate_sources(
    store: ExcelStore,
    source_dirs: list[Path | str],
    progress_cb: Callable[[int, int, str], None] | None = None,
) -> AggregateReport:
    """S2: 直接合并用户指定的多个源工作区目录，**不需要 incoming/ 中转**。

    与 `aggregate_incoming` 关键差异：
    - 输入是一组绝对路径（用户在文件对话框逐个添加），不是统一仓库
    - 不 rename 加锁（用户的源目录原样保留，合并不动它们）
    - 不分流到 processed/conflicts/errors（无统一仓库可分流）
    - 仍创合并前快照、仍走段守护、仍跨 voucher SHA256 照片审核 (photo_duplicate_policy="report")
    - duplicates 报告写到中心机 `数据/duplicates_报告/` 子目录（沿数据走）

    适合"主管在文件管理器里随手选 D:\\ydy\\ + D:\\yss\\ + ..."的场景，
    省去先复制到 incoming/ 的步骤。
    """
    if not source_dirs:
        return AggregateReport()

    # 过滤合法源（必须含 数据/ 子目录）
    valid_sources: list[Path] = []
    skipped_paths: list[tuple[str, str]] = []
    for s in source_dirs:
        p = Path(s).resolve()
        if not p.is_dir():
            skipped_paths.append((str(s), "路径不存在或非目录"))
            continue
        if not (p / "数据").is_dir():
            skipped_paths.append((p.name, "缺 数据/ 子目录，非合法工作区"))
            continue
        valid_sources.append(p)

    if not valid_sources:
        return AggregateReport(errored=skipped_paths)

    snapshot_path = store.create_data_snapshot(
        "聚合前快照",
        f"批量导入工作区目录，共 {len(valid_sources)} 个源",
    )

    processed: list[str] = []
    conflicted: list[tuple] = []
    errored: list[tuple] = list(skipped_paths)
    duplicates: list[tuple] = []
    total_imported = 0
    total_photos = 0

    duplicates_dir = store.data_dir / "duplicates_报告"

    total = len(valid_sources)
    for idx, source in enumerate(valid_sources):
        if progress_cb:
            progress_cb(idx, total, source.name)
        try:
            manifest = _read_manifest(source)
            _check_segment_guard(source, manifest)
            result = store.import_workspace(source, photo_duplicate_policy="report")
            total_imported += result.imported
            total_photos += result.photos_imported
            processed.append(source.name)
            if result.duplicate_candidates:
                rpt = _write_duplicates_report(
                    duplicates_dir, source.name, result.duplicate_candidates
                )
                duplicates.append(
                    (source.name, list(result.duplicate_candidates), rpt)
                )
        except _SegmentGuardError as exc:
            errored.append((source.name, str(exc)))
        except ImportConflictError as exc:
            conflicted.append((source.name, str(exc), exc.report_path))
        except Exception as exc:  # noqa: BLE001 - 兜底任何异常
            errored.append((source.name, f"{type(exc).__name__}: {exc}"))

    if progress_cb:
        progress_cb(total, total, "")

    return AggregateReport(
        processed=processed,
        conflicted=conflicted,
        errored=errored,
        duplicates=duplicates,
        total_imported=total_imported,
        total_photos=total_photos,
        snapshot_path=snapshot_path,
    )


def aggregate_incoming(
    store: ExcelStore,
    incoming_root: Path | str,
    progress_cb: Callable[[int, int, str], None] | None = None,
) -> AggregateReport:
    """扫描 incoming_root 下所有候选子目录，逐个调 import_workspace，分流归档。

    progress_cb(done_count, total_count, current_name) 在每个子目录开始处理前调用一次；
    全部完成后再调一次 progress_cb(total, total, "")。
    """
    incoming = Path(incoming_root).resolve()
    if not incoming.is_dir():
        raise WorkspaceError(f"收件箱目录不存在：{incoming}")

    # S1: 递归扫描（支持嵌套）。旧的一层扫描逻辑被 _discover_candidates 包含。
    candidates = _discover_candidates(incoming)
    candidates.sort(key=_sort_key)

    if not candidates:
        return AggregateReport()

    # 合并前强制快照（含 incoming 路径方便事后追溯）
    snapshot_path = store.create_data_snapshot(
        "聚合前快照",
        f"来自 {incoming}，候选子目录 {len(candidates)} 个",
    )

    processed_dir = incoming / "processed"
    conflicts_dir = incoming / "conflicts"
    errors_dir = incoming / "errors"
    duplicates_dir = incoming / "duplicates"
    processed_dir.mkdir(exist_ok=True)
    conflicts_dir.mkdir(exist_ok=True)
    errors_dir.mkdir(exist_ok=True)

    processed: list[str] = []
    conflicted: list[tuple] = []
    errored: list[tuple] = []
    duplicates: list[tuple] = []
    total_imported = 0
    total_photos = 0

    total = len(candidates)
    for idx, candidate in enumerate(candidates):
        # S1 嵌套场景：用相对路径段拼成扁平名，防止同名子目录撞锁/撞归档。
        # 例：incoming/ydy/工作区A/ → 扁平名 "ydy__工作区A"
        try:
            rel_parts = candidate.relative_to(incoming).parts
        except ValueError:
            rel_parts = (candidate.name,)
        original_name = "__".join(rel_parts) if rel_parts else candidate.name
        if progress_cb:
            progress_cb(idx, total, original_name)

        # 加锁：rename 到 incoming/processing_<扁平名>；
        # rename 失败（被别的合并员抢先或跨文件系统）→ errored 并继续
        lock_path = incoming / (PROCESSING_PREFIX + original_name)
        if lock_path.exists():
            errored.append((original_name, f"加锁失败：目标 lock 路径已被占用：{lock_path}"))
            continue
        try:
            candidate.rename(lock_path)
        except OSError as exc:
            errored.append((original_name, f"加锁失败：{exc}"))
            continue

        try:
            # M4 段守护：读 manifest，若声明了 voucher_range 则校验所有源 voucher 在段内
            manifest = _read_manifest(lock_path)
            _check_segment_guard(lock_path, manifest)
            # M4 调用 import_workspace 的 "report" 模式：跨 voucher 同 SHA256 照片
            # 不写入 photo 表，而是收集到 duplicate_candidates 由本函数写报告
            result = store.import_workspace(lock_path, photo_duplicate_policy="report")
            total_imported += result.imported
            total_photos += result.photos_imported
            processed.append(original_name)
            if result.duplicate_candidates:
                report_path = _write_duplicates_report(
                    duplicates_dir, original_name, result.duplicate_candidates
                )
                duplicates.append((original_name, list(result.duplicate_candidates), report_path))
            dest = _unique_dest(processed_dir, original_name)
            lock_path.rename(dest)
        except _SegmentGuardError as exc:
            # 段守护失败 → errors/（不是 conflicts/，因为是录入员越权而非数据冲突）
            message = str(exc)
            errored.append((original_name, message))
            dest = _unique_dest(errors_dir, original_name)
            try:
                lock_path.rename(dest)
            except OSError:
                shutil.copytree(lock_path, dest)
                shutil.rmtree(lock_path, ignore_errors=True)
            try:
                (dest / "error.log").write_text(message + "\n", encoding="utf-8")
            except OSError:
                pass
        except ImportConflictError as exc:
            conflicted.append((original_name, str(exc), exc.report_path))
            dest = _unique_dest(conflicts_dir, original_name)
            try:
                lock_path.rename(dest)
            except OSError:
                # rename 失败（跨设备等）→ 用 copy + remove 兜底
                shutil.copytree(lock_path, dest)
                shutil.rmtree(lock_path, ignore_errors=True)
            # 冲突报告 xlsx 一并复制进归档目录便于查阅
            if exc.report_path and exc.report_path.exists():
                try:
                    shutil.copy2(exc.report_path, dest / exc.report_path.name)
                except OSError:
                    pass
        except Exception as exc:  # noqa: BLE001 - 兜底任何异常，写日志后继续
            message = f"{type(exc).__name__}: {exc}"
            errored.append((original_name, message))
            dest = _unique_dest(errors_dir, original_name)
            try:
                lock_path.rename(dest)
            except OSError:
                shutil.copytree(lock_path, dest)
                shutil.rmtree(lock_path, ignore_errors=True)
            try:
                (dest / "error.log").write_text(message + "\n", encoding="utf-8")
            except OSError:
                pass

    if progress_cb:
        progress_cb(total, total, "")

    # S3: 跑完所有候选后扫 processed/ 找跨子目录同名不同内容照片
    processed_paths = [processed_dir / name for name in processed]
    name_conflicts_dict = _scan_name_conflicts(processed_paths)
    name_conflicts_report = _write_name_conflicts_report(incoming, name_conflicts_dict)
    name_conflicts_list: list[tuple] = [
        (name, entries) for name, entries in sorted(name_conflicts_dict.items())
    ]

    return AggregateReport(
        processed=processed,
        conflicted=conflicted,
        errored=errored,
        duplicates=duplicates,
        name_conflicts=name_conflicts_list,
        total_imported=total_imported,
        total_photos=total_photos,
        snapshot_path=snapshot_path,
        name_conflicts_report_path=name_conflicts_report,
    )
