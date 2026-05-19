from __future__ import annotations

import os
import signal
import subprocess
import sys
import queue
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any, Callable, Optional, TYPE_CHECKING

# 规范化软件设计 2026-05 P1 优化:PIL lazy import。
# ui.py 内仅 pil_to_qpixmap 用到 PIL.Image 实例,无 Image.* 类方法调用。
# type hint `Image.Image` 因 `from __future__ import annotations` 而仅为字符串,运行时不解析。
# 实际真正用 PIL 的代码在 image_cache.py 顶层 import,本模块顶层无需再加载,省 5-10MB 启动 RSS。
if TYPE_CHECKING:
    from PIL import Image  # type: ignore[import-not-found]
from PyQt5.QtCore import Qt, QTimer, QThread, pyqtSignal, QPoint, QSize, QByteArray, QModelIndex
from PyQt5.QtGui import QImage, QPixmap, QKeySequence, QFont, QPainter, QCursor, QFontMetrics, QColor, QStandardItem, QStandardItemModel
from PyQt5.QtWidgets import (
    QAction,
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QFrame,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QLineEdit,
    QListWidget,
    QShortcut,
    QListWidgetItem,
    QMainWindow,
    QMenu,
    QMessageBox,
    QProgressDialog,
    QPushButton,
    QScrollArea,
    QSpinBox,
    QSplitter,
    QStatusBar,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QToolBar,
    QToolButton,
    QVBoxLayout,
    QWidget,
    QCompleter,
)
from PyQt5.QtWidgets import QGraphicsView, QGraphicsScene, QGraphicsPixmapItem

from . import __version__
from .app_settings import (
    DEFAULT_PHOTO_FILENAME_FILL_SHORTCUT,
    PHOTO_MANAGEMENT_OPTIONS,
    PREVIEW_QUALITY_OPTIONS,
    PREVIEW_QUALITY_SIZES,
    load_settings,
    remember_workspace,
    save_settings,
)
from .classification_fields import (
    EDITABLE_CLASSIFICATION_COLUMNS,
    FAMILY_LOOKUP_INPUT_COLUMNS,
    SPECIES_LOOKUP_INPUT_COLUMNS,
    TAXONOMY_LOOKUP_INPUT_COLUMNS,
    classification_values_from_family_match,
    classification_values_from_species_match,
)
from .icon import get_app_icon
from .excel_store import ExcelStore
from .image_cache import ThumbnailCache
from .image_search import (
    ImageSearchIndex,
    ImageSearchResult,
    _get_or_build_search_index,
    append_images_to_index,
    clear_image_index,
    default_image_query,
    image_file_filter,
    image_index_exists,
    image_search_results,
    is_supported_image,
    suffixes_for_image_type,
)
from .accession_series import AccessionSeries, BUILTIN_PRESETS, format_series_number
from .batch_export import BatchExportDialog  # 批量导出功能
from .startup_diag import mark as _startup_mark
from .models import (
    CARRY_OVER_SPECIMEN_FIELDS,
    ImportConflictError,
    PHOTO_COUNT_COLUMN,
    SAVE_METHOD_OPTIONS,
    SPECIMEN_HEADERS,
    SUMMARY_COLUMNS,
    SUMMARY_COLUMN_SOURCE,
    SUMMARY_DEFAULT_VISIBLE_COLUMNS,
    WorkspaceLockedError,
    WorkspaceNotInitializedError,
)
from .parsing import (
    derive_specimen_fields_from_tube_number,
    extract_specimen_tube_from_filename,
    parse_voucher_serial,
)
from .release_manager import list_releases, release_roots
from .server_sync import aggregate_incoming, aggregate_sources, preview_aggregate  # M1/S2/S7
from .task_package import export_task_package, import_task_package  # M3: 任务包
from .dwc_export import export_dwc_archive  # A1: Darwin Core Archive 导出
from .exif_info import apply_exif_to_specimen  # A2: EXIF 抽取与回填
from .updater import (
    check_latest_release,
    default_download_root,
    download_release,
    download_update,
    is_newer,
)
from .species import FamilyMatch, SpeciesMatch, SpeciesMatcher
from .workspace import (
    default_workspace,
    has_workspace_data,
    initialize_workspace,
    is_generated_workspace_path,
    is_unsafe_workspace_root,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def grid_shape(count: int) -> tuple[int, int]:
    """根据照片数量计算最佳网格列数和行数。

    布局规则：
    - 1张: (1, 1)  单格填满
    - 2张: (2, 1)  水平并列
    - 3-4张: (2, 2) 2×2方阵
    - 5-6张: (3, 2) 3列2行
    - 7-8张: (4, 2) 4列2行
    """
    if count <= 2:
        return max(1, count), 1
    if count <= 4:
        return 2, 2
    if count <= 6:
        return 3, 2
    return 4, 2


def _join_display(*parts: str) -> str:
    return "  ".join(part for part in parts if part)


TAXONOMY_INSERT_TEXT_ROLE = Qt.UserRole + 1
TAXONOMY_CANDIDATE_ROW_ROLE = Qt.UserRole + 2


class FieldValueCompleter(QCompleter):
    """Show full taxonomy candidates, but insert only the active field value."""

    def pathFromIndex(self, index: QModelIndex) -> str:
        value = index.data(TAXONOMY_INSERT_TEXT_ROLE)
        return str(value or "")


def format_taxonomy_candidate_label(
    field: str,
    kind: str,
    match: SpeciesMatch | FamilyMatch,
) -> str:
    if kind == "species" and isinstance(match, SpeciesMatch):
        if field == "属名":
            return _join_display(
                match.genus_name,
                match.chinese_name,
                match.latin_name,
                match.family_name,
                match.family_latin,
            )
        if field == "种拉丁":
            return _join_display(match.latin_name, match.chinese_name, match.family_name, match.family_latin)
        return _join_display(match.chinese_name, match.latin_name, match.family_name, match.family_latin)
    if kind == "family" and isinstance(match, FamilyMatch):
        if field == "科拉丁":
            return _join_display(match.family_latin, match.family_name)
        return _join_display(match.family_name, match.family_latin)
    return ""


def classification_column_value_from_taxonomy_match(
    field: str,
    kind: str,
    match: SpeciesMatch | FamilyMatch,
) -> str:
    if kind == "species" and isinstance(match, SpeciesMatch):
        if field == "属名":
            return match.genus_name
        if field == "种拉丁":
            return match.latin_name
        return match.chinese_name
    if kind == "family" and isinstance(match, FamilyMatch):
        if field == "科拉丁":
            return match.family_latin
        return match.family_name
    return ""


PHOTO_FILENAME_FILL_FIELDS = ("管内编号*", "采集地点缩写*", "采集日期", "保存方式")


def photo_filename_source_for_specimen_fill(photo_row: dict[str, Any]) -> str:
    """Use the original name first; archived names may have suffixes like _2."""
    original = str(photo_row.get("原始文件名", "") or "").strip()
    if original:
        return original
    return str(photo_row.get("文件名", "") or "").strip()


def specimen_updates_from_photo_filename(filename: str) -> dict[str, str]:
    tube = extract_specimen_tube_from_filename(filename)
    updates: dict[str, str] = {}
    if tube:
        updates["管内编号*"] = tube
        # 原照片文件名逻辑分别解析日期、地点和保存方式；现在复用管内编号派生逻辑，避免两套规则不一致。
        updates.update(derive_specimen_fields_from_tube_number(tube))
    return {field: updates[field] for field in PHOTO_FILENAME_FILL_FIELDS if updates.get(field)}


def default_photo_filename_fill_fields(updates: dict[str, str], current: dict[str, Any]) -> list[str]:
    return [
        field
        for field in PHOTO_FILENAME_FILL_FIELDS
        if updates.get(field) and not str(current.get(field, "") or "").strip()
    ]


VIEW_MODES = [
    ("单张", 1, "1"),
    ("2宫格", 2, "2"),
    ("4宫格", 4, "4"),
    ("6宫格", 6, "6"),
    ("8宫格", 8, "8"),
]

_VIEW_BTN_STYLE = """
    QPushButton { border: 1px solid #aab; border-radius: 4px; padding: 4px 8px; font-weight: bold; }
    QPushButton:checked { background-color: #2a6fbd; color: white; }
"""


def _shorten(value: str, max_length: int) -> str:
    if len(value) <= max_length:
        return value
    return value[: max(1, max_length - 1)] + "…"


def pil_to_qpixmap(pil_image: Image.Image) -> QPixmap:
    if pil_image.mode == "RGBA":
        data = pil_image.tobytes("raw", "RGBA")
        fmt = QImage.Format_RGBA8888
    else:
        if pil_image.mode != "RGB":
            pil_image = pil_image.convert("RGB")
        data = pil_image.tobytes("raw", "RGB")
        fmt = QImage.Format_RGB888
    qimg = QImage(data, pil_image.width, pil_image.height, fmt).copy()
    return QPixmap.fromImage(qimg)


# 管理密码：原本硬编码 "123" 散落在删除入库编号流程（_context_delete_voucher /
# _context_batch_delete_vouchers）。集中为一处常量便于维护，并供「用 Excel 打开
# 数据文件」复用同一道密码门。值不变，行为完全兼容。
ADMIN_PASSWORD = "123"


def _open_path(path: Path) -> None:
    # 规范化软件设计 2026-05 P1 审查修复:
    # 旧:Windows 用 shell=True + "cmd /c start" — shell 元字符 (& | > 等) 在路径里会被解释为命令。
    # 现:用 os.startfile (Windows 原生 API,不走 shell);失败兜底走 explorer。
    try:
        if os.name == "nt":
            # os.startfile 是 Windows 标准 API,不走 cmd shell,路径含特殊字符也安全
            os.startfile(str(path))
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])  # shell=False (默认)
        else:
            subprocess.Popen(["xdg-open", str(path)])  # shell=False (默认)
    except Exception as exc:
        QMessageBox.critical(None, "打开失败", str(exc))


def open_image_external(path: Path) -> None:
    """用外部程序打开原图:设置了自定义图片查看器就用它,否则回退系统默认程序。

    用户要求「打开原图」走外部程序(避免在主界面内载入全分辨率导致卡顿),
    并可在设置里自定义查看器(settings.image_viewer_path)。

    规范化软件设计 2026-05 P1 审查修复:加 viewer 路径校验
    - 必须绝对路径(防 PATH 注入)
    - 必须文件存在
    - Linux/Mac 必须可执行 (os.access X_OK)
    - 不能含 shell 元字符(防绕过 list 形式 subprocess.Popen,极保守)
    """
    viewer = (load_settings().image_viewer_path or "").strip()
    if viewer:
        viewer_path = Path(viewer)
        valid = (
            viewer_path.is_absolute()
            and viewer_path.exists()
            and viewer_path.is_file()
        )
        if valid and sys.platform != "win32":
            valid = os.access(str(viewer_path), os.X_OK)
        if valid:
            try:
                subprocess.Popen([str(viewer_path), str(path)])  # list 形式, shell=False
                return
            except Exception as exc:
                QMessageBox.critical(None, "打开失败", f"自定义图片查看器启动失败:{exc}\n将改用系统默认程序。")
        elif viewer:
            # 设置了但路径无效 → 提示一次,继续回退
            QMessageBox.warning(None, "查看器路径无效", f"设置的查看器路径无效,改用系统默认:{viewer}")
    # 未设置 / 查看器无效 / 启动失败 -> 回退系统默认。
    _open_path(path)


# ---------------------------------------------------------------------------
# 全局界面字体缩放
# ---------------------------------------------------------------------------
# run_app 启动时记录系统默认字号，用于"恢复默认"/Ctrl+0 复位。
_default_app_font_point: int = 0


def apply_app_font_size(size: int) -> None:
    """按绝对 pt 设置全局 app 字体；size<=0 时恢复系统默认字号。

    通过 QApplication.setFont 影响所有未显式 setFont/无 stylesheet font-size 的控件。
    """
    app = QApplication.instance()
    if app is None:
        return
    font = app.font()
    target = size if size and size > 0 else (_default_app_font_point or font.pointSize())
    if target and target > 0:
        font.setPointSize(target)
    app.setFont(font)


def apply_app_cursor(style: str) -> None:
    """按 style 给所有顶层窗口设趣味光标；"default"/未知 -> unsetCursor 恢复系统箭头。

    刻意用逐顶层窗口 setCursor，而非 QApplication.setOverrideCursor —— 后者是栈式
    全局覆盖，会把文本框 I-beam、splitter 拉伸、忙碌光标全盖掉。setCursor 只替换
    "默认箭头"出现处，QLineEdit / QSplitter 等自带光标的控件照常工作。
    """
    app = QApplication.instance()
    if app is None:
        return
    from .cursors import make_cursor
    cursor = make_cursor(style)
    for widget in app.topLevelWidgets():
        if cursor is None:
            widget.unsetCursor()
        else:
            widget.setCursor(cursor)


def apply_app_icon(variant: str) -> None:
    """按 variant 给 QApplication 及所有顶层窗口设应用图标。

    找不到对应素材时 get_app_icon 自动回退程序生成的经典图标（不会崩）。
    """
    app = QApplication.instance()
    if app is None:
        return
    icon = get_app_icon(variant)
    app.setWindowIcon(icon)
    for widget in app.topLevelWidgets():
        widget.setWindowIcon(icon)


def _species_matcher() -> SpeciesMatcher:
    """用软件自带的分类预设建 SpeciesMatcher。

    旧逻辑：读 workspace_root/字段模版/表格信息预设字段.xlsx —— 工作区缺该文件时
    种名/科名自动匹配静默失效。现改读**软件自带**的 specimen_app/字段模版/ 下预设
    （随软件分发），与工作区无关。找不到时给个不存在路径占位，SpeciesMatcher 会
    优雅空载（_rows=[]）。
    """
    from .field_help import bundled_template_path
    preset = bundled_template_path("表格信息预设字段.xlsx")
    if preset:
        _startup_mark(f"taxonomy preset found: {preset}")
    else:
        _startup_mark("taxonomy preset: NOT FOUND — species autofill will be disabled")
    return SpeciesMatcher(preset or Path("__missing_taxonomy_preset__.xlsx"))


# ---------------------------------------------------------------------------
# 工具栏 / 快捷键注册表（规范化软件设计 2026-05 新增）
# ---------------------------------------------------------------------------
# 工具栏 action 注册表：action_id -> (display_label, slot_method_name, category, tooltip)
# - action_id 是稳定字符串，settings.json 持久化用这个 key，不能轻易改名
# - slot_method_name 是 SpecimenWindow 上的方法名（运行时 getattr 取）
# - category 决定主栏内默认分组顺序：file -> edit -> view -> tools；同 category 内按声明顺序
# - 自动保存（_auto_save_action）是 checkable QAction，特殊处理；不进本注册表
TOOLBAR_ACTIONS: dict[str, dict] = {
    # icon: QStyle.StandardPixmap 枚举名（_rebuild_toolbars 时用 self.style().standardIcon() 取）。
    # 选 Qt 自带的 StandardPixmap 既零资源依赖、跨平台样式一致，又能立刻给视觉锚点。
    "import_workspace":  {"label": "导入工作区", "slot": "import_workspace",       "category": "file",
                          "icon": "SP_DirOpenIcon",
                          "tooltip": "从外部目录导入工作区数据（基于指纹去重）"},
    "import_data":       {"label": "导入数据",   "slot": "import_data_file",       "category": "file",
                          "icon": "SP_FileDialogStart",
                          "tooltip": "从指定目录扫描照片并按文件名匹配凭证"},
    "export_data":       {"label": "导出数据",   "slot": "export_data",            "category": "file",
                          "icon": "SP_DialogSaveButton",
                          "tooltip": "把当前工作区数据导出"},
    "batch_export":      {"label": "批量导出",   "slot": "open_batch_export",      "category": "file",
                          "icon": "SP_FileDialogDetailedView",
                          "tooltip": "按入库编号列表批量导出（仿 NCBI Batch Entrez）"},
    "switch_workspace":  {"label": "切换工作区", "slot": "switch_workspace",       "category": "file",
                          "icon": "SP_DirIcon",
                          "tooltip": "切换当前工作区或新建工作区"},
    "undo":              {"label": "撤回",       "slot": "undo",                   "category": "edit",
                          "icon": "SP_ArrowBack",
                          "tooltip": "撤回上一次操作（Ctrl+Z）"},
    "redo":              {"label": "返回",       "slot": "redo",                   "category": "edit",
                          "icon": "SP_ArrowForward",
                          "tooltip": "重做（Ctrl+Y / Ctrl+Shift+Z）"},
    "clear_photos":      {"label": "清除照片关联","slot": "clear_photos",          "category": "edit",
                          "icon": "SP_DialogDiscardButton",
                          "tooltip": "取消当前凭证下全部照片的关联（不删原文件）"},
    "ingest_summary":    {"label": "入库汇总",   "slot": "open_ingest_summary",    "category": "view",
                          "icon": "SP_FileDialogListView",
                          "tooltip": "打开入库汇总宽表（非模态单实例）"},
    "version_manager":   {"label": "版本管理",   "slot": "open_version_manager",   "category": "view",
                          "icon": "SP_BrowserReload",
                          "tooltip": "查看 / 还原数据版本快照、检查软件更新"},
    "settings":          {"label": "设置",       "slot": "open_settings",          "category": "tools",
                          "icon": "SP_ComputerIcon",
                          "tooltip": "应用设置（界面字体、光标样式、自动保存等）"},
    "worms":             {"label": "WoRMS",      "slot": "_open_worms_match",      "category": "worms",
                          "icon": "SP_TitleBarShadeButton",
                          "tooltip": "WoRMS 物种分类匹配窗口（单实例）"},
    # 规范化软件设计 2026-05 Phase 4 (P4) 扩展:WoRMS / 入库 / 视图 子功能可拖入工具栏
    "worms_browse":      {"label": "WoRMS 查询", "slot": "_open_worms_browse",     "category": "worms",
                          "icon": "SP_FileDialogContentsView",
                          "tooltip": "WoRMS 查询(按学名搜索;原 分类浏览)"},
    "worms_db":          {"label": "WoRMS 数据库","slot": "_open_worms_db_manager", "category": "worms",
                          "icon": "SP_DriveHDIcon",
                          "tooltip": "WoRMS 本地数据库管理(导入/统计/清空)"},
    "persons_manager":   {"label": "入库人员管理","slot": "_open_persons_manager",  "category": "ingest",
                          "icon": "SP_DialogYesButton",
                          "tooltip": "入库人员库 + 工作量统计 + 任务明细 + 编号分发"},
    "workload_report":   {"label": "入库人员记录","slot": "_open_workload_report",  "category": "ingest",
                          "icon": "SP_FileDialogDetailedView",
                          "tooltip": "入库人员工作量统计 (PersonsManagerDialog Tab 2)"},
    "series_manager":    {"label": "编号系列",   "slot": "_open_series_manager",   "category": "ingest",
                          "icon": "SP_DialogApplyButton",
                          "tooltip": "入库编号系列管理(新增/编辑/删除)"},
    "batch_generate":    {"label": "批量生成编号","slot": "_open_batch_generate",   "category": "edit",
                          "icon": "SP_ArrowRight",
                          "tooltip": "批量预留入库编号段 + 导出 xlsx/csv"},
    "new_window":        {"label": "新建工作区窗口","slot": "_open_new_workspace_window", "category": "view",
                          "icon": "SP_FileDialogNewFolder",
                          "tooltip": "选另一工作区开新窗口(各自有锁)"},
    "readonly_clone":    {"label": "只读副本",   "slot": "_open_readonly_clone",   "category": "view",
                          "icon": "SP_DialogCancelButton",
                          "tooltip": "本工作区只读副本(允多个同时打开,不抢锁)"},
    "data_snapshot":     {"label": "数据快照",   "slot": "open_version_manager",   "category": "tools",
                          "icon": "SP_DialogSaveAllButton",
                          "tooltip": "数据版本管理 + 检查 GitHub 更新"},
    "excel_open":        {"label": "Excel 打开数据","slot": "_open_data_in_excel",  "category": "tools",
                          "icon": "SP_FileDialogStart",
                          "tooltip": "用 Excel 打开工作区数据文件(密码门控)"},
    # Phase 5: 手动添加入库编号 + 规则推断 + 批量生成
    "manual_voucher":    {"label": "手动添加编号", "slot": "_open_manual_voucher",  "category": "edit",
                          "icon": "SP_FileDialogNewFolder",
                          "tooltip": "手动输入入库编号(单/多条);≥2 个自动识别规则可批量生成"},
    # 升级中心 v0.8.0 (D1-D20):一级 升级 菜单 + 升级中心 dialog 7 tab。
    # 顶层 5 项(主流对齐 Claude Code/VSCode 简洁风),其余 6 项进 高级 子菜单。
    "oneclick_upgrade":  {"label": "立即升级", "slot": "_oneclick_upgrade_now", "category": "upgrade",
                          "icon": "SP_ArrowUp",
                          "tooltip": "一键升级:后台 check + download + 弹窗问重启 → 一气呵成"},
    "check_update_now":  {"label": "检查更新", "slot": "_check_update_now", "category": "upgrade",
                          "icon": "SP_BrowserReload",
                          "tooltip": "检查 GitHub 最新版本，发现新版打开升级中心"},
    "upgrade_settings":  {"label": "自动更新设置", "slot": "_open_upgrade_settings", "category": "upgrade",
                          "icon": "SP_ComputerIcon",
                          "tooltip": "自动升级模式 / channel / 间隔 / 保留版本数"},
    "upgrade_about":     {"label": "关于当前版本", "slot": "_open_upgrade_about", "category": "upgrade",
                          "icon": "SP_DialogHelpButton",
                          "tooltip": "当前版本号、安装方式、构建信息"},
    "upgrade_center":    {"label": "升级中心", "slot": "_open_upgrade_center", "category": "upgrade",
                          "icon": "SP_ArrowUp",
                          "tooltip": "升级中心:检查/下载/导入/历史/设置一站式 dialog"},
    "install_from_zip":  {"label": "从本地文件安装更新", "slot": "_install_from_zip", "category": "upgrade",
                          "icon": "SP_DirOpenIcon",
                          "tooltip": "选 zip 文件离线安装(同事 U 盘转运场景)"},
    "download_installer": {"label": "下载安装包供分发", "slot": "_download_installer", "category": "upgrade",
                          "icon": "SP_DialogSaveButton",
                          "tooltip": "下载 Windows/Linux 安装包到指定文件夹,不安装"},
    "upgrade_history":   {"label": "历史版本管理", "slot": "_open_upgrade_history", "category": "upgrade",
                          "icon": "SP_FileDialogDetailedView",
                          "tooltip": "已安装的旧版本列表 + 设为当前(回滚)"},
    "upgrade_build_remote": {"label": "远程触发 GitHub 构建", "slot": "_open_upgrade_build_remote", "category": "upgrade",
                          "icon": "SP_ArrowForward",
                          "tooltip": "通过 GitHub Actions workflow_dispatch 远程构建(v0.8.1)"},
    "upgrade_build_local": {"label": "本地重新打包并安装", "slot": "_open_upgrade_build_local", "category": "upgrade",
                          "icon": "SP_FileDialogContentsView",
                          "tooltip": "(开发者模式) 本地 PyInstaller 重打包 + 自动 debug(v0.8.1)"},
}

# 默认布局：用户未自定义时（settings.toolbar_layout 为空）用此。
# 规范化软件设计 2026-05 起主栏只放 7 项高频按钮（用户反馈 13 项太挤）；
# 低频按钮进辅栏（默认隐藏，视图菜单可勾选打开）。Photoshop / IDE 一类紧凑工具栏惯例。
TOOLBAR_DEFAULT_LAYOUT: dict[str, list[str]] = {
    "main": [
        "import_workspace", "import_data", "batch_export",
        "undo", "redo",
        "ingest_summary",
        "worms",
    ],
    "aux": [
        "export_data", "switch_workspace",
        "clear_photos",
        "version_manager",
        "settings",
    ],
}

# 可自定义快捷键的 action 注册表（E 项）：action_id -> (display_label, default_keyseq, slot_method_name)
# 重叠工具栏 action：工具栏 action 也能绑定全局 shortcut（一份注册表两种用途）。
SHORTCUTABLE_ACTIONS: dict[str, dict] = {
    "undo":               {"label": "撤回",                "default": "Ctrl+Z",        "slot": "undo"},
    "redo":               {"label": "重做",                "default": "Ctrl+Y",        "slot": "redo"},
    "select_all_voucher": {"label": "全选凭证",            "default": "Ctrl+A",        "slot": "_select_all_vouchers"},
    "fit_image":          {"label": "照片适配窗口",        "default": "F",             "slot": "fit_image"},
    "return_to_grid":     {"label": "返回照片网格",        "default": "Esc",           "slot": "return_to_grid"},
    "zoom_in":            {"label": "界面字体放大",        "default": "Ctrl+=",        "slot": "_zoom_font_in"},
    "zoom_out":           {"label": "界面字体缩小",        "default": "Ctrl+-",        "slot": "_zoom_font_out"},
    "zoom_reset":         {"label": "界面字体复位",        "default": "Ctrl+0",        "slot": "_zoom_font_reset"},
    "photo_filename_fill":{"label": "从照片文件名填充",    "default": "Ctrl+Alt+F",    "slot": "fill_current_photo_from_filename"},
    "ingest_summary":     {"label": "打开入库汇总",        "default": "",              "slot": "open_ingest_summary"},
    "batch_export":       {"label": "打开批量导出",        "default": "",              "slot": "open_batch_export"},
    "worms":              {"label": "打开 WoRMS 匹配",     "default": "",              "slot": "_open_worms_match"},
    "user_manual":        {"label": "打开使用说明",        "default": "F1",            "slot": "_open_user_manual_dialog"},
}


# ---------------------------------------------------------------------------
# Background thumbnail loader (QThread + signal — no polling)
# ---------------------------------------------------------------------------

class _ThumbnailRequest:
    __slots__ = ("path", "size", "token")

    def __init__(self, path: Path, size: tuple[int, int], token: int):
        self.path = path
        self.size = size
        self.token = token


class ThumbnailWorker(QThread):
    result_ready = pyqtSignal(int, object, object)  # token, QPixmap|None, Exception|None

    def __init__(self, cache: ThumbnailCache, parent=None, max_workers: int | None = None):
        super().__init__(parent)
        self.cache = cache
        self._queue: queue.Queue[_ThumbnailRequest | None] = queue.Queue()
        self._running = True
        # 原值 max_workers=4:最多 4 张超大图同时解码,峰值内存可达数 GB,易拖垮整机。
        # 降到 2,限制并发解码的内存峰值(解码内存上限另见 image_cache._MAX_DECODE_PIXELS)。
        # 规范化软件设计 2026-05 内存档位:max_workers 由 memory_profile 驱动 (1/2/4)。
        # settings 未配置或异常 fallback 到 is_low_memory 二档。
        if max_workers is None:
            try:
                from .app_settings import load_settings
                from .env_detect import memory_profile_params
                profile = load_settings().memory_profile
                max_workers = memory_profile_params(profile)["thumb_workers"]
            except Exception:
                try:
                    from .env_detect import is_low_memory
                    max_workers = 1 if is_low_memory() else 2
                except Exception:
                    max_workers = 2
        self._pool = ThreadPoolExecutor(max_workers=max(1, int(max_workers)), thread_name_prefix="thumb")

    def enqueue(self, path: Path, size: tuple[int, int], token: int) -> None:
        self._queue.put(_ThumbnailRequest(path, size, token))
        if not self.isRunning():
            self.start()

    def clear_pending(self) -> None:
        while True:
            try:
                req = self._queue.get_nowait()
            except queue.Empty:
                break
            if req is None:
                self._queue.put(None)
                break

    def stop(self) -> None:
        # 规范化软件设计 2026-05 P1 审查修复:
        # 旧:_running=False → clear_pending → sentinel → pool.shutdown(wait=False) → self.wait(3000)
        #     但 thread 内 run() 仍可能在 sentinel 之前提交新 _process → pool 已 shutdown → RuntimeError
        # 现:先等 thread 退出(self.wait),再 shutdown pool。run() 取到 sentinel/_running=False 后退出,
        #     不会再 submit;pool.shutdown 此刻就安全。
        self._running = False
        self.clear_pending()
        self._queue.put(None)  # sentinel to unblock get()
        # 先等 thread 主循环退出(run()),保证不再 submit 到 pool
        self.wait(3000)
        # pool 现在可安全 shutdown(wait=False:已提交的解码任务自行结束,不阻塞)
        self._pool.shutdown(wait=False)

    def run(self) -> None:
        while self._running:
            try:
                req = self._queue.get(timeout=0.5)
            except queue.Empty:
                continue
            if req is None:
                break
            # P1 审查修复:pool 已 shutdown 时 submit 会 RuntimeError → 兜底捕获
            try:
                self._pool.submit(self._process, req)
            except RuntimeError:
                break  # pool 已关闭,正常退出循环

    def _process(self, req: _ThumbnailRequest) -> None:
        try:
            pil_img = self.cache.thumbnail(req.path, req.size)
            qpixmap = pil_to_qpixmap(pil_img)
            self.result_ready.emit(req.token, qpixmap, None)
        except Exception as exc:
            self.result_ready.emit(req.token, None, exc)


# ---------------------------------------------------------------------------
# Background workers for GitHub update check / download (QThread + signal)
# ---------------------------------------------------------------------------

class UpdateCheckWorker(QThread):
    # LatestRelease|None, Exception|None
    finished_check = pyqtSignal(object, object)

    def run(self) -> None:
        try:
            self.finished_check.emit(check_latest_release(), None)
        except Exception as exc:  # 网络/解析错误统一回传，UI 决定是否提示
            self.finished_check.emit(None, exc)


class UpdateDownloadWorker(QThread):
    progress = pyqtSignal(int)               # 下载进度百分比
    # Path|None, incremental(bool), Exception|None
    finished_download = pyqtSignal(object, object, object)

    def __init__(self, release, dest_root, local_roots=None, parent=None):
        super().__init__(parent)
        self._release = release
        self._dest_root = dest_root
        self._local_roots = local_roots or []

    def run(self) -> None:
        try:
            # download_update 尽量走增量（只下应用包）；老 release 自动回退完整 zip。
            path, incremental = download_update(
                self._release, self._dest_root, self._local_roots, self.progress.emit
            )
            self.finished_download.emit(path, incremental, None)
        except Exception as exc:
            self.finished_download.emit(None, False, exc)


# ---------------------------------------------------------------------------
# Photo preview — single mode (QGraphicsView with zoom/pan/drag-drop)
# ---------------------------------------------------------------------------

class PhotoGraphicsView(QGraphicsView):
    photo_dropped = pyqtSignal(list)  # list of str paths
    zoom_changed = pyqtSignal(float)  # zoom level (1.0 = 100%)
    return_requested = pyqtSignal()  # 由宫格双击展开的单张视图里，图上再双击 -> 请求返回宫格
    context_requested = pyqtSignal(object)  # 右键单张预览 -> 携带 globalPos，由主窗口出菜单

    _ZOOM_MIN = 0.05
    _ZOOM_MAX = 20.0
    _ZOOM_STEP = 1.15

    def __init__(self, parent=None):
        super().__init__(parent)
        self._scene = QGraphicsScene(self)
        self.setScene(self._scene)
        self._pixmap_item: QGraphicsPixmapItem | None = None
        self._zoom_level = 1.0
        self._is_fit_mode = True
        # True 时（从宫格双击展开的单张状态）双击图片 -> 发 return_requested 返回宫格，
        # 而不是做 适配↔100% 切换。由 SpecimenWindow 在展开/返回宫格/切模式时设置。
        self._double_click_returns = False
        self.setRenderHints(QPainter.Antialiasing | QPainter.SmoothPixmapTransform)
        self.setDragMode(QGraphicsView.ScrollHandDrag)
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.AnchorViewCenter)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._normal_style = "background-color: #d8d8d8; border: 1px solid #667;"
        self._drag_style = "background-color: #d8d8d8; border: 2px dashed #2a6fbd;"
        self.setStyleSheet(self._normal_style)
        self.setAcceptDrops(True)

    def set_image(self, pixmap: QPixmap) -> None:
        self._scene.clear()
        self._pixmap_item = self._scene.addPixmap(pixmap)
        self._scene.setSceneRect(self._pixmap_item.boundingRect())
        self.fitInView(self._pixmap_item, Qt.KeepAspectRatio)
        self._is_fit_mode = True
        self._zoom_level = self._current_zoom()

    def clear_image(self) -> None:
        self._scene.clear()
        self._pixmap_item = None
        self.resetTransform()
        self._zoom_level = 1.0
        self._is_fit_mode = True

    def fit_to_window(self) -> None:
        if self._pixmap_item:
            self.resetTransform()
            self.fitInView(self._pixmap_item, Qt.KeepAspectRatio)
            self._is_fit_mode = True
            self._zoom_level = self._current_zoom()
            self.zoom_changed.emit(self._zoom_level)

    def zoom(self, factor: float) -> None:
        new_level = self._zoom_level * factor
        if self._ZOOM_MIN <= new_level <= self._ZOOM_MAX:
            self._zoom_level = new_level
            self._is_fit_mode = False
            self.scale(factor, factor)
            self.zoom_changed.emit(self._zoom_level)

    def zoom_percent(self) -> int:
        return int(self._zoom_level * 100)

    def _current_zoom(self) -> float:
        if not self._pixmap_item:
            return 1.0
        return self.transform().m11()

    def wheelEvent(self, event) -> None:
        factor = self._ZOOM_STEP if event.angleDelta().y() > 0 else 1 / self._ZOOM_STEP
        new_level = self._zoom_level * factor
        if self._ZOOM_MIN <= new_level <= self._ZOOM_MAX:
            self._zoom_level = new_level
            self._is_fit_mode = False
            self.scale(factor, factor)
            self.zoom_changed.emit(self._zoom_level)
        event.accept()

    def set_double_click_returns(self, enabled: bool) -> None:
        self._double_click_returns = bool(enabled)

    def contextMenuEvent(self, event) -> None:
        # 单张预览原本没有右键菜单；改为发信号让主窗口出照片管理菜单。
        self.context_requested.emit(event.globalPos())
        event.accept()

    def mouseDoubleClickEvent(self, event) -> None:
        if event.button() != Qt.LeftButton or not self._pixmap_item:
            return super().mouseDoubleClickEvent(event)
        # 从宫格双击展开的单张状态：图上双击 -> 返回宫格（不用找「返回网格」按钮）。
        if self._double_click_returns:
            self.return_requested.emit()
            event.accept()
            return
        # 旧逻辑（保留）：普通单张状态双击在 适配↔100% 间切换。
        if self._is_fit_mode:
            self.resetTransform()
            self._zoom_level = 1.0
            self._is_fit_mode = False
            self.centerOn(self._pixmap_item)
        else:
            self.fit_to_window()
        self.zoom_changed.emit(self._zoom_level)

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        if self._is_fit_mode and self._pixmap_item:
            self.resetTransform()
            self.fitInView(self._pixmap_item, Qt.KeepAspectRatio)
            self._zoom_level = self._current_zoom()
            self.zoom_changed.emit(self._zoom_level)

    # -- Drag-drop for photos --
    def dragEnterEvent(self, event) -> None:
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.setStyleSheet(self._drag_style)

    def dragMoveEvent(self, event) -> None:
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dragLeaveEvent(self, event) -> None:
        self.setStyleSheet(self._normal_style)

    def dropEvent(self, event) -> None:
        self.setStyleSheet(self._normal_style)
        paths = []
        for url in event.mimeData().urls():
            if url.isLocalFile():
                paths.append(url.toLocalFile())
        if paths:
            self.photo_dropped.emit(paths)
        event.acceptProposedAction()


# ---------------------------------------------------------------------------
# Interactive image cells for grid mode
# ---------------------------------------------------------------------------

class GridPhotoView(QGraphicsView):
    clicked = pyqtSignal()
    double_clicked = pyqtSignal()
    right_clicked = pyqtSignal(object)
    zoom_changed = pyqtSignal(float)

    _ZOOM_MIN = 0.2
    _ZOOM_MAX = 8.0
    _ZOOM_STEP = 1.15

    def __init__(self, parent=None):
        super().__init__(parent)
        self._scene = QGraphicsScene(self)
        self.setScene(self._scene)
        self._pixmap_item: QGraphicsPixmapItem | None = None
        self._zoom_level = 1.0
        self._is_fit_mode = True
        self.setRenderHints(QPainter.SmoothPixmapTransform)
        self.setDragMode(QGraphicsView.ScrollHandDrag)
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.AnchorViewCenter)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setStyleSheet("background-color: #e8edf0; border: none;")

    def set_pixmap(self, pixmap: QPixmap) -> None:
        self._scene.clear()
        self._pixmap_item = self._scene.addPixmap(pixmap)
        self._scene.setSceneRect(self._pixmap_item.boundingRect())
        self.fit_to_window()

    def show_message(self, text: str, error: bool = False) -> None:
        self._scene.clear()
        self._pixmap_item = None
        self.resetTransform()
        self._zoom_level = 1.0
        self._is_fit_mode = True
        item = self._scene.addText(text)
        item.setDefaultTextColor(Qt.darkRed if error else Qt.darkGray)
        rect = item.boundingRect()
        item.setPos(-rect.width() / 2, -rect.height() / 2)
        self._scene.setSceneRect(-120, -60, 240, 120)

    def fit_to_window(self) -> None:
        if not self._pixmap_item:
            return
        self.resetTransform()
        self.fitInView(self._pixmap_item, Qt.KeepAspectRatio)
        self._zoom_level = 1.0
        self._is_fit_mode = True
        self.zoom_changed.emit(self._zoom_level)

    def zoom(self, factor: float) -> None:
        if not self._pixmap_item:
            return
        new_level = self._zoom_level * factor
        if self._ZOOM_MIN <= new_level <= self._ZOOM_MAX:
            self._zoom_level = new_level
            self._is_fit_mode = False
            self.scale(factor, factor)
            self.zoom_changed.emit(self._zoom_level)

    def zoom_percent(self) -> int:
        return int(self._zoom_level * 100)

    def wheelEvent(self, event) -> None:
        factor = self._ZOOM_STEP if event.angleDelta().y() > 0 else 1 / self._ZOOM_STEP
        self.zoom(factor)
        event.accept()

    def mousePressEvent(self, event) -> None:
        if event.button() == Qt.LeftButton:
            self.clicked.emit()
        super().mousePressEvent(event)

    def mouseDoubleClickEvent(self, event) -> None:
        if event.button() == Qt.LeftButton:
            self.double_clicked.emit()
            event.accept()
            return
        super().mouseDoubleClickEvent(event)

    def contextMenuEvent(self, event) -> None:
        self.right_clicked.emit(event)
        event.accept()

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        if self._is_fit_mode and self._pixmap_item:
            self.fit_to_window()


_GRID_CELL_NORMAL_STYLE = "QFrame#GridPhotoCell { background-color: #eef2f3; border: 2px solid #7a858a; }"
_GRID_CELL_SELECTED_STYLE = "QFrame#GridPhotoCell { background-color: #eef2f3; border: 2px solid #2a6fbd; }"
_GRID_FILENAME_STYLE = "color: #263238; background-color: #dfe7eb; padding: 1px 4px; font-size: 11px;"


class GridPhotoCell(QFrame):
    clicked = pyqtSignal(int)
    double_clicked = pyqtSignal(int)
    right_clicked = pyqtSignal(int, object)
    zoom_changed = pyqtSignal(int, float)

    def __init__(self, index: int, show_filename: bool, parent=None):
        super().__init__(parent)
        self._index = index
        self._filename = ""
        self.setObjectName("GridPhotoCell")
        self.setStyleSheet(_GRID_CELL_NORMAL_STYLE)
        self.setMinimumSize(80, 90)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(3, 3, 3, 3)
        layout.setSpacing(2)

        self._name_label = QLabel()
        self._name_label.setAlignment(Qt.AlignCenter)
        self._name_label.setFixedHeight(20)
        self._name_label.setStyleSheet(_GRID_FILENAME_STYLE)
        self._name_label.setVisible(show_filename)
        layout.addWidget(self._name_label)

        self._view = GridPhotoView()
        self._view.clicked.connect(lambda: self.clicked.emit(self._index))
        self._view.double_clicked.connect(lambda: self.double_clicked.emit(self._index))
        self._view.right_clicked.connect(lambda event: self.right_clicked.emit(self._index, event))
        self._view.zoom_changed.connect(lambda zoom: self.zoom_changed.emit(self._index, zoom))
        layout.addWidget(self._view, stretch=1)

        self.set_loading()

    @property
    def photo_index(self) -> int:
        return self._index

    def set_filename(self, filename: str) -> None:
        self._filename = filename
        self._name_label.setToolTip(filename)
        self._update_filename_label()

    def set_filename_visible(self, visible: bool) -> None:
        self._name_label.setVisible(visible)
        self._update_filename_label()

    def set_selected(self, selected: bool) -> None:
        self.setStyleSheet(_GRID_CELL_SELECTED_STYLE if selected else _GRID_CELL_NORMAL_STYLE)

    def set_loading(self) -> None:
        self._view.show_message("加载中")

    def set_error(self, text: str) -> None:
        self._view.show_message(text, error=True)

    def set_pixmap(self, pixmap: QPixmap) -> None:
        self._view.set_pixmap(pixmap)

    def zoom(self, factor: float) -> None:
        self._view.zoom(factor)

    def fit_to_window(self) -> None:
        self._view.fit_to_window()

    def zoom_percent(self) -> int:
        return self._view.zoom_percent()

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        self._update_filename_label()

    def mousePressEvent(self, event) -> None:
        super().mousePressEvent(event)

    def mouseDoubleClickEvent(self, event) -> None:
        if event.button() == Qt.LeftButton:
            self.double_clicked.emit(self._index)
            event.accept()
            return
        super().mouseDoubleClickEvent(event)

    def contextMenuEvent(self, event) -> None:
        self.right_clicked.emit(self._index, event)
        event.accept()

    def _update_filename_label(self) -> None:
        if not self._filename:
            self._name_label.setText("")
            return
        width = max(50, self._name_label.width() - 8)
        text = QFontMetrics(self._name_label.font()).elidedText(self._filename, Qt.ElideMiddle, width)
        self._name_label.setText(text)


# ---------------------------------------------------------------------------
# Main window
# ---------------------------------------------------------------------------

# 哨兵：表示启动时没有可自动解析的工作区，窗口应以"未绑定"状态先构建并显示，
# 工作区选择推迟到 show() 之后由 _prompt_initial_workspace 驱动。
_DEFERRED_WORKSPACE = object()


class SpecimenWindow(QMainWindow):
    # voucher_table 列宽/字体基准值（系统默认字号下的原始值）。
    # 全局字体缩放时按 (当前字号 - 默认字号) 的差值同比放大，避免文字被覆盖。
    _VOUCHER_COL_BASE_WIDTHS = (85, 36, 36, 36, 52, 42)
    _VOUCHER_TABLE_BASE_PT = 10

    def __init__(self, workspace_root: Path | str | None, manager: "WindowManager | None" = None,
                 read_only: bool = False):
        """初始化主窗口。

        规范化软件设计 2026-05 多窗口:read_only=True 时 ExcelStore 也走 read_only,
        标题加"(只读)",禁所有写入 UI(新增/编辑/删除/任务/导入)。
        """
        super().__init__()
        self.manager = manager
        self.read_only = bool(read_only)
        title = "标本入库管理(只读副本)" if self.read_only else "标本入库管理"
        self.setWindowTitle(title)
        # 旧：get_app_icon() 永远用程序生成图标。现按用户选的图标变体加载。
        self.setWindowIcon(get_app_icon(load_settings().app_icon_variant))
        self.resize(1320, 820)
        self.setMinimumSize(QSize(1100, 680))

        prepared = self._prepare_start_workspace(workspace_root)
        if prepared is None:
            self.close()
            raise SystemExit
        if prepared is _DEFERRED_WORKSPACE:
            # 首次启动、没有可自动解析的工作区：以"未绑定"状态构建窗口。
            # 旧逻辑：_prepare_start_workspace 在窗口构建前阻塞弹 QFileDialog（背后无窗口）。
            # 现改为 show() 之后由 _prompt_initial_workspace 驱动选择器。
            self.workspace_root = None
            self.store = None
            self.matcher = None
        else:
            self.workspace_root, create_workspace_files = prepared
            # 只读副本不抢已存在的主窗口焦点(允许同工作区多只读副本共存)
            if not self.read_only and self.manager is not None and self.manager.focus_workspace(self.workspace_root):
                self.close()
                raise SystemExit

            try:
                # 只读副本不锁工作区(不抢锁)
                lock_flag = not self.read_only
                self.store = ExcelStore(self.workspace_root, lock=lock_flag,
                                        create_if_missing=create_workspace_files,
                                        read_only=self.read_only)
            except WorkspaceLockedError as exc:
                lock_path = self.workspace_root / "数据" / ".workspace.lock"
                msg = f'{exc}\n\n如果软件已退出但仍然被占用，可以点击"强制解锁"。'
                btn = QMessageBox.critical(self, "工作区被占用", msg, QMessageBox.Abort | QMessageBox.Retry)
                if btn == QMessageBox.Retry and lock_path.exists():
                    try:
                        lock_path.unlink()
                        self.store = ExcelStore(self.workspace_root, lock=True,
                                                create_if_missing=create_workspace_files,
                                                read_only=self.read_only)
                    except Exception:
                        self.close()
                        raise SystemExit
                else:
                    self.close()
                    raise SystemExit from exc
            except WorkspaceNotInitializedError as exc:
                QMessageBox.critical(self, "工作区未初始化", str(exc))
                self.close()
                raise SystemExit from exc

            _startup_mark("SpecimenWindow: ExcelStore ready")
            self.matcher = _species_matcher()  # 旧：读 workspace_root/字段模版/，现读软件自带预设

        self.current_voucher: str | None = None
        self.current_photos: list[dict[str, str]] = []
        self.current_photo_index = 0
        self._loading = False
        self._is_closing = False
        self._photo_load_token = 0
        self._grid_load_token = 0
        self._import_job_active = False
        self._grid_mode_before_expand = ""
        self._photo_page_size = 200
        self._photo_page = 0
        self._photo_view_states: dict[str, tuple[float, int, int]] = {}
        self._show_grid_filenames = load_settings().show_grid_filenames

        # thumbnail_cache / _thumb_worker 依赖工作区路径。
        # 旧逻辑：在 __init__ 中无条件创建。现在首次启动可能没有工作区，
        # 改为有工作区时才创建；未绑定时由 _load_workspace_into_window 首次载入时创建。
        self.search_index: ImageSearchIndex | None = None
        self._index_build_worker: IndexBuildWorker | None = None
        if self.workspace_root is not None:
            self.thumbnail_cache = ThumbnailCache(self.workspace_root)
            self._thumb_worker = ThumbnailWorker(self.thumbnail_cache, self)
            self._thumb_worker.result_ready.connect(self._on_thumbnail_ready)
            # 规范化软件设计 2026-05 启动卡死优化:
            # 旧:__init__ 内立即 start(),后台线程跟主线程 Excel 读争 I/O+内存,2GB 机卡顿。
            # 现:_finish_initial_load 末尾 QTimer.singleShot(2000, ...) 延后 2s 才 start。
            # K 章 (2026-05) 高档位快路径:high / extra_high 跳过延迟,立刻 start —
            # 64GB 机不需要让主线程 I/O 独占 CPU,首次照片预览无 2s 等待。
            try:
                from .env_detect import is_fast_profile
                if is_fast_profile():
                    self._thumb_worker.start()
            except Exception:
                pass
        else:
            self.thumbnail_cache = None
            self._thumb_worker = None

        self.specimen_widgets: dict[str, QLineEdit | QComboBox] = {}
        self.class_widgets: dict[str, QLineEdit] = {}
        self.photo_widgets: dict[str, QLineEdit] = {}
        self._taxonomy_candidate_models: dict[str, QStandardItemModel] = {}
        self._taxonomy_candidate_rows: dict[str, list[tuple[str, SpeciesMatch | FamilyMatch]]] = {}

        self._save_timers: dict[str, QTimer] = {}
        # 自动保存开关：True=输入停 0.5s 自动写；False=只在点「保存」按钮时写。工具栏可勾选切换。
        self.auto_save_enabled = load_settings().auto_save_enabled
        self._list_refresh_timer = QTimer(self)
        self._list_refresh_timer.setSingleShot(True)
        self._list_refresh_timer.timeout.connect(self.refresh_list)
        self._view_mode: str = "单张"
        self._current_qpixmap: QPixmap | None = None
        self._grid_labels: list[GridPhotoCell] = []
        self._grid_requests: dict[int, int] = {}  # token -> slot index
        self._grid_selected_indices: set[int] = set()  # 宫格多选（Ctrl/Shift），photo 索引集合
        self._photo_filename_fill_action: QAction | None = None
        # 入库汇总窗口：非模态、单实例。open_ingest_summary 据此判断是新建还是聚焦刷新。
        self._ingest_summary_dialog: "IngestSummaryDialog | None" = None
        # WoRMS 匹配窗口：非模态、单实例（v0.5.0+，原 worms_dialog.py 模态对话框已替换）。
        self._worms_window: "object | None" = None  # WormsMatchWindow
        # 使用说明弹窗：非模态、单实例（规范化软件设计 2026-05 新增）。
        self._manual_dialog: "object | None" = None  # UserManualDialog

        # 录入任务状态：开始录入任务后填充，结束后清空。
        # keys: 记录ID, 人员, 用途, 备注, 开始时间, 新增数量
        self._active_task: dict | None = None

        # 搜索数据容器：必须在 _build_ui() 之前初始化，因为 UI 构造期间
        # QComboBox.currentIndexChanged 信号可能提前触发 _apply_voucher_filter。
        self._all_vouchers: list[str] = []
        self._all_tube_numbers: dict[str, str] = {}
        self._all_photo_filenames: dict[str, list[str]] = {}

        self._build_ui()
        _startup_mark("SpecimenWindow._build_ui")

        # 新建主窗口即应用当前趣味光标设置（与字体在窗口创建时继承同理）。
        apply_app_cursor(load_settings().cursor_style)

        if self.workspace_root is not None:
            remember_workspace(self.workspace_root)
            self.statusBar().showMessage("正在加载工作区数据...")
            QTimer.singleShot(0, self._finish_initial_load)
        else:
            # 未绑定：窗口已构建，show() 之后再驱动工作区选择器。
            self.statusBar().showMessage("请选择或新建工作区")
            QTimer.singleShot(0, self._prompt_initial_workspace)

    def _finish_initial_load(self) -> None:
        """启动后初始加载。拆 3 步 + QTimer.singleShot(0) 让事件循环穿插,UI 点击立即响应。

        规范化软件设计 2026-05 Phase 4 修:
        旧:单方法同步链 ~200-500ms 阻塞事件循环 → crash hint QMessageBox 按钮 / 主窗口 click
            响应延迟。
        现:Step1 refresh_list → yield → Step2 select_voucher → yield → Step3 后续 timer。
            每 yield 让 Qt 事件循环跑一次,处理 pending click / QMessageBox 交互。
        """
        if self._is_closing or self.store is None:
            return
        self.refresh_list()
        _startup_mark("_finish_initial_load: refresh_list")
        if self.read_only:
            self._apply_read_only_ui()
        # yield 事件循环 → step2
        QTimer.singleShot(0, self._finish_initial_load_step2)

    def _finish_initial_load_step2(self) -> None:
        if self._is_closing or self.store is None:
            return
        vouchers = self._all_vouchers
        if vouchers:
            self.select_voucher(vouchers[0], defer_preview=True)
        _startup_mark("_finish_initial_load: first voucher selected")
        # yield 事件循环 → step3
        QTimer.singleShot(0, self._finish_initial_load_step3)

    def _finish_initial_load_step3(self) -> None:
        if self._is_closing or self.store is None:
            return
        self.statusBar().showMessage("工作区已加载", 2000)
        # 分类预设缺失时显示持久黄色警告条（旧：8 秒状态栏消息，极易错过）。
        if self.matcher is not None and not list(self.matcher.all_rows()):
            self._preset_warning_banner.show()
        # 旧（v0.5.0 及以前）：启动后 1.2s 自动 _build_search_index_background()，
        # 用户多数从不开图片搜索却平白多吃 ~20MB + ~1.2s。
        # 现（规范化软件设计 2026-05 起）：图片索引改为**按需**——open_image_search() 内
        # 首次打开时 dlg 自带后台 worker 建索引。本启动钩子改为仅触发 gc.collect()
        # 强制回收 openpyxl 读 Excel 用的临时对象（zip / xml DOM 等），实测可省 10–30MB。
        QTimer.singleShot(200, self._post_load_gc)
        # D3+D11+D19 升级中心启动钩子链:
        # 1. arm sentinel 清除定时器(新版活过 30s 视为健康,清掉 sentinel)
        # 2. 1.5s: 检 pending_update 弹"立即安装并重启?"
        # 3. 2s:   检 post_update_sentinel 残留弹"上次启动失败,回退?"
        # 4. 2.5s: 4 档自动检查 GitHub 更新
        self._arm_post_update_sentinel()
        QTimer.singleShot(1500, self._apply_pending_update_on_startup)
        QTimer.singleShot(2000, self._check_post_update_sentinel_on_startup)
        QTimer.singleShot(2500, self._maybe_check_updates_on_startup)
        # 规范化软件设计 2026-05 K 章:高档位快路径检测
        try:
            from .env_detect import is_fast_profile
            fast = is_fast_profile()
        except Exception:
            fast = False
        # ThumbnailWorker 延后 start(规范化软件设计 2026-05 启动卡死优化):
        # __init__ 内不立即 start,延 2s 避免与主线程 Excel 读争 CPU/内存。
        # 高档位:__init__ 已立 start,这里跳过。
        if self._thumb_worker is not None and not self._thumb_worker.isRunning():
            QTimer.singleShot(2000, self._thumb_worker.start)
        # K 章 高档位:数据预热 + 图片索引预建 (低档位不做)
        if fast:
            _startup_mark("fast profile: scheduling preheat + search index prebuild")
            QTimer.singleShot(500, self._preheat_caches)
            QTimer.singleShot(1200, self._build_search_index_background)

    def _apply_read_only_ui(self) -> None:
        """规范化软件设计 2026-05 多窗口:只读副本禁所有写入 UI。

        - 任务相关按钮 disabled
        - "新增入库编号" disabled
        - 标本/分类/照片面板 输入控件 setReadOnly
        - 工具栏写按钮 disabled (撤回/返回/清除照片关联/导入工作区/导入数据)
        - 自动保存切 off (避免误触发)
        - 状态栏永久 banner 提示
        """
        # 标题已加 (只读),保险再设
        self.setWindowTitle(f"标本入库管理 (只读) — {self.workspace_root}")
        # 禁任务按钮
        for attr in ("_task_start_btn", "_task_end_btn", "_new_voucher_btn"):
            btn = getattr(self, attr, None)
            if btn is not None:
                btn.setEnabled(False)
                btn.setToolTip("只读副本:禁用写入操作。请在主窗口操作。")
        # 禁工具栏写 action(按 id 索引)
        write_action_ids = {"import_workspace", "import_data", "export_data",
                            "batch_export", "switch_workspace", "undo", "redo",
                            "clear_photos", "version_manager", "settings"}
        for aid, action in getattr(self, "_toolbar_actions", {}).items():
            if aid in write_action_ids:
                action.setEnabled(False)
                action.setToolTip("只读副本:禁用")
        # 自动保存切 off + 禁切换
        auto = getattr(self, "_auto_save_action", None)
        if auto is not None:
            auto.setChecked(False)
            auto.setEnabled(False)
        self.auto_save_enabled = False
        # 标本/分类/照片 input setReadOnly
        for w_dict_name in ("specimen_widgets", "class_widgets", "photo_widgets"):
            w_dict = getattr(self, w_dict_name, {})
            for w in w_dict.values():
                if hasattr(w, "setReadOnly"):
                    try:
                        w.setReadOnly(True)
                    except Exception:
                        pass
                if hasattr(w, "setEditable"):
                    try:
                        w.setEditable(False)
                    except Exception:
                        pass
        # 状态栏永久横幅
        try:
            ro_label = QLabel("🔒 只读副本 — 禁所有写入操作")
            ro_label.setStyleSheet("color: #c14d4d; font-weight: bold; padding: 0 8px;")
            self.statusBar().addPermanentWidget(ro_label)
        except Exception:
            pass

    def _preheat_caches(self) -> None:
        """规范化软件设计 2026-05 K 章高档位快路径:把 specimen/classification/photo 三表
        载入 _row_cache,后续工作区切换 / 汇总 / voucher 切换 0 重读。

        非阻塞:用 QThread 后台跑 read_rows(3)。高档位用户内存足够,这点 _row_cache 撑得起
        (LRU maxsize=12/20)。
        """
        if self._is_closing or self.store is None:
            return
        _startup_mark("preheat: start")
        try:
            # 直接同步预热:read_rows 走 _row_cache,3 张表读完 ~200-500ms,在 _finish_initial_load
            # 后 500ms 触发,不影响首次交互。如果工作区超大可改 QThread,目前足够。
            self.store.read_rows("specimen")
            self.store.read_rows("classification")
            self.store.read_rows("photo")
            _startup_mark("preheat: done (3 tables cached)")
        except Exception as exc:
            _startup_mark(f"preheat: failed ({exc})")

    def _post_load_gc(self) -> None:
        """启动初始加载后强制 gc，回收 openpyxl 读 Excel 时的临时对象。

        Python 的引用计数会立刻释放，但 gc 还要清环引用；openpyxl 内部对象常有循环引用，
        启动加载后调一次 gc.collect() 可缩短"高水位线 -> 稳态 RSS"的间隔。
        """
        try:
            import gc
            gc.collect()
        except Exception:
            pass
        _startup_mark("_finish_initial_load: post-load gc done")

    def _prompt_initial_workspace(self) -> None:
        if self._is_closing:
            return
        # 旧逻辑：__init__ 在窗口构建前阻塞弹 QFileDialog（背后无窗口）。
        # 现改为窗口已 show() 之后再驱动同一选择/校验/初始化/载入链路（switch_workspace）。
        self.switch_workspace()
        if self.store is None:
            # 用户取消了首次工作区选择：保留空窗口，工具栏"切换工作区"按钮仍可用。
            self.statusBar().showMessage("尚未选择工作区，可点击工具栏“切换工作区”随时打开", 0)

    # ---- workspace preparation ----

    def _prepare_start_workspace(self, workspace_root: Path | str | None) -> "tuple[Path, bool] | object | None":
        candidate = Path(workspace_root).resolve() if workspace_root else None
        # 旧逻辑：candidate 为 None 时在 while 循环里阻塞弹 QFileDialog 选目录。
        # 现改为：没有可自动解析的工作区时直接返回 _DEFERRED_WORKSPACE，
        # 由窗口 show() 之后的 _prompt_initial_workspace 驱动选择器（此时窗口可见）。
        if candidate is None:
            return _DEFERRED_WORKSPACE
        # 传入了明确的工作区（--workspace 或自动解析到的上次工作区）：仍按原逻辑校验，
        # 无效则返回 None -> __init__ 抛 SystemExit。
        return self._prepare_workspace_candidate(candidate)

    def _prepare_workspace_candidate(self, path: Path) -> tuple[Path, bool] | None:
        if is_generated_workspace_path(path):
            QMessageBox.critical(
                self, "不能使用软件目录",
                f"不能把 build/dist/releases 等软件构建或版本目录作为工作区：\n{path}\n\n请选择实际保存数据和照片的工作目录。",
            )
            return None
        # 文件系统根 / 盘符根 / 用户主目录范围过大，全工作区扫描会遍历海量文件、拖垮电脑。
        if is_unsafe_workspace_root(path):
            QMessageBox.critical(
                self, "目录范围过大",
                f"不能把文件系统根目录、盘符根目录或用户主目录作为工作区：\n{path}\n\n"
                "这类目录过大，软件的全工作区扫描（如图片索引）会遍历海量文件、可能拖垮电脑。\n"
                "请选择实际保存数据和照片的子目录。",
            )
            return None
        if has_workspace_data(path):
            return path, False
        if QMessageBox.question(
            self, "初始化工作区",
            f"该目录还没有数据文件：\n{path}\n\n是否在此目录创建新的数据文件夹和 Excel 数据文件？",
            QMessageBox.Yes | QMessageBox.No,
        ) != QMessageBox.Yes:
            return None
        initialize_workspace(path, self._template_source_for_initialization(path))
        return path, True

    def _template_source_for_initialization(self, target: Path) -> Path | None:
        candidates = [
            self.workspace_root if hasattr(self, "workspace_root") else None,
            target,
            Path.cwd(),
            Path(sys.executable).resolve().parent,
            Path(sys.executable).resolve().parent.parent,
            Path(sys.executable).resolve().parent.parent.parent,
            Path(__file__).resolve().parents[1],
            # specimen_app/ 自身：a35358b 已把 specimen_app/字段模版/ 提交进仓库作兜底
            # 模板源；打包后由 build_release.py --add-data 带进 _internal/specimen_app/。
            Path(__file__).resolve().parent,
        ]
        for candidate in candidates:
            if not candidate:
                continue
            source = Path(candidate).resolve()
            if (source / "字段模版").exists():
                return source
        return None

    # ---- close ----

    def closeEvent(self, event) -> None:
        self._is_closing = True
        # 先停定时器,防止关闭过程中回调触碰已被释放的 store。
        timer = getattr(self, "_list_refresh_timer", None)
        if timer is not None:
            timer.stop()
        # 规范化软件设计 2026-05 新增:停内存状态定时器。
        mem_timer = getattr(self, "_memory_status_timer", None)
        if mem_timer is not None:
            try:
                mem_timer.stop()
            except Exception:
                pass

        # E3 helper：wait 超时后强杀，保证主进程一定能退（特别是 Windows 上 QThread 卡死场景）。
        def _stop_worker(worker, wait_ms: int = 2000, label: str = "") -> None:
            if worker is None or not worker.isRunning():
                return
            try:
                worker.requestInterruption()
            except Exception:
                pass
            if not worker.wait(wait_ms) and worker.isRunning():
                try:
                    worker.terminate()
                    worker.wait(500)  # terminate 后给个短暂收尾窗口
                except Exception as exc:
                    print(f"[closeEvent] terminate {label} 失败：{exc}", file=sys.stderr)

        # Stop background index builder if running
        _stop_worker(getattr(self, "_index_build_worker", None), wait_ms=5000, label="index_builder")
        # Stop thumbnail worker
        thumb = getattr(self, "_thumb_worker", None)
        if thumb is not None:
            try:
                thumb.stop()
            except Exception:
                pass
        # Stop startup update check worker (network thread, may still be polling)
        _stop_worker(getattr(self, "_startup_update_worker", None), wait_ms=2000, label="update_check")
        # Stop photo import thread if one is running
        _stop_worker(getattr(self, "_import_thread", None), wait_ms=2000, label="photo_import")
        # C1: 接管所有注册到 WindowManager 的 dialog 的 worker（如 DbManagerDialog 的 worms worker）
        if self.manager is not None:
            try:
                self.manager.stop_all_dialog_workers(wait_ms=3000)
            except Exception as exc:
                print(f"[closeEvent] stop dialog workers 失败：{exc}", file=sys.stderr)
        try:
            settings = load_settings()
            settings.window_geometry = self.saveGeometry().toBase64().data().decode()
            settings.splitter_sizes = [
                [int(x) for x in self.main_splitter.sizes()],
                [int(x) for x in self.right_splitter.sizes()],
            ]
            settings.show_grid_filenames = getattr(self, "_show_grid_filenames", settings.show_grid_filenames)
            save_settings(settings)
        except OSError as exc:
            # 磁盘满 / 无权限等环境问题：保存窗口状态失败可容忍，不阻断关闭，但记录到 stderr。
            # 原代码 except Exception: pass 会静默吞掉一切异常（含真正的 bug）。
            print(f"[closeEvent] 保存窗口状态失败：{exc}", file=sys.stderr)
        except Exception as exc:
            print(f"[closeEvent] 保存窗口状态时发生异常：{exc}", file=sys.stderr)
        # 关闭时若有活动录入任务，静默结束并写入日志。
        if getattr(self, "_active_task", None) is not None:
            try:
                self._end_task()
            except Exception as exc:
                print(f"[closeEvent] 结束录入任务失败：{exc}", file=sys.stderr)
        # 关闭前把所有待写字段全部落盘，防止 500ms 防抖还未触发时关窗口丢数据。
        try:
            self._flush_pending_saves()
        except Exception as exc:
            print(f"[closeEvent] 刷写待保存数据失败：{exc}", file=sys.stderr)
        store = getattr(self, "store", None)
        if store is not None:
            try:
                store.close()
            except Exception:
                pass
        if self.manager is not None:
            self.manager.unregister(self)
        # E1: closeEvent 末尾 touch last_exit_clean marker，下次启动据此识别是否上次正常退出。
        try:
            from .crash_log import mark_app_exiting_clean
            mark_app_exiting_clean()
        except Exception:
            pass
        event.accept()

    def eventFilter(self, obj, event) -> bool:
        # 规范化软件设计 2026-05 P1 审查修复:用 getattr 防御,_build_ui 内 splash processEvents
        # 触发的早期事件可能在 self.photo_table / grid_frame 创建前到达,直接 self.x 会 AttributeError。
        photo_table = getattr(self, "photo_table", None)
        if photo_table is not None and obj is photo_table and event.type() == event.KeyPress:
            if event.matches(QKeySequence.Copy):
                return self._copy_photo_table_selection()
            if event.modifiers() & Qt.ControlModifier and event.key() == Qt.Key_A:
                photo_table.selectAll()
                return True
            if event.key() == Qt.Key_F2:
                return self._edit_current_photo_table_item()
        # grid_frame 拖入照片视觉反馈:Qt drag 事件经此 filter,正确派发到 Python(原 lambda 赋值方式失败)。
        grid_frame = getattr(self, "grid_frame", None)
        if grid_frame is not None and grid_frame is obj:
            etype = event.type()
            if etype == event.DragEnter:
                if event.mimeData().hasUrls():
                    event.acceptProposedAction()
                    grid_frame.setStyleSheet("border: 2px dashed #2a6fbd;")
                    return True
            elif etype == event.DragMove:
                if event.mimeData().hasUrls():
                    event.acceptProposedAction()
                    return True
            elif etype == event.DragLeave:
                grid_frame.setStyleSheet("")
                return True
        return super().eventFilter(obj, event)

    # ---- UI building ----

    def _build_ui(self) -> None:
        # Toolbars: main + aux（规范化软件设计 2026-05 起，从单条工具栏拆为两条 + 可拖拽 + 可自定义）。
        # - 主工具栏（main_toolbar）默认可见，放高频按钮，按 file/edit/view/tools 四组用 separator 分隔。
        # - 辅助工具栏（aux_toolbar）默认隐藏，放低频按钮，视图菜单或自定义对话框打开。
        # - 两条工具栏都 setMovable(True)，用户可拖到左/右/底 dock area。
        # - 内容由 settings.toolbar_layout 控制；未配置时回落到 TOOLBAR_DEFAULT_LAYOUT。
        # - 自动保存（_auto_save_action）始终挂主栏末尾，不进 layout（特殊 checkable QAction）。
        self._toolbar_actions: dict[str, QAction] = {}  # action_id -> QAction，供自定义对话框 / 快捷键绑定用
        # 两条工具栏都设 ToolButtonTextBesideIcon：图标提供视觉锚点 + 中文 label 保留可读性。
        # icon 大小 18px：兼顾紧凑度与可识别（Qt 默认 24 偏大，跟标准字号搭配会显胖）。
        # Phase 5: DraggableToolBar 支持 action inline 拖换位 + 跨栏拖
        from .widgets_toolbar import DraggableToolBar
        self._main_toolbar = DraggableToolBar("主工具栏", slot="main", parent=self)
        self._main_toolbar.setObjectName("main_toolbar")
        self._main_toolbar.setMovable(True)
        self._main_toolbar.setIconSize(QSize(18, 18))
        self._main_toolbar.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.addToolBar(self._main_toolbar)
        self._aux_toolbar = DraggableToolBar("辅助工具栏", slot="aux", parent=self)
        self._aux_toolbar.setObjectName("aux_toolbar")
        self._aux_toolbar.setMovable(True)
        self._aux_toolbar.setIconSize(QSize(18, 18))
        self._aux_toolbar.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.addToolBarBreak()  # 让辅栏默认换行（主栏在第 1 行、辅栏在第 2 行）
        self.addToolBar(self._aux_toolbar)
        self._rebuild_toolbars()  # 按 settings 填充

        # 旧逻辑：工具栏首项是「＋新增标本」按钮 → 现移至左侧入库编号面板顶部，详见 voucher_layout。
        # 旧逻辑：工具栏「沿用上条信息」checkable → 现固定规则（new_specimen 内永远生效），UI 移除。

        # 自动保存开关：始终挂主栏末尾（不进 layout 自定义；checkable，特殊处理）。
        self._auto_save_action = QAction(self)
        self._auto_save_action.setCheckable(True)
        self._auto_save_action.setChecked(self.auto_save_enabled)
        self._auto_save_action.setToolTip("勾选=输入后自动保存；取消=只在点「保存」按钮时写入")
        self._auto_save_action.toggled.connect(self._on_auto_save_toggled)
        self._update_auto_save_action_text()
        self._main_toolbar.addSeparator()
        self._main_toolbar.addAction(self._auto_save_action)

        # 应用辅栏可见性（settings 持久化，默认隐藏）。
        self._aux_toolbar.setVisible(bool(load_settings().aux_toolbar_visible))

        # Workspace bar
        ws_bar = QHBoxLayout()
        if self.workspace_root is not None:
            ws_label = QLabel(f"当前工作目录：{self.workspace_root}")
        else:
            ws_label = QLabel("当前工作目录：（未选择，请点击工具栏“切换工作区”）")
        ws_label.setContentsMargins(8, 4, 8, 4)
        self.workspace_label = ws_label
        ws_bar.addWidget(ws_label)
        ws_bar.addStretch()
        self.dashboard_label = QLabel()
        self.dashboard_label.setStyleSheet(
            "color: #1a5faa; font-weight: bold; font-size: 13px; padding: 4px 12px;"
            "background-color: #e8f0fe; border-radius: 4px;"
        )
        ws_bar.addWidget(self.dashboard_label)
        self._update_dashboard()

        # Central widget — photo preview area
        central = QWidget()
        self.setCentralWidget(central)
        central_layout = QVBoxLayout(central)
        central_layout.setContentsMargins(8, 8, 8, 4)
        central_layout.addLayout(ws_bar)

        # 分类预设缺失警告条：正常启动时 hide()；bundled_template_path 找不到文件时 show()。
        # 旧：_finish_initial_load 里仅有 8 秒状态栏消息，极易错过。现改为持久黄色条。
        self._preset_warning_banner = QLabel(
            "[ ! ] 分类预设文件缺失 — 种名/科名自动匹配已停用"
            "（字段模版/表格信息预设字段.xlsx 未找到，请重新安装或更新软件）"
        )
        self._preset_warning_banner.setWordWrap(True)
        self._preset_warning_banner.setStyleSheet(
            "background-color: #fff3cd; color: #856404; padding: 6px 12px;"
            "border-bottom: 1px solid #ffc107;"
        )
        self._preset_warning_banner.hide()
        central_layout.addWidget(self._preset_warning_banner)

        # D19 升级 banner:notify 模式发现新版时显示,三按钮 [立即升级] [稍后] [跳过此版]。
        self._update_banner_release = None
        self._update_banner = QFrame()
        self._update_banner.setObjectName("_update_banner")
        self._update_banner.setStyleSheet(
            "QFrame#_update_banner { background-color: #fff3cd; color: #856404; "
            "border-bottom: 1px solid #ffc107; }"
        )
        _ub_layout = QHBoxLayout(self._update_banner)
        _ub_layout.setContentsMargins(12, 6, 12, 6)
        _ub_text = QLabel("发现新版")
        _ub_text.setObjectName("_update_banner_text")
        _ub_text.setStyleSheet("color: #856404;")
        _ub_layout.addWidget(_ub_text, stretch=1)
        _ub_install = QPushButton("立即升级")
        _ub_install.clicked.connect(self._upgrade_banner_install)
        _ub_layout.addWidget(_ub_install)
        _ub_later = QPushButton("稍后")
        _ub_later.clicked.connect(self._upgrade_banner_later)
        _ub_layout.addWidget(_ub_later)
        _ub_skip = QPushButton("跳过此版")
        _ub_skip.clicked.connect(self._upgrade_banner_skip)
        _ub_layout.addWidget(_ub_skip)
        self._update_banner.hide()
        central_layout.addWidget(self._update_banner)

        # Stacked: graphics view (single) + grid frame (grid)
        self._photo_stack_container = QWidget()
        photo_stack = QVBoxLayout(self._photo_stack_container)
        photo_stack.setContentsMargins(0, 0, 0, 0)

        self.photo_view = PhotoGraphicsView()
        self.photo_view.photo_dropped.connect(self.add_photo_paths_async)
        self.photo_view.zoom_changed.connect(self._on_zoom_changed)
        # 由宫格双击展开的单张视图里，图上再双击 -> 返回原宫格。
        self.photo_view.return_requested.connect(self.return_to_grid)
        # 单张预览右键 -> 照片管理菜单。
        self.photo_view.context_requested.connect(self._show_single_preview_context_menu)
        photo_stack.addWidget(self.photo_view)

        self.grid_frame = QFrame()
        self.grid_layout = QGridLayout(self.grid_frame)
        self.grid_layout.setSpacing(10)
        self.grid_frame.setAcceptDrops(True)
        # 规范化软件设计 2026-05 P1 审查修复:
        # 旧:`grid_frame.dragEnterEvent = lambda e: ...` — Qt C++ 事件 dispatch 不走 Python
        #    instance attr,lambda 实际不被调,drag 视觉反馈静默失效。
        # 现:用 installEventFilter + 一个 EventFilter 对象。事件 filter 走 QObject.event 链路,
        #    Qt 可正确派发到 Python 端。
        self.grid_frame.installEventFilter(self)
        self.grid_frame.dropEvent = self._on_grid_drop  # dropEvent 已用 method 引用,Qt 可识别
        self.grid_frame.hide()
        photo_stack.addWidget(self.grid_frame)

        self._placeholder_label = QLabel('点击"添加照片"关联图片')
        self._placeholder_label.setAlignment(Qt.AlignCenter)
        self._placeholder_label.setStyleSheet("color: #59666b; font-size: 14px;")
        photo_stack.addWidget(self._placeholder_label)
        self._placeholder_label.raise_()

        central_layout.addWidget(self._photo_stack_container, stretch=1)

        # Photo controls
        controls = QHBoxLayout()
        self._return_grid_btn = QPushButton("返回网格")
        self._return_grid_btn.clicked.connect(self.return_to_grid)
        self._return_grid_btn.hide()
        controls.addWidget(self._return_grid_btn)
        for label, slot in [
            ("添加照片", self.add_photo),
            ("检索图片", self.open_image_search),
            # 「替换照片」原本只在照片表右键菜单，按需求补成与添加/检索并列的可见按钮。
            ("替换照片", self._replace_current_photo),
            # 旧标签「删除照片」——实际只取消该照片与入库编号的关联（删记录行 +
            # 无引用的工作区归档副本），原始照片不动，按需求改名「取消关联」。
            ("取消关联", self.delete_photo),
            ("分配入库编号", self._assign_voucher_to_selected),
            ("上一张", lambda: self.shift_photo(-1)),
            ("下一张", lambda: self.shift_photo(1)),
            ("-", lambda: self.adjust_zoom(0.8)),
            ("适配", self.fit_image),
            ("+", lambda: self.adjust_zoom(1.25)),
            ("打开原图", self.open_current_photo_external),
        ]:
            btn = QPushButton(label)
            btn.clicked.connect(slot)
            if label in ("-", "+"):
                btn.setFixedWidth(30)
            controls.addWidget(btn)
        self.photo_counter = QLabel("0 / 0")
        controls.addWidget(self.photo_counter, alignment=Qt.AlignRight)
        self._confirm_btn = QPushButton("确定保存")
        self._confirm_btn.setStyleSheet("QPushButton { font-weight: bold; background-color: #2a6fbd; color: white; padding: 4px 12px; border-radius: 3px; }")
        self._confirm_btn.clicked.connect(self._confirm_save)
        controls.addWidget(self._confirm_btn)
        self._zoom_label = QLabel("")
        self._zoom_label.setStyleSheet("color: #59666b;")
        controls.addWidget(self._zoom_label)
        central_layout.addLayout(controls)

        # View mode
        view_row = QHBoxLayout()
        view_row.addWidget(QLabel("显示"))
        self._view_buttons = []
        for label, count, short in VIEW_MODES:
            btn = QPushButton(short)
            btn.setCheckable(True)
            btn.setFixedWidth(36)
            btn.setToolTip(label)
            btn.setProperty("grid_count", count)
            # 规范化软件设计 2026-05 P1 优化:用 class 选择器(theme.py APP_QSS 内 QPushButton[class="view-btn"])
            # 替代 inline _VIEW_BTN_STYLE,5 个按钮共享同一规则省 QSS 解析对象。
            btn.setProperty("class", "view-btn")
            btn.clicked.connect(lambda checked, c=count: self._set_view_mode(c))
            view_row.addWidget(btn)
            self._view_buttons.append(btn)
        self._view_buttons[0].setChecked(True)
        self._show_filename_check = QCheckBox("显示文件名")
        self._show_filename_check.setChecked(self._show_grid_filenames)
        self._show_filename_check.stateChanged.connect(self._on_show_grid_filenames_changed)
        view_row.addWidget(self._show_filename_check)
        view_row.addStretch()
        central_layout.addLayout(view_row)

        # ---- Panels (QSplitter layout — no QDockWidgets) ----
        # Left: voucher panel with table + filter + pagination
        voucher_content = QWidget()
        voucher_layout = QVBoxLayout(voucher_content)
        voucher_layout.setContentsMargins(0, 0, 0, 0)
        voucher_layout.setSpacing(2)
        # 录入任务指示器行（任务进行中时绿色背景 + 结束按钮，否则显示「开始录入任务」）
        self._task_indicator = QWidget()
        self._task_indicator.setObjectName("task_indicator")
        task_ind_layout = QHBoxLayout(self._task_indicator)
        task_ind_layout.setContentsMargins(4, 2, 4, 2)
        task_ind_layout.setSpacing(4)
        self._task_label = QLabel("未开始录入任务")
        self._task_label.setStyleSheet("color: #888;")
        task_ind_layout.addWidget(self._task_label, stretch=1)
        self._task_start_btn = QPushButton("▶ 开始录入任务")
        self._task_start_btn.setToolTip("开始录入任务，记录录入人员和工作时长")
        self._task_start_btn.clicked.connect(self._start_task)
        task_ind_layout.addWidget(self._task_start_btn)
        self._task_end_btn = QPushButton("结束任务")
        self._task_end_btn.setVisible(False)
        self._task_end_btn.clicked.connect(self._end_task)
        task_ind_layout.addWidget(self._task_end_btn)
        voucher_layout.addWidget(self._task_indicator)
        personnel_btn = QPushButton("查看人员记录")
        personnel_btn.setToolTip("查看录入工作量汇总，了解各录入人员的任务次数和时长")
        personnel_btn.clicked.connect(self._open_workload_report)
        voucher_layout.addWidget(personnel_btn)

        # 「＋新增入库编号」按钮 + 系列选择器行
        new_voucher_row = QHBoxLayout()
        new_voucher_row.setSpacing(4)
        self._new_voucher_btn = QPushButton("＋新增入库编号")
        self._new_voucher_btn.setEnabled(False)
        self._new_voucher_btn.setToolTip("请先开始录入任务")
        self._new_voucher_btn.clicked.connect(self.new_specimen)
        new_voucher_row.addWidget(self._new_voucher_btn, stretch=1)
        self._series_selector = QComboBox()
        self._series_selector.setToolTip("选择入库编号系列（当前系列用于新增编号）")
        self._series_selector.setMinimumWidth(70)
        self._series_selector.setMaximumWidth(110)
        self._refresh_series_selector()
        self._series_selector.currentIndexChanged.connect(self._on_series_selector_changed)
        new_voucher_row.addWidget(self._series_selector)
        manage_series_btn = QPushButton("管理")
        manage_series_btn.setToolTip("管理入库编号系列（新增/编辑/删除）")
        manage_series_btn.setFixedWidth(40)
        manage_series_btn.clicked.connect(self._open_series_manager)
        new_voucher_row.addWidget(manage_series_btn)
        voucher_layout.addLayout(new_voucher_row)
        # Search + quick filter
        filter_row = QHBoxLayout()
        self._voucher_search = QLineEdit()
        self._voucher_search.setPlaceholderText("搜索编号/照片名...")
        self._voucher_search.setClearButtonEnabled(True)
        self._voucher_search.textChanged.connect(self._apply_voucher_filter)
        filter_row.addWidget(self._voucher_search)
        self._search_scope = QComboBox()
        self._search_scope.addItems(["全部", "入库编号", "管内编号", "照片名"])
        # 旧逻辑：setFixedWidth(80) 硬固定宽度，面板拖窄时筛选行控件挤压重叠。
        # 改为允许在 56–96px 间伸缩，面板可自由拖窄而不重叠。
        # self._search_scope.setFixedWidth(80)
        self._search_scope.setMinimumWidth(56)
        self._search_scope.setMaximumWidth(96)
        self._search_scope.currentIndexChanged.connect(self._apply_voucher_filter)
        filter_row.addWidget(self._search_scope)
        # 系列筛选下拉
        self._series_filter_combo = QComboBox()
        self._series_filter_combo.setToolTip("按编号系列筛选凭证列表")
        self._series_filter_combo.setMinimumWidth(56)
        self._series_filter_combo.setMaximumWidth(90)
        self._refresh_series_filter_combo()
        self._series_filter_combo.currentIndexChanged.connect(self._apply_voucher_filter)
        filter_row.addWidget(self._series_filter_combo)
        # 复选框：显示/隐藏关联照片列
        self._show_photos_checkbox = QCheckBox("照片名")
        self._show_photos_checkbox.setToolTip("在凭证列表中显示关联的照片文件名")
        self._show_photos_checkbox.toggled.connect(self._toggle_photo_names_column)
        filter_row.addWidget(self._show_photos_checkbox)
        voucher_layout.addLayout(filter_row)
        quick_row = QHBoxLayout()
        quick_row.setSpacing(2)
        self._filter_buttons: dict[str, QPushButton] = {}
        for key, label in [("all","全部"),("claimed","已认领"),("complete","完整"),("incomplete","待补全")]:
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setFixedHeight(22)
            # 规范化软件设计 2026-05 P1 优化:class 选择器替代 inline QSS。
            btn.setProperty("class", "filter-btn")
            btn.clicked.connect(lambda checked, k=key: self._set_voucher_filter(k))
            quick_row.addWidget(btn)
            self._filter_buttons[key] = btn
        self._filter_buttons["all"].setChecked(True)
        self._active_filter = "all"
        voucher_layout.addLayout(quick_row)
        # Table: 入库编号 | 标本 | 照片 | 分类 | 认领 | 照片数
        self.voucher_table = QTableWidget(0, 7)
        self.voucher_table.setHorizontalHeaderLabels(["入库编号","标本","照片","分类","认领","照片数","关联照片"])
        self.voucher_table.setSelectionBehavior(QTableWidget.SelectRows)
        # 原代码：SingleSelection 仅单选；改为 ExtendedSelection 支持 Windows 操作习惯：
        # Ctrl+Click 多选 / Shift+Click 范围选 / 拖拽多选
        self.voucher_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.voucher_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.voucher_table.verticalHeader().setVisible(False)
        # 旧逻辑：列宽硬编码 85/36/36/36/52/42，表格字体固定 QFont("Consolas", 10)。
        # 字体放大时列宽装不下、文字被覆盖。现改为按全局字号缩放（见 _refresh_scaled_fonts）；
        # 列宽基准值集中到类常量 _VOUCHER_COL_BASE_WIDTHS。
        # self.voucher_table.setColumnWidth(0, 85)
        # self.voucher_table.setColumnWidth(1, 36)
        # self.voucher_table.setColumnWidth(2, 36)
        # self.voucher_table.setColumnWidth(3, 36)
        # self.voucher_table.setColumnWidth(4, 52)
        # self.voucher_table.setColumnWidth(5, 42)
        self.voucher_table.setColumnWidth(6, 0)  # 关联照片列：默认隐藏，通过复选框切换
        self.voucher_table.horizontalHeader().setStretchLastSection(True)
        # self.voucher_table.setFont(QFont("Consolas", 10))  # 改为 _refresh_scaled_fonts() 统一设置
        self.voucher_table.itemSelectionChanged.connect(self._on_voucher_table_selected)
        self.voucher_table.horizontalHeader().sectionClicked.connect(self._on_voucher_header_clicked)
        self.voucher_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.voucher_table.customContextMenuRequested.connect(self._voucher_context_menu)
        # 全局快捷键（规范化软件设计 2026-05 起持引用，支持 settings.custom_shortcuts 热更新）。
        # 旧：所有 QShortcut 用 `_ = QShortcut(...)` 临时变量；变量丢弃后 Qt 仍持有但无法重绑 keyseq。
        # 现：保留实例引用 _sc_xxx，启动末尾的 _apply_custom_shortcuts() 可按 settings 改 key。
        # Ctrl+Shift+Z 是 Ctrl+Y 的同义重做绑定，固定不参与自定义。
        self._sc_select_all_voucher = QShortcut(QKeySequence("Ctrl+A"), self.voucher_table, self._select_all_vouchers)
        self._sc_undo = QShortcut(QKeySequence("Ctrl+Z"), self, self.undo)
        self._sc_redo = QShortcut(QKeySequence("Ctrl+Y"), self, self.redo)
        self._sc_redo_alt = QShortcut(QKeySequence("Ctrl+Shift+Z"), self, self.redo)  # 同义绑定
        self._sc_zoom_in = QShortcut(QKeySequence("Ctrl+="), self, self._zoom_font_in)
        self._sc_zoom_in_alt = QShortcut(QKeySequence("Ctrl++"), self, self._zoom_font_in)  # 同义（按键盘布局）
        self._sc_zoom_out = QShortcut(QKeySequence("Ctrl+-"), self, self._zoom_font_out)
        self._sc_zoom_reset = QShortcut(QKeySequence("Ctrl+0"), self, self._zoom_font_reset)
        self._col_filters: dict[int, str] = {}  # col_index -> filter value
        self._col_header_labels = ["入库编号","标本","照片","分类","认领","照片数","关联照片"]
        self._show_photo_names = False
        # 按当前全局字号设置表格字体与列宽（默认字号时与旧版完全一致）。
        self._refresh_scaled_fonts()
        voucher_layout.addWidget(self.voucher_table, stretch=1)
        # Pagination
        page_row = QHBoxLayout()
        self._voucher_page_label = QLabel("")
        page_row.addWidget(self._voucher_page_label)
        page_row.addStretch()
        self._voucher_prev_btn = QPushButton("◀")
        self._voucher_prev_btn.setFixedWidth(28)
        self._voucher_prev_btn.clicked.connect(self._voucher_prev_page)
        page_row.addWidget(self._voucher_prev_btn)
        self._voucher_next_btn = QPushButton("▶")
        self._voucher_next_btn.setFixedWidth(28)
        self._voucher_next_btn.clicked.connect(self._voucher_next_page)
        page_row.addWidget(self._voucher_next_btn)
        voucher_layout.addLayout(page_row)
        self._voucher_page = 0
        self._voucher_page_size = 200
        self.voucher_panel = self._create_panel("入库编号", voucher_content, collapsible=False)
        # 旧逻辑：setMinimumWidth(290) 硬下限，拖到 290px 卡住；后降到 150。
        # 现配合 main_splitter.setChildrenCollapsible(True) 再降到 60，可拖到极窄甚至折叠，
        # 把空间让给中央图片显示区。
        # self.voucher_panel.setMinimumWidth(290)
        # self.voucher_panel.setMinimumWidth(150)
        self.voucher_panel.setMinimumWidth(60)

        # Right: specimen info panel
        specimen_content = QWidget()
        sf_layout = QFormLayout(specimen_content)
        sf_layout.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        sf_layout.setContentsMargins(0, 0, 0, 0)
        for field in SPECIMEN_HEADERS:
            if field == "保存方式":
                widget = QComboBox()
                widget.addItems(SAVE_METHOD_OPTIONS)
                widget.setEditable(True)
                widget.currentTextChanged.connect(lambda text, f=field: self.schedule_save("specimen", f))
            else:
                widget = QLineEdit()
                widget.textChanged.connect(lambda text, f=field: self.schedule_save("specimen", f))
            # 标签保持纯文字；「?」填写说明跟在输入框后面（_wrap_field_with_hint）。
            sf_layout.addRow(field, self._wrap_field_with_hint(field, widget))
            self.specimen_widgets[field] = widget
        spec_save_btn = QPushButton("保存标本信息")
        spec_save_btn.clicked.connect(lambda: self._save_panel("specimen"))
        sf_layout.addRow(spec_save_btn)
        self.specimen_panel = self._create_panel("标本信息", specimen_content)

        # Right: photo info panel
        photo_content = QWidget()
        pf_layout = QVBoxLayout(photo_content)
        pf_layout.setContentsMargins(0, 0, 0, 0)
        self.photo_table = QTableWidget(0, 4)
        self.photo_table.setHorizontalHeaderLabels(["序号", "文件名", "相对路径", "描述"])
        self.photo_table.horizontalHeader().setStretchLastSection(True)
        self.photo_table.setSelectionBehavior(QTableWidget.SelectItems)
        self.photo_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.photo_table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.EditKeyPressed | QTableWidget.SelectedClicked)
        self.photo_table.setColumnWidth(0, 44)
        self.photo_table.setColumnWidth(1, 150)
        self.photo_table.setColumnWidth(2, 210)
        self.photo_table.currentCellChanged.connect(self._on_photo_table_row_changed)
        self.photo_table.itemChanged.connect(self._on_photo_table_item_changed)
        self.photo_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.photo_table.customContextMenuRequested.connect(self._photo_table_context_menu)
        self.photo_table.installEventFilter(self)
        pf_layout.addWidget(self.photo_table, stretch=1)
        page_row = QHBoxLayout()
        self._page_label = QLabel("")
        page_row.addWidget(self._page_label, stretch=1)
        self._prev_page_btn = QPushButton("上一页")
        self._prev_page_btn.clicked.connect(self._photo_prev_page)
        page_row.addWidget(self._prev_page_btn)
        self._next_page_btn = QPushButton("下一页")
        self._next_page_btn.clicked.connect(self._photo_next_page)
        page_row.addWidget(self._next_page_btn)
        pf_layout.addLayout(page_row)
        for field in ("文件名", "相对路径", "绝对路径", "描述"):
            row = QHBoxLayout()
            row.addWidget(QLabel(field))
            widget = QLineEdit()
            if field in {"相对路径", "绝对路径"}:
                widget.setReadOnly(True)
            elif field == "文件名":
                widget.setReadOnly(False)
                widget.setClearButtonEnabled(True)
                widget.textChanged.connect(lambda text, f=field: self.schedule_save("photo", f))
                widget.setContextMenuPolicy(Qt.CustomContextMenu)
                widget.customContextMenuRequested.connect(lambda pos, w=widget: self._filename_context_menu(w, pos))
            else:
                widget.textChanged.connect(lambda text: self.schedule_save("photo", "描述"))
            self.photo_widgets[field] = widget
            # 「?」填写说明跟在输入框后面（无说明的字段 _wrap 直接返回原控件）。
            row.addWidget(self._wrap_field_with_hint(field, widget))
            pf_layout.addLayout(row)
        photo_save_btn = QPushButton("保存照片信息")
        photo_save_btn.clicked.connect(lambda: self._save_panel("photo"))
        pf_layout.addWidget(photo_save_btn)
        self.photo_panel = self._create_panel("照片信息", photo_content)
        # 存储照片面板标题 QLabel，以便切换标本时动态显示当前入库编号
        self._photo_panel_title = self.photo_panel.findChild(QLabel)

        # Right: classification panel
        class_content = QWidget()
        cf_layout = QFormLayout(class_content)
        cf_layout.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        cf_layout.setContentsMargins(0, 0, 0, 0)
        for field in EDITABLE_CLASSIFICATION_COLUMNS:
            widget = QLineEdit()
            widget.textChanged.connect(lambda text, f=field: self.schedule_save("classification", f))
            # 标签保持纯文字；「?」填写说明跟在输入框后面（_wrap_field_with_hint）。
            cf_layout.addRow(field, self._wrap_field_with_hint(field, widget))
            self.class_widgets[field] = widget
            if field in TAXONOMY_LOOKUP_INPUT_COLUMNS:
                self._attach_taxonomy_lookup_to_classification_field(field, widget)
        class_save_btn = QPushButton("保存分类信息")
        class_save_btn.clicked.connect(lambda: self._save_panel("classification"))
        cf_layout.addRow(class_save_btn)
        self.class_panel = self._create_panel("分类信息", class_content)

        # Right vertical splitter
        self.right_splitter = QSplitter(Qt.Vertical)
        self.right_splitter.addWidget(self.specimen_panel)
        self.right_splitter.addWidget(self.photo_panel)
        self.right_splitter.addWidget(self.class_panel)
        self.right_splitter.setSizes([260, 280, 200])
        # 右侧 标本/照片/分类 三块允许各自折叠，方便把空间让给图片区。
        self.right_splitter.setChildrenCollapsible(True)

        # Right side container with save-all button at top
        right_container = QWidget()
        right_layout = QVBoxLayout(right_container)
        right_layout.setContentsMargins(0, 0, 0, 4)
        right_layout.setSpacing(4)
        save_all_btn = QPushButton("一键保存所有")
        save_all_btn.setStyleSheet("QPushButton { font-weight: bold; background-color: #2a6fbd; color: white; padding: 6px; border-radius: 3px; }")
        save_all_btn.clicked.connect(self._save_all_panels)
        right_layout.addWidget(save_all_btn)
        right_layout.addWidget(self.right_splitter, stretch=1)

        # Main horizontal splitter — reuse existing 'central' widget as center
        self.main_splitter = QSplitter(Qt.Horizontal)
        self.main_splitter.addWidget(self.voucher_panel)
        self.main_splitter.addWidget(central)
        self.main_splitter.addWidget(right_container)
        self.main_splitter.setSizes([220, 580, 480])
        # 旧逻辑：setChildrenCollapsible(False) —— 左侧入库编号栏 / 右侧面板拖不窄、不能折叠，
        # 中央图片区无法占满。改为 True：侧边栏可拖到极窄甚至折叠，把空间让给图片显示区。
        self.main_splitter.setChildrenCollapsible(True)
        self.setCentralWidget(self.main_splitter)

        # Restore saved splitter sizes
        saved = load_settings()
        if saved.splitter_sizes and len(saved.splitter_sizes) >= 2:
            try:
                self.main_splitter.setSizes([int(x) for x in saved.splitter_sizes[0]])
                self.right_splitter.setSizes([int(x) for x in saved.splitter_sizes[1]])
            except Exception:
                pass
        if saved.window_geometry:
            self.restoreGeometry(QByteArray.fromBase64(saved.window_geometry.encode()))

        # ---- Menu bar ----
        # 规范化软件设计 2026-05 Phase 4:顶层菜单 4 → 6 (入库/编号/WoRMS/工具/视图/帮助)
        # Phase 5:菜单项右键弹"加到工具栏"小菜单 (ToolbarAwareMenu);
        # 每项 QAction.setData(action_id) 关联 TOOLBAR_ACTIONS 注册表。
        from .widgets_toolbar import ToolbarAwareMenu

        def _make_menu(title: str) -> ToolbarAwareMenu:
            m = ToolbarAwareMenu(title, self)
            self.menuBar().addMenu(m)
            return m

        def _add(menu, text: str, slot, action_id: str = ""):
            """加菜单项,可选 action_id(关联 TOOLBAR_ACTIONS;右键可加到工具栏)。"""
            act = QAction(text, self)
            act.triggered.connect(slot)
            if action_id:
                act.setData(action_id)
            menu.addAction(act)
            return act

        # 顶层「入库」菜单
        ingest_menu = _make_menu("入库")
        _add(ingest_menu, "入库汇总", self.open_ingest_summary, "ingest_summary")
        _add(ingest_menu, "入库编号系列管理…", self._open_series_manager, "series_manager")
        ingest_menu.addSeparator()
        _add(ingest_menu, "入库人员管理…", self._open_persons_manager, "persons_manager")
        _add(ingest_menu, "入库人员记录…", self._open_workload_report, "workload_report")

        # 顶层「编号」菜单
        number_menu = _make_menu("编号")
        _add(number_menu, "批量生成编号…", self._open_batch_generate, "batch_generate")
        _add(number_menu, "手动添加入库编号…", self._open_manual_voucher, "manual_voucher")
        number_menu.addSeparator()
        self._series_switch_menu = number_menu.addMenu("切换活动系列")
        self._series_switch_menu.aboutToShow.connect(self._populate_series_switch_menu)

        # 顶层「WoRMS」菜单
        worms_menu = _make_menu("WoRMS")
        _add(worms_menu, "分类匹配…", self._open_worms_match, "worms")
        _add(worms_menu, "查询…", self._open_worms_browse, "worms_browse")
        worms_menu.addSeparator()
        _add(worms_menu, "本地数据库管理…", self._open_worms_db_manager, "worms_db")

        # 顶层「工具」菜单
        tools_menu = _make_menu("工具")
        _add(tools_menu, "操作记录", self._open_action_log)
        _add(tools_menu, "用 Excel 打开数据文件…", self._open_data_in_excel, "excel_open")
        tools_menu.addSeparator()
        _add(tools_menu, "从收件箱聚合…", self._open_aggregate_incoming)
        _add(tools_menu, "批量导入工作区目录…", self._open_batch_import_sources)
        _add(tools_menu, "升级工作区到多人协作格式…", self._upgrade_workspace_to_multi_user)
        tools_menu.addSeparator()
        _add(tools_menu, "导出 Darwin Core Archive…", self._export_dwc_archive)
        _add(tools_menu, "从 EXIF 批量回填采集日期…", self._bulk_apply_exif)
        tools_menu.addSeparator()
        _add(tools_menu, "降低工作区兼容版本…", self._downgrade_workspace_schema)
        tools_menu.addSeparator()
        _add(tools_menu, "打开合并/导入操作示例…", self._open_import_examples)
        tools_menu.addSeparator()
        _add(tools_menu, "新建工作区窗口…", self._open_new_workspace_window, "new_window")
        if not getattr(self, "read_only", False):
            _add(tools_menu, "新建本工作区只读副本", self._open_readonly_clone, "readonly_clone")

        # 升级中心 v0.8.0 (D1-D20):一级 升级 菜单。主流对齐 Claude Code/VSCode,
        # 顶层 5 项,其余功能藏 高级 子菜单 + 升级中心 dialog tab。
        upgrade_menu = _make_menu("升级")
        _add(upgrade_menu, "立即升级到最新版", self._oneclick_upgrade_now, "oneclick_upgrade")
        upgrade_menu.addSeparator()
        _add(upgrade_menu, "检查更新…", self._check_update_now, "check_update_now")
        _add(upgrade_menu, "自动更新设置…", self._open_upgrade_settings, "upgrade_settings")
        _add(upgrade_menu, "关于当前版本…", self._open_upgrade_about, "upgrade_about")
        upgrade_advanced = upgrade_menu.addMenu("高级")
        upgrade_advanced.setToolTipsVisible(True)
        _adv_center = QAction("升级中心…", self)
        _adv_center.triggered.connect(self._open_upgrade_center)
        _adv_center.setData("upgrade_center")
        upgrade_advanced.addAction(_adv_center)
        upgrade_advanced.addSeparator()
        _adv_import = QAction("从本地文件安装更新…", self)
        _adv_import.triggered.connect(self._install_from_zip)
        _adv_import.setData("install_from_zip")
        upgrade_advanced.addAction(_adv_import)
        _adv_dl = QAction("下载安装包供分发…", self)
        _adv_dl.triggered.connect(self._download_installer)
        _adv_dl.setData("download_installer")
        upgrade_advanced.addAction(_adv_dl)
        upgrade_advanced.addSeparator()
        _adv_hist = QAction("历史版本管理…", self)
        _adv_hist.triggered.connect(self._open_upgrade_history)
        _adv_hist.setData("upgrade_history")
        upgrade_advanced.addAction(_adv_hist)
        upgrade_advanced.addSeparator()
        _adv_remote = QAction("远程触发 GitHub 构建…", self)
        _adv_remote.triggered.connect(self._open_upgrade_build_remote)
        _adv_remote.setData("upgrade_build_remote")
        upgrade_advanced.addAction(_adv_remote)
        # 本地重打包仅 dev 模式可见 (D6)。
        try:
            from .install_kind import installation_kind
            if installation_kind() == "source":
                _adv_local = QAction("本地重新打包并安装…", self)
                _adv_local.triggered.connect(self._open_upgrade_build_local)
                _adv_local.setData("upgrade_build_local")
                upgrade_advanced.addAction(_adv_local)
        except Exception:
            pass

        view_menu = _make_menu("视图")
        for panel_name, panel_ref in [
            ("入库编号", self.voucher_panel),
            ("标本信息", self.specimen_panel),
            ("照片信息", self.photo_panel),
            ("分类信息", self.class_panel),
        ]:
            action = QAction(panel_name, self, checkable=True, checked=True)
            action.toggled.connect(lambda checked, p=panel_ref: p.setVisible(checked))
            view_menu.addAction(action)
        view_menu.addSeparator()
        # 辅助工具栏可见性切换（规范化软件设计 2026-05 新增）：状态持久化到 settings.aux_toolbar_visible
        self._aux_toolbar_action = QAction("辅助工具栏", self, checkable=True)
        self._aux_toolbar_action.setChecked(load_settings().aux_toolbar_visible)
        self._aux_toolbar_action.toggled.connect(self._on_aux_toolbar_toggled)
        view_menu.addAction(self._aux_toolbar_action)
        # 自定义工具栏 / 自定义快捷键入口（D / E）
        view_menu.addAction("自定义工具栏…", self._open_toolbar_customize)
        view_menu.addAction("自定义快捷键…", self._open_shortcuts_customize)
        view_menu.addSeparator()
        reset_layout_action = QAction("重置窗口布局", self)
        reset_layout_action.triggered.connect(self._reset_window_layout)
        view_menu.addAction(reset_layout_action)

        # 顶层「帮助」菜单（规范化软件设计 2026-05 新增）：
        # 旧：无统一 Help / 关于入口；仅状态栏显示版本号。
        # 现：菜单栏右端固定「帮助」菜单，下分使用说明 / 字段速查 / 快捷键速查 / 检查更新 /
        # 打开崩溃日志目录 / 关于。各 slot 详见 _open_user_manual_dialog 等方法。
        help_menu = _make_menu("帮助")
        help_menu.addAction("使用说明…", self._open_user_manual_dialog)
        help_menu.addAction("字段填写说明速查…", self._open_field_help_index)
        help_menu.addAction("快捷键速查…", self._open_shortcuts_dialog)
        help_menu.addSeparator()
        help_menu.addAction("检查更新…", self._check_github_update_from_help)
        help_menu.addAction("打开崩溃日志目录…", self._open_crash_log_dir)
        help_menu.addSeparator()
        help_menu.addAction("关于…", self._open_about_dialog)

        # ---- Status bar ----
        self.statusBar().showMessage("就绪")
        self._status_dashboard = QLabel()
        self._status_dashboard.setStyleSheet("color: #1a5faa; font-weight: bold; padding: 0 8px;")
        self.statusBar().addPermanentWidget(self._status_dashboard)
        # 规范化软件设计 2026-05 新增:内存档位 + 实时 RSS 状态栏 (每 5s 刷新)
        self._memory_status_label = QLabel()
        self._memory_status_label.setStyleSheet("color: #59666b; padding: 0 8px;")
        self._memory_status_label.setToolTip("当前内存档位与进程 RSS;改档位在「设置 → 内存档位」")
        self.statusBar().addPermanentWidget(self._memory_status_label)
        # 入库人员管理 2026-05:当前录入员下拉(状态栏)
        from .widgets_persons import PersonComboBox
        self._current_recorder_combo = PersonComboBox(allow_manage=True)
        self._current_recorder_combo.setToolTip(
            "当前录入员;点开切换或新增。任务进行中切换会弹确认。"
        )
        self._current_recorder_combo.setFixedWidth(220)
        self._current_recorder_combo.member_changed.connect(self._on_current_recorder_changed)
        # 加载团队库 + 预选 settings.current_recorder
        try:
            current = load_settings().current_recorder
            self._current_recorder_combo.refresh(preselect=current)
        except Exception:
            self._current_recorder_combo.refresh()
        self.statusBar().addPermanentWidget(self._current_recorder_combo)
        self.statusBar().addPermanentWidget(QLabel(f"软件版本：v{__version__}"))
        # 启动定时器 5s 刷新 RSS;立即刷一次
        self._memory_status_timer = QTimer(self)
        self._memory_status_timer.timeout.connect(self._refresh_memory_status)
        self._memory_status_timer.start(5000)
        self._refresh_memory_status()

        # ---- Keyboard shortcuts ----
        # 规范化软件设计 2026-05 起持引用为 self._sc_fit / self._sc_esc，支持 settings.custom_shortcuts 重绑。
        self._sc_fit = QAction(self)
        self._sc_fit.setShortcut(QKeySequence("F"))
        self._sc_fit.triggered.connect(self.fit_image)
        self.addAction(self._sc_fit)

        self._sc_esc = QAction(self)
        self._sc_esc.setShortcut(QKeySequence("Esc"))
        self._sc_esc.triggered.connect(self.return_to_grid)
        self.addAction(self._sc_esc)

        self._photo_filename_fill_action = QAction("从照片文件名填充标本信息", self)
        self._photo_filename_fill_action.triggered.connect(self.fill_current_photo_from_filename)
        self.addAction(self._photo_filename_fill_action)
        self._apply_photo_filename_fill_shortcut()

        # 应用 settings.custom_shortcuts 到所有可自定义快捷键的 action / shortcut（E 项）。
        self._apply_custom_shortcuts()

    # ---- 工具栏构建 / 重建（规范化软件设计 2026-05 新增） ----

    def _rebuild_toolbars(self) -> None:
        """按 settings.toolbar_layout 填充主/辅工具栏。

        - settings 缺/空时回落 TOOLBAR_DEFAULT_LAYOUT。
        - 同 category 间自动加 separator（按 file/edit/view/tools 顺序）。
        - 未知 action_id 静默跳过（向后兼容：日后删 action 不会让旧 settings 崩）。
        - 已挂的 QAction 保留在 self._toolbar_actions（id → QAction），自定义对话框 / 快捷键绑定用。

        本方法可在自定义对话框保存后重复调用，达到热更新。
        """
        # 清空两栏（保留 _auto_save_action — 它不在 layout 内，由 _build_ui 另挂）
        for tb in (self._main_toolbar, self._aux_toolbar):
            # 备份 _auto_save_action（如果在该栏内）
            actions_to_remove = [
                a for a in tb.actions()
                if a is not getattr(self, "_auto_save_action", None)
            ]
            for a in actions_to_remove:
                tb.removeAction(a)
        self._toolbar_actions.clear()

        layout = load_settings().toolbar_layout or {}
        main_ids = layout.get("main") or list(TOOLBAR_DEFAULT_LAYOUT["main"])
        aux_ids = layout.get("aux") or list(TOOLBAR_DEFAULT_LAYOUT["aux"])

        from PyQt5.QtWidgets import QStyle
        style = self.style()

        def _add_ids(toolbar: QToolBar, action_ids: list[str]) -> None:
            last_category = None
            for action_id in action_ids:
                spec = TOOLBAR_ACTIONS.get(action_id)
                if spec is None:
                    continue  # 未知 id，跳（向后兼容老 settings）
                slot = getattr(self, spec["slot"], None)
                if slot is None:
                    continue  # slot 方法不存在，跳
                # 同 category 内连排；跨 category 加 separator
                category = spec.get("category")
                if last_category is not None and category != last_category:
                    toolbar.addSeparator()
                last_category = category
                action = QAction(spec["label"], self)
                # Phase 5: 关联 action_id 供 DraggableToolBar drag mime + 菜单右键加快捷
                action.setData(action_id)
                tip = spec.get("tooltip")
                if tip:
                    action.setToolTip(tip)
                # 给 action 加 Qt 自带 StandardPixmap 图标（零资源依赖、跨平台）。
                icon_name = spec.get("icon")
                if icon_name:
                    pix = getattr(QStyle, icon_name, None)
                    if pix is not None:
                        try:
                            action.setIcon(style.standardIcon(pix))
                        except Exception:
                            pass  # 主题不支持时静默跳过
                action.triggered.connect(slot)
                toolbar.addAction(action)
                self._toolbar_actions[action_id] = action

        # _auto_save_action 已在 _build_ui 末尾挂主栏；这里清空时已排除它，重排时它仍在原位
        # 但顺序可能错乱：先清完再重排时它实际已被 removeAction 排除外，仍在主栏。
        # 为简化：忽略它的位置，每次重建后 _auto_save_action 会出现在主栏的某处；
        # 实际更准确的做法是把 _auto_save_action 也 removeAction 再重新 addAction 到末尾。
        auto_save = getattr(self, "_auto_save_action", None)
        if auto_save is not None and auto_save in self._main_toolbar.actions():
            self._main_toolbar.removeAction(auto_save)

        _add_ids(self._main_toolbar, main_ids)
        _add_ids(self._aux_toolbar, aux_ids)

        if auto_save is not None:
            self._main_toolbar.addSeparator()
            self._main_toolbar.addAction(auto_save)

    def _apply_custom_shortcuts(self) -> None:
        """把 settings.custom_shortcuts 应用到对应的 QShortcut / QAction（E 项）。

        - 仅作用于本注册表内 action_id；未知 id 静默跳过。
        - 空字符串 keyseq 表示用 SHORTCUTABLE_ACTIONS 默认值；不显式清除已绑的。
        - 冲突由 ShortcutsCustomizeDialog 在录入阶段保证（不绑重复 keyseq）。
        """
        try:
            custom = (load_settings().custom_shortcuts or {})
        except Exception:
            return

        def _seq_for(action_id: str) -> Optional[QKeySequence]:
            raw = custom.get(action_id)
            if raw is None:
                # 未自定义 -> 保持创建时的默认 keyseq，不动
                return None
            if not raw.strip():
                # 空 -> 用注册表 default 还原
                spec = SHORTCUTABLE_ACTIONS.get(action_id)
                if spec and spec.get("default"):
                    return QKeySequence(spec["default"])
                return QKeySequence()  # 完全清空
            return QKeySequence(raw)

        # 各 action_id 对应的 setter（QShortcut 用 setKey，QAction 用 setShortcut）。
        # 缺失引用（未在本类里创建 QShortcut/QAction 的）跳过：日后补绑时只需在此添加分支。
        setters = {
            "undo":                lambda ks: self._sc_undo.setKey(ks),
            "redo":                lambda ks: self._sc_redo.setKey(ks),
            "select_all_voucher":  lambda ks: self._sc_select_all_voucher.setKey(ks),
            "fit_image":           lambda ks: self._sc_fit.setShortcut(ks),
            "return_to_grid":      lambda ks: self._sc_esc.setShortcut(ks),
            "zoom_in":             lambda ks: self._sc_zoom_in.setKey(ks),
            "zoom_out":            lambda ks: self._sc_zoom_out.setKey(ks),
            "zoom_reset":          lambda ks: self._sc_zoom_reset.setKey(ks),
            "photo_filename_fill": lambda ks: self._photo_filename_fill_action.setShortcut(ks),
            "ingest_summary":      lambda ks: self._toolbar_actions["ingest_summary"].setShortcut(ks)
                                       if "ingest_summary" in self._toolbar_actions else None,
            "batch_export":        lambda ks: self._toolbar_actions["batch_export"].setShortcut(ks)
                                       if "batch_export" in self._toolbar_actions else None,
            "worms":               lambda ks: self._toolbar_actions["worms"].setShortcut(ks)
                                       if "worms" in self._toolbar_actions else None,
            # user_manual: 帮助菜单内 QAction 暂未持引用 -> 跳过，由 ShortcutsCustomizeDialog 提示
        }
        for action_id, setter in setters.items():
            ks = _seq_for(action_id)
            if ks is None:
                continue
            try:
                setter(ks)
            except Exception:
                pass  # 单个 action 绑定失败不影响其他

    def _wrap_field_with_hint(self, field: str, widget: QWidget) -> QWidget:
        """把输入控件包一层，在其**后面**放一个低调的「?」信息按钮（参考专业软件做法）。

        旧实现 _make_field_label 把「?」放标签旁、还带 stretch，太显眼。现改为：
        「?」跟在输入框后、固定 18px、灰色小字、hover 才变蓝 —— hover 看摘要、点击弹完整说明。
        无填写说明的字段直接返回原控件（不包，不加「?」）。
        """
        from .field_help import field_help_for
        info = field_help_for(field)
        if not info:
            return widget
        parts: list[str] = []
        if info.get("示例"):
            parts.append(f"填写示例：{info['示例']}")
        if info.get("说明"):
            parts.append(f"说明：{info['说明']}")
        if info.get("其他要求"):
            parts.append(f"其他要求：{info['其他要求']}")
        tip = "\n".join(parts)
        container = QWidget()
        row = QHBoxLayout(container)
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(2)
        row.addWidget(widget, stretch=1)
        hint = QToolButton()
        hint.setText("?")
        hint.setAutoRaise(True)
        hint.setFixedSize(18, 18)
        hint.setFocusPolicy(Qt.NoFocus)
        # 规范化软件设计 2026-05 P1 优化:用 class 选择器(theme.py APP_QSS 内 QToolButton[class="hint"])
        # 替代 inline 灰色 + hover 蓝;每个字段一个按钮,N 个字段共享同一规则省 N×160 字节字符串。
        hint.setProperty("class", "hint")
        hint.setToolTip(tip)
        hint.clicked.connect(
            lambda _=False, f=field, t=tip: QMessageBox.information(self, f"「{f}」填写说明", t)
        )
        row.addWidget(hint)
        return container

    def _apply_photo_filename_fill_shortcut(self) -> None:
        if self._photo_filename_fill_action is None:
            return
        settings = load_settings()
        shortcut = settings.photo_filename_fill_shortcut or DEFAULT_PHOTO_FILENAME_FILL_SHORTCUT
        self._photo_filename_fill_action.setShortcut(QKeySequence(shortcut))
        self._photo_filename_fill_action.setShortcutContext(Qt.ApplicationShortcut)

    # ---- Taxonomy lookup ----

    def _attach_taxonomy_lookup_to_classification_field(self, field: str, line_edit: QLineEdit) -> None:
        completer = FieldValueCompleter(self)
        model = QStandardItemModel(self)
        completer.setModel(model)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setCompletionMode(QCompleter.UnfilteredPopupCompletion)
        completer.setMaxVisibleItems(20)
        completer.activated[QModelIndex].connect(
            lambda index, f=field: self._fill_fields_from_selected_taxonomy_candidate(f, index)
        )
        line_edit.setCompleter(completer)
        line_edit.textChanged.connect(
            lambda text, f=field, w=line_edit: self._refresh_taxonomy_lookup_candidates(f, text, w)
        )
        line_edit.editingFinished.connect(lambda f=field: self._fill_fields_from_unique_taxonomy_match(f))
        self._taxonomy_candidate_models[field] = model
        self._taxonomy_candidate_rows[field] = []

    def _refresh_taxonomy_lookup_candidates(self, field: str, text: str, line_edit: QLineEdit) -> None:
        model = self._taxonomy_candidate_models.get(field)
        if model is None:
            return
        model.clear()
        candidates: list[tuple[str, SpeciesMatch | FamilyMatch]] = []
        self._taxonomy_candidate_rows[field] = candidates
        if self._loading or not text.strip():
            return
        # 未绑定工作区的窗口没有物种预设匹配器，跳过候选计算。
        if self.matcher is None:
            return

        if field in SPECIES_LOOKUP_INPUT_COLUMNS:
            for match in self.matcher.species_matches(text):
                self._add_taxonomy_candidate_to_model(model, candidates, field, "species", match)
        elif field in FAMILY_LOOKUP_INPUT_COLUMNS:
            for match in self.matcher.family_matches(text):
                self._add_taxonomy_candidate_to_model(model, candidates, field, "family", match)

        completer = line_edit.completer()
        if candidates and completer is not None and line_edit.hasFocus():
            completer.complete()

    def _add_taxonomy_candidate_to_model(
        self,
        model: QStandardItemModel,
        candidates: list[tuple[str, SpeciesMatch | FamilyMatch]],
        field: str,
        kind: str,
        match: SpeciesMatch | FamilyMatch,
    ) -> None:
        insert_value = classification_column_value_from_taxonomy_match(field, kind, match)
        if not insert_value:
            return
        item = QStandardItem(format_taxonomy_candidate_label(field, kind, match))
        item.setEditable(False)
        item.setData(insert_value, TAXONOMY_INSERT_TEXT_ROLE)
        item.setData(len(candidates), TAXONOMY_CANDIDATE_ROW_ROLE)
        model.appendRow(item)
        candidates.append((kind, match))

    def _fill_fields_from_selected_taxonomy_candidate(self, field: str, index: QModelIndex) -> None:
        raw_row = index.data(TAXONOMY_CANDIDATE_ROW_ROLE)
        if raw_row is None:
            return
        try:
            row = int(raw_row)
        except (TypeError, ValueError):
            return
        candidates = self._taxonomy_candidate_rows.get(field, [])
        if not 0 <= row < len(candidates):
            return
        kind, match = candidates[row]
        QTimer.singleShot(
            0,
            lambda k=kind, m=match: self._fill_classification_fields_from_taxonomy_match(k, m),
        )

    def _fill_fields_from_unique_taxonomy_match(self, field: str) -> None:
        if self._loading or not self.current_voucher:
            return
        widget = self.class_widgets.get(field)
        if widget is None:
            return
        text = widget.text().strip()
        if not text:
            return
        if field in SPECIES_LOOKUP_INPUT_COLUMNS:
            match = self.matcher.resolve_unique_species(text)
            if match is not None:
                self._fill_classification_fields_from_taxonomy_match("species", match)
        elif field in FAMILY_LOOKUP_INPUT_COLUMNS:
            match = self.matcher.resolve_unique_family(text)
            if match is not None:
                self._fill_classification_fields_from_taxonomy_match("family", match)

    def _fill_classification_fields_from_taxonomy_match(self, kind: str, match: SpeciesMatch | FamilyMatch) -> None:
        if kind == "species" and isinstance(match, SpeciesMatch):
            updates = classification_values_from_species_match(match)
        elif kind == "family" and isinstance(match, FamilyMatch):
            updates = classification_values_from_family_match(match)
        else:
            updates = {}
        self._apply_classification_updates(updates)

    def _apply_classification_updates(self, updates: dict[str, str]) -> None:
        if not self.current_voucher:
            return
        updates = {field: value for field, value in updates.items() if field in self.class_widgets}
        if not updates:
            return
        existing = self.store.get_classification(self.current_voucher) or {}
        conflicts = {
            k: (str(existing[k]), v)
            for k, v in updates.items()
            if existing.get(k) and str(existing[k]) != v
        }
        if conflicts:
            lines = "\n".join(f"  {k}：{old} → {new}" for k, (old, new) in conflicts.items())
            reply = QMessageBox.question(
                self, "确认覆盖",
                f"以下字段已有内容，确认覆盖？\n{lines}",
                QMessageBox.Yes | QMessageBox.No,
            )
            if reply != QMessageBox.Yes:
                return
        self._cancel_pending_classification_saves(updates)
        try:
            self.store.set_fields(
                "classification",
                self.current_voucher,
                updates,
                action_type="classification_autofill",
            )
        except Exception as exc:
            QMessageBox.critical(self, "保存失败", str(exc))
            return

        self._loading = True
        try:
            for field, value in updates.items():
                widget = self.class_widgets.get(field)
                if widget is None:
                    continue
                widget.blockSignals(True)
                widget.setText(value)
                widget.blockSignals(False)
        finally:
            self._loading = False
        self.refresh_list()

    def _cancel_pending_classification_saves(self, updates: dict[str, str]) -> None:
        for field in updates:
            timer = self._save_timers.pop(f"classification:{field}", None)
            if timer is not None:
                timer.stop()

    # ---- Voucher list ----

    def _schedule_list_refresh(self) -> None:
        self._list_refresh_timer.start(300)

    def refresh_list(self) -> None:
        if self.store is None:
            return
        current = self.current_voucher
        overview = self.store.workspace_overview()
        self._all_vouchers = list(overview["vouchers"])
        self._all_flags = dict(overview["flags"])
        self._all_photo_counts = dict(overview["photo_counts"])
        self._all_tube_numbers = dict(overview["tube_numbers"])
        self._all_photo_filenames = dict(overview["photo_filenames"])
        self._refresh_series_selector()
        self._refresh_series_filter_combo()
        self._apply_voucher_filter()
        if current and current in self._all_flags:
            self._select_voucher_in_table(current)
        self._update_dashboard()

    def _refresh_series_selector(self) -> None:
        """刷新系列选择器下拉（新增入库编号时用）。"""
        if not hasattr(self, "_series_selector") or self.store is None:
            return
        combo = self._series_selector
        combo.blockSignals(True)
        combo.clear()
        active = self.store.get_active_series_name()
        for name in self.store.get_all_series_names():
            combo.addItem(name)
        idx = combo.findText(active)
        if idx >= 0:
            combo.setCurrentIndex(idx)
        combo.blockSignals(False)

    def _refresh_series_filter_combo(self) -> None:
        """刷新系列筛选下拉（凭证列表过滤用）。"""
        if not hasattr(self, "_series_filter_combo") or self.store is None:
            return
        combo = self._series_filter_combo
        prev_data = combo.currentData()
        combo.blockSignals(True)
        combo.clear()
        combo.addItem("全部系列", "__all__")
        for name in self.store.get_all_series_names():
            combo.addItem(name, name)
        idx = combo.findData(prev_data)
        combo.setCurrentIndex(idx if idx >= 0 else 0)
        combo.blockSignals(False)

    def _on_series_selector_changed(self, index: int) -> None:
        """切换活跃系列（影响下一个新增入库编号的格式）。"""
        if self.store is None or not hasattr(self, "_series_selector"):
            return
        name = self._series_selector.currentText()
        if name:
            self.store.set_active_series(name)

    def _open_series_manager(self) -> None:
        """打开系列管理对话框。"""
        if self.store is None:
            return
        dlg = AccessionSeriesDialog(self.store, self)
        dlg.exec_()
        self._refresh_series_selector()
        self._refresh_series_filter_combo()

    def _populate_series_switch_menu(self) -> None:
        """动态填充「切换活动系列」子菜单（每次打开前重建，确保与当前配置一致）。"""
        menu = self._series_switch_menu
        menu.clear()
        if self.store is None:
            return
        active = self.store.get_active_series_name()
        all_names = ["YZZ"] + self.store.get_all_series_names()
        for name in all_names:
            act = QAction(name, self, checkable=True)
            act.setChecked(name == active)
            act.triggered.connect(lambda checked, n=name: self._switch_active_series(n))
            menu.addAction(act)

    def _switch_active_series(self, name: str) -> None:
        """从菜单切换活动系列，同步刷新左侧下拉框。"""
        if self.store is None:
            return
        self.store.set_active_series(name)
        self._refresh_series_selector()

    # ── 录入任务门控 ──────────────────────────────────────────────────────────

    def _start_task(self) -> None:
        if self.store is None:
            return
        dlg = _StartTaskDialog(self)
        if dlg.exec_() != QDialog.Accepted:
            return
        import uuid
        from datetime import datetime as _dt
        task_id = uuid.uuid4().hex[:12]
        now = _dt.now().isoformat(timespec="seconds")
        self._active_task = {
            "记录ID": task_id,
            "人员": dlg.person,
            "用途": dlg.purpose,
            "备注": dlg.note,
            "开始时间": now,
            "新增数量": 0,
        }
        self.store.log_alloc_event({
            "记录ID": task_id,
            "时间": now,
            "类型": "任务开始",
            "人员": dlg.person,
            "用途": dlg.purpose,
            "备注": dlg.note,
        })
        self._update_task_indicator()

    def _end_task(self) -> None:
        if self._active_task is None:
            return
        from datetime import datetime as _dt
        self.store.log_alloc_event({
            "记录ID": self._active_task["记录ID"] + "_end",
            "时间": _dt.now().isoformat(timespec="seconds"),
            "类型": "任务结束",
            "人员": self._active_task["人员"],
            "数量": str(self._active_task["新增数量"]),
            "关联任务ID": self._active_task["记录ID"],
        })
        self._active_task = None
        self._update_task_indicator()
        from .models import ALLOC_LOG_FILE
        log_path = self.store.data_dir / ALLOC_LOG_FILE
        self.statusBar().showMessage(f"任务已结束，记录保存至：{log_path}", 8000)

    def _update_task_indicator(self) -> None:
        if not hasattr(self, "_task_label"):
            return
        if self._active_task:
            person = self._active_task["人员"]
            purpose = self._active_task["用途"]
            count = self._active_task["新增数量"]
            self._task_label.setText(f"● {person} · {purpose} · {count}条")
            self._task_label.setStyleSheet("color: #1a7a1a; font-weight: bold;")
            self._task_indicator.setStyleSheet("#task_indicator { background: #d4edda; border-radius: 3px; }")
            self._task_start_btn.setVisible(False)
            self._task_end_btn.setVisible(True)
            self._new_voucher_btn.setEnabled(True)
            self._new_voucher_btn.setToolTip("")
        else:
            self._task_label.setText("未开始录入任务")
            self._task_label.setStyleSheet("color: #888;")
            self._task_indicator.setStyleSheet("")
            self._task_start_btn.setVisible(True)
            self._task_end_btn.setVisible(False)
            self._new_voucher_btn.setEnabled(False)
            self._new_voucher_btn.setToolTip("请先开始录入任务")

    def _open_batch_generate(self) -> None:
        if self.store is None:
            return
        dlg = BatchGenerateDialog(self.store, self)
        dlg.exec_()

    def _open_manual_voucher(self) -> None:
        """Phase 5: 手动添加入库编号 + 规则推断 + 批量生成。"""
        if self.store is None:
            QMessageBox.information(self, "未选择工作区", "请先选择工作区。")
            return
        from .manual_voucher_dialog import ManualVoucherDialog
        dlg = ManualVoucherDialog(self.store, self)
        dlg.exec_()

    # ----------------------------------------------------------------- #
    # 升级中心 v0.8.0 (D1-D20) slot 占位实现。
    # T10/T11/T12 会把这些 stub 替换成 UpgradeCenterDialog 7 tab 接线。
    # 现阶段先让菜单可点 + 显示 "WIP" 提示,方便用户验证菜单生效。
    # ----------------------------------------------------------------- #

    def _open_upgrade_center(self) -> None:
        try:
            from .ui_upgrade import UpgradeCenterDialog
        except ImportError:
            self._upgrade_wip("升级中心")
            return
        dlg = UpgradeCenterDialog(self, initial_tab="overview")
        dlg.exec_()

    def _check_update_now(self) -> None:
        try:
            from .ui_upgrade import UpgradeCenterDialog
        except ImportError:
            self._upgrade_wip("检查更新")
            return
        dlg = UpgradeCenterDialog(self, initial_tab="check")
        dlg.exec_()

    def _open_upgrade_settings(self) -> None:
        try:
            from .ui_upgrade import UpgradeCenterDialog
        except ImportError:
            self._upgrade_wip("自动更新设置")
            return
        dlg = UpgradeCenterDialog(self, initial_tab="settings")
        dlg.exec_()

    def _open_upgrade_about(self) -> None:
        try:
            from .ui_upgrade import open_about_dialog
        except ImportError:
            self._upgrade_about_fallback()
            return
        open_about_dialog(self)

    def _install_from_zip(self) -> None:
        try:
            from .ui_upgrade import UpgradeCenterDialog
        except ImportError:
            self._upgrade_wip("从本地文件安装更新")
            return
        dlg = UpgradeCenterDialog(self, initial_tab="import")
        dlg.exec_()

    def _download_installer(self) -> None:
        try:
            from .ui_upgrade import UpgradeCenterDialog
        except ImportError:
            self._upgrade_wip("下载安装包供分发")
            return
        dlg = UpgradeCenterDialog(self, initial_tab="distribute")
        dlg.exec_()

    def _open_upgrade_history(self) -> None:
        try:
            from .ui_upgrade import UpgradeCenterDialog
        except ImportError:
            self._upgrade_wip("历史版本管理")
            return
        dlg = UpgradeCenterDialog(self, initial_tab="history")
        dlg.exec_()

    def _open_upgrade_build_remote(self) -> None:
        QMessageBox.information(
            self, "升级 → 远程触发 GitHub 构建",
            "远程触发构建预计在 v0.8.1 上线。\n\n"
            "现阶段构建走 push tag v00N 自动触发 GitHub Actions release.yml。",
        )

    def _open_upgrade_build_local(self) -> None:
        QMessageBox.information(
            self, "升级 → 本地重新打包并安装",
            "本地重新打包 + 自动 debug 预计在 v0.8.1 上线。\n\n"
            "现阶段开发者请直接跑 build_release.py:\n"
            "python build_release.py --version 0.8.0",
        )

    def _upgrade_wip(self, label: str) -> None:
        QMessageBox.information(
            self, label,
            f"{label} 功能正在开发中（v0.8.0 进行中）。\n\n"
            f"模块 ui_upgrade.py 尚未实现,菜单已就位但 dialog 待填充。",
        )

    def _upgrade_about_fallback(self) -> None:
        from . import __version__
        from .install_kind import installation_kind, kind_description, upgrade_advice
        kind = installation_kind()
        QMessageBox.information(
            self, "关于当前版本",
            f"标本入库管理 v{__version__}\n\n"
            f"安装方式：{kind_description(kind)}\n"
            f"{upgrade_advice(kind)}",
        )

    def _open_workload_report(self) -> None:
        """工具菜单 → 入库人员记录 = PersonsManagerDialog 默认打开"工作量统计" Tab (Phase 2 复用)。

        旧 WorkloadReportDialog 类仍保留向后兼容,但本入口走 PersonsManagerDialog。
        优势:统一 UI、复用 SpreadsheetPreviewWidget(排序/筛选/复制/Excel+CSV)、
        含明细+汇总+编号分发 三 Tab,数据维度也加了照片 / 首次/末次。
        """
        if self.store is None:
            return
        from .persons_dialog import PersonsManagerDialog
        # initial_tab=1 → 直接显"工作量统计"
        dlg = PersonsManagerDialog(self, workspace=self.workspace_root,
                                   store=self.store, initial_tab=1)
        dlg.exec_()

    def _on_voucher_header_clicked(self, col: int) -> None:
        """Cycle column filter: all -> √ -> × -> all (or all -> 已认领 -> 未认领 -> all for col 4)."""
        if col in (0, 5, 6):  # 入库编号、照片数、关联照片 — no per-column filter
            return
        # Clear quick filter when using column filters
        self._active_filter = "all"
        for btn in self._filter_buttons.values():
            btn.setChecked(False)
        self._filter_buttons["all"].setChecked(True)
        current = self._col_filters.get(col, "")
        if col == 4:  # 认领 column cycles: all -> 已认领 -> 未认领
            cycle = {"": "已认领", "已认领": "未认领", "未认领": ""}
        else:  # specimen/photo/class columns cycle: all -> √ -> ×
            cycle = {"": "√", "√": "×", "×": ""}
        self._col_filters[col] = cycle.get(current, "")
        self._update_header_labels()
        self._voucher_page = 0
        self._apply_voucher_filter()

    def _update_header_labels(self) -> None:
        labels = list(self._col_header_labels)
        for col, val in self._col_filters.items():
            if val:
                labels[col] = f"{self._col_header_labels[col]} {val}"
        self.voucher_table.setHorizontalHeaderLabels(labels)

    def _apply_voucher_filter(self) -> None:
        search = self._voucher_search.text().strip().lower()
        vouchers = self._all_vouchers
        # Apply series filter
        series_sel = self._series_filter_combo.currentData() if hasattr(self, "_series_filter_combo") else None
        if series_sel and series_sel != "__all__":
            from .parsing import parse_voucher_serial
            from .accession_series import series_prefix_of
            if series_sel == "YZZ":
                vouchers = [v for v in vouchers if parse_voucher_serial(v) is not None]
            else:
                vouchers = [v for v in vouchers if series_prefix_of(v) == series_sel]
        # Apply search
        if search:
            scope = self._search_scope.currentText()
            if scope == "入库编号":
                vouchers = [v for v in vouchers if search in v.lower()]
            elif scope == "管内编号":
                vouchers = [v for v in vouchers if search in self._all_tube_numbers.get(v, "").lower()]
            elif scope == "照片名":
                vouchers = [v for v in vouchers if any(search in fn.lower() for fn in self._all_photo_filenames.get(v, []))]
            else:  # 全部：搜索入库编号 + 管内编号 + 照片名
                vouchers = [
                    v for v in vouchers
                    if search in v.lower()
                    or search in self._all_tube_numbers.get(v, "").lower()
                    or any(search in fn.lower() for fn in self._all_photo_filenames.get(v, []))
                ]
        # Apply quick filter
        af = self._active_filter
        if af == "claimed":
            vouchers = [v for v in vouchers if self._all_photo_counts.get(v, 0) > 0]
        elif af == "complete":
            vouchers = [v for v in vouchers if self._all_flags.get(v) and self._all_flags[v].label() == "√√√"]
        elif af == "incomplete":
            vouchers = [v for v in vouchers if self._all_flags.get(v) and self._all_flags[v].label() != "√√√"]
        # Apply per-column filters
        for col, val in self._col_filters.items():
            if not val:
                continue
            if col == 4:  # 认领
                vouchers = [v for v in vouchers if (self._all_photo_counts.get(v, 0) > 0) == (val == "已认领")]
            elif col in (1, 2, 3):  # 标本/照片/分类: filter by √ or ×
                idx = col - 1  # 0=specimen, 1=photo, 2=class
                vouchers = [v for v in vouchers if self._all_flags.get(v) and self._all_flags[v].label()[idx] == val]
        # Paginate
        self._filtered_vouchers = vouchers
        total = len(self._all_vouchers)
        shown = len(vouchers)
        if shown < total:
            self.statusBar().showMessage(f"筛选中：显示 {shown}/{total} 条", 0)
        else:
            self.statusBar().clearMessage()
        total_pages = max(1, (len(vouchers) + self._voucher_page_size - 1) // self._voucher_page_size)
        self._voucher_page = min(self._voucher_page, total_pages - 1)
        start = self._voucher_page * self._voucher_page_size
        page = vouchers[start:start + self._voucher_page_size]
        # Populate table
        self.voucher_table.blockSignals(True)
        self.voucher_table.setRowCount(len(page))
        for i, v in enumerate(page):
            f = self._all_flags.get(v)
            label = f.label() if f else "×××"
            pc = self._all_photo_counts.get(v, 0)
            claimed = "已认领" if pc > 0 else "未认领"
            self.voucher_table.setItem(i, 0, QTableWidgetItem(v))
            self.voucher_table.setItem(i, 1, QTableWidgetItem(label[0]))
            self.voucher_table.setItem(i, 2, QTableWidgetItem(label[1]))
            self.voucher_table.setItem(i, 3, QTableWidgetItem(label[2]))
            self.voucher_table.setItem(i, 4, QTableWidgetItem(claimed))
            self.voucher_table.setItem(i, 5, QTableWidgetItem(str(pc)))
            # 关联照片列：显示逗号分隔的照片文件名（可通过复选框隐藏）
            fnames = self._all_photo_filenames.get(v, [])
            photo_item = QTableWidgetItem("，".join(fnames) if fnames else "")
            photo_item.setToolTip("\n".join(fnames) if fnames else "")
            self.voucher_table.setItem(i, 6, photo_item)
        self.voucher_table.blockSignals(False)
        self._voucher_page_label.setText(f"第 {self._voucher_page + 1}/{total_pages} 页")
        self._voucher_prev_btn.setEnabled(self._voucher_page > 0)
        self._voucher_next_btn.setEnabled(self._voucher_page < total_pages - 1)

    def _set_voucher_filter(self, key: str) -> None:
        for k, btn in self._filter_buttons.items():
            btn.setChecked(k == key)
        self._active_filter = key
        self._col_filters.clear()
        self._update_header_labels()
        self._voucher_page = 0
        self._apply_voucher_filter()

    def _toggle_photo_names_column(self, show: bool) -> None:
        """切换凭证列表中'关联照片'列的显示/隐藏。"""
        self._show_photo_names = show
        if show:
            self.voucher_table.setColumnWidth(6, 200)
        else:
            self.voucher_table.setColumnWidth(6, 0)

    def _voucher_prev_page(self) -> None:
        if self._voucher_page > 0:
            self._voucher_page -= 1
            self._apply_voucher_filter()

    def _voucher_next_page(self) -> None:
        total_pages = max(1, (len(self._filtered_vouchers) + self._voucher_page_size - 1) // self._voucher_page_size)
        if self._voucher_page < total_pages - 1:
            self._voucher_page += 1
            self._apply_voucher_filter()

    def _on_voucher_table_selected(self) -> None:
        rows = self.voucher_table.selectionModel().selectedRows()
        if not rows:
            return
        # 单选时自动跳转到该标本（保持原有行为）；多选时不跳转（避免混乱）
        if len(rows) == 1:
            voucher = self.voucher_table.item(rows[0].row(), 0).text()
            self.select_voucher(voucher)

    def _select_voucher_in_table(self, voucher: str) -> None:
        for row in range(self.voucher_table.rowCount()):
            item = self.voucher_table.item(row, 0)
            if item and item.text() == voucher:
                self.voucher_table.selectRow(row)
                return

    def reveal_voucher(self, voucher: str) -> None:
        """从入库汇总等处跳转过来：必要时清搜索/筛选、翻到目标页、选中并滚动到该行。

        旧问题：调用方只调 select_voucher，只填表单、不动主列表 —— 排序后目标行在别页/被
        筛选掉时看不出跳转。这里把列表也同步过去：selectRow 会触发
        _on_voucher_table_selected -> select_voucher，复用既有加载链路。
        """
        if voucher not in self._all_vouchers:
            return
        filtered = getattr(self, "_filtered_vouchers", self._all_vouchers)
        if voucher not in filtered:
            # 被搜索/快速筛选/列筛选挡住 —— 清掉让目标行可见。
            self._voucher_search.blockSignals(True)
            self._voucher_search.clear()
            self._voucher_search.blockSignals(False)
            self._set_voucher_filter("all")  # 重置快速+列筛选，内部已 _apply_voucher_filter
        if voucher in self._filtered_vouchers:
            self._voucher_page = self._filtered_vouchers.index(voucher) // self._voucher_page_size
            self._apply_voucher_filter()
        for row in range(self.voucher_table.rowCount()):
            item = self.voucher_table.item(row, 0)
            if item and item.text() == voucher:
                self.voucher_table.selectRow(row)
                self.voucher_table.scrollToItem(item)
                return

    def _select_all_vouchers(self) -> None:
        """Ctrl+A 全选凭证列表中当前可见的所有行。"""
        self.voucher_table.selectAll()

    def _update_dashboard(self) -> None:
        if not hasattr(self, "_dashboard_timer"):
            self._dashboard_timer = QTimer(self)
            self._dashboard_timer.setSingleShot(True)
            self._dashboard_timer.timeout.connect(self._compute_dashboard)
        self._dashboard_timer.start(1000)  # debounce 1s

    def _on_current_recorder_changed(self, name_or_special: str) -> None:
        """状态栏当前录入员下拉变化:持久化 + 任务活跃时弹确认。"""
        from .widgets_persons import PersonComboBox
        # 处理 "+ 管理人员…" 特殊项
        if name_or_special == PersonComboBox.SPECIAL_MANAGE:
            self._open_persons_manager()
            # 刷新自身,保留之前选中
            try:
                prev = load_settings().current_recorder
                self._current_recorder_combo.refresh(preselect=prev)
            except Exception:
                self._current_recorder_combo.refresh()
            return
        if not name_or_special:
            return
        # 任务活跃时弹确认
        if self._active_task is not None and self._active_task.get("人员") != name_or_special:
            ret = QMessageBox.question(
                self, "切换录入员",
                f"当前任务录入员是「{self._active_task.get('人员')}」。\n"
                f"切换到「{name_or_special}」会结束当前任务并开始新任务,确认?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
            )
            if ret != QMessageBox.Yes:
                # 还原下拉到旧值
                try:
                    self._current_recorder_combo.refresh(
                        preselect=self._active_task.get("人员", "")
                    )
                except Exception:
                    pass
                return
            # 结束当前任务 + 开新任务
            self._end_task()
            # 自动开始新任务(简化:用相同用途/备注)
            # TODO:可改弹 _StartTaskDialog 让用户确认用途
        # 持久化 current_recorder + update last_used_at
        try:
            settings = load_settings()
            settings.current_recorder = name_or_special
            save_settings(settings)
            from .persons_store import update_last_used
            update_last_used(name_or_special, self.workspace_root)
        except Exception:
            pass

    def _open_new_workspace_window(self) -> None:
        """规范化软件设计 2026-05 多窗口:选不同工作区开新窗口(可写)。

        同工作区已开则 focus 现有,不新开 (WindowManager 内置)。
        """
        if self.manager is None:
            QMessageBox.information(self, "不支持", "无法新建窗口。")
            return
        directory = QFileDialog.getExistingDirectory(self, "选择要打开的工作区")
        if not directory:
            return
        self.manager.open_workspace(Path(directory), read_only=False)

    def _open_readonly_clone(self) -> None:
        """规范化软件设计 2026-05 多窗口:本工作区只读副本(可同时多个,不抢锁)。"""
        if self.manager is None or self.workspace_root is None:
            QMessageBox.information(self, "不支持", "尚未选择工作区。")
            return
        self.manager.open_workspace(self.workspace_root, read_only=True)

    def _open_persons_manager(self) -> None:
        """工具菜单入口 / 状态栏 "管理人员…" 选项 → 打开 PersonsManagerDialog。

        Phase 2 (2026-05): 传 store 让 Tab 2/3/4 显示工作量统计 / 任务明细 / 编号分发。
        """
        from .persons_dialog import PersonsManagerDialog
        dlg = PersonsManagerDialog(self, workspace=self.workspace_root, store=self.store)
        dlg.exec_()
        # 关闭后刷新状态栏下拉
        try:
            prev = load_settings().current_recorder
            self._current_recorder_combo.refresh(preselect=prev)
        except Exception:
            self._current_recorder_combo.refresh()

    def _refresh_memory_status(self) -> None:
        """状态栏:档位 + RSS。规范化软件设计 2026-05 新增,每 5s 刷新。"""
        label = getattr(self, "_memory_status_label", None)
        if label is None:
            return
        try:
            from .env_detect import current_rss_mb
            from .app_settings import load_settings, MEMORY_PROFILE_OPTIONS
            profile = load_settings().memory_profile
            display_full = MEMORY_PROFILE_OPTIONS.get(profile, profile)
            # 取首段(空格前) — "极低 / 低 / 自动 / 高 / 极高"
            display = display_full.split(" ")[0] if display_full else profile
            rss = current_rss_mb()
            rss_txt = f"{rss}MB" if rss is not None else "?"
            label.setText(f"档位:{display} | RSS:{rss_txt}")
        except Exception:
            pass

    def _compute_dashboard(self) -> None:
        if self.store is None:
            return
        all_v = self.store.list_vouchers()
        photo_counts = self.store.voucher_photo_counts()
        total_photos = sum(photo_counts.values())
        claimed = sum(1 for v in all_v if photo_counts.get(v, 0) > 0)
        flags = self.store.all_status_flags()
        complete = sum(1 for v in all_v if flags.get(v) and flags[v].label() == "√√√")
        text = f"照片：{total_photos} 张 | 已认领：{claimed} 个 | 完整：{complete} 个"
        self.dashboard_label.setText(text)
        if hasattr(self, "_status_dashboard"):
            self._status_dashboard.setText(text)

    def select_voucher(self, voucher: str, defer_preview: bool = False) -> None:
        self._save_current_photo_view_state()
        self._loading = True
        self.current_voucher = voucher
        # 动态更新照片面板标题，显示当前入库编号
        if hasattr(self, "_photo_panel_title") and self._photo_panel_title:
            self._photo_panel_title.setText(f"照片信息 — {voucher}")
        specimen = self.store.get_specimen(voucher) or {}
        classification = self.store.get_classification(voucher) or {}
        for field, widget in self.specimen_widgets.items():
            widget.blockSignals(True)
            if isinstance(widget, QComboBox):
                widget.setCurrentText(str(specimen.get(field, "")))
            else:
                widget.setText(str(specimen.get(field, "")))
            widget.blockSignals(False)
        for field, widget in self.class_widgets.items():
            widget.blockSignals(True)
            widget.setText(str(classification.get(field, "")))
            widget.blockSignals(False)
        self.current_photos = self.store.get_photos(voucher)
        self.current_photo_index = min(self.current_photo_index, max(0, len(self.current_photos) - 1))
        self._photo_page = 0
        self._loading = False
        self.refresh_photo_table()
        if defer_preview:
            QTimer.singleShot(250, self.load_current_photo)
        else:
            self.load_current_photo()

    # ---- Field save (debounced) ----

    def _update_auto_save_action_text(self) -> None:
        """工具栏「自动保存」按钮文字直接显示开/关状态（勾选态本身不够醒目）。"""
        action = getattr(self, "_auto_save_action", None)
        if action is not None:
            action.setText("自动保存：开" if self.auto_save_enabled else "自动保存：关")

    def _on_auto_save_toggled(self, checked: bool) -> None:
        """工具栏「自动保存」勾选切换：更新内存开关 + 按钮文字 + 持久化到 settings.json。"""
        self.auto_save_enabled = checked
        self._update_auto_save_action_text()
        settings = load_settings()
        settings.auto_save_enabled = checked
        save_settings(settings)
        self.statusBar().showMessage(
            "已开启自动保存" if checked else "已关闭自动保存（改动需点「保存」按钮写入）", 3000
        )

    def _on_aux_toolbar_toggled(self, checked: bool) -> None:
        """视图菜单「辅助工具栏」勾选切换：显示/隐藏辅栏 + 持久化（规范化软件设计 2026-05 新增）。"""
        if hasattr(self, "_aux_toolbar"):
            self._aux_toolbar.setVisible(bool(checked))
        settings = load_settings()
        settings.aux_toolbar_visible = bool(checked)
        save_settings(settings)

    def _open_toolbar_customize(self) -> None:
        """打开「自定义工具栏」对话框（视图菜单入口）。"""
        from .toolbar_customize import ToolbarCustomizeDialog
        dlg = ToolbarCustomizeDialog(self)
        dlg.exec_()

    def _open_shortcuts_customize(self) -> None:
        """打开「自定义快捷键」对话框（视图菜单入口）。"""
        from .shortcuts_customize import ShortcutsCustomizeDialog
        dlg = ShortcutsCustomizeDialog(self)
        dlg.exec_()

    def schedule_save(self, category: str, field: str) -> None:
        if self._loading or not self.current_voucher:
            return
        voucher = self.current_voucher
        key = f"{category}:{field}"
        if key not in self._save_timers:
            timer = QTimer(self)
            timer.setSingleShot(True)
            timer.timeout.connect(lambda v=voucher, c=category, f=field: self.save_field(c, f, v))
            self._save_timers[key] = timer
        # 旧逻辑：无条件 start(500) 自动保存。现按开关：自动保存关时 timer 仅登记不启动，
        # 这样手动「保存」按钮的 _flush_pending_saves 仍能把改动捞出来写（关开关照样能存）。
        if self.auto_save_enabled:
            self._save_timers[key].start(500)

    def _cancel_pending_saves(self) -> None:
        for timer in self._save_timers.values():
            timer.stop()
        self._save_timers.clear()

    def _flush_pending_saves(self, category: str | None = None) -> int:
        saved = 0
        for key, timer in list(self._save_timers.items()):
            if category is not None and not key.startswith(category + ":"):
                continue
            timer.stop()
            parts = key.split(":", 1)
            if len(parts) == 2:
                self.save_field(parts[0], parts[1], self.current_voucher)
                saved += 1
            self._save_timers.pop(key, None)
        return saved

    def save_field(self, category: str, field: str, voucher: str | None = None) -> None:
        if voucher is None:
            voucher = self.current_voucher
        if not voucher:
            return
        # If the user switched vouchers since the timer was scheduled, discard
        if voucher != self.current_voucher:
            return
        try:
            if category == "specimen":
                widget = self.specimen_widgets[field]
                value = widget.currentText() if isinstance(widget, QComboBox) else widget.text()
                changed = self.store.set_fields("specimen", voucher, {field: value})
                if changed:
                    specimen = self.store.get_specimen(voucher) or {}
                    self._loading = True
                    # 原代码只刷新“采集日期”和“采集地点缩写*”；现在管内编号也会派生“保存方式”。
                    for auto_field in ("采集日期", "采集地点缩写*", "保存方式"):
                        w = self.specimen_widgets[auto_field]
                        w.blockSignals(True)
                        value = str(specimen.get(auto_field, ""))
                        if isinstance(w, QComboBox):
                            w.setCurrentText(value)
                        else:
                            w.setText(value)
                        w.blockSignals(False)
                    self._loading = False
            elif category == "classification":
                self.store.set_fields("classification", voucher, {field: self.class_widgets[field].text()})
            elif category == "photo":
                if field == "描述":
                    self.store.set_photo_description(voucher, self.current_photo_index, self.photo_widgets["描述"].text())
                elif field == "文件名":
                    self.store.set_photo_filename(voucher, self.current_photo_index, self.photo_widgets["文件名"].text())
                self.current_photos = self.store.get_photos(voucher)
                self.refresh_photo_table()
                if field == "文件名":
                    if self.current_photo_index < len(self.current_photos):
                        actual = str(self.current_photos[self.current_photo_index].get("文件名", ""))
                        self.photo_widgets["文件名"].blockSignals(True)
                        self.photo_widgets["文件名"].setText(actual)
                        self.photo_widgets["文件名"].blockSignals(False)
                    cell = self._current_grid_cell()
                    if cell is not None and self.current_photo_index < len(self.current_photos):
                        cell.set_filename(str(self.current_photos[self.current_photo_index].get("文件名", "")))
                    self._refresh_image_index_after_photo_change()
            self._schedule_list_refresh()
        except Exception as exc:
            # Roll back widget to last-known-good value from the store
            try:
                if category == "specimen":
                    row = self.store.get_specimen(voucher) or {}
                    widget = self.specimen_widgets[field]
                    widget.blockSignals(True)
                    stored = str(row.get(field, ""))
                    if isinstance(widget, QComboBox):
                        widget.setCurrentText(stored)
                    else:
                        widget.setText(stored)
                    widget.blockSignals(False)
                elif category == "classification":
                    row = self.store.get_classification(voucher) or {}
                    widget = self.class_widgets[field]
                    widget.blockSignals(True)
                    widget.setText(str(row.get(field, "")))
                    widget.blockSignals(False)
            except Exception:
                pass
            QMessageBox.critical(self, "保存失败", str(exc))

    # ---- Specimen CRUD ----

    # 旧逻辑：工具栏"沿用上条信息"开关变化时静默持久化到 settings.json。
    # 该开关已移除（沿用是固定模式规则），此槽函数不再被任何信号连接，注释保留备查。
    # def _on_carry_over_toggled(self, checked: bool) -> None:
    #     """工具栏"沿用上条信息"开关变化时，静默持久化到 settings.json。"""
    #     try:
    #         settings = load_settings()
    #         settings.carry_over_specimen_fields = bool(checked)
    #         save_settings(settings)
    #     except Exception:
    #         pass  # 持久化失败不影响本次会话内的开关状态

    def new_specimen(self) -> None:
        # 原逻辑：create_specimen() -> refresh_list() -> select_voucher()，新记录字段全空。
        # 现在：新增入库编号时把上一条的标本信息字段（CARRY_OVER_SPECIMEN_FIELDS）
        # 带入新记录，减少重复录入。沿用是本工作模式的固定规则。
        # 需先「开始录入任务」才能新增编号（门控）。
        if not self._active_task:
            return
        try:
            carry: dict[str, str] = {}
            # 旧逻辑：仅当工具栏「沿用上条信息」开关勾选时才沿用。
            # 现在：开关已移除，沿用是固定模式规则，只要存在当前标本就沿用。
            # if (
            #     getattr(self, "_carry_over_action", None) is not None
            #     and self._carry_over_action.isChecked()
            #     and self.current_voucher
            # ):
            if self.current_voucher:
                prev = self.store.get_specimen(self.current_voucher) or {}
                carry = {
                    field: str(prev.get(field, ""))
                    for field in CARRY_OVER_SPECIMEN_FIELDS
                    if str(prev.get(field, "")).strip()
                }
            voucher = self.store.create_specimen()
            if carry:
                self.store.set_fields("specimen", voucher, carry)
            if self._active_task:
                self._active_task["新增数量"] += 1
                self._update_task_indicator()
            self.refresh_list()
            self.select_voucher(voucher)
        except Exception as exc:
            QMessageBox.critical(self, "新增失败", str(exc))

    def clear_photos(self) -> None:
        if not self.current_voucher:
            return
        photos = self.store.get_photos(self.current_voucher)
        if not photos:
            QMessageBox.information(self, "清除照片关联", f"{self.current_voucher} 没有关联的照片。")
            return
        if QMessageBox.question(
            self,
            "清除照片关联",
            f"确定清除 {self.current_voucher} 的全部 {len(photos)} 张照片关联吗？\n\n"
            "入库编号、标本信息和分类信息将保留不变；未被其他记录引用的工作区归档照片文件也会删除。",
            QMessageBox.Yes | QMessageBox.No,
        ) != QMessageBox.Yes:
            return
        count = self.store.clear_photos(self.current_voucher)
        self.statusBar().showMessage(f"已清除 {self.current_voucher} 的 {count} 张照片关联", 3000)
        self._refresh_image_index_after_photo_change()
        self.reload_current()

    def _voucher_context_menu(self, pos) -> None:
        """右键菜单：单行操作 + 多选批量删除。"""
        index = self.voucher_table.indexAt(pos)
        if not index.isValid():
            return
        voucher = self.voucher_table.item(index.row(), 0)
        if not voucher:
            return
        voucher_text = voucher.text().strip()
        if not voucher_text:
            return

        # 获取当前选中的所有行（支持 ExtendedSelection 多选）
        selected_rows = self.voucher_table.selectionModel().selectedRows()
        selected_vouchers: list[str] = []
        for model_index in selected_rows:
            item = self.voucher_table.item(model_index.row(), 0)
            if item and item.text().strip():
                selected_vouchers.append(item.text().strip())

        menu = QMenu(self)
        # 批量导出：将选中的入库编号带入对话框（单行也支持，方便快捷导出当前标本）
        if len(selected_vouchers) >= 1:
            menu.addAction(
                f"批量导出选中 ({len(selected_vouchers)}个)",
                lambda: self._context_batch_export(selected_vouchers),
            )
            menu.addAction(
                f"批量设置标本信息 ({len(selected_vouchers)}个)",
                lambda: self._batch_set_specimen_fields(selected_vouchers),
            )
            menu.addSeparator()
        menu.addAction("清除照片关联", lambda: self._context_clear_photos(voucher_text))
        menu.addSeparator()

        # 单行删除（右键点击的行）
        has_photos = bool(self.store.get_photos(voucher_text))
        if not has_photos:
            menu.addAction("删除入库编号", lambda: self._context_delete_voucher(voucher_text))
        else:
            menu.addAction("删除入库编号（需先清除照片关联）").setEnabled(False)

        # 多选批量删除：仅当选中 >= 2 行时显示
        if len(selected_vouchers) >= 2:
            menu.addSeparator()
            # 统计可删除的（无照片关联的）数量
            eligible = [v for v in selected_vouchers if not self.store.get_photos(v)]
            skipped = len(selected_vouchers) - len(eligible)
            label = f"删除选中的入库编号 ({len(eligible)}个)"
            if skipped:
                label += f"，跳过{skipped}个有照片的"
            action = menu.addAction(label)
            if not eligible:
                action.setEnabled(False)
            else:
                action.triggered.connect(lambda: self._context_batch_delete_vouchers(eligible))

        menu.exec_(self.voucher_table.viewport().mapToGlobal(pos))

    def _context_clear_photos(self, voucher: str) -> None:
        photos = self.store.get_photos(voucher)
        if not photos:
            QMessageBox.information(self, "清除照片关联", f"{voucher} 没有关联的照片。")
            return
        if QMessageBox.question(
            self,
            "清除照片关联",
            f"确定清除 {voucher} 的全部 {len(photos)} 张照片关联吗？\n\n"
            "入库编号、标本信息和分类信息将保留不变；未被其他记录引用的工作区归档照片文件也会删除。",
            QMessageBox.Yes | QMessageBox.No,
        ) != QMessageBox.Yes:
            return
        count = self.store.clear_photos(voucher)
        self.statusBar().showMessage(f"已清除 {voucher} 的 {count} 张照片关联", 3000)
        self._refresh_image_index_after_photo_change()
        self.reload_current()

    def _open_data_in_excel(self) -> None:
        """用 Excel 打开数据目录里的 xlsx（密码门 + 风险提示 + 可选快照）。

        本程序不监控外部改动，且保存时整文件重写会覆盖外部修改 —— 故先输管理密码，
        再明确警告风险、可选建数据快照，最后只打开「数据」目录让用户自己挑文件。
        """
        store = getattr(self, "store", None)
        if store is None:
            QMessageBox.information(self, "未选择工作区", "请先选择工作区再使用此功能。")
            return
        password, ok = QInputDialog.getText(
            self, "用 Excel 打开数据文件",
            "用 Excel 直接编辑数据文件有风险，需管理员操作。\n\n请输入管理密码：",
            QLineEdit.Password,
        )
        if not ok or not password:
            return
        if password != ADMIN_PASSWORD:
            QMessageBox.warning(self, "密码错误", "密码不正确，操作已取消。")
            return
        answer = QMessageBox.warning(
            self, "风险提示",
            "直接用 Excel 修改数据文件存在风险：\n"
            "· 本程序不会自动检测外部改动；\n"
            "· 程序内保存时会整文件重写，可能覆盖你在 Excel 里的修改；\n"
            "· 改动字段名 / 表头 / 编号可能破坏数据兼容性。\n\n"
            "建议：编辑期间不要在本程序内改同一数据；改完后「重新打开工作区」或重启程序。\n\n"
            "是否先创建一个数据快照以便回退？",
            QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
            QMessageBox.Yes,
        )
        if answer == QMessageBox.Cancel:
            return
        if answer == QMessageBox.Yes:
            try:
                snapshot = store.create_data_snapshot(
                    "Excel 外部编辑前快照", "用户用 Excel 打开数据文件前自动快照"
                )
                self.statusBar().showMessage(f"已创建数据快照：{snapshot.name}", 5000)
            except Exception as exc:
                if QMessageBox.warning(
                    self, "快照失败",
                    f"创建数据快照失败：{exc}\n\n仍要继续打开数据目录吗？",
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
                ) != QMessageBox.Yes:
                    return
        _open_path(store.data_dir)
        QMessageBox.information(
            self, "已打开数据目录",
            "已打开「数据」目录，请用 Excel 打开其中的 .xlsx 文件编辑。\n\n"
            "编辑保存后，务必在本程序「重新打开工作区」或重启程序，才能加载最新数据。",
        )

    def _open_worms_match(self) -> None:
        """工具菜单入口：打开 WoRMS 分类匹配窗口（非模态，单实例）。"""
        store = getattr(self, "store", None)
        if store is None:
            QMessageBox.information(self, "未选择工作区", "请先选择工作区再使用此功能。")
            return
        # 缓存守卫(规范化软件设计 2026-05 启动卡死优化):
        # run_app() 现把 WoRMS bootstrap 延后 3s 后台触发(不阻塞启动)。
        # 如用户 3s 内就开 WoRMS,这里同步触发一次 ensure_bootstrap_cache —— 幂等,
        # 已就绪时立刻返回,首次需 ~100ms 解压 188KB sqlite.gz。
        self._ensure_worms_cache_ready()
        from .worms_match import WormsMatchWindow
        # Single-instance: reuse existing visible window or create new one.
        if self._worms_window is None or not self._worms_window.isVisible():
            self._worms_window = WormsMatchWindow(self, store)
        self._worms_window.show()
        self._worms_window.raise_()
        self._worms_window.activateWindow()

    def _ensure_worms_cache_ready(self) -> None:
        """WoRMS slot 守卫:首次调用时同步触发 ensure_bootstrap_cache(幂等)。"""
        try:
            from .worms_client import ensure_bootstrap_cache
            self.statusBar().showMessage("准备 WoRMS 分类数据…", 2000)
            QApplication.processEvents()  # 让 statusBar 消息可见
            if ensure_bootstrap_cache():
                self.statusBar().showMessage("已加载内置 WoRMS 分类缓存", 3000)
        except Exception:
            pass  # bootstrap 失败不阻塞 WoRMS 功能(WoRMS REST API 仍可联网用)

    def _open_worms_db_manager(self) -> None:
        """WoRMS 菜单入口：直接打开本地数据库管理对话框。"""
        self._ensure_worms_cache_ready()  # 守卫:首启 3s 内点开本入口时同步准备缓存
        from .worms_match import _DbManagerDialog
        dlg = _DbManagerDialog(parent=self)
        # C1: 注册到 WindowManager，让主窗口关闭时能接管 dlg 的后台 worker
        if self.manager is not None:
            self.manager.register_dialog_stopper(dlg, dlg._stop_worker)
        try:
            dlg.exec_()
        finally:
            # exec 返回后注销（dlg 自己也会注销，这里二次保证）
            if self.manager is not None:
                self.manager.unregister_dialog_stopper(dlg)

    def _open_worms_browse(self) -> None:
        """打开 WoRMS 窗口并切换到「分类浏览」Tab（index 1）。"""
        self._open_worms_match()
        if self._worms_window:
            from PyQt5.QtWidgets import QTabWidget
            for child in self._worms_window.children():
                if isinstance(child, QTabWidget):
                    child.setCurrentIndex(1)
                    break

    # ---------------------------------------------------------------
    # 帮助菜单 slot 集合（规范化软件设计 2026-05 新增）
    # ---------------------------------------------------------------
    def _open_user_manual_dialog(self) -> None:
        """打开使用说明（单实例 UserManualDialog，QTextBrowser + markdown 渲染）。"""
        from .help_dialog import UserManualDialog, manual_root
        if manual_root() is None:
            QMessageBox.information(
                self,
                "用户手册未找到",
                "未找到 docs/manual/ 目录。若你从源码运行，请确认仓库已拉取完整；\n"
                "若你从安装包运行，请尝试重新安装或联系管理员。",
            )
            return
        if self._manual_dialog is None or not self._manual_dialog.isVisible():
            self._manual_dialog = UserManualDialog(self)
        self._manual_dialog.show()
        self._manual_dialog.raise_()
        self._manual_dialog.activateWindow()

    def _open_field_help_index(self) -> None:
        """打开字段填写说明速查（全字段表格）。"""
        from .help_dialog import FieldHelpIndexDialog
        dlg = FieldHelpIndexDialog(self)
        dlg.exec_()

    def _open_shortcuts_dialog(self) -> None:
        """打开快捷键速查。"""
        from .help_dialog import ShortcutsDialog
        dlg = ShortcutsDialog(self)
        dlg.exec_()

    def _check_github_update_from_help(self) -> None:
        """帮助菜单「检查更新…」：仿 VersionManagerDialog._check_github_update。
        若检测到新版，弹是否打开版本管理对话框下载。
        """
        try:
            from . import updater
        except Exception as exc:
            QMessageBox.warning(self, "检查更新失败", f"更新模块加载失败：{exc}")
            return
        try:
            release = updater.check_latest_release()
        except Exception as exc:
            QMessageBox.warning(
                self,
                "检查更新失败",
                f"无法连接 GitHub Release：{exc}\n请检查网络后重试。",
            )
            return
        if release is None:
            QMessageBox.information(self, "检查更新", "未找到可用的发布版本。")
            return
        if not updater.is_newer(release.version, __version__):
            QMessageBox.information(
                self, "已是最新", f"当前版本 v{__version__} 已是最新。"
            )
            return
        ret = QMessageBox.question(
            self,
            "发现新版本",
            f"GitHub 上有新版 v{release.version}（当前 v{__version__}）。\n"
            "是否打开「版本管理」对话框查看与下载？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )
        if ret == QMessageBox.Yes:
            self.open_version_manager()

    def _open_crash_log_dir(self) -> None:
        """打开崩溃日志目录（~/.specimen_inventory/ 或 %APPDATA%/标本入库管理/）。"""
        try:
            from .app_settings import app_config_dir
            target = app_config_dir()
        except Exception as exc:
            QMessageBox.warning(self, "打开失败", f"无法定位崩溃日志目录：{exc}")
            return
        if target is None:
            QMessageBox.warning(self, "打开失败", "未配置崩溃日志目录。")
            return
        _open_path(Path(target))

    def _open_about_dialog(self) -> None:
        """打开关于对话框。"""
        from .help_dialog import AboutDialog
        dlg = AboutDialog(self)
        dlg.exec_()

    def _downgrade_workspace_schema(self) -> None:
        """将工作区兼容版本降回 1.0.0，以便旧版软件可以打开此工作区。

        旧：无此功能。用户用新版打开工作区后数据版本升至 1.1.x，旧软件看到版本高于其支持的
        1.0.0 就锁死写入，且无任何恢复途径。现：新版软件提供本工具，让用户自助降级。
        降级只改 工作区配置.json 里的版本号，不回滚任何数据内容。
        """
        TARGET_VERSION = "1.0.0"
        store = getattr(self, "store", None)
        if store is None:
            QMessageBox.information(self, "未选择工作区", "请先选择工作区再使用此功能。")
            return
        current_version = str(store.config.get("data_schema_version", "1.0.0"))
        def _ver(v: str) -> tuple[int, ...]:
            parts = []
            for p in str(v).split(".")[:3]:
                try:
                    parts.append(int(p))
                except ValueError:
                    parts.append(0)
            return tuple(parts)

        if _ver(current_version) <= _ver(TARGET_VERSION):
            QMessageBox.information(
                self, "无需操作",
                f"当前工作区兼容版本已为 {current_version}，旧版软件可以直接打开。",
            )
            return
        password, ok = QInputDialog.getText(
            self, "降低工作区兼容版本",
            "此操作将工作区数据版本降至 1.0.0，以便旧版软件（v0.3.x 及以下）可以打开。\n\n"
            "数据内容不会改变，可随时用新版软件重新打开（会自动升回最新版本）。\n\n"
            "请输入管理密码确认：",
            QLineEdit.Password,
        )
        if not ok or not password:
            return
        if password != ADMIN_PASSWORD:
            QMessageBox.warning(self, "密码错误", "密码不正确，操作已取消。")
            return
        answer = QMessageBox.question(
            self, "是否先创建数据快照？",
            f"即将把工作区数据版本从 {current_version} 降至 {TARGET_VERSION}。\n\n"
            "建议先创建数据快照，万一出现问题可以恢复。是否先建快照？",
            QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
            QMessageBox.Yes,
        )
        if answer == QMessageBox.Cancel:
            return
        if answer == QMessageBox.Yes:
            try:
                snapshot = store.create_data_snapshot(
                    "降级兼容版本前快照", f"用户将数据版本从 {current_version} 降至 {TARGET_VERSION} 前自动快照"
                )
                self.statusBar().showMessage(f"已创建数据快照：{snapshot.name}", 5000)
            except Exception as exc:
                if QMessageBox.warning(
                    self, "快照失败",
                    f"创建数据快照失败：{exc}\n\n仍要继续降级吗？",
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
                ) != QMessageBox.Yes:
                    return
        try:
            store.downgrade_schema_version(TARGET_VERSION)
        except Exception as exc:
            QMessageBox.critical(self, "降级失败", f"操作失败：{exc}")
            return
        QMessageBox.information(
            self, "降级完成",
            f"工作区兼容版本已降至 {TARGET_VERSION}。\n\n"
            "旧版软件现在可以打开此工作区。\n"
            "下次用新版软件打开时，版本会自动升回最新。",
        )

    def _context_delete_voucher(self, voucher: str) -> None:
        password, ok = QInputDialog.getText(
            self, "删除入库编号",
            f"删除 {voucher} 将永久移除该编号及其标本信息、分类信息。\n\n请输入管理密码确认删除：",
            QLineEdit.Password,
        )
        if not ok or not password:
            return
        if password != ADMIN_PASSWORD:  # 旧：password != "123"，改引用常量，值不变
            QMessageBox.warning(self, "密码错误", "密码不正确，操作已取消。")
            return
        answer = QMessageBox.warning(
            self, "确认删除",
            f"密码验证通过。\n\n确定要永久删除 {voucher} 吗？\n删除后可通过「撤回」恢复，但超出撤回深度后将永久丢失。",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            return
        self.store.delete_specimen(voucher)
        self.current_voucher = None
        self.refresh_list()
        vouchers = self.store.list_vouchers()
        if vouchers:
            self.select_voucher(vouchers[0])
        self.statusBar().showMessage(f"已删除 {voucher}", 3000)

    def _context_batch_delete_vouchers(self, vouchers: list[str]) -> None:
        """批量删除选中的入库编号（不含照片关联的凭证）。"""
        if not vouchers:
            return
        # 密码验证（一次）
        password, ok = QInputDialog.getText(
            self, "批量删除入库编号",
            f"将删除 {len(vouchers)} 个入库编号及对应的标本信息、分类信息。\n\n"
            f"编号列表：{', '.join(vouchers[:10])}"
            + (f" ...等共{len(vouchers)}个" if len(vouchers) > 10 else "")
            + "\n\n请输入管理密码确认删除：",
            QLineEdit.Password,
        )
        if not ok or not password:
            return
        if password != ADMIN_PASSWORD:  # 旧：password != "123"，改引用常量，值不变
            QMessageBox.warning(self, "密码错误", "密码不正确，操作已取消。")
            return
        # 二次确认
        answer = QMessageBox.warning(
            self, "确认批量删除",
            f"密码验证通过。\n\n确定要永久删除以下 {len(vouchers)} 个入库编号吗？\n"
            + "\n".join(f"  · {v}" for v in vouchers[:20])
            + ("\n  ..." if len(vouchers) > 20 else "")
            + "\n\n删除后可通过「撤回」逐条恢复，但超出撤回深度后将永久丢失。",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            return
        # 逐一删除（每条记录独立可撤回）
        deleted = 0
        for voucher in vouchers:
            self.store.delete_specimen(voucher)
            deleted += 1
        self.current_voucher = None
        self.refresh_list()
        vouchers_remaining = self.store.list_vouchers()
        if vouchers_remaining:
            self.select_voucher(vouchers_remaining[0])
        self.statusBar().showMessage(f"已批量删除 {deleted} 个入库编号", 5000)

    def undo(self) -> None:
        action = self.store.undo_last()
        if action:
            self.statusBar().showMessage(f"已撤回：{action}", 3000)
            self.reload_current()
        else:
            self.statusBar().showMessage("没有可撤回的操作", 2000)

    def redo(self) -> None:
        action = self.store.redo_last()
        if action:
            self.statusBar().showMessage(f"已重做：{action}", 3000)
            self.reload_current()
        else:
            self.statusBar().showMessage("没有可重做的操作", 2000)

    def _undo_redo_counts(self) -> tuple[int, int]:
        """Return (undo_count, redo_count) for display."""
        rows = self.store._read_plain_rows(self.store.data_dir / "操作记录.xlsx")
        if not rows:
            return 0, 0
        undone = sum(1 for r in rows if self.store._value(r, "是否撤销") == "是")
        # Find last non-undone index
        depth = int(self.store.config.get("undo_depth", 200))
        candidates = [r for r in rows[-depth:] if self.store._value(r, "是否撤销") != "是"]
        undo_count = len(candidates)
        redo_count = undone
        return undo_count, redo_count

    def reload_current(self) -> None:
        vouchers = self.store.list_vouchers()
        if self.current_voucher not in vouchers:
            self.current_voucher = vouchers[0] if vouchers else None
        self.refresh_list()
        if self.current_voucher:
            self.select_voucher(self.current_voucher)

    # ---- Photo management ----

    def add_photo(self) -> None:
        if not self.current_voucher:
            QMessageBox.information(self, "请选择标本", "请先选择或新增一个标本。")
            return
        paths, _ = QFileDialog.getOpenFileNames(
            self, "选择照片", "",
            image_file_filter(),
        )
        if paths:
            self.add_photo_paths_async(paths)

    def add_photo_paths(self, paths: list[str], ask_for_outside: bool = True) -> int:
        if not self.current_voucher:
            QMessageBox.information(self, "请选择标本", "请先选择或新增一个标本。")
            return 0
        photo_paths, allow_outside, skipped = self._prepare_photo_paths(paths, ask_for_outside)
        if not photo_paths:
            self._show_skipped_photos(skipped)
            return 0
        photo_paths = self._check_photo_conflicts(photo_paths, self.current_voucher)
        if not photo_paths:
            return 0
        if not self._confirm_archive_name_conflicts(photo_paths):
            return 0
        try:
            mode, library_path = self._photo_management_settings()
            added_rows = self.store.add_photos(
                self.current_voucher,
                photo_paths,
                allow_outside=allow_outside,
                photo_management_mode=mode,
                photo_library_path=library_path,
            )
            added = len(added_rows)
        except Exception as exc:
            QMessageBox.critical(self, "照片关联失败", str(exc))
            return 0
        self._append_added_photos_to_image_index(added_rows)
        self.current_photos = self.store.get_photos(self.current_voucher)
        self.current_photo_index = max(0, len(self.current_photos) - 1)
        self.refresh_photo_table()
        self.load_current_photo()
        self.refresh_list()
        self._show_skipped_photos(skipped)
        return added

    def add_photo_paths_async(self, paths: list[str], ask_for_outside: bool = True) -> None:
        if self._import_job_active:
            QMessageBox.information(self, "照片导入中", "当前已有照片导入任务正在执行。")
            return
        if not self.current_voucher:
            QMessageBox.information(self, "请选择标本", "请先选择或新增一个标本。")
            return
        voucher = self.current_voucher
        photo_paths, allow_outside, skipped = self._prepare_photo_paths(paths, ask_for_outside)
        if not photo_paths:
            self._show_skipped_photos(skipped)
            return
        photo_paths = self._check_photo_conflicts(photo_paths, voucher)
        if not photo_paths:
            return
        if not self._confirm_archive_name_conflicts(photo_paths):
            return
        # Show batch confirmation for 3+ photos
        if len(photo_paths) >= 3:
            dlg = PhotoBatchDialog(voucher, photo_paths, self)
            if dlg.exec_() != QDialog.Accepted or not dlg.confirmed:
                return
        try:
            mode, library_path = self._photo_management_settings()
        except Exception as exc:
            QMessageBox.critical(self, "照片关联失败", str(exc))
            return
        self._import_job_active = True
        self.statusBar().showMessage(f"正在关联 {len(photo_paths)} 张照片...")

        class ImportThread(QThread):
            finished = pyqtSignal(list, str)
            def __init__(self, store, voucher, paths, allow_outside, mode, library_path):
                super().__init__()
                self.store = store
                self.voucher = voucher
                self.paths = paths
                self.allow_outside = allow_outside
                self.mode = mode
                self.library_path = library_path
            def run(self):
                try:
                    rows = self.store.add_photos(
                        self.voucher,
                        self.paths,
                        allow_outside=self.allow_outside,
                        photo_management_mode=self.mode,
                        photo_library_path=self.library_path,
                    )
                    self.finished.emit(rows, "")
                except Exception as e:
                    self.finished.emit([], str(e))

        self._import_thread = ImportThread(self.store, voucher, photo_paths, allow_outside, mode, library_path)
        self._import_thread.finished.connect(
            lambda rows, err: self._on_import_done(voucher, rows, err, skipped)
        )
        self._import_thread.start()

    def _on_import_done(self, voucher: str, rows: list, err: str, skipped: list[str]) -> None:
        self._import_job_active = False
        if err:
            self.statusBar().showMessage("照片关联失败")
            QMessageBox.critical(self, "照片关联失败", err)
            return
        if self.current_voucher == voucher:
            self.current_photos = self.store.get_photos(voucher)
            self.current_photo_index = max(0, len(self.current_photos) - 1)
            self.refresh_photo_table()
            self.load_current_photo()
        self._append_added_photos_to_image_index(rows)
        self.refresh_list()
        self._show_skipped_photos(skipped)
        self.statusBar().showMessage(f"已关联 {len(rows)} 张照片")

    def _append_added_photos_to_image_index(self, rows: list[dict[str, Any]]) -> None:
        paths: list[Path] = []
        for row in rows:
            try:
                path = self.store.resolve_photo_path(row)
            except Exception:
                continue
            if path.exists():
                paths.append(path)
        if paths:
            append_images_to_index(self.workspace_root, paths)
            self.search_index = None
            QTimer.singleShot(200, self._build_search_index_background)

    def _refresh_image_index_after_photo_change(self) -> None:
        clear_image_index()
        self.search_index = None
        QTimer.singleShot(200, lambda: self._build_search_index_background(force_rebuild=True))

    def _prepare_photo_paths(self, paths: list[str], ask_for_outside: bool = True) -> tuple[list[Path], bool, list[str]]:
        skipped: list[str] = []
        photo_paths: list[Path] = []
        for raw in paths:
            path = Path(raw).resolve()
            if not is_supported_image(path):
                skipped.append(path.name)
                continue
            photo_paths.append(path)
        return photo_paths, True, skipped

    def _photo_management_settings(self) -> tuple[str, str]:
        settings = load_settings()
        mode = settings.photo_management_mode if settings.photo_management_mode in PHOTO_MANAGEMENT_OPTIONS else "copy_with_absolute"
        library_path = settings.photo_library_path.strip()
        if mode == "copy_to_custom_library" and not library_path:
            raise ValueError("请先在设置中选择自定义照片库目录。")
        return mode, library_path

    def _show_skipped_photos(self, skipped: list[str]) -> None:
        if skipped:
            QMessageBox.information(self, "部分文件已跳过", "以下文件不是支持的图片格式：\n" + "\n".join(skipped[:20]))

    def _check_photo_conflicts(self, photo_paths: list[Path], voucher: str) -> list[Path]:
        conflicts = self.store.find_photo_conflicts(photo_paths, voucher)
        if not conflicts:
            return photo_paths
        conflict_resolved = {Path(p).resolve() for p in conflicts}
        names = ", ".join(Path(p).name for p in list(conflicts)[:5])
        extra = f" 等{len(conflicts)}张" if len(conflicts) > 5 else ""
        self.statusBar().showMessage(f"已跳过已分配照片：{names}{extra}", 5000)
        return [p for p in photo_paths if p.resolve() not in conflict_resolved]

    def _confirm_archive_name_conflicts(self, photo_paths: list[Path]) -> bool:
        conflicts = self.store.find_archive_name_conflicts(photo_paths)
        if not conflicts:
            return True
        names = "\n".join(f"{Path(src).name}  ->  已存在 {Path(dst).name}" for src, dst in list(conflicts.items())[:10])
        extra = f"\n等 {len(conflicts)} 个同名文件。" if len(conflicts) > 10 else ""
        reply = QMessageBox.question(
            self,
            "照片同名提醒",
            "工作区照片目录中已存在同名但内容不同的照片。\n\n"
            f"{names}{extra}\n\n"
            "继续导入时，软件会自动使用 _2、_3 后缀保存新照片，不会覆盖已有照片。是否继续？",
            QMessageBox.Yes | QMessageBox.No,
        )
        return reply == QMessageBox.Yes

    def delete_photo(self) -> None:
        if not self.current_voucher or not self.current_photos:
            return
        self.delete_photo_at(self.current_photo_index)

    def delete_photo_at(self, photo_index: int) -> None:
        if not self.current_voucher or not self.current_photos:
            return
        if not (0 <= photo_index < len(self.current_photos)):
            return
        # 旧逻辑：单张取消关联点了立即生效、无确认。按需求加确认弹窗，
        # 文案说明只取消关联、不删原始照片（所有单张路径都经此方法，统一一处）。
        if QMessageBox.question(
            self, "取消关联",
            "确定取消关联这张照片吗？\n\n仅取消该照片与本入库编号的关联，不会删除原始照片。",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        ) != QMessageBox.Yes:
            return
        self.store.delete_photo(self.current_voucher, photo_index)
        self.current_photos = self.store.get_photos(self.current_voucher)
        self.current_photo_index = min(photo_index, max(0, len(self.current_photos) - 1))
        if not self.current_photos:
            self._grid_mode_before_expand = ""
        self.refresh_photo_table()
        self.load_current_photo()
        self.refresh_list()
        self._refresh_image_index_after_photo_change()

    def shift_photo(self, delta: int) -> None:
        if not self.current_photos:
            return
        self._save_current_photo_view_state()
        step = delta * self._grid_count() if self._is_grid_mode() else delta
        self.current_photo_index = (self.current_photo_index + step) % len(self.current_photos)
        self.load_current_photo()

    # ---- Photo table ----

    def refresh_photo_table(self) -> None:
        total = len(self.current_photos)
        max_page = max(0, (total - 1) // self._photo_page_size) if total else 0
        self._photo_page = min(self._photo_page, max_page)
        start = self._photo_page * self._photo_page_size
        end = min(start + self._photo_page_size, total)
        page_photos = self.current_photos[start:end]

        self.photo_table.blockSignals(True)
        self.photo_table.setRowCount(len(page_photos))
        for idx, row in enumerate(page_photos):
            real_idx = start + idx
            self.photo_table.setItem(idx, 0, self._photo_table_item(str(real_idx + 1), editable=False))
            self.photo_table.setItem(idx, 1, self._photo_table_item(str(row.get("文件名", "")), editable=True))
            self.photo_table.setItem(idx, 2, self._photo_table_item(str(row.get("相对路径", "")), editable=False))
            self.photo_table.setItem(idx, 3, self._photo_table_item(str(row.get("描述", "")), editable=True))
        self.photo_table.blockSignals(False)
        self._select_photo_table_row()

        if total > self._photo_page_size:
            self._page_label.setText(f"显示 {start + 1}-{end} / 共 {total} 张")
            self._prev_page_btn.setEnabled(self._photo_page > 0)
            self._next_page_btn.setEnabled(self._photo_page < max_page)
            self._page_label.show()
            self._prev_page_btn.show()
            self._next_page_btn.show()
        else:
            self._page_label.setText(f"共 {total} 张")
            self._prev_page_btn.hide()
            self._next_page_btn.hide()

    def _photo_table_item(self, text: str, editable: bool) -> QTableWidgetItem:
        item = QTableWidgetItem(text)
        flags = item.flags() | Qt.ItemIsSelectable | Qt.ItemIsEnabled
        if editable:
            flags |= Qt.ItemIsEditable
        else:
            flags &= ~Qt.ItemIsEditable
        item.setFlags(flags)
        return item

    def _photo_prev_page(self) -> None:
        if self._photo_page > 0:
            self._photo_page -= 1
            self.refresh_photo_table()

    def _photo_next_page(self) -> None:
        total = len(self.current_photos)
        max_page = max(0, (total - 1) // self._photo_page_size) if total else 0
        if self._photo_page < max_page:
            self._photo_page += 1
            self.refresh_photo_table()

    def _select_photo_table_row(self) -> None:
        if not self.current_photos:
            return
        real_idx = min(self.current_photo_index, len(self.current_photos) - 1)
        page_row = real_idx - self._photo_page * self._photo_page_size
        if page_row < 0 or page_row >= self.photo_table.rowCount():
            return
        self.photo_table.blockSignals(True)
        self.photo_table.selectRow(page_row)
        self.photo_table.scrollToItem(self.photo_table.item(page_row, 0))
        self.photo_table.blockSignals(False)

    def _on_photo_table_row_changed(self, row: int, _col: int, _prev_row: int, _prev_col: int) -> None:
        if self._loading or row < 0:
            return
        real_idx = self._photo_page * self._photo_page_size + row
        if real_idx >= len(self.current_photos):
            return
        self._flush_pending_saves()
        self._save_current_photo_view_state()
        self.current_photo_index = real_idx
        self.load_current_photo()

    def _on_photo_table_item_changed(self, item: QTableWidgetItem) -> None:
        if self._loading or not self.current_voucher:
            return
        column_fields = {1: "文件名", 3: "描述"}
        field = column_fields.get(item.column())
        if not field:
            return
        real_idx = self._photo_page * self._photo_page_size + item.row()
        if real_idx < 0 or real_idx >= len(self.current_photos):
            return
        value = item.text()
        try:
            # 表格内编辑与下方照片信息栏共用同一套保存逻辑，避免两处显示不一致。
            if field == "文件名":
                self.store.set_photo_filename(self.current_voucher, real_idx, value)
            else:
                self.store.set_photo_description(self.current_voucher, real_idx, value)
            self.current_photos = self.store.get_photos(self.current_voucher)
            self.refresh_photo_table()
            if real_idx == self.current_photo_index:
                widget = self.photo_widgets[field]
                widget.blockSignals(True)
                widget.setText(str(self.current_photos[real_idx].get(field, "")))
                widget.blockSignals(False)
                if field == "文件名":
                    cell = self._current_grid_cell()
                    if cell is not None:
                        cell.set_filename(str(self.current_photos[real_idx].get("文件名", "")))
                    self._refresh_image_index_after_photo_change()
            self._schedule_list_refresh()
        except Exception as exc:
            self.current_photos = self.store.get_photos(self.current_voucher)
            fallback = ""
            if 0 <= real_idx < len(self.current_photos):
                fallback = str(self.current_photos[real_idx].get(field, ""))
            self.photo_table.blockSignals(True)
            item.setText(fallback)
            self.photo_table.blockSignals(False)
            QMessageBox.critical(self, "保存失败", str(exc))

    def _photo_table_context_menu(self, pos) -> None:
        clicked = self.photo_table.itemAt(pos)
        if clicked is not None and not clicked.isSelected():
            self.photo_table.clearSelection()
            self.photo_table.setCurrentItem(clicked)
            clicked.setSelected(True)
        menu = QMenu(self)
        rows = sorted(set(idx.row() for idx in self.photo_table.selectedItems()))
        if not rows:
            return
        real_rows = [self._photo_page * self._photo_page_size + r for r in rows if r < self.photo_table.rowCount()]
        menu.addAction("复制选中内容", self._copy_photo_table_selection)
        if real_rows:
            menu.addAction("复制文件名", lambda: self._copy_photo_filename(real_rows[0]))
            menu.addAction("复制相对路径", lambda: self.copy_photo_relative_path(real_rows[0]))
            menu.addAction("复制绝对路径", lambda: self.copy_photo_absolute_path(real_rows[0]))
            fill_action = menu.addAction(
                "从照片文件名填充标本信息",
                lambda idx=real_rows[0]: self.fill_photo_from_filename(idx),
            )
            if self._photo_filename_fill_action is not None:
                fill_action.setShortcut(self._photo_filename_fill_action.shortcut())
        menu.addSeparator()
        if len(rows) == 1:
            menu.addAction("编辑文件名", lambda: self._edit_photo_table_cell(real_rows[0], 1))
            menu.addAction("编辑描述", lambda: self._edit_photo_table_cell(real_rows[0], 3))
            menu.addSeparator()
            menu.addAction("打开原图", lambda: self.open_current_photo_external(real_rows[0]))
            menu.addAction("打开原图所在位置", lambda: self.open_photo_location(real_rows[0]))
            menu.addAction("替换此照片", self._replace_current_photo)
            # 旧标签「删除此照片」/「删除选中的 N 张照片」——实为取消关联，原始照片不动。
            menu.addAction("取消关联此照片", self.delete_photo)
        else:
            menu.addAction(f"取消关联选中的 {len(rows)} 张照片", self._delete_selected_photos)
        # 「移动到其他编号」：原本只有工具栏「分配入库编号」入口，入库汇总精简后
        # 把照片「移动到」能力补进右键菜单（复用已存在的 _assign_voucher_to_selected）。
        menu.addSeparator()
        menu.addAction("移动到其他编号", self._assign_voucher_to_selected)
        menu.exec_(self.photo_table.viewport().mapToGlobal(pos))

    def _copy_photo_filename(self, real_idx: int) -> None:
        if 0 <= real_idx < len(self.current_photos):
            name = str(self.current_photos[real_idx].get("文件名", ""))
            QApplication.clipboard().setText(name)

    def _filename_context_menu(self, widget: QLineEdit, pos) -> None:
        menu = widget.createStandardContextMenu()
        menu.addSeparator()
        menu.addAction("复制完整文件名", lambda: QApplication.clipboard().setText(widget.text()))
        menu.addAction("全选文件名", widget.selectAll)
        menu.addSeparator()
        action = menu.addAction("从照片文件名填充标本信息", self.fill_current_photo_from_filename)
        if self._photo_filename_fill_action is not None:
            action.setShortcut(self._photo_filename_fill_action.shortcut())
        menu.exec_(widget.mapToGlobal(pos))

    def _copy_photo_table_selection(self) -> bool:
        indexes = self.photo_table.selectedIndexes()
        if not indexes:
            item = self.photo_table.currentItem()
            if item is None:
                return False
            QApplication.clipboard().setText(item.text())
            return True
        rows = sorted({idx.row() for idx in indexes})
        cols = sorted({idx.column() for idx in indexes})
        lines: list[str] = []
        for row in rows:
            values: list[str] = []
            for col in cols:
                if any(idx.row() == row and idx.column() == col for idx in indexes):
                    cell = self.photo_table.item(row, col)
                    values.append(cell.text() if cell else "")
                else:
                    values.append("")
            lines.append("\t".join(values))
        QApplication.clipboard().setText("\n".join(lines))
        return True

    def _edit_current_photo_table_item(self) -> bool:
        item = self.photo_table.currentItem()
        if item is None:
            return False
        if item.column() not in (1, 3):
            return False
        self.photo_table.editItem(item)
        return True

    def _edit_photo_table_cell(self, real_idx: int, column: int) -> None:
        row = real_idx - self._photo_page * self._photo_page_size
        if row < 0 or row >= self.photo_table.rowCount():
            return
        item = self.photo_table.item(row, column)
        if item is None:
            return
        self.photo_table.setCurrentItem(item)
        self.photo_table.editItem(item)

    def _replace_current_photo(self) -> None:
        if not self.current_voucher or self.current_photo_index >= len(self.current_photos):
            return
        paths, _ = QFileDialog.getOpenFileNames(self, "选择替换照片", "", image_file_filter())
        if not paths:
            return
        new_path = Path(paths[0]).resolve()
        if not is_supported_image(new_path):
            QMessageBox.warning(self, "格式不支持", "仅支持图片文件。")
            return
        if not self._confirm_archive_name_conflicts([new_path]):
            return
        try:
            mode, library_path = self._photo_management_settings()
            new_row = self.store.replace_photo(
                self.current_voucher,
                self.current_photo_index,
                new_path,
                allow_outside=True,
                photo_management_mode=mode,
                photo_library_path=library_path,
            )
        except Exception as exc:
            QMessageBox.critical(self, "替换失败", str(exc))
            return
        if new_row:
            self.current_photos = self.store.get_photos(self.current_voucher)
            self.refresh_photo_table()
            self.load_current_photo()
            self.refresh_list()
            self._append_added_photos_to_image_index([new_row])
            self._refresh_image_index_after_photo_change()

    def _delete_selected_photos(self) -> None:
        if not self.current_voucher or not self.current_photos:
            return
        # 宫格模式：用宫格多选集合；列表模式：用 photo_table 选中行（旧逻辑）。
        if self._is_grid_mode():
            real_indices = [i for i in sorted(self._grid_selected_indices) if 0 <= i < len(self.current_photos)]
        else:
            rows = sorted(set(idx.row() for idx in self.photo_table.selectedItems()), reverse=True)
            if not rows:
                return
            real_indices = [self._photo_page * self._photo_page_size + r for r in rows]
            real_indices = [i for i in real_indices if 0 <= i < len(self.current_photos)]
        if not real_indices:
            return
        # 旧文案：标题「批量删除」/「确定删除选中的 N 张照片？」——措辞像删文件。
        # 实为取消关联（原始照片不动），改名澄清。
        answer = QMessageBox.question(
            self, "批量取消关联",
            f"确定取消关联选中的 {len(real_indices)} 张照片吗？\n\n仅取消关联，不会删除原始照片。",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            return
        all_photos = self.store.get_photos(self.current_voucher)
        for idx in sorted(real_indices, reverse=True):
            if idx < len(all_photos):
                self.store.delete_photo(self.current_voucher, idx)
        self.current_photos = self.store.get_photos(self.current_voucher)
        self.current_photo_index = min(self.current_photo_index, max(0, len(self.current_photos) - 1))
        self.refresh_photo_table()
        self.load_current_photo()
        self.refresh_list()

    # ---- Photo rendering ----

    def _preview_size(self) -> tuple[int, int] | None:
        if not hasattr(self, '_cached_preview_quality'):
            self._cached_preview_quality = load_settings().preview_quality
        return PREVIEW_QUALITY_SIZES.get(self._cached_preview_quality, (800, 600))

    def _is_grid_mode(self) -> bool:
        return self._view_mode != "单张"

    def _grid_count(self) -> int:
        for count in (2, 4, 6, 8):
            if self._view_mode.startswith(str(count)):
                return count
        return 1

    def _on_view_mode_changed(self, text: str) -> None:
        self._grid_mode_before_expand = ""
        self._return_grid_btn.hide()
        # 手动切模式后，单张视图双击恢复为 适配↔100%（不再返回宫格）。
        self.photo_view.set_double_click_returns(False)
        self._view_mode = text
        self._save_current_photo_view_state()
        self.load_current_photo()

    def _set_view_mode(self, count: int) -> None:
        for btn in self._view_buttons:
            btn.setChecked(btn.property("grid_count") == count)
        mode = "单张" if count == 1 else f"{count}宫格"
        self._on_view_mode_changed(mode)

    def _sync_view_buttons(self, count: int) -> None:
        for btn in self._view_buttons:
            btn.blockSignals(True)
            btn.setChecked(btn.property("grid_count") == count)
            btn.blockSignals(False)

    def _on_show_grid_filenames_changed(self, state: int) -> None:
        self._show_grid_filenames = state == Qt.Checked
        for cell in self._grid_labels:
            cell.set_filename_visible(self._show_grid_filenames)
        try:
            settings = load_settings()
            settings.show_grid_filenames = self._show_grid_filenames
            save_settings(settings)
        except Exception:
            pass

    def load_current_photo(self) -> None:
        self._loading = True
        for widget in self.photo_widgets.values():
            widget.blockSignals(True)
            widget.setText("")
            widget.blockSignals(False)

        self._placeholder_label.hide()
        self.photo_view.clear_image()
        self._clear_grid()
        self._current_qpixmap = None
        self._zoom_label.setText("")

        if not self.current_photos:
            self.photo_counter.setText("0 / 0")
            self._placeholder_label.setText("暂无照片")
            self._placeholder_label.show()
            self._loading = False
            return

        row = self.current_photos[self.current_photo_index]
        self._select_photo_table_row()
        mode_suffix = f" · {self._view_mode}" if self._is_grid_mode() else ""
        self.photo_counter.setText(f"{self.current_photo_index + 1} / {len(self.current_photos)}{mode_suffix}")
        self.photo_widgets["文件名"].blockSignals(True)
        self.photo_widgets["文件名"].setText(str(row.get("文件名", "")))
        self.photo_widgets["文件名"].blockSignals(False)
        self.photo_widgets["相对路径"].blockSignals(True)
        self.photo_widgets["相对路径"].setText(str(row.get("相对路径", "")))
        self.photo_widgets["相对路径"].blockSignals(False)
        self.photo_widgets["绝对路径"].blockSignals(True)
        self.photo_widgets["绝对路径"].setText(str(row.get("绝对路径", "")))
        self.photo_widgets["绝对路径"].blockSignals(False)
        self.photo_widgets["描述"].blockSignals(True)
        self.photo_widgets["描述"].setText(str(row.get("描述", "")))
        self.photo_widgets["描述"].blockSignals(False)

        if self._is_grid_mode():
            self._render_grid()
            self._loading = False
            return

        path = self.store.resolve_photo_path(row)
        self._photo_load_token += 1
        token = self._photo_load_token
        self._placeholder_label.setText("正在加载预览...")
        self._placeholder_label.show()
        self.statusBar().showMessage(f"正在加载照片：{path.name}")
        size = self._preview_size() or (2800, 2200)
        self._thumb_worker.clear_pending()
        self._thumb_worker.enqueue(path, size, token)
        self._loading = False

    def _on_thumbnail_ready(self, token: int, qpixmap: QPixmap | None, exc: Exception | None) -> None:
        # Check if this is a single-photo load
        if token == self._photo_load_token and not self._is_grid_mode():
            self._finish_single_photo_load(token, qpixmap, exc)
            return
        # Check grid loads
        if token in self._grid_requests:
            slot = self._grid_requests.pop(token)
            if token >= self._grid_load_token * 100:
                self._finish_grid_thumbnail(slot, qpixmap, exc)

    def _finish_single_photo_load(self, token: int, qpixmap: QPixmap | None, exc: Exception | None) -> None:
        self._placeholder_label.hide()
        if exc:
            self._current_qpixmap = None
            self._placeholder_label.setText(f"无法预览照片\n{exc}")
            self._placeholder_label.setStyleSheet("color: #8b2f2f; font-size: 12px;")
            self._placeholder_label.show()
            self.statusBar().showMessage("照片预览加载失败")
            return
        if qpixmap is None:
            return
        self._current_qpixmap = qpixmap
        self.photo_view.set_image(qpixmap)
        self.photo_view.show()
        self.grid_frame.hide()
        self.statusBar().showMessage("就绪")

    def _render_grid(self) -> None:
        """Render photo grid view.

        根据实际显示的照片数量自适应网格形状，而非固定使用按钮计数。
        例如：只有1张照片时选6宫格 → 1×1；2张时 → 2×1。
        """
        self._clear_grid()
        self._grid_load_token += 1
        token = self._grid_load_token
        if not self.current_photos:
            return
        self._thumb_worker.clear_pending()
        count = self._grid_count()
        # 计算当前页的照片索引范围
        page_start = (self.current_photo_index // count) * count
        page_indices = list(range(page_start, min(page_start + count, len(self.current_photos))))
        # 原代码：grid_shape(count) 始终基于按钮数量，1张照片也按3×2显示 → 变形
        # 改为基于实际照片数自适应：1→1×1, 2→2×1, 3-4→2×2, 5-6→3×2, 7-8→4×2
        cols, rows = grid_shape(len(page_indices))

        # 重新渲染宫格时，多选状态重置为当前张（翻页/切 voucher/切宫格数后无残留）。
        self._grid_selected_indices = {self.current_photo_index}

        self.photo_view.hide()
        self._placeholder_label.hide()
        self.grid_frame.show()

        for slot, photo_index in enumerate(page_indices):
            r = slot // cols
            c = slot % cols
            photo = self.current_photos[photo_index]
            label = GridPhotoCell(photo_index, self._show_grid_filenames)
            label.clicked.connect(self._on_grid_cell_clicked)
            label.double_clicked.connect(self._on_grid_cell_double_clicked)
            label.right_clicked.connect(self._show_grid_context_menu)
            label.zoom_changed.connect(self._on_grid_cell_zoom_changed)
            # 旧逻辑：label.set_selected(photo_index == self.current_photo_index)
            # 现支持多选，按集合判断高亮。
            label.set_selected(photo_index in self._grid_selected_indices)
            path = self.store.resolve_photo_path(photo)
            filename = str(photo.get("文件名", ""))
            label.set_filename(filename)
            if path.exists():
                self._grid_requests[token * 100 + slot] = slot
                # 缩略图尺寸与网格单元格匹配，使用固定的预览尺寸
                self._thumb_worker.enqueue(path, (320, 240), token * 100 + slot)
            else:
                label.set_error("照片不存在")
            self.grid_layout.addWidget(label, r, c)
            self._grid_labels.append(label)

        page_text = f"{page_start + 1}-{page_indices[-1] + 1} / {len(self.current_photos)}" if page_indices else f"0 / {len(self.current_photos)}"
        self.photo_counter.setText(f"{page_text} · {self._view_mode}")
        self._zoom_label.setText("")

    def _clear_grid(self) -> None:
        self._grid_requests.clear()
        for label in self._grid_labels:
            self.grid_layout.removeWidget(label)
            label.deleteLater()
        self._grid_labels.clear()

    def _finish_grid_thumbnail(self, slot: int, qpixmap: QPixmap | None, exc: Exception | None) -> None:
        if slot >= len(self._grid_labels):
            return
        label = self._grid_labels[slot]
        if not label:
            return
        if exc or qpixmap is None:
            label.set_error("无法预览")
            return
        label.set_pixmap(qpixmap)

    def _on_grid_cell_clicked(self, photo_index: int) -> None:
        if photo_index < 0 or photo_index >= len(self.current_photos):
            return
        self._flush_pending_saves()
        # 宫格多选：Ctrl 切换单格、Shift 范围选（限当前宫格页）、无修饰键单选。
        mods = QApplication.keyboardModifiers()
        count = self._grid_count()
        page_start = (self.current_photo_index // count) * count
        page_end = min(page_start + count, len(self.current_photos))
        page_range = set(range(page_start, page_end))
        if mods & Qt.ControlModifier:
            if photo_index in self._grid_selected_indices:
                self._grid_selected_indices.discard(photo_index)
            else:
                self._grid_selected_indices.add(photo_index)
            if not self._grid_selected_indices:  # 不允许空集
                self._grid_selected_indices.add(photo_index)
        elif mods & Qt.ShiftModifier:
            lo, hi = sorted((self.current_photo_index, photo_index))
            self._grid_selected_indices |= (set(range(lo, hi + 1)) & page_range)
        else:
            self._grid_selected_indices = {photo_index}
        self.current_photo_index = photo_index
        self._select_photo_table_row()
        self._populate_photo_fields(self.current_photos[photo_index])
        self._refresh_grid_selection()

    def _on_grid_cell_double_clicked(self, photo_index: int) -> None:
        self.enlarge_photo_from_grid(photo_index)

    def _on_grid_cell_zoom_changed(self, photo_index: int, zoom: float) -> None:
        if photo_index == self.current_photo_index:
            self._zoom_label.setText(f"宫格 {int(zoom * 100)}%")

    def _populate_photo_fields(self, row: dict[str, Any]) -> None:
        for field in ("文件名", "相对路径", "绝对路径", "描述"):
            widget = self.photo_widgets[field]
            widget.blockSignals(True)
            widget.setText(str(row.get(field, "")))
            widget.blockSignals(False)

    def _refresh_grid_selection(self) -> None:
        # 旧逻辑：仅高亮 current_photo_index 一格；现按多选集合高亮。
        for cell in self._grid_labels:
            cell.set_selected(cell.photo_index in self._grid_selected_indices)

    def _current_grid_cell(self) -> GridPhotoCell | None:
        for cell in self._grid_labels:
            if cell.photo_index == self.current_photo_index:
                return cell
        return None

    def _show_grid_context_menu(self, photo_index: int, event) -> None:
        # 右键未选中的格：仿 photo_table，先把它设为唯一选中再出菜单。
        if photo_index not in self._grid_selected_indices:
            self._grid_selected_indices = {photo_index}
            self.current_photo_index = photo_index
            self._refresh_grid_selection()
        indices = sorted(self._grid_selected_indices)
        menu = QMenu(self)
        if len(indices) == 1:
            idx = indices[0]
            self.current_photo_index = idx  # 替换/放大依赖 current_photo_index
            menu.addAction("放大显示", lambda: self.enlarge_photo_from_grid(idx))
            menu.addAction("打开原图", lambda: self.open_current_photo_external(idx))
            menu.addAction("打开原图所在位置", lambda: self.open_photo_location(idx))
            menu.addAction("替换此照片", self._replace_current_photo)
            # 旧标签「删除此照片」——实为取消关联（原始照片不动）。
            menu.addAction("取消关联此照片", lambda: self.delete_photo_at(idx))
            menu.addAction("移动到其他编号", self._assign_voucher_to_selected)
            menu.addSeparator()
            fill_action = menu.addAction("从照片文件名填充标本信息", lambda: self.fill_photo_from_filename(idx))
            if self._photo_filename_fill_action is not None:
                fill_action.setShortcut(self._photo_filename_fill_action.shortcut())
            menu.addAction("复制相对路径", lambda: self.copy_photo_relative_path(idx))
            menu.addAction("复制绝对路径", lambda: self.copy_photo_absolute_path(idx))
        else:
            menu.addAction("打开原图", lambda: self.open_current_photo_external(indices[0]))
            menu.addAction(f"取消关联选中的 {len(indices)} 张照片", self._delete_selected_photos)
            menu.addAction("移动到其他编号", self._assign_voucher_to_selected)
        menu.exec_(event.globalPos())

    def _show_single_preview_context_menu(self, global_pos) -> None:
        """单张预览右键菜单：对 current_photo_index 出照片管理菜单。"""
        if not self.current_photos:
            return
        idx = self.current_photo_index
        if idx < 0 or idx >= len(self.current_photos):
            return
        menu = QMenu(self)
        menu.addAction("打开原图", lambda: self.open_current_photo_external(idx))
        menu.addAction("打开原图所在位置", lambda: self.open_photo_location(idx))
        menu.addAction("替换此照片", self._replace_current_photo)
        # 旧标签「删除此照片」——实为取消关联（原始照片不动）。
        menu.addAction("取消关联此照片", lambda: self.delete_photo_at(idx))
        menu.addAction("移动到其他编号", self._assign_voucher_to_selected)
        menu.addSeparator()
        fill_action = menu.addAction("从照片文件名填充标本信息", lambda: self.fill_photo_from_filename(idx))
        if self._photo_filename_fill_action is not None:
            fill_action.setShortcut(self._photo_filename_fill_action.shortcut())
        menu.addAction("复制相对路径", lambda: self.copy_photo_relative_path(idx))
        menu.addAction("复制绝对路径", lambda: self.copy_photo_absolute_path(idx))
        if self._grid_mode_before_expand:
            menu.addSeparator()
            menu.addAction("返回宫格", self.return_to_grid)
        menu.exec_(global_pos)

    def _get_selected_photo_indices(self) -> list[int]:
        """Return list of currently selected photo indices across all views."""
        if not self.current_photos:
            return []
        if self._is_grid_mode():
            # 旧逻辑：宫格单选，按 page 窗口取 current_photo_index 一个。
            # 现宫格支持 Ctrl/Shift 多选，返回选中集合（空则回退当前张）。
            return sorted(self._grid_selected_indices) or [self.current_photo_index]
        else:
            return [self.current_photo_index]

    def _assign_voucher_to_selected(self) -> None:
        """Assign selected photos to a (new or existing) voucher number."""
        if not self.current_photos:
            QMessageBox.information(self, "提示", "当前没有可分配的照片。")
            return
        indices = self._get_selected_photo_indices()
        if not indices:
            QMessageBox.information(self, "提示", "请先在预览区选择要分配的照片。")
            return
        # Show dialog
        voucher, ok = QInputDialog.getText(
            self, "分配入库编号", "输入目标入库编号（留空自动新建）：", QLineEdit.Normal, ""
        )
        if not ok:
            return
        voucher = voucher.strip()
        if not voucher:
            # Create new voucher
            try:
                voucher = self.store.create_specimen()
            except Exception as exc:
                QMessageBox.critical(self, "创建失败", str(exc))
                return
        elif voucher not in self.store.list_vouchers():
            QMessageBox.warning(self, "编号不存在", f"入库编号 {voucher} 不存在，请先创建或使用已有编号。")
            return
        # 原代码逐张 add_photo 后 delete_photo；add 失败时仍可能删除源记录。
        moved = self.store.move_photos(self.current_voucher or "", voucher, indices)
        self.current_photos = self.store.get_photos(self.current_voucher or "")
        self.current_photo_index = min(self.current_photo_index, max(0, len(self.current_photos) - 1))
        self.refresh_photo_table()
        self.load_current_photo()
        self.refresh_list()
        self.statusBar().showMessage(f"已将 {moved} 张照片分配给 {voucher}", 5000)

    def _save_panel(self, category: str) -> None:
        """Save all pending changes for a specific panel."""
        if not self.current_voucher:
            return
        saved = self._flush_pending_saves(category)
        names = {"specimen": "标本信息", "photo": "照片信息", "classification": "分类信息"}
        self.statusBar().showMessage(f"{names.get(category, category)}已保存 ({saved} 项)", 2000)

    def _save_all_panels(self) -> None:
        """Save all pending changes across all panels."""
        for category in ("specimen", "photo", "classification"):
            self._save_panel(category)
        self.statusBar().showMessage("所有信息已保存", 3000)

    def _confirm_save(self) -> None:
        """Force-save any pending field changes and show confirmation."""
        self._save_all_panels()

    def fill_current_photo_from_filename(self) -> None:
        self.fill_photo_from_filename(self.current_photo_index)

    def _photo_filename_info(self, photo_index: int) -> tuple[dict[str, str], str]:
        if photo_index < 0 or photo_index >= len(self.current_photos):
            return {}, ""
        filename = photo_filename_source_for_specimen_fill(self.current_photos[photo_index])
        return specimen_updates_from_photo_filename(filename), filename

    def fill_photo_from_filename(self, photo_index: int) -> bool:
        if not self.current_voucher:
            return False
        self._flush_pending_saves()
        self.current_photos = self.store.get_photos(self.current_voucher)
        updates, filename = self._photo_filename_info(photo_index)
        if not updates:
            QMessageBox.information(self, "无法填充", "该照片文件名中未找到规范的编号、日期或保存方式。")
            return False

        current_specimen = self.store.get_specimen(self.current_voucher) or {}
        dialog = PhotoFilenameFillDialog(filename, updates, current_specimen, self)
        if dialog.exec_() != QDialog.Accepted:
            return False
        selected = dialog.selected_updates()
        if not selected:
            self.statusBar().showMessage("未选择要填充的标本字段", 2500)
            return False

        try:
            changed = self.store.set_fields(
                "specimen",
                self.current_voucher,
                selected,
                action_type="photo_filename_fill",
                auto_derive_specimen_fields=False,
            )
        except Exception as exc:
            QMessageBox.critical(self, "填充失败", str(exc))
            return False

        specimen = self.store.get_specimen(self.current_voucher) or {}
        self._loading = True
        try:
            for field in PHOTO_FILENAME_FILL_FIELDS:
                widget = self.specimen_widgets.get(field)
                if widget is None:
                    continue
                value = str(specimen.get(field, "") or "")
                widget.blockSignals(True)
                if isinstance(widget, QComboBox):
                    widget.setCurrentText(value)
                else:
                    widget.setText(value)
                widget.blockSignals(False)
        finally:
            self._loading = False

        if changed:
            self.refresh_list()
            fields = "、".join(selected.keys())
            self.statusBar().showMessage(f"已从照片文件名填充：{fields}", 3000)
        else:
            self.statusBar().showMessage("标本信息没有变化", 2500)
        return changed

    def adjust_zoom(self, factor: float) -> None:
        if self._is_grid_mode():
            cell = self._current_grid_cell()
            if cell:
                cell.zoom(factor)
                self._zoom_label.setText(f"宫格 {cell.zoom_percent()}%")
            return
        self.photo_view.zoom(factor)

    def fit_image(self) -> None:
        if self._is_grid_mode():
            cell = self._current_grid_cell()
            if cell:
                cell.fit_to_window()
                self._zoom_label.setText("宫格 100%")
            return
        self.photo_view.fit_to_window()

    def open_current_photo_external(self, photo_index: int | None = None) -> None:
        if not self.current_photos:
            return
        index = self.current_photo_index if photo_index is None else photo_index
        if index < 0 or index >= len(self.current_photos):
            return
        path = self.store.resolve_photo_path(self.current_photos[index])
        if not path.exists():
            QMessageBox.critical(self, "照片不存在", str(path))
            return
        if not is_supported_image(path):
            QMessageBox.critical(self, "安全限制", "仅支持打开图片文件。")
            return
        # 旧逻辑：_open_path(path) 直接系统默认程序；现走 open_image_external —— 优先用
        # 设置里的自定义图片查看器，未设/无效则回退系统默认。
        open_image_external(path)

    def enlarge_photo_from_grid(self, photo_index: int) -> None:
        if photo_index < 0 or photo_index >= len(self.current_photos):
            return
        if self._is_grid_mode():
            self._grid_mode_before_expand = self._view_mode
        self.current_photo_index = photo_index
        self._view_mode = "单张"
        self._sync_view_buttons(1)
        self._return_grid_btn.show()
        # 此状态下单张视图双击 -> 返回原宫格（见 PhotoGraphicsView.return_requested）。
        self.photo_view.set_double_click_returns(True)
        self.refresh_photo_table()
        self.load_current_photo()

    def return_to_grid(self) -> None:
        if not self._grid_mode_before_expand:
            return
        mode = self._grid_mode_before_expand
        self._grid_mode_before_expand = ""
        self._view_mode = mode
        count = 1
        for _, c, _ in VIEW_MODES:
            if (f"{c}宫格" if c > 1 else "单张") == mode:
                count = c
                break
        self._sync_view_buttons(count)
        self._return_grid_btn.hide()
        # 回到宫格，单张视图双击恢复为 适配↔100%。
        self.photo_view.set_double_click_returns(False)
        self._save_current_photo_view_state()
        self.load_current_photo()

    def copy_photo_relative_path(self, photo_index: int) -> None:
        if photo_index < 0 or photo_index >= len(self.current_photos):
            return
        QApplication.clipboard().setText(str(self.current_photos[photo_index].get("相对路径", "")))

    def copy_photo_absolute_path(self, photo_index: int) -> None:
        if photo_index < 0 or photo_index >= len(self.current_photos):
            return
        row = self.current_photos[photo_index]
        QApplication.clipboard().setText(str(row.get("绝对路径", "") or self.store.resolve_photo_path(row)))

    def open_photo_location(self, photo_index: int) -> None:
        if photo_index < 0 or photo_index >= len(self.current_photos):
            return
        directory = self.store.resolve_photo_path(self.current_photos[photo_index]).parent
        if not directory.exists():
            QMessageBox.critical(self, "目录不存在", str(directory))
            return
        _open_path(directory)

    def _on_grid_drop(self, event) -> None:
        self.grid_frame.setStyleSheet("")
        paths = []
        for url in event.mimeData().urls():
            if url.isLocalFile():
                paths.append(url.toLocalFile())
        if paths:
            self.add_photo_paths_async(paths)
        event.acceptProposedAction()

    def _on_zoom_changed(self, level: float) -> None:
        if not self._is_grid_mode():
            self._zoom_label.setText(f"| {int(level * 100)}%")

    # ---- Photo view state persistence ----

    def _photo_state_key(self) -> str:
        if not self.current_photos:
            return ""
        row = self.current_photos[self.current_photo_index]
        return f"{row.get('入库编号*','')}|{row.get('相对路径','')}|{self.current_photo_index}"

    def _save_current_photo_view_state(self) -> None:
        key = self._photo_state_key()
        if key:
            # 防止长时间浏览后字典无限增长：超 500 条时淘汰前半部旧条目。
            if len(self._photo_view_states) >= 500:
                for _k in list(self._photo_view_states)[:250]:
                    del self._photo_view_states[_k]
            self._photo_view_states[key] = (1.0, 0, 0)

    def _view_state_for_current_photo(self) -> tuple[float, int, int]:
        return self._photo_view_states.get(self._photo_state_key(), (1.0, 0, 0))

    # ---- Workspace operations ----

    def import_workspace(self) -> None:
        source = QFileDialog.getExistingDirectory(self, "选择要导入的旧工作区")
        if not source:
            return
        # 规范化软件设计 2026-05 P1 审查修复:加 QProgressDialog + processEvents 让 UI 不卡。
        progress_dlg = QProgressDialog("正在导入工作区…", "", 0, 0, self)
        progress_dlg.setWindowTitle("导入工作区")
        progress_dlg.setCancelButton(None)
        progress_dlg.setMinimumDuration(0)
        progress_dlg.setModal(True)
        progress_dlg.show()
        QApplication.processEvents()
        try:
            result = self.store.import_workspace(source)
            progress_dlg.close()
            # 原代码导入后保留旧图片索引;新导入照片或图谱目录需要重新建索引才能被检索到。
            self.search_index = None
            clear_image_index()
            QTimer.singleShot(200, self._build_search_index_background)
            message = f"导入 {result.imported} 个标本,跳过 {result.skipped} 个重复记录,关联照片 {result.photos_imported} 张。"
            if result.report_path:
                message += f"\n缺失照片报告:{result.report_path}"
            QMessageBox.information(self, "导入完成", message)
            self.refresh_list()
        except ImportConflictError as exc:
            progress_dlg.close()
            detail = str(exc)
            if exc.report_path:
                detail += f"\n冲突报告:{exc.report_path}"
            QMessageBox.critical(self, "导入已阻止", detail)
        except Exception as exc:
            progress_dlg.close()
            QMessageBox.critical(self, "导入失败", str(exc))

    def _open_aggregate_incoming(self) -> None:
        """工具菜单入口：从 incoming/ 一键聚合所有子目录到中心机（M1 含 P1 降级模式）。

        合并核心走现有 `ExcelStore.import_workspace`，所以指纹冲突 / 编号唯一性 / 照片
        物理去重 全部沿用已有保护。本入口只是"批量循环 + 分流归档 + 跨机锁"封装。
        无 `manifest.json` 的子目录同样能被吃（P1 等价：任何含 `数据/` 子目录即合法源）。
        密码门控复用 ADMIN_PASSWORD，与"用 Excel 打开数据文件…"等管理操作一致。
        """
        store = getattr(self, "store", None)
        if store is None:
            QMessageBox.information(self, "未选择工作区", "请先选择中心机工作区再使用此功能。")
            return
        incoming_dir = QFileDialog.getExistingDirectory(self, "选择收件箱目录（incoming）")
        if not incoming_dir:
            return
        incoming_path = Path(incoming_dir)
        password, ok = QInputDialog.getText(
            self, "从收件箱聚合",
            "本操作会把收件箱里所有「含 数据/ 子目录」的文件夹合并到当前中心机。\n"
            "合并前会自动创建快照，可一键回退。\n\n"
            "请输入管理密码以继续：",
            QLineEdit.Password,
        )
        if not ok or not password:
            return
        if password != ADMIN_PASSWORD:
            QMessageBox.warning(self, "密码错误", "密码不正确，操作已取消。")
            return
        # S7: 先 dry-run 预览,让用户看到预计结果再决定是否真合并
        # 规范化软件设计 2026-05 P1 审查修复:大工作区预扫可能耗时,加 QProgressDialog 防 UI 冻。
        preview_dlg = QProgressDialog("正在预扫收件箱…", "", 0, 0, self)
        preview_dlg.setWindowTitle("预扫")
        preview_dlg.setCancelButton(None)
        preview_dlg.setMinimumDuration(0)
        preview_dlg.setModal(True)
        preview_dlg.show()
        QApplication.processEvents()
        try:
            preview = preview_aggregate(store, incoming_path)
        except Exception as exc:
            preview_dlg.close()
            QMessageBox.critical(self, "预览失败", f"预扫出错:{exc}")
            return
        preview_dlg.close()
        if preview.total_candidates == 0:
            QMessageBox.information(
                self, "预览结果",
                "在该目录下未发现可合并的子目录（要求子目录含 数据/）。",
            )
            return
        # 拼预览文本（统计 + 头 10 个候选明细）
        lines = [
            f"候选子目录：{preview.total_candidates}",
            "",
            f"预计新增 voucher：{preview.predicted_new_vouchers}",
            f"预计重复跳过：{preview.predicted_skipped_vouchers}",
            f"预计冲突：{preview.predicted_conflicts}",
            f"预计照片新增：{preview.predicted_photos}",
            f"预计跨 voucher 同 SHA256 照片：{preview.predicted_cross_voucher_duplicates}",
            f"预计同名不同内容照片：{preview.predicted_name_conflicts}",
            "",
            "明细：",
        ]
        for name, outcome, _count, note in preview.candidates[:10]:
            lines.append(f"  · {name} [{outcome}]: {note}")
        if len(preview.candidates) > 10:
            lines.append(f"  …（其余 {len(preview.candidates) - 10} 个省略）")
        lines.append("")
        lines.append("以上是只读预览 — 数据未被修改。是否继续真正合并？")
        answer = QMessageBox.question(
            self, "合并预览（只读，未动数据）",
            "\n".join(lines),
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            return
        # 规范化软件设计 2026-05 P1 审查修复:
        # 旧:aggregate_incoming 主线程同步,大工作区合并卡 UI 30s+。
        # 现:用 QProgressDialog + processEvents,通过 progress_cb 回调让事件循环跑,
        #    UI 保持响应。完整 QThread 重构留后续(成本高)。
        progress_dlg = QProgressDialog("正在聚合收件箱…", "", 0, 0, self)
        progress_dlg.setWindowTitle("从收件箱聚合")
        progress_dlg.setCancelButton(None)  # 不支持中途取消(底层不可中断)
        progress_dlg.setMinimumDuration(0)
        progress_dlg.setModal(True)
        progress_dlg.show()
        QApplication.processEvents()

        def _on_progress(stage: str, current: int, total: int) -> None:
            if total > 0:
                progress_dlg.setMaximum(total)
                progress_dlg.setValue(current)
            progress_dlg.setLabelText(f"{stage}({current}/{total if total > 0 else '?'})")
            QApplication.processEvents()

        try:
            # aggregate_incoming 支持 progress_cb;传入让 UI 周期 processEvents
            try:
                report = aggregate_incoming(store, incoming_path, progress_cb=_on_progress)
            except TypeError:
                # 老版签名无 progress_cb 时回落
                report = aggregate_incoming(store, incoming_path)
        except Exception as exc:
            progress_dlg.close()
            QMessageBox.critical(self, "聚合失败", f"聚合过程出错:{exc}")
            return
        progress_dlg.close()
        # 导入后旧图片索引可能不再覆盖新照片，刷新主窗口并后台重建索引。
        self.search_index = None
        clear_image_index()
        QTimer.singleShot(200, self._build_search_index_background)
        self.refresh_list()
        lines: list[str] = []
        if not report.processed and not report.conflicted and not report.errored:
            lines.append("收件箱里没有可合并的子目录（要求子目录含 数据/）。")
        else:
            lines.append(
                f"成功合并：{len(report.processed)} 个子目录 — "
                f"共 {report.total_imported} 条 voucher、{report.total_photos} 张照片。"
            )
        if report.conflicted:
            lines.append(
                f"\n冲突待人工处理：{len(report.conflicted)} 个 — 已移至 conflicts/，"
                "请查看冲突报告 xlsx。"
            )
        if report.errored:
            lines.append(
                f"出错待排查：{len(report.errored)} 个 — 已移至 errors/，"
                "请查看 error.log。"
            )
        if report.duplicates:
            lines.append(
                f"\n跨 voucher 同 SHA256 照片审核：{len(report.duplicates)} 个子目录命中 — "
                "已写报告到 duplicates/，主管审核后决定是否补登。"
            )
        if report.name_conflicts:
            lines.append(
                f"\n同名不同内容照片：{len(report.name_conflicts)} 个文件名 — "
                f"已写报告到 name_conflicts/{report.name_conflicts_report_path.name if report.name_conflicts_report_path else ''}。"
            )
        if report.snapshot_path is not None:
            lines.append(f"\n已自动快照：{report.snapshot_path.name}（可在 工具→数据版本… 回退）")
        QMessageBox.information(self, "聚合完成", "\n".join(lines))

    def _open_import_examples(self) -> None:
        """打开合并/导入示例文档 — 4 种合并场景操作示例。

        旧（v0.5.0+）：定位 `docs/import-merge-examples.md`。
        现（规范化软件设计 2026-05）：docs 重排后路径变 `docs/manual/import-merge.md`；
        保留旧路径作 fallback 以兼容老安装包（升级前的 release 仍带旧文件名）。
        多根解析顺序：源 / PyInstaller `_MEIPASS` / `_internal/` / cwd。
        """
        import sys as _sys
        # 主路径（规范化后）；旧路径作 fallback
        rel_paths = [
            Path("docs") / "manual" / "import-merge.md",
            Path("docs") / "import-merge-examples.md",
        ]
        candidates: list[Path] = []
        for rel in rel_paths:
            # 源代码根（specimen_app/../<rel>）
            candidates.append(Path(__file__).resolve().parent.parent / rel)
            # PyInstaller 打包根
            meipass = getattr(_sys, "_MEIPASS", None)
            if meipass:
                candidates.append(Path(meipass) / rel)
            # onedir frozen
            if getattr(_sys, "frozen", False):
                candidates.append(Path(_sys.executable).resolve().parent / "_internal" / rel)
            # 工作目录
            candidates.append(Path.cwd() / rel)
            # 工作区目录（用户放了一份）
            if getattr(self, "workspace_root", None):
                candidates.append(Path(self.workspace_root) / rel)
        for p in candidates:
            if p.exists():
                _open_path(p)  # 模块级 helper（不是 self.method）
                return
        QMessageBox.information(
            self, "文档未找到",
            "未在以下位置找到合并/导入示例：\n" + "\n".join(str(c) for c in candidates),
        )

    def _open_batch_import_sources(self) -> None:
        """S2：弹 BatchImportSourcesDialog 让主管多选源工作区目录直接合并。

        与「从收件箱聚合」区别：不需要把源目录先复制到一个统一 incoming/，
        而是支持在文件管理器里直接逐个选 D:\\ydy\\ 和 D:\\yss\\。源目录原样保留。
        """
        store = getattr(self, "store", None)
        if store is None:
            QMessageBox.information(self, "未选择工作区", "请先打开中心机工作区再使用此功能。")
            return
        dlg = BatchImportSourcesDialog(store, self)
        dlg.exec_()
        # 完成后刷新主窗口（dlg 内部有自己的结果弹窗）
        self.search_index = None
        clear_image_index()
        QTimer.singleShot(200, self._build_search_index_background)
        self.refresh_list()

    def _upgrade_workspace_to_multi_user(self) -> None:
        """M5：把当前旧版工作区升级到多人协作格式。

        升级范围 = 工作区配置.json 加两个键（`multi_user_protocol_version` /
        `legacy_yzz_segment`），不动任何 Excel 数据 / 照片 / 已有快照。
        升级前强制 snapshot，可一键回退（工具菜单 → 操作记录 → 撤销）。
        升级后旧工作区即可被「从收件箱聚合」吃下，或与新工作区一起合并。
        """
        store = getattr(self, "store", None)
        if store is None:
            QMessageBox.information(self, "未选择工作区", "请先打开工作区再使用此功能。")
            return
        # 已升级则只提示，不重复操作
        if store.config.get("multi_user_protocol_version"):
            QMessageBox.information(
                self, "无需升级",
                f"该工作区已是多人协作格式（v{store.config.get('multi_user_protocol_version')}）。",
            )
            return
        # 旧工作区 + 含数据 才提示用户升级；空工作区直接帮加标记
        is_legacy = store.detect_legacy_workspace()
        if is_legacy:
            answer = QMessageBox.question(
                self, "升级到多人协作格式",
                "将把当前工作区升级到多人协作格式。\n\n"
                "升级内容：\n"
                "  · 仅在 工作区配置.json 写入两个标记键（不动任何已有 Excel / 照片）；\n"
                "  · 自动创建升级前快照，可一键回退；\n"
                "  · 升级后工作区可用「从收件箱聚合」吃外部子目录，或作为子目录发给主管聚合。\n\n"
                "是否继续？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes,
            )
            if answer != QMessageBox.Yes:
                return
        try:
            summary = store.upgrade_to_multi_user_protocol()
        except Exception as exc:
            QMessageBox.critical(self, "升级失败", f"升级出错：{exc}")
            return
        if summary.get("already_upgraded"):
            QMessageBox.information(self, "无需升级", "该工作区已是多人协作格式。")
            return
        snap = summary.get("snapshot_path")
        seg = summary.get("legacy_yzz_segment", [1, 0])
        QMessageBox.information(
            self, "升级完成",
            f"工作区已升级到多人协作格式（v{summary.get('multi_user_protocol_version')}）。\n\n"
            f"已自动快照：{snap.name if snap else '（无）'}\n"
            f"历史 YZZ 段范围：[{seg[0]}, {seg[1]}]（升级前已分配的连号区间，便于追溯）\n\n"
            f"可在「工具 → 操作记录」找到本次 upgrade_to_multi_user_protocol 行。",
        )

    def _bulk_apply_exif(self) -> None:
        """A2：扫所有"有照片但采集日期为空"的 voucher，从第一张照片读 EXIF 自动回填。

        - 仅填空字段（已有「采集日期」的 voucher 不动）
        - 跳过没有照片的 voucher
        - 照片无 EXIF / EXIF 无 DateTimeOriginal → 跳过该 voucher
        """
        store = getattr(self, "store", None)
        if store is None:
            QMessageBox.information(self, "未选择工作区", "请先打开工作区再使用此功能。")
            return
        # 找候选
        candidates: list[tuple[str, dict]] = []
        for voucher in store.list_vouchers():
            spec = store.get_specimen(voucher) or {}
            if str(spec.get("采集日期", "") or "").strip():
                continue
            photos = store.get_photos(voucher)
            if not photos:
                continue
            candidates.append((voucher, photos[0]))
        if not candidates:
            QMessageBox.information(
                self, "无需回填",
                "所有 voucher 的采集日期都已填，或没有可用照片。",
            )
            return
        answer = QMessageBox.question(
            self, "批量回填确认",
            f"将从照片 EXIF 的 DateTimeOriginal 读取拍摄日期，\n"
            f"回填到 {len(candidates)} 条 voucher 的「采集日期」字段。\n\n"
            f"仅填空字段，不覆盖已有值。\n\n是否继续？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )
        if answer != QMessageBox.Yes:
            return
        filled = 0
        no_exif = 0
        unreachable = 0
        for voucher, photo_row in candidates:
            try:
                photo_path = store.resolve_photo_path(photo_row)
            except Exception:
                unreachable += 1
                continue
            if not photo_path or not photo_path.exists():
                unreachable += 1
                continue
            try:
                result = apply_exif_to_specimen(store, voucher, photo_path)
            except Exception:
                no_exif += 1
                continue
            if "采集日期" in result.get("filled_fields", []):
                filled += 1
            else:
                no_exif += 1
        self.refresh_list()
        QMessageBox.information(
            self, "EXIF 批量回填完成",
            f"已回填采集日期：{filled} 条\n"
            f"照片无 EXIF / 无 DateTimeOriginal：{no_exif} 条\n"
            f"照片无法访问：{unreachable} 条",
        )

    def _export_dwc_archive(self) -> None:
        """A1：把当前工作区导出成 Darwin Core Archive (DwC-A) zip。

        DwC-A 是 TDWG/GBIF 通用生物多样性数据交换格式，含 meta.xml + occurrence.txt +
        multimedia.txt + eml.xml。可直接用 GBIF IPT、iDigBio Validator 验证，或上传到
        GBIF 节点。**纯只读导出，不动工作区**。
        """
        store = getattr(self, "store", None)
        if store is None:
            QMessageBox.information(self, "未选择工作区", "请先打开工作区再使用此功能。")
            return
        voucher_count = len(store.list_vouchers())
        if voucher_count == 0:
            answer = QMessageBox.question(
                self, "工作区为空",
                "当前工作区没有任何 voucher 记录。仍要导出空 archive 吗？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if answer != QMessageBox.Yes:
                return
        # 默认文件名带日期与计数
        from datetime import datetime as _dt
        default_name = f"dwca_{_dt.now().strftime('%Y%m%d')}_{voucher_count}records.zip"
        zip_path_str, _ = QFileDialog.getSaveFileName(
            self, "导出 Darwin Core Archive",
            default_name,
            "Darwin Core Archive (*.zip)",
        )
        if not zip_path_str:
            return
        # 让用户填可选 dataset title / creator（默认用工作区配置）
        title, ok1 = QInputDialog.getText(
            self, "数据集标题",
            "数据集标题（出现在 eml.xml，可空，默认 \"Specimen Inventory Workspace Export\"）：",
        )
        if not ok1:
            return
        creator, ok2 = QInputDialog.getText(
            self, "数据集创建者",
            "数据集创建者（可空）：",
        )
        if not ok2:
            return
        try:
            result_path = export_dwc_archive(
                store, zip_path_str,
                dataset_title=title.strip() or "Specimen Inventory Workspace Export",
                dataset_creator=creator.strip(),
            )
        except FileExistsError as exc:
            QMessageBox.warning(self, "目标已存在", str(exc))
            return
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", f"DwC 导出出错：{exc}")
            return
        QMessageBox.information(
            self, "导出完成",
            f"已生成 Darwin Core Archive：\n{result_path}\n\n"
            f"包含 {voucher_count} 条 occurrence。可用 GBIF IPT / iDigBio Validator 校验，\n"
            f"或上传到 GBIF / iDigBio 节点。",
        )

    def export_data(self) -> None:
        """原有导出功能：导出全部数据到单个 Excel 文件。"""
        default_path = str(self.workspace_root / "标本数据导出.xlsx")
        path, _ = QFileDialog.getSaveFileName(self, "导出数据", default_path, "Excel 文件 (*.xlsx)")
        if not path:
            return
        try:
            count = self.store.export_all_data(Path(path))
            QMessageBox.information(self, "导出完成", f"已导出 {count} 条记录到\n{path}")
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", str(exc))

    def open_batch_export(self, preselected: list[str] | None = None) -> None:
        """打开批量导出对话框（工具栏按钮入口）。

        类似 NCBI Batch Entrez：可选择导出标本信息、分类信息、照片路径和照片文件。
        支持从当前凭证列表多选带入编号，也可在对话框中手动粘贴。
        """
        # 如果没有传入预选编号，尝试从当前凭证表获取选中的行
        if preselected is None:
            rows = self.voucher_table.selectionModel().selectedRows()
            preselected = []
            for model_index in rows:
                item = self.voucher_table.item(model_index.row(), 0)
                if item and item.text().strip():
                    preselected.append(item.text().strip())

        dlg = BatchExportDialog(self.store, preselected=preselected, parent=self)
        dlg.exec_()

    def _context_batch_export(self, vouchers: list[str]) -> None:
        """右键菜单入口：将选中的入库编号带入批量导出对话框。"""
        self.open_batch_export(preselected=vouchers)

    def _batch_set_specimen_fields(self, vouchers: list[str]) -> None:
        """右键菜单入口：对多选的入库编号批量设置标本信息字段。"""
        if not vouchers:
            return
        dlg = BatchSpecimenFieldsDialog(self)
        if dlg.exec_() != QDialog.Accepted:
            return
        updates = dlg.selected_updates()
        if not updates:
            QMessageBox.information(self, "批量设置", "未勾选任何要修改的字段。")
            return
        fields_label = "、".join(updates.keys())
        if QMessageBox.question(
            self, "确认批量设置",
            f"将对 {len(vouchers)} 个入库编号写入字段：{fields_label}\n确定吗？",
            QMessageBox.Yes | QMessageBox.No,
        ) != QMessageBox.Yes:
            return
        # 复用 store.set_fields：每个编号写一条操作记录，可逐条撤回（与批量删除一致）。
        try:
            changed = 0
            for voucher in vouchers:
                if self.store.set_fields("specimen", voucher, updates):
                    changed += 1
        except Exception as exc:
            QMessageBox.critical(self, "批量设置失败", str(exc))
            return
        self.refresh_list()
        if self.current_voucher in vouchers:
            self.select_voucher(self.current_voucher)
        self.statusBar().showMessage(f"已批量设置 {changed} 个入库编号的标本信息", 5000)

    def import_data_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "导入数据", "", "Excel 文件 (*.xlsx)")
        if not path:
            return
        try:
            result = self.store.import_from_file(Path(path))
            QMessageBox.information(self, "导入完成", f"导入 {result.imported} 个标本，跳过 {result.skipped} 个重复记录。")
            self.refresh_list()
        except ImportConflictError as exc:
            detail = str(exc)
            if exc.report_path:
                detail += f"\n冲突报告：{exc.report_path}"
            QMessageBox.critical(self, "导入已阻止", detail)
        except Exception as exc:
            QMessageBox.critical(self, "导入失败", str(exc))

    def switch_workspace(self) -> None:
        target = QFileDialog.getExistingDirectory(self, "选择工作区目录")
        if not target:
            return
        target_path = Path(target).resolve()
        # self.store 为 None 时（未绑定的首次启动窗口）没有"当前工作区"，跳过相等判断。
        if self.store is not None and target_path == self.workspace_root:
            return
        if self.manager is not None and self.manager.focus_workspace(target_path, exclude=self):
            self.statusBar().showMessage("该工作区已在其他窗口打开", 3000)
            return
        if is_generated_workspace_path(target_path):
            QMessageBox.critical(
                self, "不能使用软件目录",
                f"不能把 build/dist/releases 等软件构建或版本目录作为工作区：\n{target_path}\n\n请选择实际保存数据和照片的工作目录。",
            )
            return
        # 与 _prepare_workspace_candidate 对齐：拒绝文件系统根/盘符根/用户主目录。
        # 旧逻辑：switch_workspace 缺此检查，仅 __init__ 的工作区准备路径有。
        if is_unsafe_workspace_root(target_path):
            QMessageBox.critical(
                self, "目录范围过大",
                f"不能把文件系统根目录、盘符根目录或用户主目录作为工作区：\n{target_path}\n\n"
                "这类目录过大，软件的全工作区扫描（如图片索引）会遍历海量文件、可能拖垮电脑。\n"
                "请选择实际保存数据和照片的子目录。",
            )
            return
        create_workspace_files = False
        if not has_workspace_data(target_path):
            if QMessageBox.question(
                self, "初始化工作区",
                f"该目录还没有数据文件：\n{target_path}\n\n是否在此目录创建新的数据文件夹和 Excel 数据文件？",
                QMessageBox.Yes | QMessageBox.No,
            ) != QMessageBox.Yes:
                return
            initialize_workspace(target_path, self._template_source_for_initialization(target_path))
            create_workspace_files = True
        self._load_workspace_into_window(target_path, create_workspace_files)

    def _load_workspace_into_window(self, target_path: Path, create_files: bool) -> bool:
        """把工作区载入当前窗口：构建 ExcelStore 并替换 self.store、刷新所有相关状态。

        由 switch_workspace（切换工作区）和首次启动的 _prompt_initial_workspace 共用。
        返回是否成功载入。
        """
        # ExcelStore 构建 + 锁处理。锁处理逻辑原本只在 __init__ 里，这里并入，
        # 让首次启动的工作区选择也能处理"工作区被占用"。
        try:
            new_store = ExcelStore(target_path, lock=True, create_if_missing=create_files)
        except WorkspaceLockedError as exc:
            lock_path = target_path / "数据" / ".workspace.lock"
            msg = f'{exc}\n\n如果软件已退出但仍然被占用，可以点击"强制解锁"。'
            btn = QMessageBox.critical(self, "工作区被占用", msg, QMessageBox.Abort | QMessageBox.Retry)
            if btn == QMessageBox.Retry and lock_path.exists():
                try:
                    lock_path.unlink()
                    new_store = ExcelStore(target_path, lock=True, create_if_missing=create_files)
                except Exception as exc2:
                    QMessageBox.critical(self, "切换失败", str(exc2))
                    return False
            else:
                return False
        except WorkspaceNotInitializedError as exc:
            QMessageBox.critical(self, "工作区未初始化", str(exc))
            return False
        except Exception as exc:
            QMessageBox.critical(self, "切换失败", str(exc))
            return False
        # 未绑定窗口此前没有 store、也未注册过工作区，跳过 unregister/close。
        if self.store is not None:
            if self.manager is not None:
                self.manager.unregister(self)
            self.store.close()
        self.store = new_store
        self.workspace_root = target_path
        self.matcher = _species_matcher()  # 旧：读 workspace_root/字段模版/，现读软件自带预设
        # thumbnail_cache：未绑定启动时为 None，这里首次创建。
        if self.thumbnail_cache is None:
            self.thumbnail_cache = ThumbnailCache(self.workspace_root)
        else:
            self.thumbnail_cache.set_workspace(self.workspace_root)
        self.search_index = None
        clear_image_index()
        QTimer.singleShot(300, self._build_search_index_background)
        if self._thumb_worker is not None:
            self._thumb_worker.stop()
        self._thumb_worker = ThumbnailWorker(self.thumbnail_cache, self)
        self._thumb_worker.result_ready.connect(self._on_thumbnail_ready)
        self._thumb_worker.start()
        self.workspace_label.setText(f"当前工作目录：{self.workspace_root}")
        remember_workspace(self.workspace_root)
        if self.manager is not None:
            self.manager.register(self)
        self.current_voucher = None
        self.current_photos = []
        self.current_photo_index = 0
        self._photo_view_states.clear()
        self.refresh_list()
        vouchers = self._all_vouchers
        if vouchers:
            self.select_voucher(vouchers[0])
        else:
            self._loading = True
            for w in list(self.specimen_widgets.values()) + list(self.class_widgets.values()) + list(self.photo_widgets.values()):
                w.blockSignals(True)
                if isinstance(w, QComboBox):
                    w.setCurrentIndex(0)
                else:
                    w.setText("")
                w.blockSignals(False)
            self._loading = False
            self.load_current_photo()
        return True

    def open_new_window(self) -> None:
        if self.manager is not None:
            self.manager.open_workspace(None)
            return
        window = SpecimenWindow(None)
        window.show()

    # ---- Image search ----

    def open_image_search(self) -> None:
        if not self.current_voucher:
            QMessageBox.information(self, "请选择标本", "请先选择或新增一个标本。")
            return
        dlg = ImageSearchDialog(self)
        dlg.exec_()

    def open_ingest_summary(self) -> None:
        # 旧逻辑：dlg = IngestSummaryDialog(self); dlg.exec_()（模态，每次新建）。
        # 现改为非模态、单实例：已开着就刷新并聚焦，否则新建 show()。
        dlg = getattr(self, "_ingest_summary_dialog", None)
        if dlg is not None and dlg.isVisible():
            dlg._refresh()
            dlg.raise_()
            dlg.activateWindow()
            return
        dlg = IngestSummaryDialog(self)
        self._ingest_summary_dialog = dlg
        dlg.show()

    def _open_action_log(self) -> None:
        dlg = ActionLogDialog(self)
        dlg.exec_()

    # ---- Search index ----

    def _build_search_index_background(self, force_rebuild: bool = False) -> None:
        """Build the image search index in the background after startup."""
        if self._is_closing or self._index_build_worker is not None:
            return
        self._index_build_worker = IndexBuildWorker(self.workspace_root, self, force_rebuild=force_rebuild)
        self._index_build_worker.index_ready.connect(self._on_index_ready)
        self._index_build_worker.finished.connect(
            lambda: setattr(self, "_index_build_worker", None)
        )
        self._index_build_worker.start()

    def _on_index_ready(self, index: ImageSearchIndex | None) -> None:
        self.search_index = index
        _startup_mark("image search index ready")

    # ---- Version manager ----

    def open_version_manager(self) -> None:
        dlg = VersionManagerDialog(self)
        dlg.exec_()

    # ---- Startup update check ----

    def _maybe_check_updates_on_startup(self) -> None:
        """启动后根据 auto_update_mode 4 档执行检查。

        - off:      不检查
        - notify:   后台查 → 有新版用顶部黄条提示（D19 banner）
        - download: 后台查 → 有新版自动下载到 releases/ + write pending
        - install:  同 download(swap 仍仅在下次启动应用，避免会话中切版本)

        所有模式都受 ``auto_update_interval_hours`` 限频。
        """
        if self._is_closing:
            return
        settings = load_settings()
        mode = (settings.auto_update_mode or "off").lower()
        if mode == "off":
            return
        last = settings.last_update_check
        if last:
            try:
                from datetime import datetime
                interval_s = max(1, int(settings.auto_update_interval_hours or 24)) * 3600
                if (datetime.now() - datetime.fromisoformat(last)).total_seconds() < interval_s:
                    return
            except ValueError:
                pass  # 时间戳损坏则照常检查
        # Channel-aware check goes through ui_upgrade._ChannelCheckWorker
        from .ui_upgrade import _ChannelCheckWorker
        worker = _ChannelCheckWorker(
            channel=settings.auto_update_channel or "stable",
            parent=self,
        )
        worker.finished_check.connect(self._on_startup_update_checked)
        worker.finished.connect(lambda: setattr(self, "_startup_update_worker", None))
        self._startup_update_worker = worker
        worker.start()

    def _on_startup_update_checked(self, release, error) -> None:
        from datetime import datetime
        settings = load_settings()
        settings.last_update_check = datetime.now().isoformat(timespec="seconds")
        save_settings(settings)
        if error is not None or release is None:
            return  # 启动检查失败保持静默，不打扰用户
        if not is_newer(release.version):
            return
        # 已跳过该版本 → 静默
        if release.version in (settings.auto_update_skipped_versions or []):
            return
        mode = (settings.auto_update_mode or "off").lower()
        if mode == "notify":
            self._show_update_banner(release)
        elif mode in ("download", "install"):
            self._show_update_banner(release)
            self._start_background_download_for_pending(release)

    def _show_update_banner(self, release) -> None:
        """D19 启动 banner:notify 模式有新版时主窗口顶部黄条提示。

        Banner widget 在 _build_ui 阶段已挂到 central_layout 中,默认 hide。
        """
        banner = getattr(self, "_update_banner", None)
        if banner is None:
            # banner widget 未注册(向后兼容):退回状态栏短消息
            self.statusBar().showMessage(
                f'发现新版本 v{release.version},可在"升级"菜单中下载更新', 15000
            )
            return
        self._update_banner_release = release
        label = banner.findChild(QLabel, "_update_banner_text")
        if label is not None:
            label.setText(
                f"🔔 发现新版 v{release.version}（当前 v{__version__}）"
            )
        banner.show()

    def _start_background_download_for_pending(self, release) -> None:
        """download/install 模式:启动后台下载,写 pending,下次启动安装。"""
        from .ui_upgrade import _locate_bundle
        from .updater_pending import PendingUpdate, now_iso, write_pending
        workspace = getattr(self, "workspace_root", None)
        dest_root = default_download_root(workspace) if workspace else Path.cwd() / "releases"
        local_roots = release_roots(workspace) if workspace else []

        def _done(target_dir, incremental, error):
            self._bg_download_worker = None
            if error is not None or target_dir is None:
                return
            bundle_dir, exe_name = _locate_bundle(Path(target_dir))
            if bundle_dir is None or not exe_name:
                return
            pending = PendingUpdate(
                version=release.version,
                bundle_dir=str(bundle_dir),
                exe_name=exe_name,
                from_version=__version__,
                staged_at=now_iso(),
                incremental=bool(incremental),
                workspace=str(workspace or ""),
            )
            write_pending(pending)
            self.statusBar().showMessage(
                f"v{release.version} 已下载,下次启动时弹窗确认安装", 15000
            )

        worker = UpdateDownloadWorker(release, dest_root, local_roots, parent=self)
        worker.finished_download.connect(_done)
        self._bg_download_worker = worker
        worker.start()

    def _upgrade_banner_install(self) -> None:
        """Banner 按钮:一键升级流程 — 后台 check + download + 问重启 + swap。

        相当于 VSCode "Restart to Update" / Chrome 静默后台升级用户视角的"一键"。
        失败时回落开升级中心对话框让用户手动操作。
        """
        self._oneclick_upgrade_now(source="banner")

    def _oneclick_upgrade_now(self, *, source: str = "menu") -> None:
        """一键升级:静默 check → 静默 download → "立即重启升级?" 弹窗 → swap。

        所有阶段在后台 worker 跑,主窗口仅一条 statusBar 进度提示。
        失败任一步 → 提示并打开升级中心让用户接手。
        """
        if getattr(self, "_oneclick_in_progress", False):
            self.statusBar().showMessage("一键升级已在进行中…", 5000)
            return
        self._oneclick_in_progress = True

        # Hide banner immediately so user doesn't double-click.
        banner = getattr(self, "_update_banner", None)
        if banner is not None:
            banner.hide()

        settings = load_settings()
        channel = settings.auto_update_channel or "stable"
        self.statusBar().showMessage(f"一键升级:正在查询最新版本（{channel}）…", 0)

        from .ui_upgrade import _ChannelCheckWorker
        check = _ChannelCheckWorker(channel=channel, parent=self)
        check.finished_check.connect(
            lambda release, error: self._oneclick_after_check(release, error)
        )
        self._oneclick_check_worker = check
        check.start()

    def _oneclick_after_check(self, release, error) -> None:
        if error is not None:
            self._oneclick_in_progress = False
            self.statusBar().clearMessage()
            QMessageBox.warning(
                self, "一键升级失败",
                f"检查更新失败：{error}\n\n请稍后重试。",
            )
            return
        if release is None or not is_newer(release.version):
            self._oneclick_in_progress = False
            self.statusBar().clearMessage()
            QMessageBox.information(
                self, "已是最新",
                f"当前 v{__version__} 已是该 channel 最新版本。",
            )
            return
        # 已下载的 pending 直接重用,不重复下。
        from .updater_pending import read_pending
        pending = read_pending()
        if pending and not pending.is_stale() and pending.version == release.version:
            self._oneclick_in_progress = False
            self.statusBar().clearMessage()
            self._launch_pending_swap_with_confirm(pending)
            return
        # 启动下载。
        from pathlib import Path as _P
        workspace = getattr(self, "workspace_root", None)
        dest_root = default_download_root(workspace) if workspace else _P.cwd() / "releases"
        local_roots = release_roots(workspace) if workspace else []

        self.statusBar().showMessage(
            f"一键升级:正在下载 v{release.version} … 0%", 0
        )

        worker = UpdateDownloadWorker(release, dest_root, local_roots, parent=self)
        worker.progress.connect(
            lambda pct, ver=release.version: self.statusBar().showMessage(
                f"一键升级:正在下载 v{ver} … {pct}%", 0
            )
        )
        worker.finished_download.connect(
            lambda target, inc, err, rel=release: self._oneclick_after_download(rel, target, inc, err)
        )
        self._oneclick_dl_worker = worker
        worker.start()

    def _oneclick_after_download(self, release, target_dir, incremental, error) -> None:
        self._oneclick_in_progress = False
        if error is not None:
            self.statusBar().clearMessage()
            QMessageBox.warning(
                self, "下载失败",
                f"v{release.version} 下载失败：{error}",
            )
            return
        if target_dir is None:
            self.statusBar().clearMessage()
            return
        from pathlib import Path as _P
        from .ui_upgrade import _locate_bundle
        from .updater_pending import PendingUpdate, now_iso, write_pending
        bundle_dir, exe_name = _locate_bundle(_P(target_dir))
        if bundle_dir is None or not exe_name:
            self.statusBar().clearMessage()
            QMessageBox.warning(
                self, "下载完成但未识别 bundle",
                f"下载到 {target_dir},但未在其中找到可执行文件。",
            )
            return
        pending = PendingUpdate(
            version=release.version,
            bundle_dir=str(bundle_dir),
            exe_name=exe_name,
            from_version=__version__,
            staged_at=now_iso(),
            incremental=bool(incremental),
            workspace=str(getattr(self, "workspace_root", "") or ""),
        )
        write_pending(pending)
        # VSCode 风一键升级:下完直接走 swap,不二次确认。snapshot 在 _launch_pending_swap
        # 内部强制创建,数据安全。statusBar 短倒计时让用户最后一秒能 Ctrl+C 终止。
        self._auto_swap_countdown = 3
        self._auto_swap_pending = pending
        self._auto_swap_release = release
        self.statusBar().showMessage(
            f"✅ v{release.version} 已下载完成,3 秒后自动重启升级 …", 0
        )
        from PyQt5.QtCore import QTimer
        self._auto_swap_timer = QTimer(self)
        self._auto_swap_timer.timeout.connect(self._auto_swap_tick)
        self._auto_swap_timer.start(1000)

    def _arm_auto_swap_countdown(self, pending, release=None) -> None:
        """让其他来源 (UpgradeCenterDialog 检查/导入 tab) 走相同的 3 秒倒计时
        → swap 流水。pending 必填,release 可选(没有就用 pending.version 显示)。
        """
        from PyQt5.QtCore import QTimer as _QT
        self._auto_swap_pending = pending
        # 没有 release 对象时构造一个最小占位,只用它的 .version 属性。
        class _RelStub:
            def __init__(self, version):
                self.version = version
        self._auto_swap_release = release or _RelStub(pending.version)
        self._auto_swap_countdown = 3
        self.statusBar().showMessage(
            f"✅ v{self._auto_swap_release.version} 准备就绪,3 秒后自动重启升级 …", 0
        )
        timer = _QT(self)
        timer.timeout.connect(self._auto_swap_tick)
        self._auto_swap_timer = timer
        timer.start(1000)

    def _auto_swap_tick(self) -> None:
        self._auto_swap_countdown -= 1
        if self._auto_swap_countdown <= 0:
            self._auto_swap_timer.stop()
            pending = self._auto_swap_pending
            self.statusBar().showMessage("正在准备升级:创建数据快照 + 释放工作区锁 …", 0)
            self._launch_pending_swap(pending)
            return
        release = self._auto_swap_release
        self.statusBar().showMessage(
            f"✅ v{release.version} 已下载完成,{self._auto_swap_countdown} 秒后自动重启升级 …", 0
        )

    def _launch_pending_swap_with_confirm(self, pending) -> None:
        """统一确认 + 走 _launch_pending_swap。给一键升级 / banner / oneclick 复用。"""
        kind = self._install_kind_safe()
        msg = (
            f"✅ v{pending.version} 已下载完成。\n\n"
            f"位置：{pending.bundle_dir}\n\n"
            "立即关闭并安装新版？\n"
            "（升级会先创建数据快照；新版仅在重启后生效。）"
        )
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Question)
        box.setWindowTitle("准备升级")
        box.setText(msg)
        btn_now = box.addButton("立即升级（重启）", QMessageBox.YesRole)
        box.addButton("稍后", QMessageBox.NoRole)
        box.exec_()
        if box.clickedButton() is btn_now:
            self._launch_pending_swap(pending)

    def _upgrade_banner_later(self) -> None:
        banner = getattr(self, "_update_banner", None)
        if banner is not None:
            banner.hide()

    def _upgrade_banner_skip(self) -> None:
        release = getattr(self, "_update_banner_release", None)
        banner = getattr(self, "_update_banner", None)
        if release is not None:
            settings = load_settings()
            if release.version not in settings.auto_update_skipped_versions:
                settings.auto_update_skipped_versions = list(
                    settings.auto_update_skipped_versions
                ) + [release.version]
                save_settings(settings)
        if banner is not None:
            banner.hide()

    # ---- D3+D11 启动入口:apply pending + sentinel 健康检查 ----

    def _apply_pending_update_on_startup(self) -> None:
        """启动时检测 pending_update.json,有就弹"立即安装并重启?"三选。"""
        from .updater_pending import clear_pending, read_pending
        pending = read_pending()
        if pending is None:
            return
        if pending.is_stale():
            clear_pending()
            return
        kind = self._install_kind_safe()
        if kind not in ("frozen-current", "frozen-direct"):
            # source / appimage / system-package — pending state is meaningless here.
            # Don't clear — user might be testing; just note it.
            self.statusBar().showMessage(
                f"已下载 v{pending.version} 但当前运行模式 ({kind}) 不支持自动应用,请手动启动", 12000
            )
            return
        msg = (
            f"已下载新版 v{pending.version}（从 v{pending.from_version}）。\n\n"
            f"位置：{pending.bundle_dir}\n\n"
            "是否立即关闭软件并安装？\n"
            "（升级会先创建数据快照；新版仅在重启后生效。）"
        )
        btn_install = QMessageBox.Yes
        btn_later = QMessageBox.No
        btn_discard = QMessageBox.Discard
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Question)
        box.setWindowTitle("已下载更新")
        box.setText(msg)
        box.addButton("立即安装并重启", QMessageBox.YesRole)
        box.addButton("稍后", QMessageBox.NoRole)
        box.addButton("丢弃此次更新", QMessageBox.DestructiveRole)
        ret = box.exec_()
        clicked = box.clickedButton()
        if clicked is None:
            return
        role = box.buttonRole(clicked)
        if role == QMessageBox.YesRole:
            self._launch_pending_swap(pending)
        elif role == QMessageBox.DestructiveRole:
            clear_pending()
            self.statusBar().showMessage(f"已丢弃待安装 v{pending.version}", 5000)
        # NoRole: keep pending for next launch

    def _install_kind_safe(self) -> str:
        try:
            from .install_kind import installation_kind
            return installation_kind()
        except Exception:
            return "unknown"

    def _launch_pending_swap(self, pending) -> None:
        """走 D2 swap 脚本:snapshot → 释锁 → detached swap → quit。

        Only the ``frozen-current`` path can do a real junction swap. For
        ``frozen-direct`` we still launch the new exe and let the user
        decide whether to manually enable ``current/`` later.
        """
        import os
        from pathlib import Path as _P
        from .updater_pending import clear_pending
        from .release_manager import current_install_root

        # Flush any debounced auto-saves so in-progress edits are persisted
        # before snapshot. _flush_pending_saves is a no-op when no saves
        # are pending — safe to call unconditionally.
        try:
            flusher = getattr(self, "_flush_pending_saves", None)
            if callable(flusher):
                flusher()
        except Exception:
            pass

        # Mandatory pre-swap snapshot (D9 default — strict; opt-in skip later).
        try:
            if self.store is not None:
                self.store.create_data_snapshot(
                    "自动更新前快照",
                    f"v{pending.from_version}→v{pending.version}",
                )
        except Exception:
            pass

        # Release the workspace lock before swap so the new process can acquire it.
        try:
            if self.store is not None:
                self.store.release_lock()
        except Exception:
            pass

        # Pick the current/ junction location.
        install_root = current_install_root()
        bundle_dir = _P(pending.bundle_dir)
        if install_root is None:
            install_root = bundle_dir.parent.parent  # releases/v0.8.0/ → releases → parent
        current_link = install_root / "current"

        try:
            from .updater_swap import launch_swap_detached
            launch_swap_detached(
                pid=os.getpid(),
                new_bundle=bundle_dir,
                current_link=current_link,
                new_exe_name=pending.exe_name,
                workspace=_P(pending.workspace) if pending.workspace else None,
            )
        except Exception as exc:
            QMessageBox.warning(
                self, "启动 swap 失败",
                f"无法启动升级脚本：{exc}\n\n"
                f"请手动启动 {bundle_dir / pending.exe_name}",
            )
            return

        clear_pending()
        # Quit the current process so the swap script can take over.
        QApplication.instance().quit()

    def _check_post_update_sentinel_on_startup(self) -> None:
        """D11 健康回滚:若上次启动留下 sentinel(说明 30s 内挂了)弹回退。"""
        from .updater_pending import (
            clear_post_update_sentinel,
            read_post_update_sentinel,
        )
        sentinel = read_post_update_sentinel()
        if sentinel is None:
            return
        # If the sentinel was written by an earlier successful run that
        # _did_ clear at the 30 s mark, it's gone — so any sentinel found
        # at startup means the previous launch did not survive 30 s.
        reply = QMessageBox.warning(
            self, "升级后启动失败",
            f"上次启动 v{sentinel.current_version} 似乎在 30 秒内异常退出。\n"
            f"是否回退到 v{sentinel.from_version}？",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            # Best-effort reverse swap — actual swap-back wiring deferred to v0.8.1.
            QMessageBox.information(
                self, "回退",
                "回退流程将在 v0.8.1 完整实现。\n"
                f"请在 升级 → 高级 → 历史版本管理 选 v{sentinel.from_version} 手动启动。",
            )
        clear_post_update_sentinel()

    def _arm_post_update_sentinel(self) -> None:
        """新版启动 30s 内未崩 → 清 sentinel。配合 D11 健康回滚机制。"""
        from .updater_pending import clear_post_update_sentinel
        QTimer.singleShot(30000, clear_post_update_sentinel)

    # ---- Settings ----

    def open_settings(self) -> None:
        dlg = SettingsDialog(self)
        if dlg.exec_() == QDialog.Accepted:
            self.store.set_undo_depth(dlg.undo_depth)
            current_settings = load_settings()
            quality_keys = list(PREVIEW_QUALITY_OPTIONS.keys())
            idx = dlg.quality_combo.currentIndex()
            current_settings.preview_quality = quality_keys[idx] if 0 <= idx < len(quality_keys) else "compressed"
            mode_key = dlg.photo_management_combo.currentData()
            current_settings.photo_management_mode = mode_key if mode_key in PHOTO_MANAGEMENT_OPTIONS else "copy_with_absolute"
            current_settings.photo_library_path = dlg.photo_library_edit.text().strip()
            current_settings.image_viewer_path = dlg.image_viewer_path
            current_settings.photo_filename_fill_shortcut = dlg.photo_filename_fill_shortcut
            current_settings.check_updates_on_startup = dlg.check_updates_box.isChecked()
            current_settings.ui_font_size = dlg.font_size
            current_settings.cursor_style = dlg.cursor_style
            current_settings.app_icon_variant = dlg.app_icon_variant
            current_settings.memory_profile = dlg.memory_profile
            save_settings(current_settings)
            self._cached_preview_quality = current_settings.preview_quality
            self._apply_photo_filename_fill_shortcut()
            # 应用全局字体大小并刷新所有窗口的表格字体/列宽。
            apply_app_font_size(current_settings.ui_font_size)
            self._refresh_all_windows_fonts()
            # 应用趣味光标 + 应用图标变体(对所有已打开窗口即时生效)。
            apply_app_cursor(current_settings.cursor_style)
            apply_app_icon(current_settings.app_icon_variant)
            # 规范化软件设计 2026-05 内存档位:即时热应用 ThumbnailCache + _row_cache_maxsize
            # (ThumbnailWorker.max_workers 由 ThreadPoolExecutor 构造时定,需重启生效)。
            try:
                from .env_detect import memory_profile_params
                params = memory_profile_params(current_settings.memory_profile)
                if self.thumbnail_cache is not None:
                    self.thumbnail_cache.memory_limit_bytes = params["thumb_cache_bytes"]
                if self.store is not None:
                    self.store._row_cache_maxsize = params["row_cache_maxsize"]
                    self.store._enforce_row_cache_size()
            except Exception:
                pass
            # 立即刷新状态栏档位显示
            if hasattr(self, "_refresh_memory_status"):
                self._refresh_memory_status()

    # ---- 全局字体缩放 ----

    def _refresh_scaled_fonts(self) -> None:
        """按当前全局字号设置 voucher_table 字体与列宽（默认字号时与旧版一致）。"""
        table = getattr(self, "voucher_table", None)
        if table is None:
            return
        app = QApplication.instance()
        cur_pt = app.font().pointSize() if app is not None else self._VOUCHER_TABLE_BASE_PT
        base_pt = _default_app_font_point or cur_pt
        delta = cur_pt - base_pt
        table_pt = max(6, self._VOUCHER_TABLE_BASE_PT + delta)
        table.setFont(QFont("Consolas", table_pt))
        scale = table_pt / self._VOUCHER_TABLE_BASE_PT
        for col, base in enumerate(self._VOUCHER_COL_BASE_WIDTHS):
            table.setColumnWidth(col, max(base, int(base * scale)))

    def _refresh_all_windows_fonts(self) -> None:
        """刷新所有打开窗口的缩放字体（app.setFont 已全局生效，这里补表格字体/列宽）。"""
        windows = list(self.manager._windows.values()) if self.manager is not None else [self]
        for win in windows:
            try:
                win._refresh_scaled_fonts()
            except Exception:
                pass

    def _zoom_font(self, delta: int) -> None:
        app = QApplication.instance()
        if app is None:
            return
        new_size = max(7, min(24, app.font().pointSize() + delta))
        apply_app_font_size(new_size)
        try:
            settings = load_settings()
            settings.ui_font_size = new_size
            save_settings(settings)
        except Exception:
            pass  # 持久化失败不影响本次会话
        self._refresh_all_windows_fonts()
        self.statusBar().showMessage(f"界面字体大小：{new_size} pt", 2000)

    # _zoom_font_in / _zoom_font_out 是 _zoom_font 的零参 wrapper，
    # 提供给 SHORTCUTABLE_ACTIONS 注册表用稳定 slot 名（规范化软件设计 2026-05 新增）。
    def _zoom_font_in(self) -> None:
        self._zoom_font(1)

    def _zoom_font_out(self) -> None:
        self._zoom_font(-1)

    def _zoom_font_reset(self) -> None:
        apply_app_font_size(0)  # 0 -> 恢复系统默认字号
        try:
            settings = load_settings()
            settings.ui_font_size = 0
            save_settings(settings)
        except Exception:
            pass
        self._refresh_all_windows_fonts()
        self.statusBar().showMessage("界面字体大小已恢复默认", 2000)

    # ---- Panel helpers ----

    @staticmethod
    def _create_panel(title: str, content: QWidget, collapsible: bool = True) -> QFrame:
        frame = QFrame()
        frame.setFrameStyle(QFrame.StyledPanel | QFrame.Raised)
        layout = QVBoxLayout(frame)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(4)
        title_bar = QHBoxLayout()
        title_label = QLabel(title)
        title_label.setStyleSheet("font-weight: bold; font-size: 12px; padding: 2px 0;")
        title_bar.addWidget(title_label)
        title_bar.addStretch()
        if collapsible:
            collapse_btn = QPushButton("−")
            collapse_btn.setFixedSize(22, 22)
            collapse_btn.setToolTip("折叠/展开面板")
            collapse_btn.clicked.connect(lambda: _toggle_content(collapse_btn, content))
            title_bar.addWidget(collapse_btn)
        layout.addLayout(title_bar)
        layout.addWidget(content, stretch=1)
        # Store toggle function in closure
        def _toggle_content(btn, w):
            visible = w.isVisible()
            w.setVisible(not visible)
            btn.setText("−" if not visible else "+")
        return frame

    def _reset_window_layout(self) -> None:
        if QMessageBox.question(
            self, "重置窗口布局",
            "确定要重置窗口布局为默认状态吗？",
            QMessageBox.Yes | QMessageBox.No,
        ) != QMessageBox.Yes:
            return
        try:
            settings = load_settings()
            settings.window_geometry = ""
            settings.splitter_sizes = []
            save_settings(settings)
        except Exception:
            pass
        self.main_splitter.setSizes([220, 580, 480])
        self.right_splitter.setSizes([260, 280, 200])
        self.voucher_panel.show()
        self.specimen_panel.show()
        self.photo_panel.show()
        self.class_panel.show()
        self.resize(1320, 820)
        QMessageBox.information(self, "布局已重置", "窗口布局已恢复为默认状态。")


# ---------------------------------------------------------------------------
# Image search dialog
# ---------------------------------------------------------------------------

DEFAULT_PHOTO_SCOPE = "工作区/照片"
WORKSPACE_SCOPE = "整个工作区"
IMAGE_TYPE_CHOICES = [
    ("全部图片", "all"),
    ("TIF", "tif"),
    ("JPG", "jpg"),
    ("TIF+JPG", "tif_jpg"),
]


class IndexBuildWorker(QThread):
    """Background worker that builds the image search index at startup."""

    index_ready = pyqtSignal(object)  # ImageSearchIndex | None

    def __init__(self, workspace_root: Path, parent=None, force_rebuild: bool = False):
        super().__init__(parent)
        self.workspace_root = workspace_root
        self.force_rebuild = force_rebuild

    def run(self) -> None:
        try:
            photo_dir = Path(self.workspace_root).resolve() / "照片"
            if photo_dir.is_dir():
                roots = [photo_dir]
                scan_depth = 0  # 专用照片目录，深度不限
            else:
                # 原代码：照片目录不存在时 fallback 扫整个工作区且 max_depth=0（无限深度）。
                # 若工作区本身是个巨大目录，会遍历海量文件、拖垮整机。fallback 改为有界深度。
                roots = [Path(self.workspace_root).resolve()]
                scan_depth = 4
            if self.isInterruptionRequested():
                self.index_ready.emit(None)
                return
            index = _get_or_build_search_index(
                roots,
                max_depth=scan_depth,
                should_stop=self.isInterruptionRequested,
                force_rebuild=self.force_rebuild,
                cache_root=self.workspace_root,
            )
            self.index_ready.emit(index)
        except Exception:
            self.index_ready.emit(None)


class ImageSearchWorker(QThread):
    result_ready = pyqtSignal(int, object, object)  # token, list[ImageSearchResult], Exception|None

    def __init__(
        self,
        token: int,
        workspace_root: Path,
        voucher: str,
        specimen: dict[str, Any],
        classification: dict[str, Any],
        linked_paths: list[Path],
        query: str,
        search_roots: list[str] | None,
        image_type: str,
        search_index: ImageSearchIndex | None = None,
        force_rebuild: bool = False,
        limit: int = 50,
        path_to_vouchers: dict[str, list[str]] | None = None,
        parent=None,
    ):
        super().__init__(parent)
        self.token = token
        self.workspace_root = workspace_root
        self.voucher = voucher
        self.specimen = specimen
        self.classification = classification
        self.linked_paths = linked_paths
        self.query = query
        self.search_roots = search_roots
        self.image_type = image_type
        self.search_index = search_index
        self.force_rebuild = force_rebuild
        self.limit = limit
        self.path_to_vouchers = path_to_vouchers

    def run(self) -> None:
        try:
            results = image_search_results(
                self.workspace_root,
                self.voucher,
                self.specimen,
                self.classification,
                self.linked_paths,
                query=self.query,
                extra_roots=self.search_roots,
                suffixes=suffixes_for_image_type(self.image_type),
                limit=self.limit,
                should_stop=self.isInterruptionRequested,
                search_index=self.search_index,
                force_rebuild=self.force_rebuild,
                path_to_vouchers=self.path_to_vouchers,
            )
        except Exception as exc:
            self.result_ready.emit(self.token, [], exc)
        else:
            self.result_ready.emit(self.token, results, None)


class ImageSearchDialog(QDialog):
    def __init__(self, app: SpecimenWindow):
        super().__init__(app)
        self.app = app
        self.setWindowTitle("图片检索")
        self.resize(980, 680)
        self.setMinimumSize(QSize(760, 520))

        self.results: list[ImageSearchResult] = []
        self.selected_indices: set[int] = set()
        self.last_selected_index: int | None = None
        self.result_limit: int = 50
        self._render_token = 0
        self._search_token = 0
        self._search_workers: list[ImageSearchWorker] = []
        self._card_labels: dict[int, QLabel] = {}
        self._cards: list[QFrame] = []
        self._thumb_worker = ThumbnailWorker(app.thumbnail_cache, self)
        self._thumb_worker.result_ready.connect(self._on_thumbnail_ready)
        self._thumb_worker.start()

        self._build()
        self.refresh_results()

    def _build(self) -> None:
        layout = QVBoxLayout(self)

        # Top bar
        top = QHBoxLayout()
        top.addWidget(QLabel(f"当前标本：{self.app.current_voucher or ''}"))
        top.addWidget(QLabel("核心编号"))
        self.query_edit = QLineEdit(self._default_query())
        self.query_edit.setPlaceholderText("输入关键词搜索，如 QD-C、CK、SC008")
        self.query_edit.textChanged.connect(self.schedule_refresh)
        top.addWidget(self.query_edit, stretch=1)
        top.addWidget(QLabel("显示"))
        self.limit_spin = QSpinBox()
        self.limit_spin.setRange(10, 500)
        self.limit_spin.setValue(50)
        self.limit_spin.setSuffix(" 张")
        self.limit_spin.setToolTip("搜索结果最大显示数量")
        self.limit_spin.valueChanged.connect(self._on_limit_changed)
        top.addWidget(self.limit_spin)
        top.addWidget(self._make_button("重新扫描", self.rescan_results))
        top.addWidget(self._make_button("添加选中图片", self.add_selected))
        layout.addLayout(top)

        # Search paths row
        path_row = QHBoxLayout()
        path_row.addWidget(QLabel("搜索范围"))
        self.path_combo = QComboBox()
        self.path_combo.setEditable(True)
        self.path_combo.addItems(self._default_search_paths())
        self.path_combo.setToolTip("输入自定义路径（可用 ; 分隔多个路径），或选择预设")
        self.path_combo.currentTextChanged.connect(self.schedule_refresh)
        path_row.addWidget(self.path_combo, stretch=1)
        path_row.addWidget(QLabel("类型"))
        self.type_combo = QComboBox()
        for label, key in IMAGE_TYPE_CHOICES:
            self.type_combo.addItem(label, key)
        self.type_combo.currentIndexChanged.connect(self.schedule_refresh)
        path_row.addWidget(self.type_combo)
        path_row.addWidget(self._make_button("添加目录", self._add_search_dir))
        path_row.addWidget(self._make_button("恢复默认", self._reset_search_paths))
        path_row.addWidget(self._make_button("清除索引", self._clear_index_cache))
        layout.addLayout(path_row)

        # Results scroll area
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setStyleSheet("QScrollArea { background-color: #f4f6f8; border: 1px solid #a9b3bd; }")
        self.result_container = QWidget()
        self.result_grid = QGridLayout(self.result_container)
        self.result_grid.setSpacing(6)
        self.scroll_area.setWidget(self.result_container)
        layout.addWidget(self.scroll_area, stretch=1)

        # Bottom bar
        bottom = QHBoxLayout()
        self.status_label = QLabel()
        bottom.addWidget(self.status_label, stretch=1)
        bottom.addWidget(self._make_button("添加", self.add_selected))
        bottom.addWidget(self._make_button("关闭", self.close))
        layout.addLayout(bottom)

    @staticmethod
    def _make_button(text: str, slot) -> QPushButton:
        btn = QPushButton(text)
        btn.clicked.connect(slot)
        return btn

    def _default_query(self) -> str:
        if not self.app.current_voucher:
            return ""
        specimen = self.app.store.get_specimen(self.app.current_voucher) or {}
        return default_image_query(specimen)

    def _default_search_paths(self) -> list[str]:
        paths = [DEFAULT_PHOTO_SCOPE, WORKSPACE_SCOPE]
        settings = load_settings()
        for p in settings.search_paths:
            if p not in paths:
                paths.append(p)
        return paths

    def _add_search_dir(self) -> None:
        d = QFileDialog.getExistingDirectory(self, "选择搜索目录")
        if d:
            if self.path_combo.findText(d) < 0:
                self.path_combo.addItem(d)
            self.path_combo.setCurrentText(d)

    def _reset_search_paths(self) -> None:
        self.path_combo.setCurrentText(DEFAULT_PHOTO_SCOPE)

    def _clear_index_cache(self) -> None:
        clear_image_index()
        self.app.search_index = None
        self.status_label.setText("索引缓存已清除，下次检索将重新扫描。")
        QTimer.singleShot(500, self.refresh_results)

    def _parse_search_roots(self) -> list[str] | None:
        text = self.path_combo.currentText().strip()
        if not text or text == DEFAULT_PHOTO_SCOPE:
            return None
        if text == WORKSPACE_SCOPE:
            return [str(self.app.workspace_root)]
        roots = []
        for part in text.replace("\n", ";").split(";"):
            part = part.strip()
            if part and part not in (DEFAULT_PHOTO_SCOPE, WORKSPACE_SCOPE) and Path(part).is_dir():
                roots.append(part)
        return roots or None

    def _selected_image_type(self) -> str:
        key = self.type_combo.currentData()
        return str(key or "tif")

    def _on_limit_changed(self, value: int) -> None:
        self.result_limit = value
        self.schedule_refresh()

    def schedule_refresh(self) -> None:
        if not hasattr(self, "_refresh_timer"):
            self._refresh_timer = QTimer(self)
            self._refresh_timer.setSingleShot(True)
            self._refresh_timer.timeout.connect(self.refresh_results)
        self._refresh_timer.start(250)

    def rescan_results(self) -> None:
        self.app.search_index = None
        self.refresh_results(force_rebuild=True)

    def refresh_results(self, force_rebuild: bool = False) -> None:
        if not self.app.current_voucher:
            self.close()
            return
        self.selected_indices.clear()
        self.last_selected_index = None
        query = self.query_edit.text().strip()
        if not query:
            self.results = []
            self._clear_cards()
            self.status_label.setText("输入关键词搜索图片，如 QD-C、CK、SC008")
            return
        search_roots = self._parse_search_roots()
        if force_rebuild:
            self.status_label.setText(f"正在重新建立图片索引，并检索 {query}...")
        elif self.app.search_index is not None:
            self.status_label.setText(f"正在检索：{query}...")
        elif image_index_exists(self.app.workspace_root, search_roots):
            self.status_label.setText(f"正在检索索引：{query}...")
        else:
            self.status_label.setText(f"正在建立图片索引，并检索 {query}...")

        linked_paths = [self.app.store.resolve_photo_path(row) for row in self.app.current_photos]
        path_to_vouchers = self.app.store.get_all_photo_voucher_map()
        specimen = self.app.store.get_specimen(self.app.current_voucher) or {}
        classification = self.app.store.get_classification(self.app.current_voucher) or {}
        self._search_token += 1
        token = self._search_token
        for old_worker in list(self._search_workers):
            old_worker.requestInterruption()
        self._clear_cards()
        worker = ImageSearchWorker(
            token=token,
            workspace_root=self.app.workspace_root,
            voucher=self.app.current_voucher,
            specimen=specimen,
            classification=classification,
            linked_paths=linked_paths,
            query=query,
            search_roots=search_roots,
            image_type=self._selected_image_type(),
            # 原代码所有范围都传启动索引；自定义目录和整个工作区应使用对应范围的新索引。
            search_index=self.app.search_index if search_roots is None and not force_rebuild else None,
            force_rebuild=force_rebuild,
            limit=self.result_limit,
            path_to_vouchers=path_to_vouchers,
            parent=self,
        )
        worker.result_ready.connect(self._on_search_ready)
        worker.finished.connect(lambda w=worker: self._discard_search_worker(w))
        self._search_workers.append(worker)
        worker.start()

    def _on_search_ready(self, token: int, results: list[ImageSearchResult], exc: Exception | None) -> None:
        if token != self._search_token:
            return
        if exc is not None:
            self.results = []
            self._clear_cards()
            self.status_label.setText(f"检索失败：{exc}")
            return
        self.results = results
        self.render_results()

    def _discard_search_worker(self, worker: ImageSearchWorker) -> None:
        if worker in self._search_workers:
            self._search_workers.remove(worker)

    def _clear_cards(self) -> None:
        self._thumb_worker.clear_pending()
        for child in self.result_container.children():
            if isinstance(child, QWidget):
                self.result_grid.removeWidget(child)
                child.deleteLater()
        self._render_token += 1
        self._card_labels.clear()
        self._cards = []

    def render_results(self) -> None:
        self._clear_cards()
        columns = 4

        if not self.results:
            query = self.query_edit.text().strip()
            self.status_label.setText(f'未找到与 "{query}" 匹配的图片')
            return

        for index, result in enumerate(self.results):
            r = index // columns
            c = index % columns
            card = self._create_card(index, result)
            self.result_grid.addWidget(card, r, c)
            self._cards.append(card)

        self._update_status()

    def _create_card(self, index: int, result: ImageSearchResult) -> QFrame:
        """创建单张搜索结果卡片。

        卡片背景：
        - 已关联到当前标本（is_linked=True）：灰色 #eceff2
        - 已关联到其他标本（linked_vouchers 非空但 is_linked=False）：浅黄 #fef9e7
        - 未关联：白色 #ffffff
        """
        selected = index in self.selected_indices
        # 判断是否关联到任意标本（当前或其他）
        has_any_link = bool(result.linked_vouchers)
        bg = self._card_background(selected, result.is_linked, has_any_link)
        card = QFrame()
        card.setStyleSheet(f"QFrame {{ background-color: {bg}; border: 1px solid #ccc; padding: 6px; }}")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(6, 6, 6, 6)

        image_label = QLabel("加载中")
        image_label.setAlignment(Qt.AlignCenter)
        image_label.setFixedSize(180, 135)
        image_label.setStyleSheet("color: #59666b; background-color: #e8eaed;")
        card_layout.addWidget(image_label)

        # 标题：文件名 + 关联信息（不限当前标本，显示所有关联到的入库编号）
        title = result.file_name
        if result.linked_vouchers:
            title += f"  已关联: {', '.join(result.linked_vouchers)}"
        # 原代码只显示 _shorten(title,36)，长文件名被截断看不全。
        # 改为：仍截断显示以保持卡片整齐，但加 tooltip 显示完整文件名，并允许选中复制。
        title_label = QLabel(_shorten(title, 36))
        title_label.setStyleSheet("font-weight: bold;")
        title_label.setToolTip(result.file_name)
        title_label.setWordWrap(True)
        title_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        card_layout.addWidget(title_label)
        # 可点击的入库编号链接（任何关联都显示，不限当前标本）
        if result.linked_vouchers:
            voucher_widget = QWidget()
            voucher_layout = QHBoxLayout(voucher_widget)
            voucher_layout.setContentsMargins(0, 2, 0, 0)
            voucher_layout.setSpacing(4)
            for voucher in result.linked_vouchers:
                link = QLabel(f'<a href="{voucher}" style="color:#2a6fbd;">{voucher}</a>')
                link.setCursor(Qt.PointingHandCursor)
                link.setToolTip(f"点击跳转到 {voucher}")
                link.linkActivated.connect(self._navigate_to_voucher)
                voucher_layout.addWidget(link)
            voucher_layout.addStretch()
            card_layout.addWidget(voucher_widget)
        path_label = QLabel(_shorten(result.relative_path, 32))
        path_label.setStyleSheet("color: #59666b;")
        # tooltip 显示完整相对路径 + 绝对路径，避免截断后看不全。
        path_label.setToolTip(f"相对路径：{result.relative_path}\n绝对路径：{result.path}")
        card_layout.addWidget(path_label)
        match_text = "核心编号：" + "、".join(result.matched_keywords[:3]) if result.matched_keywords else "核心编号：无"
        match_label = QLabel(match_text)
        match_label.setStyleSheet("color: #3f4b57;")
        card_layout.addWidget(match_label)

        # Load thumbnail
        token = self._render_token * 1000 + index
        self._thumb_worker.enqueue(result.path, (180, 135), token)
        self._card_labels[token] = image_label

        # Click handling
        card.mousePressEvent = lambda event, idx=index: self._on_card_click(event, idx)
        card.mouseDoubleClickEvent = lambda event, idx=index: self._open_preview(idx)
        card.setContextMenuPolicy(Qt.CustomContextMenu)
        card.customContextMenuRequested.connect(lambda pos, idx=index: self._show_context_menu(idx))

        return card

    def _on_thumbnail_ready(self, token: int, qpixmap: QPixmap | None, exc: Exception | None) -> None:
        label = getattr(self, "_card_labels", {}).get(token)
        if not label or not qpixmap:
            return
        scaled = qpixmap.scaled(label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
        label.setPixmap(scaled)
        label.setText("")

    def _on_card_click(self, event, index: int) -> None:
        modifiers = event.modifiers()
        if modifiers & Qt.ShiftModifier and self.last_selected_index is not None:
            start, end = sorted((self.last_selected_index, index))
            self.selected_indices.update(range(start, end + 1))
        elif modifiers & Qt.ControlModifier:
            if index in self.selected_indices:
                self.selected_indices.remove(index)
            else:
                self.selected_indices.add(index)
            self.last_selected_index = index
        else:
            self.selected_indices = {index}
            self.last_selected_index = index
        self._update_selection_visuals()

    def _update_selection_visuals(self) -> None:
        for idx, card in enumerate(self._cards):
            result = self.results[idx] if idx < len(self.results) else None
            is_linked = result.is_linked if result else False
            bg = self._card_background(
                idx in self.selected_indices,
                is_linked,
                bool(result.linked_vouchers) if result else False,  # 保持黄标（已入库但非当前标本）
            )
            card.setStyleSheet(f"QFrame {{ background-color: {bg}; border: 1px solid #ccc; padding: 6px; }}")
        self._update_status()

    def _show_context_menu(self, index: int) -> None:
        if index not in self.selected_indices:
            self.selected_indices = {index}
            self.last_selected_index = index
            self._update_selection_visuals()
        menu = QMenu(self)
        menu.addAction("添加选中图片", self.add_selected)
        menu.addAction("打开原图", lambda: self._open_preview(index))
        menu.addAction("查看详情", lambda: self._show_image_detail(index))
        menu.addSeparator()
        menu.addAction("复制文件名", lambda: self._copy_file_name(index))
        menu.addAction("复制相对路径", lambda: self._copy_relative_path(index))
        menu.exec_(QCursor.pos())

    def _copy_file_name(self, index: int) -> None:
        if 0 <= index < len(self.results):
            QApplication.clipboard().setText(self.results[index].file_name)

    def _show_image_detail(self, index: int) -> None:
        """弹窗显示图片完整信息，解决卡片上文件名/路径截断看不全的问题。"""
        if not (0 <= index < len(self.results)):
            return
        result = self.results[index]
        linked = "、".join(result.linked_vouchers) if result.linked_vouchers else "无"
        keywords = "、".join(result.matched_keywords) if result.matched_keywords else "无"
        detail = (
            f"文件名：{result.file_name}\n\n"
            f"相对路径：{result.relative_path}\n\n"
            f"绝对路径：{result.path}\n\n"
            f"已关联入库编号：{linked}\n\n"
            f"匹配核心编号：{keywords}"
        )
        box = QMessageBox(self)
        box.setWindowTitle("图片详情")
        box.setText(detail)
        box.setTextInteractionFlags(Qt.TextSelectableByMouse)
        box.exec_()

    def _open_preview(self, index: int) -> None:
        if 0 <= index < len(self.results):
            _open_path(self.results[index].path)

    def _copy_relative_path(self, index: int) -> None:
        if 0 <= index < len(self.results):
            QApplication.clipboard().setText(self.results[index].relative_path)

    def add_selected(self) -> None:
        if not self.selected_indices:
            QMessageBox.information(self, "请选择图片", "请先选择要关联的图片。")
            return
        paths = [
            str(self.results[idx].path)
            for idx in sorted(self.selected_indices)
            if 0 <= idx < len(self.results) and not self.results[idx].is_linked
        ]
        if not paths:
            QMessageBox.information(self, "无需添加", "选中的图片已经关联到当前标本。")
            return
        added = self.app.add_photo_paths(paths, ask_for_outside=False)
        self.selected_indices.clear()
        self.last_selected_index = None
        self.refresh_results()
        self.status_label.setText(f"已添加 {added} 张图片。")

    def _navigate_to_voucher(self, link: str) -> None:
        """Navigate the main window to the given voucher, keeping this dialog open."""
        voucher = link.strip()
        if voucher:
            self.app.select_voucher(voucher)

    def _card_background(self, selected: bool, linked: bool, has_any_link: bool = False) -> str:
        """搜索结果卡片背景色。

        - 选中：蓝色 #d8ebff
        - 已关联到当前标本：灰色 #eceff2
        - 已关联到其他标本：浅黄 #fef9e7（提醒用户注意核对）
        - 未关联：白色 #ffffff
        """
        if selected:
            return "#d8ebff"
        if linked:
            return "#eceff2"
        # 照片已关联到其他标本（非当前打开的标本）
        if has_any_link:
            return "#fef9e7"
        return "#ffffff"

    def _update_status(self) -> None:
        # 统计所有已关联的照片（不限当前标本，linked_vouchers 非空即为已关联到任意标本）
        linked_count = sum(1 for item in self.results if item.linked_vouchers)
        current_linked = sum(1 for item in self.results if item.is_linked)
        self.status_label.setText(f"结果 {len(self.results)} 张；已关联 {linked_count} 张（当前{current_linked}）；已选 {len(self.selected_indices)} 张")

    def keyPressEvent(self, event) -> None:
        if event.modifiers() & Qt.ControlModifier and event.key() == Qt.Key_A:
            self.selected_indices = set(range(len(self.results)))
            self.last_selected_index = len(self.results) - 1 if self.results else None
            self._update_selection_visuals()
            event.accept()
        else:
            super().keyPressEvent(event)

    def closeEvent(self, event) -> None:
        # 规范化软件设计 2026-05 P1 审查修复:无论 wait 是否成功,显式清空 _search_workers 列表,
        # 防 worker.finished 信号未发(异常路径) 导致 worker 永留 list 内存泄漏。
        self._search_token += 1
        for worker in list(self._search_workers):
            try:
                worker.requestInterruption()
                worker.wait(3000)
            except Exception:
                pass
        self._search_workers.clear()  # 兜底清空,即使 finished 未触发
        self._thumb_worker.stop()
        super().closeEvent(event)


# ---------------------------------------------------------------------------
# Batch photo confirmation dialog
# ---------------------------------------------------------------------------

class PhotoBatchDialog(QDialog):
    """Show a list of photos before batch-assigning them to a voucher."""

    def __init__(self, voucher: str, photo_paths: list[Path], parent=None):
        super().__init__(parent)
        self.setWindowTitle("确认批量关联照片")
        self.resize(520, 400)
        self._paths = photo_paths
        self._confirmed = False

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            f"即将为 <b>{voucher}</b> 关联以下 {len(photo_paths)} 张照片："
        ))
        layout.addWidget(QLabel(
            "（如需取消某些照片，请先关闭此窗口，重新选择）"
        ))

        list_widget = QListWidget()
        for path in photo_paths:
            list_widget.addItem(f"  {Path(path).name}")
        layout.addWidget(list_widget, stretch=1)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(cancel_btn)
        confirm_btn = QPushButton("确定关联")
        confirm_btn.setStyleSheet("QPushButton { font-weight: bold; }")
        confirm_btn.clicked.connect(self._on_confirm)
        btn_row.addWidget(confirm_btn)
        layout.addLayout(btn_row)

    def _on_confirm(self) -> None:
        self._confirmed = True
        self.accept()

    @property
    def confirmed(self) -> bool:
        return self._confirmed


# ---------------------------------------------------------------------------
# Ingest summary dialog — paginated voucher-photo audit with password gate
# ---------------------------------------------------------------------------

class IngestSummaryDialog(QDialog):
    """入库汇总预览 — 非模态、单实例的只读宽表（Excel 式汇总视图）。

    重构历史（详见 git）：原本左=可编辑宽表、右上=只读文字详情面板、右下=照片网格+
    照片操作（替换/取消关联/移动/打开原图）+拖拽替换。按需求改为：
    - 删除右上详情面板（与宽表冗余）；
    - 删除右下照片网格及全部照片操作 —— 照片管理改由主窗口照片面板右键菜单负责；
    - 宽表改只读，双击行 -> 跳主窗口编辑器（_on_voucher_double_clicked）；
    - 改非模态、单实例（见 SpecimenWindow.open_ingest_summary）。
    随之删除的方法：_populate_detail / _on_voucher_selected / _on_cell_changed /
    _update_row_cells / _require_password / _render_photo_grid / _set_thumb_size /
    _toggle_photo_selection / _on_overview_thumbnail / _open_selected_original /
    _unlink_selected / _move_selected / _replace_selected_photo / _jump_to_specimen /
    _overview_drag_* / _overview_drop。
    """

    PAGE_SIZE = 100

    def __init__(self, app: SpecimenWindow):
        super().__init__(app)
        self.app = app
        self.store = app.store
        self.setWindowTitle("入库汇总预览")
        self.resize(1060, 720)
        self.setMinimumSize(QSize(840, 540))

        self.all_vouchers: list[str] = self.store.list_vouchers()
        self._photo_counts: dict[str, int] = {}
        self._total_photos = 0
        self._build_photo_counts()

        # 入库汇总宽表：把分散在多个 Excel 的字段 join 成一张表（纯内存视图）。
        self._summary_records: list[dict] = self.store.summary_records()
        self._record_by_voucher: dict[str, dict] = {
            r["入库编号*"]: r for r in self._summary_records
        }
        # 按列筛选条件：column -> 允许的显示值集合；与顶部全局搜索叠加。
        self._summary_filters: dict[str, set[str]] = {}
        self._search_text = ""
        # 导入的入库编号列表筛选：None=未启用；非 None=只显示集合内编号。
        self._voucher_list_filter: set[str] | None = None

        self.filtered_vouchers: list[str] = list(self.all_vouchers)
        self.current_page = 0

        self._build_ui()
        self._load_page(0)

    def _build_photo_counts(self) -> None:
        """Build voucher→photo-count dict using the store cache (avoids direct Excel I/O)."""
        counts: dict[str, int] = {}
        for row in self.store.read_rows("photo"):
            v = str(row.get("入库编号*", "")).strip()
            if v:
                counts[v] = counts.get(v, 0) + 1
        self._photo_counts = counts
        self._total_photos = sum(counts.values())

    @property
    def total_pages(self) -> int:
        return max(1, (len(self.filtered_vouchers) + self.PAGE_SIZE - 1) // self.PAGE_SIZE)

    # ---- UI build ----

    def _build_ui(self) -> None:
        self.setObjectName("ingestSummaryDialog")
        layout = QVBoxLayout(self)
        # 旧布局：默认边距/间距，搜索导航全挤一行。现统一边距 + 拆行（见下），更清爽。
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        # Stats bar
        stats = QHBoxLayout()
        self.stats_label = QLabel()
        stats.addWidget(self.stats_label)
        stats.addStretch()
        layout.addLayout(stats)

        # 旧布局：搜索框/范围/导入/页码/上下页/列设置全挤一行。现拆两行更清爽：
        # 第 1 行 = 搜索区，第 2 行 = 分页 + 列设置。控件、信号全部不变。
        search_row = QHBoxLayout()
        search_row.setSpacing(6)
        search_row.addWidget(QLabel("搜索"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("输入关键词筛选...")
        self.search_edit.textChanged.connect(self._on_search_changed)
        search_row.addWidget(self.search_edit, stretch=1)
        # 搜索范围选择器：与主凭证列表一致的 4 个选项
        self._search_scope = QComboBox()
        self._search_scope.addItem("全部")
        for _col in SUMMARY_COLUMNS:
            self._search_scope.addItem(_col)
        self._search_scope.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self._search_scope.setMinimumWidth(90)
        self._search_scope.currentIndexChanged.connect(lambda: self._on_search_changed(self.search_edit.text()))
        search_row.addWidget(self._search_scope)
        # 「导入编号列表」：粘贴 / 从 txt·csv·xlsx 载入一份入库编号，汇总表只显示列表内编号。
        self._make_btn("导入编号列表", self._open_voucher_list_filter, search_row)
        layout.addLayout(search_row)

        nav = QHBoxLayout()
        nav.setSpacing(6)
        self._make_btn("上一页", self._prev_page, nav)
        self.page_label = QLabel()
        nav.addWidget(self.page_label)
        self._make_btn("下一页", self._next_page, nav)
        nav.addStretch()
        # 「列设置」：开关右侧列选择面板（默认隐藏，可关闭）。
        self._make_btn("列设置", self._toggle_column_panel, nav)
        layout.addLayout(nav)

        # 汇总宽表：SUMMARY_COLUMNS 全字段，占满窗口。
        # 旧布局：左=可编辑宽表 + 右侧详情面板/照片网格的 QSplitter；
        # 现改为只读宽表占满（详情冗余、照片管理移至主窗口）。
        # 旧 edit triggers：DoubleClicked | EditKeyPressed（内联编辑）；现 NoEditTriggers（只读）。
        self.voucher_table = QTableWidget(0, len(SUMMARY_COLUMNS))
        self.voucher_table.setHorizontalHeaderLabels(SUMMARY_COLUMNS)
        self.voucher_table.setSelectionBehavior(QTableWidget.SelectRows)
        # 旧：SingleSelection（只读宽表只单选）。按需求改 ExtendedSelection，与 Windows
        # 资源管理器一致：Ctrl 点选加选、Shift 范围选、拖选范围；双击跳主窗口行为不变。
        self.voucher_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.voucher_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.voucher_table.setSortingEnabled(True)
        # 旧：itemSelectionChanged -> 详情面板/照片网格；itemChanged -> 内联回写。
        # 现：双击行 -> 跳主窗口编辑器。
        self.voucher_table.cellDoubleClicked.connect(self._on_voucher_double_clicked)
        _header = self.voucher_table.horizontalHeader()
        _header.setContextMenuPolicy(Qt.CustomContextMenu)
        # 旧逻辑：只有表头能右键出菜单。现表头 + 表体都接到统一的 _voucher_column_menu。
        _header.customContextMenuRequested.connect(
            lambda pos: self._voucher_column_menu(pos, from_header=True)
        )
        self.voucher_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.voucher_table.customContextMenuRequested.connect(
            lambda pos: self._voucher_column_menu(pos, from_header=False)
        )
        self._apply_visible_columns()

        # 表格 + 右侧列选择面板（面板默认隐藏，「列设置」按钮开关）。
        table_row = QHBoxLayout()
        table_row.addWidget(self.voucher_table, stretch=1)
        table_row.addWidget(self._build_column_panel())
        layout.addLayout(table_row, stretch=1)

        # Bottom：状态提示靠左；操作簇（刷新/导出）与「关闭」用 addSpacing 拉开分区。
        bottom = QHBoxLayout()
        bottom.setSpacing(6)
        self.status_label = QLabel("双击某行可跳转到主窗口编辑该入库编号")
        bottom.addWidget(self.status_label, stretch=1)
        self._make_btn("刷新", self._refresh, bottom)
        self._make_btn("导出 Excel", self._export_excel, bottom)
        self._make_btn("导出 CSV", self._export_csv, bottom)
        bottom.addSpacing(16)
        self._make_btn("关闭", self.close, bottom)
        layout.addLayout(bottom)

        self._update_stats()

    @staticmethod
    def _make_btn(text: str, slot, layout: QHBoxLayout) -> QPushButton:
        btn = QPushButton(text)
        btn.clicked.connect(slot)
        layout.addWidget(btn)
        return btn

    # ---- Pagination ----

    def _update_stats(self) -> None:
        self.stats_label.setText(
            f"入库编号：{len(self.all_vouchers)} 个  |  已入库照片：{self._total_photos} 张"
        )

    @staticmethod
    def _summary_cell_text(value) -> str:
        """统一把记录值转成显示字符串：照片聚合列是 list -> "; " 拼接，其余 str()。"""
        if isinstance(value, list):
            return "; ".join(str(v) for v in value)
        return str(value)

    def _make_cell(self, col: str, record: dict) -> QTableWidgetItem:
        """构造一个汇总宽表单元格；宽表已改全只读，照片数列用数值排序。

        旧逻辑：按 SUMMARY_COLUMN_SOURCE 的 category 决定单元格是否带 ItemIsEditable；
        现在宽表整体只读（编辑改到主窗口），所有单元格统一去掉 ItemIsEditable。
        """
        value = record.get(col, "")
        item = QTableWidgetItem()
        if col == PHOTO_COUNT_COLUMN:
            try:
                item.setData(Qt.DisplayRole, int(value or 0))  # int -> 数值排序
            except (TypeError, ValueError):
                item.setText(str(value))
        else:
            # 照片聚合列是 list -> 拼接显示，并用 tooltip 逐行展示全量。
            item.setText(self._summary_cell_text(value))
            if isinstance(value, list) and value:
                item.setToolTip("\n".join(str(v) for v in value))
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        return item

    def _load_page(self, page: int) -> None:
        self.current_page = max(0, min(page, self.total_pages - 1))
        start = self.current_page * self.PAGE_SIZE
        end = start + self.PAGE_SIZE
        page_vouchers = self.filtered_vouchers[start:end]

        # 填表期间关排序，避免排序错乱。（旧代码另置 _loading 抑制 itemChanged 回写，
        # 宽表已改只读、无 itemChanged 连接，_loading 不再需要。）
        self.voucher_table.setSortingEnabled(False)
        self.voucher_table.setRowCount(0)
        self.voucher_table.setRowCount(len(page_vouchers))
        for i, voucher in enumerate(page_vouchers):
            record = self._record_by_voucher.get(voucher, {})
            for col_idx, col in enumerate(SUMMARY_COLUMNS):
                self.voucher_table.setItem(i, col_idx, self._make_cell(col, record))
        self.voucher_table.setSortingEnabled(True)

        self.voucher_table.resizeColumnsToContents()
        self.page_label.setText(
            f"第 {self.current_page + 1} / {self.total_pages} 页"
            f"（筛选 {len(self.filtered_vouchers)} 条）"
        )

    def _prev_page(self) -> None:
        if self.current_page > 0:
            self._load_page(self.current_page - 1)

    def _next_page(self) -> None:
        if self.current_page < self.total_pages - 1:
            self._load_page(self.current_page + 1)

    def _on_search_changed(self, text: str) -> None:
        """顶部全局搜索：仅记录文本，过滤逻辑统一走 _apply_filters（与按列筛选叠加）。"""
        self._search_text = text
        self._apply_filters()

    def _apply_filters(self) -> None:
        """全局搜索 + 按列筛选叠加，重算 filtered_vouchers 并回到第一页。

        原代码 _on_search_changed 直接算 filtered_vouchers；现在搜索与表头列筛选两套条件
        统一在此合成，数据源改为 self._summary_records（已 join 的汇总记录）。
        """
        query = self._search_text.strip().lower()
        scope = self._search_scope.currentText()
        result: list[str] = []
        for record in self._summary_records:
            voucher = record["入库编号*"]
            # 导入的编号列表筛选：与全局搜索、列筛选叠加；None=未启用。
            if self._voucher_list_filter is not None and voucher not in self._voucher_list_filter:
                continue
            if query:
                # 原：硬编码 4 个范围（入库编号/管内编号/照片名/全部）。
                # 现：下拉含全部 SUMMARY_COLUMNS，统一用 _summary_cell_text 文本化后子串匹配。
                if scope == "全部":
                    ok = any(
                        query in self._summary_cell_text(record.get(col, "")).lower()
                        for col in SUMMARY_COLUMNS
                    )
                else:
                    ok = query in self._summary_cell_text(record.get(scope, "")).lower()
                if not ok:
                    continue
            # 按列筛选：每个被筛选列的显示值必须落在允许集合内
            if not all(
                self._summary_cell_text(record.get(col, "")) in allowed
                for col, allowed in self._summary_filters.items()
            ):
                continue
            result.append(voucher)
        self.filtered_vouchers = result
        self._load_page(0)

    def _reload_summary(self) -> None:
        """重建汇总记录缓存（主窗口改过数据后由 _refresh 调用）。"""
        self._summary_records = self.store.summary_records()
        self._record_by_voucher = {r["入库编号*"]: r for r in self._summary_records}

    def _refresh(self) -> None:
        """重新从 store 拉取最新数据并重算筛选/分页/统计。

        非模态后主窗口可随时改数据，靠「刷新」按钮 / 重新打开工具栏入口同步。
        """
        self.all_vouchers = self.store.list_vouchers()
        self._build_photo_counts()
        self._reload_summary()
        self._apply_filters()  # 重算 filtered_vouchers 并 _load_page(0)
        self._update_stats()
        self.status_label.setText("已刷新")

    # ---- Voucher double-click -> jump to main window ----

    def _on_voucher_double_clicked(self, row: int, col: int) -> None:
        """双击行 -> 主窗口选中该入库编号并聚焦主窗口编辑器（汇总窗口保持打开）。"""
        item = self.voucher_table.item(row, 0)  # 入库编号* 始终第 0 列
        if item is None:
            return
        voucher = item.text()
        # 旧逻辑：self.app.select_voucher(voucher) —— 只填表单，不动主列表；
        # 排序后目标行在别页/被筛选时看不出跳转。改用 reveal_voucher：
        # 清主窗口搜索/筛选 + 翻到目标页 + 选中并滚动到该行。
        self.app.reveal_voucher(voucher)
        self.app.raise_()
        self.app.activateWindow()
        self.status_label.setText(f"已在主窗口选中：{voucher}")

    # ---- 行多选 -> 导出选中入库编号关联的照片 ----

    def _selected_vouchers(self) -> list[str]:
        """汇总表当前选中行对应的入库编号，去重保序（第 0 列始终是 入库编号*）。"""
        result: list[str] = []
        seen: set[str] = set()
        for index in self.voucher_table.selectionModel().selectedRows(0):
            item = self.voucher_table.item(index.row(), 0)
            if item is None:
                continue
            voucher = item.text().strip()
            if voucher and voucher not in seen:
                seen.add(voucher)
                result.append(voucher)
        return result

    def _export_selected_photos(self, vouchers: list[str]) -> None:
        """选中入库编号 -> 复用 BatchExportDialog（photo_focus 模式）导出其关联照片。"""
        if not vouchers:
            return
        dlg = BatchExportDialog(
            self.store, preselected=vouchers, parent=self, photo_focus=True
        )
        dlg.exec_()

    def _open_worms_match_for_vouchers(self, vouchers: list[str]) -> None:
        """入库汇总右键入口：委托主窗口单实例，预填选中标本，应用后刷新汇总。"""
        if not vouchers:
            return
        # Delegate to the parent SpecimenWindow to keep single-instance guarantee.
        app = self.app
        app._open_worms_match()
        if app._worms_window is not None:
            app._worms_window.prefill_vouchers(vouchers)
        self._refresh()

    # ---- 导入入库编号列表筛选 ----

    def _open_voucher_list_filter(self) -> None:
        """导入一份入库编号列表（粘贴文本 + 从 txt·csv·xlsx 载入），汇总表只显示列表内编号。"""
        from .batch_export import _parse_voucher_numbers

        dlg = QDialog(self)
        dlg.setWindowTitle("按入库编号列表筛选")
        dlg.resize(420, 420)
        v = QVBoxLayout(dlg)
        v.addWidget(QLabel("粘贴入库编号（支持换行 / 逗号 / 空格 / 分号分隔）："))
        text_edit = QTextEdit()
        text_edit.setPlaceholderText("例如：\nYZZ000001\nYZZ000042")
        # 已有列表筛选时回填，方便在原基础上增删。
        if self._voucher_list_filter:
            text_edit.setPlainText("\n".join(sorted(self._voucher_list_filter)))
        v.addWidget(text_edit, stretch=1)

        def _load_from_file() -> None:
            path, _ = QFileDialog.getOpenFileName(
                dlg, "从文件载入入库编号", "",
                "编号列表 (*.txt *.csv *.xlsx);;所有文件 (*)",
            )
            if not path:
                return
            try:
                if path.lower().endswith(".xlsx"):
                    # xlsx：取首列所有非空单元格（兼容带/不带表头，_parse 再去重）。
                    from openpyxl import load_workbook
                    wb = load_workbook(path, read_only=True, data_only=True)
                    ws = wb.active
                    cells = [
                        str(row[0])
                        for row in ws.iter_rows(values_only=True)
                        if row and row[0] not in (None, "")
                    ]
                    wb.close()
                    loaded = "\n".join(cells)
                else:
                    loaded = Path(path).read_text(encoding="utf-8-sig")
            except Exception as exc:
                QMessageBox.critical(dlg, "载入失败", f"读取文件失败：{exc}")
                return
            existing = text_edit.toPlainText()
            text_edit.setPlainText(
                (existing + "\n" + loaded) if existing.strip() else loaded
            )

        btn_row = QHBoxLayout()
        self._make_btn("从文件载入…", _load_from_file, btn_row)

        def _clear() -> None:
            # 清除筛选：直接生效并关窗（reject -> 下方不再走解析分支）。
            self._voucher_list_filter = None
            self._apply_filters()
            self.status_label.setText("已清除入库编号列表筛选")
            dlg.reject()

        self._make_btn("清除筛选", _clear, btn_row)
        btn_row.addStretch()
        v.addLayout(btn_row)
        box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        box.accepted.connect(dlg.accept)
        box.rejected.connect(dlg.reject)
        v.addWidget(box)
        if dlg.exec_() != QDialog.Accepted:
            return
        vouchers = _parse_voucher_numbers(text_edit.toPlainText())
        if not vouchers:
            self._voucher_list_filter = None
            self.status_label.setText("编号列表为空，已不按列表筛选")
        else:
            self._voucher_list_filter = set(vouchers)
            self.status_label.setText(f"已按编号列表筛选：{len(vouchers)} 个入库编号")
        self._apply_filters()

    # ---- 列选择面板（默认隐藏、可关闭、多选） ----

    def _build_column_panel(self) -> QWidget:
        """右侧列选择面板：每列一个 QCheckBox，可连续多选、即时显隐、持久化。

        旧逻辑：列显隐靠右键「显示列」可勾选子菜单 —— 勾一列菜单就关，要重复打开。
        现改为可关闭的常驻面板（默认隐藏，「列设置」按钮开关）。
        """
        panel = QWidget()
        pv = QVBoxLayout(panel)
        pv.setContentsMargins(4, 4, 4, 4)
        header_row = QHBoxLayout()
        header_row.addWidget(QLabel("选择显示列"))
        header_row.addStretch()
        close_btn = QPushButton("关闭")
        close_btn.setFixedWidth(48)
        close_btn.clicked.connect(lambda: self._column_panel.setVisible(False))
        header_row.addWidget(close_btn)
        pv.addLayout(header_row)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        inner = QWidget()
        iv = QVBoxLayout(inner)
        iv.setContentsMargins(2, 2, 2, 2)
        self._column_checkboxes: list[QCheckBox] = []
        for idx, col in enumerate(SUMMARY_COLUMNS):
            cb = QCheckBox(col)
            cb.setChecked(not self.voucher_table.isColumnHidden(idx))
            if idx == 0:
                cb.setEnabled(False)  # 入库编号* 锁定可见
            cb.stateChanged.connect(
                lambda state, i=idx: self._toggle_column(i, state == Qt.Checked)
            )
            iv.addWidget(cb)
            self._column_checkboxes.append(cb)
        iv.addStretch()
        scroll.setWidget(inner)
        pv.addWidget(scroll, stretch=1)
        panel.setFixedWidth(180)
        panel.setVisible(False)  # 默认隐藏
        self._column_panel = panel
        return panel

    def _toggle_column_panel(self) -> None:
        self._column_panel.setVisible(not self._column_panel.isVisible())

    # ---- Column show/hide & per-column filter ----

    def _apply_visible_columns(self) -> None:
        """按设置应用列显隐；入库编号* 始终可见（选择/编辑定位依赖第 0 列）。"""
        visible = load_settings().summary_visible_columns or SUMMARY_DEFAULT_VISIBLE_COLUMNS
        visible_set = set(visible)
        for idx, col in enumerate(SUMMARY_COLUMNS):
            hidden = col != "入库编号*" and col not in visible_set
            self.voucher_table.setColumnHidden(idx, hidden)

    def _save_visible_columns(self) -> None:
        visible = [
            col
            for idx, col in enumerate(SUMMARY_COLUMNS)
            if not self.voucher_table.isColumnHidden(idx)
        ]
        settings = load_settings()
        settings.summary_visible_columns = visible
        save_settings(settings)

    def _toggle_column(self, col_idx: int, visible: bool) -> None:
        if not visible and col_idx == 0:
            return  # 入库编号* 不允许隐藏
        self.voucher_table.setColumnHidden(col_idx, not visible)
        self._save_visible_columns()

    def _voucher_column_menu(self, pos: QPoint, from_header: bool) -> None:
        """表头 / 表体右键统一菜单：按列筛选 + 清除筛选 + 列设置。

        旧逻辑：只有 _header_context_menu（仅表头），且「显示列」是可勾选子菜单。
        现：表头和表体右键都可出菜单；列显隐改由「列设置」面板负责，菜单只留入口。
        """
        if from_header:
            header = self.voucher_table.horizontalHeader()
            col_idx = header.logicalIndexAt(pos)
            global_pos = header.mapToGlobal(pos)
        else:
            col_idx = self.voucher_table.columnAt(pos.x())
            global_pos = self.voucher_table.viewport().mapToGlobal(pos)
        menu = QMenu(self)
        # 表体右键且有选中行：顶部加行操作 —— 导出选中入库编号所关联的照片。
        # （原菜单只有列筛选项；列筛选项保留在下方，逻辑不变。）
        if not from_header:
            sel_vouchers = self._selected_vouchers()
            if sel_vouchers:
                menu.addAction(
                    f"导出选中照片 ({len(sel_vouchers)}个)…",
                    lambda: self._export_selected_photos(sel_vouchers),
                )
                menu.addAction(
                    f"用 WoRMS 更新分类 ({len(sel_vouchers)}个)…",
                    lambda sv=sel_vouchers: self._open_worms_match_for_vouchers(sv),
                )
                menu.addSeparator()
        if 0 <= col_idx < len(SUMMARY_COLUMNS):
            col = SUMMARY_COLUMNS[col_idx]
            menu.addAction(f"按「{col}」筛选…", lambda: self._open_column_filter(col))
            if col in self._summary_filters:
                menu.addAction(f"清除「{col}」筛选", lambda: self._clear_column_filter(col))
        # 旧条件：仅 self._summary_filters；现导入的编号列表筛选也算「有筛选」。
        if self._summary_filters or self._voucher_list_filter is not None:
            menu.addAction("清除所有筛选", self._clear_all_filters)
        menu.addSeparator()
        menu.addAction("列设置…", lambda: self._column_panel.setVisible(True))
        menu.exec_(global_pos)

    def _clear_column_filter(self, col: str) -> None:
        self._summary_filters.pop(col, None)
        self._apply_filters()

    def _clear_all_filters(self) -> None:
        self._summary_filters.clear()
        # 旧逻辑只清列筛选；现「清除所有筛选」连导入的编号列表筛选一并清掉。
        self._voucher_list_filter = None
        self._apply_filters()

    def _open_column_filter(self, col: str) -> None:
        """弹出该列去重值多选 + 文本包含框；全选 = 无筛选。"""
        values = sorted({self._summary_cell_text(r.get(col, "")) for r in self._summary_records})
        dlg = QDialog(self)
        dlg.setWindowTitle(f"筛选「{col}」")
        dlg.resize(320, 440)
        v = QVBoxLayout(dlg)
        contains = QLineEdit()
        contains.setPlaceholderText("输入文本快速过滤候选值…")
        v.addWidget(contains)
        listw = QListWidget()
        listw.setSelectionMode(QListWidget.NoSelection)
        current = self._summary_filters.get(col)
        for val in values:
            it = QListWidgetItem(val if val != "" else "（空）")
            it.setFlags(it.flags() | Qt.ItemIsUserCheckable)
            it.setCheckState(Qt.Checked if (current is None or val in current) else Qt.Unchecked)
            it.setData(Qt.UserRole, val)
            listw.addItem(it)

        def _on_contains(text: str) -> None:
            t = text.strip().lower()
            for i in range(listw.count()):
                it = listw.item(i)
                it.setHidden(bool(t) and t not in str(it.data(Qt.UserRole)).lower())

        contains.textChanged.connect(_on_contains)
        v.addWidget(listw, stretch=1)

        def _set_all(state) -> None:
            for i in range(listw.count()):
                if not listw.item(i).isHidden():
                    listw.item(i).setCheckState(state)

        btn_row = QHBoxLayout()
        self._make_btn("全选", lambda: _set_all(Qt.Checked), btn_row)
        self._make_btn("全不选", lambda: _set_all(Qt.Unchecked), btn_row)
        btn_row.addStretch()
        v.addLayout(btn_row)
        box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        box.accepted.connect(dlg.accept)
        box.rejected.connect(dlg.reject)
        v.addWidget(box)
        if dlg.exec_() != QDialog.Accepted:
            return
        checked = {
            str(listw.item(i).data(Qt.UserRole))
            for i in range(listw.count())
            if listw.item(i).checkState() == Qt.Checked
        }
        if len(checked) == len(values):
            self._summary_filters.pop(col, None)  # 全选 = 无筛选
        else:
            self._summary_filters[col] = checked
        self._apply_filters()

    # ---- Export ----

    def _collect_export_data(self) -> tuple[list[str], list[dict]]:
        """导出范围 = 当前筛选结果（全部页）；列 = 当前可见列。"""
        columns = [
            col
            for idx, col in enumerate(SUMMARY_COLUMNS)
            if not self.voucher_table.isColumnHidden(idx)
        ]
        rows: list[dict] = []
        for voucher in self.filtered_vouchers:
            record = self._record_by_voucher.get(voucher, {})
            rows.append({col: self._summary_cell_text(record.get(col, "")) for col in columns})
        return columns, rows

    def _export_excel(self) -> None:
        columns, rows = self._collect_export_data()
        if not rows:
            QMessageBox.information(self, "无数据", "当前没有可导出的记录。")
            return
        path, _ = QFileDialog.getSaveFileName(self, "导出 Excel", "入库汇总.xlsx", "Excel 文件 (*.xlsx)")
        if not path:
            return
        try:
            from openpyxl import Workbook
            from .batch_export import _auto_width, _write_header_row

            wb = Workbook()
            ws = wb.active
            ws.title = "入库汇总"
            _write_header_row(ws, columns)
            for r, row in enumerate(rows, 2):
                for c, col in enumerate(columns, 1):
                    ws.cell(row=r, column=c, value=row.get(col, ""))
            _auto_width(ws)
            wb.save(path)
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", str(exc))
            return
        self.status_label.setText(f"已导出 {len(rows)} 条到 {path}")

    def _export_csv(self) -> None:
        columns, rows = self._collect_export_data()
        if not rows:
            QMessageBox.information(self, "无数据", "当前没有可导出的记录。")
            return
        path, _ = QFileDialog.getSaveFileName(self, "导出 CSV", "入库汇总.csv", "CSV 文件 (*.csv)")
        if not path:
            return
        try:
            import csv

            # utf-8-sig：带 BOM，Excel 打开中文不乱码。
            with open(path, "w", encoding="utf-8-sig", newline="") as fh:
                writer = csv.DictWriter(fh, fieldnames=columns)
                writer.writeheader()
                writer.writerows(rows)
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", str(exc))
            return
        self.status_label.setText(f"已导出 {len(rows)} 条到 {path}")

    def closeEvent(self, event) -> None:
        # 非模态单实例：关闭时清掉主窗口持有的引用，让下次点工具栏能新建。
        # 旧 closeEvent 还做 self._thumb_worker.stop() —— 照片网格已移除，无 thumb worker。
        if getattr(self.app, "_ingest_summary_dialog", None) is self:
            self.app._ingest_summary_dialog = None
        super().closeEvent(event)


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Action log dialog
# ---------------------------------------------------------------------------

class ActionLogDialog(QDialog):
    """Show recent undo/redo operations."""

    def __init__(self, app: SpecimenWindow):
        super().__init__(app)
        self.app = app
        self.setWindowTitle("操作记录")
        self.resize(700, 450)
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("最近操作记录（可撤回/返回）"))
        table = QTableWidget(0, 4)
        table.setHorizontalHeaderLabels(["时间", "操作类型", "入库编号", "详情"])
        table.horizontalHeader().setStretchLastSection(True)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        rows = app.store._read_plain_rows(app.store.data_dir / "操作记录.xlsx")
        depth = int(app.store.config.get("undo_depth", 200))
        recent = rows[-depth:] if len(rows) > depth else rows
        table.setRowCount(len(recent))
        for i, row in enumerate(reversed(recent)):
            undone = app.store._value(row, "是否撤销") == "是"
            atype = app.store._value(row, "操作类型")
            voucher = app.store._value(row, "入库编号")
            cat = app.store._value(row, "信息类别")
            field = app.store._value(row, "字段名")
            prefix = "[已撤销] " if undone else ""
            table.setItem(i, 0, QTableWidgetItem(app.store._value(row, "时间")))
            table.setItem(i, 1, QTableWidgetItem(prefix + atype))
            table.setItem(i, 2, QTableWidgetItem(voucher))
            detail = f"{cat} {field}" if field else cat
            table.setItem(i, 3, QTableWidgetItem(detail))
        table.resizeColumnsToContents()
        layout.addWidget(table, stretch=1)
        undo_count, redo_count = app._undo_redo_counts()
        layout.addWidget(QLabel(f"可撤回：{undo_count} 条  |  可返回：{redo_count} 条"))
        layout.addWidget(QPushButton("关闭", clicked=self.close))


# Version manager dialog
# ---------------------------------------------------------------------------

class VersionManagerDialog(QDialog):
    def __init__(self, app: SpecimenWindow):
        super().__init__(app)
        self.app = app
        self.setWindowTitle("版本管理")
        self.resize(840, 560)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(f"软件版本：v{__version__}"))
        layout.addWidget(QLabel(f"工作区：{app.workspace_root}"))

        tabs = QTabWidget()

        # Data versions tab
        data_tab = QWidget()
        data_layout = QVBoxLayout(data_tab)
        data_layout.addWidget(QPushButton("创建当前数据快照", clicked=self._create_snapshot))
        self.data_table = QTableWidget(0, 4)
        self.data_table.setHorizontalHeaderLabels(["版本ID", "时间", "操作类型", "摘要"])
        self.data_table.horizontalHeader().setStretchLastSection(True)
        self.data_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.data_table.setEditTriggers(QTableWidget.NoEditTriggers)
        data_layout.addWidget(self.data_table, stretch=1)
        btn_row = QHBoxLayout()
        btn_row.addWidget(QPushButton("回退到选中版本", clicked=self._restore_snapshot))
        btn_row.addWidget(QPushButton("打开版本目录", clicked=self._open_snapshot_dir))
        data_layout.addLayout(btn_row)
        tabs.addTab(data_tab, "工作区数据版本")
        self._populate_data_versions()

        # Release tab
        release_tab = QWidget()
        release_layout = QVBoxLayout(release_tab)
        self.release_table = QTableWidget(0, 3)
        self.release_table.setHorizontalHeaderLabels(["版本", "exe", "目录"])
        self.release_table.horizontalHeader().setStretchLastSection(True)
        self.release_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.release_table.setEditTriggers(QTableWidget.NoEditTriggers)
        release_layout.addWidget(self.release_table, stretch=1)
        rel_btn_row = QHBoxLayout()
        rel_btn_row.addWidget(QPushButton("检查 GitHub 更新", clicked=self._check_github_update))
        rel_btn_row.addWidget(QPushButton("打开版本目录", clicked=self._open_release_dir))
        rel_btn_row.addWidget(QPushButton("启动选中版本", clicked=self._launch_release))
        release_layout.addLayout(rel_btn_row)
        self._update_status_label = QLabel("")
        release_layout.addWidget(self._update_status_label)
        tabs.addTab(release_tab, "软件版本")
        self._populate_releases()

        # 后台 worker 引用，防止 QThread 被 GC
        self._update_worker: UpdateCheckWorker | None = None
        self._download_worker: UpdateDownloadWorker | None = None

        layout.addWidget(tabs, stretch=1)
        layout.addWidget(QPushButton("关闭", clicked=self.close), alignment=Qt.AlignRight)

    def _populate_data_versions(self) -> None:
        rows = self.app.store.list_data_versions()
        self.data_table.setRowCount(len(rows))
        for idx, row in enumerate(rows):
            self.data_table.setItem(idx, 0, QTableWidgetItem(str(row.get("版本ID", ""))))
            self.data_table.setItem(idx, 1, QTableWidgetItem(str(row.get("时间", ""))))
            self.data_table.setItem(idx, 2, QTableWidgetItem(str(row.get("操作类型", ""))))
            self.data_table.setItem(idx, 3, QTableWidgetItem(str(row.get("摘要", ""))))
            snapshot_path = row.get("快照路径", "")
            if snapshot_path:
                self.data_table.item(idx, 0).setData(Qt.UserRole, str(snapshot_path))

    def _populate_releases(self) -> None:
        releases = list_releases(self.app.workspace_root)
        self.release_table.setRowCount(len(releases))
        for idx, release in enumerate(releases):
            self.release_table.setItem(idx, 0, QTableWidgetItem(release.version))
            self.release_table.setItem(idx, 1, QTableWidgetItem(release.exe_path.name if release.exe_path else ""))
            self.release_table.setItem(idx, 2, QTableWidgetItem(str(release.directory)))
            self.release_table.item(idx, 0).setData(Qt.UserRole, str(release.directory))
            self.release_table.item(idx, 1).setData(Qt.UserRole, str(release.exe_path or ""))

    def _selected_snapshot_path(self) -> Path | None:
        rows = self.data_table.selectionModel().selectedRows()
        if not rows:
            QMessageBox.information(self, "请选择版本", "请先选择一个数据版本。")
            return None
        item = self.data_table.item(rows[0].row(), 0)
        path_str = item.data(Qt.UserRole) if item else ""
        return Path(path_str) if path_str else None

    def _create_snapshot(self) -> None:
        path = self.app.store.create_data_snapshot("手动快照", "用户在版本管理窗口创建")
        QMessageBox.information(self, "快照已创建", str(path))
        self.accept()
        self.app.open_version_manager()

    def _restore_snapshot(self) -> None:
        snapshot = self._selected_snapshot_path()
        if not snapshot:
            return
        if QMessageBox.question(
            self, "确认回退",
            f"回退前会自动保存当前状态。\n确定恢复到 {snapshot.name} 吗？",
            QMessageBox.Yes | QMessageBox.No,
        ) != QMessageBox.Yes:
            return
        try:
            self.app.store.restore_data_snapshot(snapshot)
        except Exception as exc:
            QMessageBox.critical(self, "回退失败", str(exc))
            return
        self.accept()
        self.app.reload_current()

    def _open_snapshot_dir(self) -> None:
        snapshot = self._selected_snapshot_path()
        if snapshot:
            _open_path(snapshot)

    def _selected_release_paths(self) -> tuple[Path | None, Path | None]:
        rows = self.release_table.selectionModel().selectedRows()
        if not rows:
            QMessageBox.information(self, "请选择版本", "请先选择一个软件版本。")
            return None, None
        r = rows[0].row()
        dir_item = self.release_table.item(r, 0)
        exe_item = self.release_table.item(r, 1)
        directory = Path(dir_item.data(Qt.UserRole)) if dir_item and dir_item.data(Qt.UserRole) else None
        exe = Path(exe_item.data(Qt.UserRole)) if exe_item and exe_item.data(Qt.UserRole) else None
        return directory, exe

    def _open_release_dir(self) -> None:
        directory, _ = self._selected_release_paths()
        if directory:
            _open_path(directory)

    def _launch_release(self) -> None:
        _directory, exe = self._selected_release_paths()
        if not exe:
            QMessageBox.critical(self, "无法启动", "该版本目录中没有可执行文件。")
            return
        if not exe.exists():
            QMessageBox.critical(self, "文件不存在", str(exe))
            return
        from .release_manager import release_roots
        trusted_roots = {r.resolve() for r in release_roots(self.app.workspace_root)}
        if not any(str(exe.resolve()).startswith(str(r)) for r in trusted_roots):
            QMessageBox.critical(self, "安全限制", "可执行文件不在受信任的版本目录内。")
            return
        # 切换版本前提示创建数据快照：新版本若与旧数据不兼容，数据层仍可回退。
        if QMessageBox.question(
            self, "数据快照",
            '启动该版本前是否先创建当前数据快照？\n（新版本出现问题时可在"工作区数据版本"中回退数据）',
            QMessageBox.Yes | QMessageBox.No,
        ) == QMessageBox.Yes:
            try:
                self.app.store.create_data_snapshot("切换版本前快照", f"启动 {exe.name} 前自动创建")
            except Exception as exc:
                QMessageBox.warning(self, "快照失败", f"快照创建失败，仍可继续启动：\n{exc}")
        try:
            if sys.platform != "win32":
                exe.chmod(exe.stat().st_mode | 0o111)
            subprocess.Popen([str(exe), "--workspace", str(self.app.workspace_root)])
        except Exception as exc:
            QMessageBox.critical(self, "启动失败", str(exc))

    # ---- GitHub update check / download ----

    def _check_github_update(self) -> None:
        if self._update_worker is not None or self._download_worker is not None:
            return  # 已有任务在跑
        self._update_status_label.setText("正在检查 GitHub 更新...")
        worker = UpdateCheckWorker(self)
        worker.finished_check.connect(self._on_update_checked)
        worker.finished.connect(lambda: setattr(self, "_update_worker", None))
        self._update_worker = worker
        worker.start()

    def _on_update_checked(self, release, error) -> None:
        self._update_status_label.setText("")
        if error is not None:
            QMessageBox.critical(self, "检查更新失败", str(error))
            return
        if release is None:
            QMessageBox.information(self, "检查更新", "未在 GitHub 上获取到发布信息。")
            return
        if not is_newer(release.version):
            QMessageBox.information(self, "检查更新", f"当前已是最新版本（v{__version__}）。")
            return
        notes = (release.notes or "").strip()
        if len(notes) > 600:
            notes = notes[:600] + "…"
        msg = (
            f"发现新版本 v{release.version}（当前 v{__version__}）。\n\n"
            "是否下载？下载后请在列表中选择该版本启动，旧版本会保留以便回退。"
        )
        if notes:
            msg += f"\n\n更新说明：\n{notes}"
        if QMessageBox.question(
            self, "发现新版本", msg, QMessageBox.Yes | QMessageBox.No,
        ) != QMessageBox.Yes:
            return
        self._start_update_download(release)

    def _start_update_download(self, release) -> None:
        dest_root = default_download_root(self.app.workspace_root)
        # local_roots：增量更新扫描这些 releases 根目录，找运行时可复用的本地版本。
        local_roots = release_roots(self.app.workspace_root)
        self._update_status_label.setText(f"正在下载 v{release.version} … 0%")
        worker = UpdateDownloadWorker(release, dest_root, local_roots, self)
        worker.progress.connect(
            lambda pct: self._update_status_label.setText(f"正在下载 v{release.version} … {pct}%")
        )
        worker.finished_download.connect(self._on_update_downloaded)
        worker.finished.connect(lambda: setattr(self, "_download_worker", None))
        self._download_worker = worker
        worker.start()

    def _on_update_downloaded(self, path, incremental, error) -> None:
        self._update_status_label.setText("")
        if error is not None:
            QMessageBox.critical(self, "下载失败", str(error))
            return
        self._populate_releases()
        mode_note = (
            "本次为增量更新（仅下载应用包，运行时已复用本地版本）。\n\n"
            if incremental else ""
        )
        QMessageBox.information(
            self, "下载完成",
            f"{mode_note}新版本已下载到：\n{path}\n\n"
            '请在上方列表中选择该版本并点击"启动选中版本"。\n'
            "启动前会提示创建数据快照，建议保留。",
        )


# ---------------------------------------------------------------------------
# Settings dialog
# ---------------------------------------------------------------------------

class PhotoFilenameFillDialog(QDialog):
    def __init__(self, filename: str, updates: dict[str, str], current: dict[str, Any], parent=None):
        super().__init__(parent)
        self.setWindowTitle("从照片文件名填充标本信息")
        self._checks: dict[str, tuple[QCheckBox, str]] = {}

        layout = QVBoxLayout(self)
        source = QLabel(f"文件名：{filename}")
        source.setWordWrap(True)
        source.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(source)

        note = QLabel("已有值默认不覆盖；需要覆盖时请手动勾选。")
        note.setWordWrap(True)
        layout.addWidget(note)

        grid = QGridLayout()
        for column, label in enumerate(("填充", "字段", "当前值", "识别值")):
            header = QLabel(label)
            header.setStyleSheet("font-weight: bold;")
            grid.addWidget(header, 0, column)

        default_fields = set(default_photo_filename_fill_fields(updates, current))
        row = 1
        for field in PHOTO_FILENAME_FILL_FIELDS:
            new_value = str(updates.get(field, "") or "")
            if not new_value:
                continue
            current_value = str(current.get(field, "") or "")
            checkbox = QCheckBox()
            checkbox.setChecked(field in default_fields)
            if current_value == new_value:
                checkbox.setChecked(False)
                checkbox.setEnabled(False)
            self._checks[field] = (checkbox, new_value)
            grid.addWidget(checkbox, row, 0, alignment=Qt.AlignCenter)
            grid.addWidget(QLabel(field), row, 1)
            grid.addWidget(QLabel(current_value or "（空）"), row, 2)
            grid.addWidget(QLabel(new_value), row, 3)
            row += 1
        layout.addLayout(grid)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        ok_button = buttons.button(QDialogButtonBox.Ok)
        if ok_button is not None:
            ok_button.setText("填充选中字段")
        cancel_button = buttons.button(QDialogButtonBox.Cancel)
        if cancel_button is not None:
            cancel_button.setText("取消")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def selected_updates(self) -> dict[str, str]:
        return {
            field: value
            for field, (checkbox, value) in self._checks.items()
            if checkbox.isChecked()
        }


class BatchSpecimenFieldsDialog(QDialog):
    """多选入库编号后，批量设置标本信息字段（CARRY_OVER_SPECIMEN_FIELDS）。

    每行一个勾选框：只有勾选的字段才会写入；勾选但留空表示批量清空该字段。
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("批量设置标本信息")
        self._rows: dict[str, tuple[QCheckBox, QLineEdit | QComboBox]] = {}

        layout = QVBoxLayout(self)
        note = QLabel("只勾选并填写需要批量修改的字段；未勾选的字段保持不变。\n勾选但留空表示批量清空该字段。")
        note.setWordWrap(True)
        layout.addWidget(note)

        grid = QGridLayout()
        for column, label in enumerate(("修改", "字段", "值")):
            header = QLabel(label)
            header.setStyleSheet("font-weight: bold;")
            grid.addWidget(header, 0, column)

        for row, field in enumerate(CARRY_OVER_SPECIMEN_FIELDS, start=1):
            checkbox = QCheckBox()
            if field == "保存方式":
                editor: QLineEdit | QComboBox = QComboBox()
                editor.addItems(SAVE_METHOD_OPTIONS)
                editor.setEditable(True)
                editor.setCurrentText("")
            else:
                editor = QLineEdit()
            grid.addWidget(checkbox, row, 0, alignment=Qt.AlignCenter)
            grid.addWidget(QLabel(field), row, 1)
            grid.addWidget(editor, row, 2)
            self._rows[field] = (checkbox, editor)
        layout.addLayout(grid)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        ok_button = buttons.button(QDialogButtonBox.Ok)
        if ok_button is not None:
            ok_button.setText("批量设置")
        cancel_button = buttons.button(QDialogButtonBox.Cancel)
        if cancel_button is not None:
            cancel_button.setText("取消")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def selected_updates(self) -> dict[str, str]:
        updates: dict[str, str] = {}
        for field, (checkbox, editor) in self._rows.items():
            if not checkbox.isChecked():
                continue
            value = editor.currentText() if isinstance(editor, QComboBox) else editor.text()
            updates[field] = value.strip()
        return updates


class SettingsDialog(QDialog):
    def __init__(self, app: SpecimenWindow):
        super().__init__(app)
        self.app = app
        self.setWindowTitle("设置")

        layout = QFormLayout(self)

        self.undo_spin = QSpinBox()
        self.undo_spin.setRange(1, 1000)
        self.undo_spin.setValue(int(app.store.config.get("undo_depth", 200)))
        layout.addRow("操作记录保存步数", self.undo_spin)

        self.quality_combo = QComboBox()
        quality_keys = list(PREVIEW_QUALITY_OPTIONS.keys())
        for key in quality_keys:
            self.quality_combo.addItem(PREVIEW_QUALITY_OPTIONS[key], key)
        current_settings = load_settings()
        current_idx = quality_keys.index(current_settings.preview_quality) if current_settings.preview_quality in quality_keys else 0
        self.quality_combo.setCurrentIndex(current_idx)
        layout.addRow("图片预览质量", self.quality_combo)

        self.photo_management_combo = QComboBox()
        management_keys = list(PHOTO_MANAGEMENT_OPTIONS.keys())
        for key in management_keys:
            self.photo_management_combo.addItem(PHOTO_MANAGEMENT_OPTIONS[key], key)
        management_idx = management_keys.index(current_settings.photo_management_mode) if current_settings.photo_management_mode in management_keys else 0
        self.photo_management_combo.setCurrentIndex(management_idx)
        layout.addRow("照片管理方式", self.photo_management_combo)

        library_row = QHBoxLayout()
        self.photo_library_edit = QLineEdit(current_settings.photo_library_path)
        library_row.addWidget(self.photo_library_edit, stretch=1)
        browse_btn = QPushButton("选择")
        browse_btn.clicked.connect(self._choose_photo_library)
        library_row.addWidget(browse_btn)
        layout.addRow("自定义照片库", library_row)

        # 自定义图片查看器：留空 = 用系统默认程序打开原图。
        viewer_row = QHBoxLayout()
        self.image_viewer_edit = QLineEdit(current_settings.image_viewer_path)
        self.image_viewer_edit.setPlaceholderText("留空 = 系统默认程序")
        viewer_row.addWidget(self.image_viewer_edit, stretch=1)
        viewer_btn = QPushButton("选择")
        viewer_btn.clicked.connect(self._choose_image_viewer)
        viewer_row.addWidget(viewer_btn)
        layout.addRow("自定义图片查看器", viewer_row)

        self.photo_fill_shortcut_edit = QLineEdit(current_settings.photo_filename_fill_shortcut)
        self.photo_fill_shortcut_edit.setPlaceholderText(DEFAULT_PHOTO_FILENAME_FILL_SHORTCUT)
        layout.addRow("照片名填充快捷键", self.photo_fill_shortcut_edit)

        self.check_updates_box = QCheckBox("启动时自动检查 GitHub 更新")
        self.check_updates_box.setChecked(current_settings.check_updates_on_startup)
        layout.addRow("软件更新", self.check_updates_box)

        # 界面字体大小：影响所有主体字体（列表/表单/标签/按钮等）。
        # 0=系统默认；这里展示绝对 pt 值，复位时回到系统默认字号。
        self.font_size_spin = QSpinBox()
        self.font_size_spin.setRange(7, 24)
        self.font_size_spin.setSuffix(" pt")
        self.font_size_spin.setToolTip("调整界面主体字体大小；也可用 Ctrl+加/减/0 快捷键")
        self.font_size_spin.setValue(
            current_settings.ui_font_size if current_settings.ui_font_size > 0
            else (_default_app_font_point or self.font().pointSize())
        )
        layout.addRow("界面字体大小", self.font_size_spin)

        # 规范化软件设计 2026-05 内存档位:让用户在低内存机锁定"低",大数据汇总锁定"高/极高"。
        from .app_settings import MEMORY_PROFILE_OPTIONS
        self.memory_profile_combo = QComboBox()
        mp_keys = list(MEMORY_PROFILE_OPTIONS.keys())
        for key in mp_keys:
            self.memory_profile_combo.addItem(MEMORY_PROFILE_OPTIONS[key], key)
        mp_idx = (
            mp_keys.index(current_settings.memory_profile)
            if current_settings.memory_profile in mp_keys else mp_keys.index("auto")
        )
        self.memory_profile_combo.setCurrentIndex(mp_idx)
        self.memory_profile_combo.setToolTip(
            "调整缩略图缓存 / Excel 缓存 / 并发解码档位\n"
            "低档省内存,高档加速大数据汇总"
        )
        layout.addRow("内存档位", self.memory_profile_combo)
        # hint label (灰色小字警告 + 重启提示)
        mp_hint = QLabel(
            "调高档位可加速大数据汇总但需要更多内存;\n"
            "极高档不建议在 < 8GB 机器使用,会触发系统 swap 反而卡顿。\n"
            "缩略图缓存与 Excel 缓存即时生效;并发解码线程数需重启应用生效。"
        )
        mp_hint.setStyleSheet("color: #888; font-size: 11px;")
        mp_hint.setWordWrap(True)
        layout.addRow("", mp_hint)

        # 趣味光标样式：替代默认箭头，可选卡通食指/手掌/钢笔/爪印/星星。
        from .cursors import CURSOR_STYLE_OPTIONS
        self.cursor_combo = QComboBox()
        cursor_keys = list(CURSOR_STYLE_OPTIONS.keys())
        for key in cursor_keys:
            self.cursor_combo.addItem(CURSOR_STYLE_OPTIONS[key], key)
        cursor_idx = (
            cursor_keys.index(current_settings.cursor_style)
            if current_settings.cursor_style in cursor_keys else 0
        )
        self.cursor_combo.setCurrentIndex(cursor_idx)
        self.cursor_combo.setToolTip("把鼠标默认箭头换成趣味光标；文本框/拉伸等专用光标不受影响")
        layout.addRow("光标样式", self.cursor_combo)

        # 应用图标变体：4 款预生成图标可切换（窗口/任务栏图标）。
        from .icon import APP_ICON_VARIANTS
        self.icon_combo = QComboBox()
        icon_keys = list(APP_ICON_VARIANTS.keys())
        for key in icon_keys:
            self.icon_combo.addItem(APP_ICON_VARIANTS[key], key)
        icon_idx = (
            icon_keys.index(current_settings.app_icon_variant)
            if current_settings.app_icon_variant in icon_keys else 0
        )
        self.icon_combo.setCurrentIndex(icon_idx)
        self.icon_combo.setToolTip("切换应用图标（窗口与任务栏）；exe 文件图标在打包时固定")
        layout.addRow("应用图标", self.icon_combo)

        btn_row = QHBoxLayout()
        btn_row.addWidget(QPushButton("保存", clicked=self.accept))
        restore_btn = QPushButton("恢复默认设置")
        restore_btn.clicked.connect(self._restore_defaults)
        btn_row.addWidget(restore_btn)
        layout.addRow(btn_row)

    def _restore_defaults(self) -> None:
        from .app_settings import AppSettings
        defaults = AppSettings()
        self.undo_spin.setValue(200)
        quality_keys = list(PREVIEW_QUALITY_OPTIONS.keys())
        default_idx = quality_keys.index(defaults.preview_quality) if defaults.preview_quality in quality_keys else 0
        self.quality_combo.setCurrentIndex(default_idx)
        management_keys = list(PHOTO_MANAGEMENT_OPTIONS.keys())
        management_idx = management_keys.index(defaults.photo_management_mode)
        self.photo_management_combo.setCurrentIndex(management_idx)
        self.photo_library_edit.setText(defaults.photo_library_path)
        self.image_viewer_edit.setText(defaults.image_viewer_path)
        self.photo_fill_shortcut_edit.setText(defaults.photo_filename_fill_shortcut)
        self.check_updates_box.setChecked(defaults.check_updates_on_startup)
        # 字体大小复位到系统默认（spinbox 展示系统默认 pt，设置值存 0）。
        self.font_size_spin.setValue(_default_app_font_point or self.font().pointSize())
        # 光标样式复位到默认箭头。
        from .cursors import CURSOR_STYLE_OPTIONS
        cursor_keys = list(CURSOR_STYLE_OPTIONS.keys())
        self.cursor_combo.setCurrentIndex(
            cursor_keys.index(defaults.cursor_style) if defaults.cursor_style in cursor_keys else 0
        )
        # 应用图标复位到默认变体。
        from .icon import APP_ICON_VARIANTS
        icon_keys = list(APP_ICON_VARIANTS.keys())
        self.icon_combo.setCurrentIndex(
            icon_keys.index(defaults.app_icon_variant) if defaults.app_icon_variant in icon_keys else 0
        )
        # 规范化软件设计 2026-05 内存档位:复位到 "auto"。
        from .app_settings import MEMORY_PROFILE_OPTIONS
        mp_keys = list(MEMORY_PROFILE_OPTIONS.keys())
        self.memory_profile_combo.setCurrentIndex(
            mp_keys.index(defaults.memory_profile) if defaults.memory_profile in mp_keys else mp_keys.index("auto")
        )
        # Apply immediately
        self.app.store.set_undo_depth(200)
        save_settings(defaults)
        self.app._cached_preview_quality = defaults.preview_quality
        self.app._show_grid_filenames = defaults.show_grid_filenames
        self.app._show_filename_check.setChecked(defaults.show_grid_filenames)
        self.app._apply_photo_filename_fill_shortcut()
        apply_app_font_size(defaults.ui_font_size)  # defaults.ui_font_size == 0 -> 系统默认
        self.app._refresh_all_windows_fonts()
        apply_app_cursor(defaults.cursor_style)  # 恢复系统箭头
        apply_app_icon(defaults.app_icon_variant)  # 恢复默认图标变体
        QMessageBox.information(self, "已恢复", "所有设置已恢复为默认值。")

    def _choose_photo_library(self) -> None:
        current = self.photo_library_edit.text().strip()
        directory = QFileDialog.getExistingDirectory(self, "选择自定义照片库", current or str(self.app.workspace_root))
        if directory:
            self.photo_library_edit.setText(directory)

    def _choose_image_viewer(self) -> None:
        current = self.image_viewer_edit.text().strip()
        path, _ = QFileDialog.getOpenFileName(self, "选择图片查看器程序", current or "")
        if path:
            self.image_viewer_edit.setText(path)

    @property
    def undo_depth(self) -> int:
        return self.undo_spin.value()

    @property
    def image_viewer_path(self) -> str:
        return self.image_viewer_edit.text().strip()

    @property
    def cursor_style(self) -> str:
        key = self.cursor_combo.currentData()
        return key if isinstance(key, str) and key else "default"

    @property
    def app_icon_variant(self) -> str:
        from .icon import DEFAULT_APP_ICON_VARIANT
        key = self.icon_combo.currentData()
        return key if isinstance(key, str) and key else DEFAULT_APP_ICON_VARIANT

    @property
    def photo_filename_fill_shortcut(self) -> str:
        value = self.photo_fill_shortcut_edit.text().strip()
        return value or DEFAULT_PHOTO_FILENAME_FILL_SHORTCUT

    @property
    def font_size(self) -> int:
        return int(self.font_size_spin.value())

    @property
    def memory_profile(self) -> str:
        from .app_settings import MEMORY_PROFILE_OPTIONS
        key = self.memory_profile_combo.currentData()
        return key if isinstance(key, str) and key in MEMORY_PROFILE_OPTIONS else "auto"


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

class WindowManager:
    def __init__(self, app: QApplication):
        self.app = app
        self._windows: dict[Path, SpecimenWindow] = {}
        # 规范化软件设计 2026-05 多窗口:只读副本不进 _windows dict (允许同工作区多副本),
        # 但需 keep_alive 持引用防 GC,放 _ro_windows list。
        self._ro_windows: list[SpecimenWindow] = []
        # C1: 已注册的子对话框 + 它们的"停止 worker"方法。
        self._dialog_stop_handlers: list[tuple[object, "callable"]] = []

    def register(self, window: SpecimenWindow) -> None:
        # 只读副本不入主 dict (允许同工作区多副本共存),改入 _ro_windows
        if getattr(window, "read_only", False):
            if window not in self._ro_windows:
                self._ro_windows.append(window)
            return
        # 未绑定工作区的窗口（首次启动尚未选工作区）不注册；
        # 载入工作区时 _load_workspace_into_window 会再次调用 register。
        if getattr(window, "workspace_root", None) is None:
            return
        self._windows[window.workspace_root.resolve()] = window

    def unregister(self, window: SpecimenWindow) -> None:
        # 只读副本从 _ro_windows 移除
        if getattr(window, "read_only", False):
            try:
                self._ro_windows.remove(window)
            except ValueError:
                pass
            return
        key = getattr(window, "workspace_root", None)
        if key is None:
            return
        key = key.resolve()
        if self._windows.get(key) is window:
            self._windows.pop(key, None)

    def register_dialog_stopper(self, dialog: object, stop_handler) -> None:
        """C1: 让对话框（如 DbManagerDialog）注册一个"停止后台 worker"回调。

        主窗口 closeEvent 会遍历调用 — `stop_handler(wait_ms: int)`，由 dialog 实现限时
        优雅停 + terminate fallback。
        """
        self._dialog_stop_handlers.append((dialog, stop_handler))

    def unregister_dialog_stopper(self, dialog: object) -> None:
        self._dialog_stop_handlers = [
            (d, h) for d, h in self._dialog_stop_handlers if d is not dialog
        ]

    def stop_all_dialog_workers(self, wait_ms: int = 3000) -> None:
        """主窗口关闭时调用：让所有注册的对话框停止它们的后台 worker。"""
        for dialog, handler in list(self._dialog_stop_handlers):
            try:
                handler(wait_ms)
            except Exception as exc:
                print(f"[WindowManager] stop dialog worker 失败：{exc}", file=sys.stderr)
        self._dialog_stop_handlers.clear()

    def focus_workspace(self, workspace_root: Path | str, exclude: SpecimenWindow | None = None) -> bool:
        key = Path(workspace_root).resolve()
        window = self._windows.get(key)
        if window is None or window is exclude:
            return False
        window.show()
        window.raise_()
        window.activateWindow()
        return True

    def open_workspace(self, workspace_root: Path | str | None,
                       read_only: bool = False) -> SpecimenWindow | None:
        """打开工作区窗口。

        规范化软件设计 2026-05 多窗口:
        - read_only=False (默认): 若同工作区主窗口已开 → focus 现有不新建
        - read_only=True: 直接新建只读副本,允许同工作区多副本共存
        """
        if not read_only and workspace_root is not None and self.focus_workspace(workspace_root):
            return self._windows.get(Path(workspace_root).resolve())
        try:
            window = SpecimenWindow(workspace_root, manager=self, read_only=read_only)
        except SystemExit:
            return None
        except Exception as exc:
            QMessageBox.critical(None, "启动失败", str(exc))
            return None
        self.register(window)
        window.show()
        return window


def run_app(workspace_root: Path | str | None) -> None:
    # 规范化软件设计 2026-05 启动卡死优化:
    # 1. QApplication 先创建,Splash 立刻可见 (无视觉反馈是"卡死"误判主因)
    # 2. 各启动阶段 splash.show_stage(text, percent) 给用户进度
    # 3. WoRMS bootstrap / ThumbnailWorker 等延后 (见 _finish_initial_load)

    # E1: 第一时间装异常 hook,让"启动过程中"的崩溃也能写 crash log。
    from .crash_log import (
        install_excepthook,
        mark_app_started,
        list_recent_crash_logs,
    )
    install_excepthook()
    _last_exit_was_clean = mark_app_started()

    if workspace_root is None:
        workspace_root = default_workspace()
    app = QApplication.instance() or QApplication(sys.argv)
    # 记录系统默认字号,并应用用户保存的全局字体大小(窗口创建前完成,新窗口即继承)。
    global _default_app_font_point
    _default_app_font_point = app.font().pointSize()
    apply_app_font_size(load_settings().ui_font_size)
    # 主题 QSS (主题不含 font 规则,不影响字体缩放)。
    from .theme import apply_app_theme
    apply_app_theme(app)
    # ---- Splash Screen 立刻可见 (规范化软件设计 2026-05 新增) ----
    splash = None
    try:
        from .splash import SplashScreen
        splash = SplashScreen()
        splash.show()
        app.processEvents()  # 让 splash 实际出现在屏幕上
        splash.show_stage("初始化…", 5)
    except Exception:
        splash = None  # splash 失败不影响启动主流程
    # ---- 落环境快照到 startup_diag (新增) ----
    try:
        from .env_detect import env_snapshot, is_low_memory, is_wsl, is_fast_profile
        from .startup_diag import mark as _mark
        _mark(f"env: {env_snapshot()}")
        if is_low_memory():
            _mark("env: LOW MEMORY MODE (< 3GB RAM)")
        if is_wsl():
            _mark("env: WSL detected, software rendering set in run_app.py")
        # 规范化软件设计 2026-05 K 章:高档位预 import openpyxl + PIL,后续首次读 Excel /
        # 缩略图无 lazy 导入延迟。失败不阻断 (try/except 已包,会 fall back lazy 路径)。
        if is_fast_profile():
            _mark("env: FAST PROFILE (high/extra_high) — preloading heavy libs")
            try:
                from .excel_store import _ensure_openpyxl
                _ensure_openpyxl()  # 预触发 openpyxl 顶层 import + numpy 阻塞逻辑
            except Exception:
                pass
            try:
                from PIL import Image  # noqa: F401 — 预加载,Python sys.modules 缓存生效
            except Exception:
                pass
            try:
                # 同时预触发 ImageOps (batch_export / image_cache 用)
                from PIL import ImageOps  # noqa: F401
            except Exception:
                pass
            _mark("env: FAST PROFILE preload done")
    except Exception:
        pass
    if splash is not None:
        splash.show_stage("打开工作区…", 25)
    manager = WindowManager(app)
    window = manager.open_workspace(workspace_root)
    # 行为变化:首次启动无工作区时 open_workspace 现在返回未绑定窗口(非 None),
    # 不再走此退出分支 —— 窗口已显示并会提示选择工作区。
    # window is None 现在仅表示真正的启动失败(如 --workspace 指向无效目录,
    # SpecimenWindow 抛 SystemExit)。
    if window is None:
        if splash is not None:
            splash.close()
        QMessageBox.warning(None, "标本入库管理", "工作区无效,程序将退出。")
        return
    if splash is not None:
        splash.show_stage("加载界面…", 80)
        app.processEvents()
    # 旧逻辑:无此调用,光标始终系统箭头。现按用户设置应用趣味光标(窗口已创建后调用)。
    _startup_settings = load_settings()
    apply_app_cursor(_startup_settings.cursor_style)
    apply_app_icon(_startup_settings.app_icon_variant)  # 应用用户选的图标变体
    if splash is not None:
        splash.show_stage("准备就绪", 100)
        app.processEvents()
        splash.finish(window)  # 主窗口可见后关闭 splash
    # WoRMS bootstrap 延后 3s 后台触发 (规范化软件设计 2026-05 启动卡死优化):
    # 旧:run_app() 末尾同步 ensure_bootstrap_cache 解压 sqlite.gz,加 0.5-1s 黑屏。
    # 现:延后到 app.exec_() 后 3s,首启不阻塞启动。若用户在 3s 内就开 WoRMS 菜单,
    # _open_worms_match 内部的守卫会触发"同步准备 + 进度条",见 S10。
    # K 章 高档位快路径 (2026-05):高档位 delay=0,启动后立刻触发,用户开 WoRMS 无等待。
    def _bootstrap_worms_later() -> None:
        try:
            from .worms_client import ensure_bootstrap_cache
            if ensure_bootstrap_cache():
                window.statusBar().showMessage("已加载内置 WoRMS 分类缓存", 3000)
        except Exception:
            pass  # bootstrap 不可用不阻塞启动
    try:
        from .env_detect import is_fast_profile
        _worms_delay = 0 if is_fast_profile() else 3000
    except Exception:
        _worms_delay = 3000
    QTimer.singleShot(_worms_delay, _bootstrap_worms_later)
    # WSL / Unix 兼容：注册 SIGINT/SIGTERM 处理，使 Ctrl+C 和终端关闭触发干净退出。
    # Qt 默认不处理 Python 信号；需要一个空 QTimer 定时让事件循环回到 Python 捡信号。
    def _handle_os_signal(signum, frame):
        a = QApplication.instance()
        if a:
            a.quit()
    signal.signal(signal.SIGINT, _handle_os_signal)
    signal.signal(signal.SIGTERM, _handle_os_signal)
    _sig_keepalive = QTimer()
    _sig_keepalive.start(500)
    _sig_keepalive.timeout.connect(lambda: None)

    # E1: 启动后 ~2s 异步提示用户"上次未正常退出"。延迟以避开启动期繁忙。
    if not _last_exit_was_clean:
        def _show_crash_hint() -> None:
            recent = list_recent_crash_logs(limit=3)
            if recent:
                paths_text = "\n".join(f"  · {p.name}" for p in recent)
                detail = (
                    f"检测到上次应用未正常退出。最近的崩溃日志：\n{paths_text}\n\n"
                    f"位置：{recent[0].parent}\n\n"
                    "如反复出现，请把日志反馈给开发者。"
                )
            else:
                detail = (
                    "检测到上次应用未正常退出（可能是任务管理器结束 / 强制重启 / 断电）。\n"
                    "没有崩溃日志说明属于系统级中止，本次启动一切正常。"
                )
            try:
                QMessageBox.information(None, "上次异常退出", detail)
            except Exception:
                pass

        QTimer.singleShot(2000, _show_crash_hint)

    app.exec_()


# ─────────────────────────────────────────────────────────────────────────────
# 系列管理对话框
# ─────────────────────────────────────────────────────────────────────────────

class AccessionSeriesDialog(QDialog):
    """管理入库编号系列：新增 / 编辑起始号 / 删除。YZZ 为系统固定系列，只读展示。"""

    _YZZ_ROW = 0  # YZZ 固定占第 0 行

    def __init__(self, store: "ExcelStore", parent: QWidget | None = None):
        super().__init__(parent)
        self.store = store
        self.setWindowTitle("入库编号系列管理")
        self.setMinimumWidth(560)
        self._build_ui()
        self._refresh()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        # 系列列表：5 列（名称 | 示例编号 | 已分发 | 下一号 | 步长）
        self._list = QTableWidget(0, 5)
        self._list.setHorizontalHeaderLabels(["名称", "示例编号", "已分发", "下一号", "步长"])
        self._list.setSelectionBehavior(QTableWidget.SelectRows)
        self._list.setSelectionMode(QTableWidget.SingleSelection)
        self._list.setEditTriggers(QTableWidget.NoEditTriggers)
        self._list.verticalHeader().setVisible(False)
        self._list.horizontalHeader().setStretchLastSection(True)
        self._list.itemSelectionChanged.connect(self._on_selection_changed)
        layout.addWidget(self._list)

        # 按钮行
        btn_row = QHBoxLayout()
        self._add_btn = QPushButton("新增系列…")
        self._add_btn.clicked.connect(self._add_series)
        btn_row.addWidget(self._add_btn)
        self._edit_btn = QPushButton("编辑起始号…")
        self._edit_btn.clicked.connect(self._edit_counter)
        btn_row.addWidget(self._edit_btn)
        self._del_btn = QPushButton("删除")
        self._del_btn.clicked.connect(self._delete_series)
        btn_row.addWidget(self._del_btn)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.rejected.connect(self.accept)
        layout.addWidget(buttons)

    def _refresh(self) -> None:
        from .accession_series import AccessionSeries, format_series_number
        from .parsing import format_voucher
        series_list = self.store.config.get("accession_series", [])
        # YZZ 固定第 0 行 + 自定义系列
        self._list.setRowCount(1 + len(series_list))

        # YZZ 行（只读，灰色背景）
        yzz_next = self.store.config.get("next_serial", 1)
        yzz_example = format_voucher(yzz_next)
        yzz_distributed = self.store.count_vouchers_by_series("YZZ")
        gray = QColor(220, 220, 220)
        for col, text in enumerate([
            "YZZ（系统默认，不可删除）",
            yzz_example,
            str(yzz_distributed),
            str(yzz_next),
            "1",
        ]):
            item = QTableWidgetItem(text)
            item.setBackground(gray)
            self._list.setItem(self._YZZ_ROW, col, item)

        # 自定义系列行
        for idx, item_dict in enumerate(series_list):
            row = idx + 1
            s = AccessionSeries.from_dict(item_dict)
            example = format_series_number(s, s.next_counter)
            distributed = self.store.count_vouchers_by_series(s.name)
            self._list.setItem(row, 0, QTableWidgetItem(s.name))
            self._list.setItem(row, 1, QTableWidgetItem(example))
            self._list.setItem(row, 2, QTableWidgetItem(str(distributed)))
            self._list.setItem(row, 3, QTableWidgetItem(str(s.next_counter)))
            self._list.setItem(row, 4, QTableWidgetItem(str(s.step)))

        self._on_selection_changed()

    def _on_selection_changed(self) -> None:
        """选中 YZZ 行时禁用删除和编辑起始号按钮。"""
        cur = self._list.currentRow()
        is_yzz = cur == self._YZZ_ROW
        has_sel = cur >= 0
        self._del_btn.setEnabled(has_sel and not is_yzz)
        self._edit_btn.setEnabled(has_sel and not is_yzz)

    def _selected_name(self) -> str | None:
        cur = self._list.currentRow()
        if cur < 0 or cur == self._YZZ_ROW:
            return None
        item = self._list.item(cur, 0)
        return item.text() if item else None

    def _add_series(self) -> None:
        dlg = _SeriesEditDialog(self.store, parent=self)
        if dlg.exec_() == QDialog.Accepted and dlg.result_series is not None:
            self.store.add_series(dlg.result_series)
            self._refresh()

    def _edit_counter(self) -> None:
        name = self._selected_name()
        if not name:
            QMessageBox.information(self, "提示", "请先选择一个非 YZZ 系列。")
            return
        series = self.store._get_series_config(name)
        if series is None:
            return
        val, ok = QInputDialog.getInt(
            self, "编辑起始号",
            f"设置「{name}」下一个编号的流水号（跳过已用编号）：",
            series.next_counter, 1, 999_999_999,
        )
        if ok:
            self.store.update_series_counter(name, val)
            self._refresh()

    def _delete_series(self) -> None:
        name = self._selected_name()
        if not name:
            QMessageBox.information(self, "提示", "请先选择一个非 YZZ 系列。")
            return
        reply = QMessageBox.question(
            self, "确认删除",
            f"删除系列「{name}」的配置？\n已录入的编号数据不受影响。",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            self.store.remove_series(name)
            self._refresh()


class _SeriesEditDialog(QDialog):
    """新增或编辑一个入库编号系列配置。"""

    def __init__(self, store: "ExcelStore", series: AccessionSeries | None = None, parent: QWidget | None = None):
        super().__init__(parent)
        self.store = store
        self.result_series: AccessionSeries | None = None
        self._editing = series
        self.setWindowTitle("新增编号系列" if series is None else "编辑编号系列")
        self.setMinimumWidth(360)
        self._build_ui(series)

    def _build_ui(self, series: AccessionSeries | None) -> None:
        layout = QVBoxLayout(self)
        form = QFormLayout()
        form.setSpacing(8)

        # 预设选择
        preset_row = QHBoxLayout()
        self._preset_combo = QComboBox()
        self._preset_combo.addItem("（自定义）", None)
        for p in BUILTIN_PRESETS:
            self._preset_combo.addItem(p["label"], p)
        self._preset_combo.currentIndexChanged.connect(self._apply_preset)
        preset_row.addWidget(QLabel("使用预设："))
        preset_row.addWidget(self._preset_combo, stretch=1)
        layout.addLayout(preset_row)

        # 表单字段
        self._name_edit = QLineEdit(series.name if series else "")
        self._name_edit.setPlaceholderText("如：BMNH 大英自然历史博物馆")
        form.addRow("系列名称：", self._name_edit)

        self._prefix_edit = QLineEdit(series.prefix if series else "")
        self._prefix_edit.setPlaceholderText("如：BMNH")
        form.addRow("前缀：", self._prefix_edit)

        self._digits_spin = QSpinBox()
        self._digits_spin.setRange(3, 12)
        self._digits_spin.setValue(series.digits if series else 6)
        form.addRow("流水号位数：", self._digits_spin)

        self._sep_combo = QComboBox()
        for label, val in [("横线 -", "-"), ("点 .", "."), ("斜线 /", "/"), ("下划线 _", "_"), ("无分隔", "")]:
            self._sep_combo.addItem(label, val)
        if series:
            idx = self._sep_combo.findData(series.separator)
            if idx >= 0:
                self._sep_combo.setCurrentIndex(idx)
        form.addRow("分隔符：", self._sep_combo)

        self._year_combo = QComboBox()
        for label, val in [("不含年份", "none"), ("年份在前 (2025-PREFIX-000001)", "before"), ("年份在后 (PREFIX-2025-000001)", "after")]:
            self._year_combo.addItem(label, val)
        if series:
            idx = self._year_combo.findData(series.year_pos)
            if idx >= 0:
                self._year_combo.setCurrentIndex(idx)
        form.addRow("年份位置：", self._year_combo)

        self._counter_spin = QSpinBox()
        self._counter_spin.setRange(1, 999_999_999)
        self._counter_spin.setValue(series.next_counter if series else 1)
        form.addRow("起始流水号：", self._counter_spin)

        self._step_spin = QSpinBox()
        self._step_spin.setRange(1, 100)
        self._step_spin.setValue(series.step if series else 1)
        form.addRow("步长：", self._step_spin)

        layout.addLayout(form)

        # 预览
        self._preview_label = QLabel()
        self._preview_label.setStyleSheet("color: #2a6fbd; font-weight: bold;")
        layout.addWidget(self._preview_label)
        for w in (self._prefix_edit, self._digits_spin, self._sep_combo, self._year_combo, self._counter_spin, self._step_spin):
            if hasattr(w, "textChanged"):
                w.textChanged.connect(self._update_preview)
            elif hasattr(w, "valueChanged"):
                w.valueChanged.connect(self._update_preview)
            elif hasattr(w, "currentIndexChanged"):
                w.currentIndexChanged.connect(self._update_preview)
        self._update_preview()

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self._accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _apply_preset(self, index: int) -> None:
        preset = self._preset_combo.currentData()
        if preset is None:
            return
        self._prefix_edit.setText(preset.get("prefix", ""))
        self._digits_spin.setValue(preset.get("digits", 6))
        sep_idx = self._sep_combo.findData(preset.get("separator", "-"))
        if sep_idx >= 0:
            self._sep_combo.setCurrentIndex(sep_idx)
        year_idx = self._year_combo.findData(preset.get("year_pos", "none"))
        if year_idx >= 0:
            self._year_combo.setCurrentIndex(year_idx)
        if not self._name_edit.text().strip():
            self._name_edit.setText(preset.get("label", ""))

    def _update_preview(self) -> None:
        try:
            s = self._build_series()
            preview = format_series_number(s, s.next_counter)
            self._preview_label.setText(f"预览：{preview}")
        except Exception:
            self._preview_label.setText("预览：（配置无效）")

    def _build_series(self) -> AccessionSeries:
        return AccessionSeries(
            name=self._name_edit.text().strip(),
            prefix=self._prefix_edit.text().strip().upper(),
            digits=self._digits_spin.value(),
            separator=self._sep_combo.currentData() or "",
            year_pos=self._year_combo.currentData() or "none",
            next_counter=self._counter_spin.value(),
            step=self._step_spin.value(),
        )

    def _accept(self) -> None:
        series = self._build_series()
        if not series.name:
            QMessageBox.warning(self, "提示", "请填写系列名称。")
            return
        if not series.prefix:
            QMessageBox.warning(self, "提示", "请填写前缀。")
            return
        if series.prefix.upper() == "YZZ":
            QMessageBox.warning(self, "提示", "YZZ 系列由系统专属管理，不可在此配置。")
            return
        # 检查名称重复（新增时）
        if self._editing is None:
            existing = [s.get("name") for s in self.store.config.get("accession_series", [])]
            if series.name in existing:
                QMessageBox.warning(self, "提示", f"系列名称「{series.name}」已存在，请换一个名称。")
                return
        self.result_series = series
        self.accept()


# ── 录入任务 / 编号分发 对话框 ────────────────────────────────────────────────


class _StartTaskDialog(QDialog):
    """开始录入任务:收集录入人员、用途、备注。

    规范化软件设计 2026-05 入库人员管理:录入人员 QLineEdit → PersonComboBox(智能下拉)。
    旧:每次手输姓名。现:从团队库选,可"管理人员…"打开 PersonsManagerDialog。
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("开始录入任务")
        self.setMinimumWidth(400)

        # 父窗口的工作区路径 (供 PersonComboBox refresh 用)
        self._parent_window = parent
        self._workspace = getattr(parent, "workspace_root", None) if parent is not None else None

        layout = QFormLayout(self)

        # 录入人员:智能下拉
        from .widgets_persons import PersonComboBox
        person_row = QHBoxLayout()
        self._person_combo = PersonComboBox(allow_manage=False)
        self._person_combo.refresh()  # 加载团队库
        # 预选 settings.current_recorder
        try:
            current_name = load_settings().current_recorder
            if current_name:
                self._person_combo.refresh(preselect=current_name)
        except Exception:
            pass
        self._person_combo.member_changed.connect(self._on_person_changed)
        person_row.addWidget(self._person_combo, 1)
        self._manage_btn = QPushButton("管理人员…")
        self._manage_btn.clicked.connect(self._open_manage)
        person_row.addWidget(self._manage_btn)
        person_widget = QWidget()
        person_widget.setLayout(person_row)
        layout.addRow("录入人员*", person_widget)

        self._purpose_combo = QComboBox()
        self._purpose_combo.addItems(["入库", "整理", "核查", "其他"])
        layout.addRow("用途", self._purpose_combo)

        self._note_edit = QLineEdit()
        self._note_edit.setPlaceholderText("可选")
        layout.addRow("备注", self._note_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self._ok_btn = buttons.button(QDialogButtonBox.Ok)
        if self._ok_btn:
            self._ok_btn.setText("开始任务")
            self._ok_btn.setEnabled(bool(self._person_combo.current_name()))
        cancel_btn = buttons.button(QDialogButtonBox.Cancel)
        if cancel_btn:
            cancel_btn.setText("取消")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def _on_person_changed(self, name_or_special: str) -> None:
        if self._ok_btn:
            # 仅当选中真人员(非空,非 SPECIAL_MANAGE)时才能开始
            m = self._person_combo.current_member()
            self._ok_btn.setEnabled(m is not None)
        # 若默认用途有,自动切
        m = self._person_combo.current_member()
        if m and m.default_purpose:
            idx = self._purpose_combo.findText(m.default_purpose)
            if idx >= 0:
                self._purpose_combo.setCurrentIndex(idx)

    def _open_manage(self) -> None:
        from .persons_dialog import PersonsManagerDialog
        store = getattr(self._parent_window, "store", None) if self._parent_window else None
        dlg = PersonsManagerDialog(self, workspace=self._workspace, store=store)
        dlg.exec_()
        # 刷新下拉,保留之前选中
        prev = self._person_combo.current_name()
        self._person_combo.refresh(preselect=prev)
        # 若 prev 被删,_ok_btn 自动 disable (通过 _on_person_changed)
        self._on_person_changed("")

    def accept(self) -> None:
        # 选中即写 last_used_at + 持久化 current_recorder
        name = self.person
        if not name:
            return
        try:
            from .persons_store import update_last_used
            update_last_used(name, self._workspace)
        except Exception:
            pass
        try:
            settings = load_settings()
            settings.current_recorder = name
            save_settings(settings)
        except Exception:
            pass
        super().accept()

    @property
    def person(self) -> str:
        m = self._person_combo.current_member()
        return m.name if m else ""

    @property
    def purpose(self) -> str:
        return self._purpose_combo.currentText()

    @property
    def note(self) -> str:
        return self._note_edit.text().strip()


class BatchImportSourcesDialog(QDialog):
    """S2：批量导入工作区目录 — 主管手动添加多个源目录，一键合并到中心机。"""

    def __init__(self, store, parent=None):
        super().__init__(parent)
        self._store = store
        self.setWindowTitle("批量导入工作区目录")
        self.setMinimumSize(560, 380)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            "选择多个工作区目录（每个必须含 数据/ 子目录），\n"
            "点「开始合并」一键并入当前中心工作区。源目录原样保留，不会被移动。"
        ))

        self._list_widget = QListWidget()
        self._list_widget.setSelectionMode(QListWidget.ExtendedSelection)
        layout.addWidget(self._list_widget)

        btn_row = QHBoxLayout()
        add_btn = QPushButton("添加目录…")
        add_btn.clicked.connect(self._add_dir)
        remove_btn = QPushButton("移除选中")
        remove_btn.clicked.connect(self._remove_selected)
        btn_row.addWidget(add_btn)
        btn_row.addWidget(remove_btn)
        btn_row.addStretch(1)
        layout.addLayout(btn_row)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self._ok_btn = buttons.button(QDialogButtonBox.Ok)
        if self._ok_btn:
            self._ok_btn.setText("开始合并")
            self._ok_btn.setEnabled(False)
        cancel_btn = buttons.button(QDialogButtonBox.Cancel)
        if cancel_btn:
            cancel_btn.setText("取消")
        buttons.accepted.connect(self._on_confirm)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _add_dir(self) -> None:
        d = QFileDialog.getExistingDirectory(self, "选择工作区目录")
        if not d:
            return
        existing = [self._list_widget.item(i).text() for i in range(self._list_widget.count())]
        if d in existing:
            return
        self._list_widget.addItem(d)
        if self._ok_btn:
            self._ok_btn.setEnabled(self._list_widget.count() > 0)

    def _remove_selected(self) -> None:
        for item in self._list_widget.selectedItems():
            self._list_widget.takeItem(self._list_widget.row(item))
        if self._ok_btn:
            self._ok_btn.setEnabled(self._list_widget.count() > 0)

    def _on_confirm(self) -> None:
        # 密码门控（与「从收件箱聚合」一致复用 ADMIN_PASSWORD）
        password, ok = QInputDialog.getText(
            self, "批量导入工作区",
            "本操作会合并所选目录到当前中心机。\n合并前自动快照，可一键回退。\n\n请输入管理密码：",
            QLineEdit.Password,
        )
        if not ok or not password:
            return
        if password != ADMIN_PASSWORD:
            QMessageBox.warning(self, "密码错误", "密码不正确。")
            return
        sources = [
            Path(self._list_widget.item(i).text())
            for i in range(self._list_widget.count())
        ]
        # S7 预览：先 dry-run 让用户看预计结果
        # 复用 preview_aggregate 需要 incoming 形态的目录；这里跳过预览（多选场景下
        # preview_aggregate 暂不支持 list 输入，留作后续优化）。直接确认即跑。
        confirm = QMessageBox.question(
            self, "确认合并",
            f"将合并 {len(sources)} 个源目录到当前中心工作区。\n继续？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )
        if confirm != QMessageBox.Yes:
            return
        try:
            report = aggregate_sources(self._store, sources)
        except Exception as exc:
            QMessageBox.critical(self, "合并失败", f"出错：{exc}")
            return
        lines: list[str] = []
        if not report.processed and not report.conflicted and not report.errored:
            lines.append("未发现可合并的工作区目录（每个源必须含 数据/ 子目录）。")
        else:
            lines.append(
                f"成功合并：{len(report.processed)} 个 — "
                f"共 {report.total_imported} 条 voucher、{report.total_photos} 张照片。"
            )
        if report.conflicted:
            lines.append(f"\n冲突：{len(report.conflicted)} 个 — 见冲突报告 xlsx。")
        if report.errored:
            lines.append(f"出错：{len(report.errored)} 个 — 见 错误描述。")
        if report.duplicates:
            lines.append(
                f"\n跨 voucher 同 SHA256 照片审核：{len(report.duplicates)} 个源命中 — "
                "报告在中心 数据/duplicates_报告/。"
            )
        if report.snapshot_path:
            lines.append(f"\n已自动快照：{report.snapshot_path.name}")
        QMessageBox.information(self, "批量合并完成", "\n".join(lines))
        self.accept()


class BatchGenerateDialog(QDialog):
    """批量预留入库编号，用于打印外贴标签。不创建标本记录。

    M2 新增"录入员独立前缀"开关（前缀分人，避免多人离线同时录入撞号）：
    - 未勾：走原 YZZ 单系列逻辑（向后兼容，旧调用零变化）
    - 勾选：自动调 `store.ensure_assignee_series(assignee, prefix)` 取/建该录入员
      专属系列，再 `batch_reserve_vouchers(n, series_name=该系列)` 预留段。
    """

    def __init__(self, store, parent=None):
        super().__init__(parent)
        self._store = store
        self.setWindowTitle("批量生成编号")
        self.setMinimumWidth(440)

        layout = QFormLayout(self)

        note = QLabel("预留连续编号段，不创建标本记录。可导出编号列表用于打印标签。")
        note.setWordWrap(True)
        layout.addRow(note)

        self._count_spin = QSpinBox()
        self._count_spin.setRange(1, 9999)
        self._count_spin.setValue(100)
        layout.addRow("生成数量", self._count_spin)

        self._person_edit = QLineEdit()
        self._person_edit.setPlaceholderText("必填")
        layout.addRow("领取人*", self._person_edit)

        self._purpose_combo = QComboBox()
        self._purpose_combo.addItems(["标签打印", "入库", "核查", "其他"])
        layout.addRow("用途", self._purpose_combo)

        self._note_edit = QLineEdit()
        self._note_edit.setPlaceholderText("可选")
        layout.addRow("备注", self._note_edit)

        # M2：为录入员创建独立编号前缀（前缀分人）。默认未勾保持旧行为。
        self._assignee_series_chk = QCheckBox("为该录入员使用独立编号前缀（多人协作时强烈建议）")
        self._assignee_series_chk.setToolTip(
            "勾选后，发给该录入员的编号会使用独立前缀（如 ZS-000001），\n"
            "与其他录入员的编号互不冲突；离线录入回传时不会撞号。\n"
            "未勾则沿用原 YZZ 单系列计数。"
        )
        layout.addRow(self._assignee_series_chk)
        self._prefix_edit = QLineEdit()
        self._prefix_edit.setPlaceholderText("仅 ASCII 字母/数字/横线/下划线，例：ZS-、LS-、A1-")
        self._prefix_edit.setEnabled(False)
        layout.addRow("录入员前缀", self._prefix_edit)
        self._assignee_series_chk.toggled.connect(self._prefix_edit.setEnabled)

        # M3：附带任务包 zip。勾选后 OK 按钮文字变为「生成并打包」，
        # 不再单独导出编号列表 xlsx，而是直接产出可发给录入员的任务包 zip
        # （内含空工作区骨架 + 字段模版 + manifest，使录入员可直接打开应用录入）。
        # 必须先勾「录入员独立前缀」；二者绑定 enable/disable。
        self._task_package_chk = QCheckBox("同时生成任务包 zip（含空工作区 + manifest，可直接发给录入员）")
        self._task_package_chk.setToolTip(
            "勾选后会产出一个 zip，录入员解压即可作为工作区使用。\n"
            "必须先勾选「录入员独立前缀」，否则任务包没有意义。"
        )
        self._task_package_chk.setEnabled(False)
        layout.addRow(self._task_package_chk)
        self._assignee_series_chk.toggled.connect(self._task_package_chk.setEnabled)
        self._assignee_series_chk.toggled.connect(
            lambda checked: None if checked else self._task_package_chk.setChecked(False)
        )

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self._ok_btn = buttons.button(QDialogButtonBox.Ok)
        if self._ok_btn:
            self._ok_btn.setText("生成并导出")
            self._ok_btn.setEnabled(False)
        cancel_btn = buttons.button(QDialogButtonBox.Cancel)
        if cancel_btn:
            cancel_btn.setText("取消")
        buttons.accepted.connect(self._on_confirm)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

        self._person_edit.textChanged.connect(
            lambda t: self._ok_btn.setEnabled(bool(t.strip())) if self._ok_btn else None
        )

    def _on_confirm(self) -> None:
        person = self._person_edit.text().strip()
        if not person:
            QMessageBox.warning(self, "提示", "请填写领取人。")
            return
        n = self._count_spin.value()

        # M3：勾选「同时生成任务包」走独立分支 — export_task_package 内部已包含
        # ensure_assignee_series + batch_reserve_vouchers + alloc_log（任务开始行）
        # + 写工作区骨架 + zip 打包；不再走下面的 batch_reserve_vouchers 通用分支
        # 避免双重预留。
        if self._task_package_chk.isChecked() and self._assignee_series_chk.isChecked():
            prefix = self._prefix_edit.text().strip()
            if not prefix:
                QMessageBox.warning(self, "提示", "勾选了「录入员独立前缀」后，请填写前缀。")
                return
            zip_path_str, _ = QFileDialog.getSaveFileName(
                self, "保存任务包",
                f"任务包_{person}_{n}个_{prefix.rstrip('-_')}.zip",
                "Zip 文件 (*.zip)",
            )
            if not zip_path_str:
                return
            try:
                result_path = export_task_package(
                    self._store, person, n, prefix,
                    dest_zip_path=zip_path_str,
                    purpose=self._purpose_combo.currentText(),
                    note=self._note_edit.text().strip(),
                )
            except ValueError as exc:
                QMessageBox.warning(self, "导出失败", str(exc))
                return
            except FileExistsError as exc:
                QMessageBox.warning(self, "目标已存在", str(exc))
                return
            except Exception as exc:
                QMessageBox.critical(self, "导出失败", f"任务包导出出错：{exc}")
                return
            QMessageBox.information(
                self, "任务包已生成",
                f"已为 {person} 生成任务包：\n{result_path}\n\n"
                f"预留段：{prefix}…（共 {n} 个编号）\n"
                f"中心机已记录任务开始；录入员可解压后直接打开工作区录入。",
            )
            self.accept()
            return

        # M2：勾选"录入员独立前缀"时先确保系列存在，再用该系列预留。
        series_name: str | None = None
        if self._assignee_series_chk.isChecked():
            prefix = self._prefix_edit.text().strip()
            if not prefix:
                QMessageBox.warning(self, "提示", "勾选了「录入员独立前缀」后，请填写前缀。")
                return
            try:
                series_name = self._store.ensure_assignee_series(person, prefix)
            except ValueError as exc:
                QMessageBox.warning(self, "前缀错误", str(exc))
                return
            except Exception as exc:
                QMessageBox.critical(self, "错误", f"创建录入员系列失败：{exc}")
                return

        try:
            numbers = self._store.batch_reserve_vouchers(n, series_name=series_name)
        except Exception as exc:
            QMessageBox.critical(self, "错误", f"生成编号失败：{exc}")
            return

        from datetime import datetime as _dt
        import uuid
        now = _dt.now().isoformat(timespec="seconds")
        record_id = uuid.uuid4().hex[:12]
        self._store.log_alloc_event({
            "记录ID": record_id,
            "时间": now,
            "类型": "批量领取",
            "人员": person,
            "用途": self._purpose_combo.currentText(),
            "备注": self._note_edit.text().strip(),
            # M2：填入"编号系列"列，便于事后审计与 M4 段守护校验。
            "编号系列": series_name or "YZZ",
            "编号起始": numbers[0],
            "编号结束": numbers[-1],
            "数量": str(n),
        })

        path, _ = QFileDialog.getSaveFileName(
            self, "保存编号列表",
            f"编号列表_{numbers[0]}-{numbers[-1]}.xlsx",
            "Excel 文件 (*.xlsx);;CSV 文件 (*.csv);;所有文件 (*)",
        )
        if path:
            try:
                self._export(numbers, person, path)
            except Exception as exc:
                QMessageBox.warning(self, "导出失败", str(exc))

        QMessageBox.information(
            self, "完成",
            f"已预留 {n} 个编号：{numbers[0]} → {numbers[-1]}\n分发记录已写入操作日志。",
        )
        self.accept()

    def _export(self, numbers: list, person: str, path: str) -> None:
        from datetime import datetime as _dt
        purpose = self._purpose_combo.currentText()
        now_str = _dt.now().strftime("%Y-%m-%d %H:%M:%S")
        rows = [[i, num, person, purpose, now_str] for i, num in enumerate(numbers, 1)]
        header = ["序号", "入库编号", "领取人", "用途", "生成时间"]

        if path.lower().endswith(".csv"):
            import csv
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                csv.writer(f).writerows([header] + rows)
        else:
            from openpyxl import Workbook as _WB
            wb = _WB()
            ws = wb.active
            ws.title = "编号列表"
            ws.append(header)
            for row in rows:
                ws.append(row)
            wb.save(path)
            # 同步导出同名 csv
            csv_path = path[:-5] + ".csv" if path.lower().endswith(".xlsx") else path + ".csv"
            import csv
            with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                csv.writer(f).writerows([header] + rows)


class WorkloadReportDialog(QDialog):
    """录入工作量报告：汇总 + 明细两视图，支持人员/时间段筛选和 Excel 导出。"""

    def __init__(self, store, parent=None):
        super().__init__(parent)
        self._store = store
        self.setWindowTitle("入库人员记录")
        self.setMinimumSize(640, 480)

        layout = QVBoxLayout(self)

        # 顶部：数据文件路径（让用户知道记录保存在哪里）
        from .models import ALLOC_LOG_FILE
        log_path = store.data_dir / ALLOC_LOG_FILE
        path_label = QLabel(f"数据文件：{log_path}")
        path_label.setStyleSheet("color: #555;")
        path_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(path_label)

        # ── 筛选栏 ──
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("人员:"))
        self._person_filter = QComboBox()
        self._person_filter.addItem("全部")
        filter_row.addWidget(self._person_filter)
        filter_row.addSpacing(12)
        filter_row.addWidget(QLabel("开始日期:"))
        self._date_from = QLineEdit()
        self._date_from.setPlaceholderText("YYYY-MM-DD")
        self._date_from.setFixedWidth(100)
        filter_row.addWidget(self._date_from)
        filter_row.addWidget(QLabel("至"))
        self._date_to = QLineEdit()
        self._date_to.setPlaceholderText("YYYY-MM-DD")
        self._date_to.setFixedWidth(100)
        filter_row.addWidget(self._date_to)
        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self._refresh)
        filter_row.addWidget(refresh_btn)
        filter_row.addStretch()
        layout.addLayout(filter_row)

        # ── 标签页 ──
        self._tabs = QTabWidget()
        self._summary_table = QTableWidget()
        self._summary_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._summary_table.setAlternatingRowColors(True)
        self._summary_table.horizontalHeader().setStretchLastSection(True)
        self._detail_table = QTableWidget()
        self._detail_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._detail_table.setAlternatingRowColors(True)
        self._detail_table.horizontalHeader().setStretchLastSection(True)
        self._tabs.addTab(self._summary_table, "汇总")
        self._tabs.addTab(self._detail_table, "明细")
        layout.addWidget(self._tabs)

        # ── 底部按钮 ──
        btn_row = QHBoxLayout()
        export_btn = QPushButton("导出 Excel")
        export_btn.clicked.connect(self._export_excel)
        btn_row.addWidget(export_btn)
        btn_row.addStretch()
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.reject)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

        self._refresh()

    # ── 数据处理 ──────────────────────────────────────────────────────────────

    def _parse_log(self) -> tuple[list[dict], list[str]]:
        """配对任务开始/结束记录，返回 (tasks, sorted_persons)。"""
        from datetime import datetime as _dt
        rows = self._store.read_alloc_log()
        starts = {r["记录ID"]: r for r in rows if r.get("类型") == "任务开始"}
        ends = [r for r in rows if r.get("类型") == "任务结束"]

        tasks = []
        for end in ends:
            start_id = end.get("关联任务ID", "")
            start = starts.get(start_id)
            if not start:
                continue
            try:
                t0 = _dt.fromisoformat(start["时间"])
                t1 = _dt.fromisoformat(end["时间"])
                duration_sec = max(0, int((t1 - t0).total_seconds()))
            except (ValueError, KeyError):
                duration_sec = 0
            tasks.append({
                "开始时间": start["时间"],
                "人员": start.get("人员", ""),
                "用途": start.get("用途", ""),
                "录入量": int(end.get("数量", 0) or 0),
                "时长秒": duration_sec,
            })
        persons = sorted({t["人员"] for t in tasks})
        return tasks, persons

    def _apply_filters(self, tasks: list[dict]) -> list[dict]:
        person = self._person_filter.currentText()
        date_from = self._date_from.text().strip()
        date_to = self._date_to.text().strip()
        result = []
        for t in tasks:
            if person != "全部" and t["人员"] != person:
                continue
            ts = t["开始时间"]
            if date_from and ts < date_from:
                continue
            if date_to and ts[:10] > date_to:
                continue
            result.append(t)
        return result

    @staticmethod
    def _fmt_duration(seconds: int) -> str:
        h, rem = divmod(seconds, 3600)
        m, _ = divmod(rem, 60)
        return f"{h}h {m:02d}m" if h else f"{m}m"

    # ── UI 刷新 ───────────────────────────────────────────────────────────────

    def _refresh(self) -> None:
        tasks, persons = self._parse_log()

        current_person = self._person_filter.currentText()
        self._person_filter.blockSignals(True)
        self._person_filter.clear()
        self._person_filter.addItem("全部")
        for p in persons:
            self._person_filter.addItem(p)
        idx = self._person_filter.findText(current_person)
        self._person_filter.setCurrentIndex(idx if idx >= 0 else 0)
        self._person_filter.blockSignals(False)

        filtered = self._apply_filters(tasks)
        self._fill_summary(filtered)
        self._fill_detail(filtered)

    def _fill_summary(self, filtered: list[dict]) -> None:
        from collections import defaultdict
        summary: dict[str, dict] = defaultdict(lambda: {"任务次数": 0, "录入标本数": 0, "时长秒": 0})
        for t in filtered:
            s = summary[t["人员"]]
            s["任务次数"] += 1
            s["录入标本数"] += t["录入量"]
            s["时长秒"] += t["时长秒"]

        headers = ["录入人", "任务次数", "录入标本数", "累计时长"]
        self._summary_table.setColumnCount(len(headers))
        self._summary_table.setHorizontalHeaderLabels(headers)
        rows = sorted(summary.items())
        self._summary_table.setRowCount(len(rows))
        for row, (person_name, stats) in enumerate(rows):
            self._summary_table.setItem(row, 0, QTableWidgetItem(person_name))
            self._summary_table.setItem(row, 1, QTableWidgetItem(str(stats["任务次数"])))
            self._summary_table.setItem(row, 2, QTableWidgetItem(str(stats["录入标本数"])))
            self._summary_table.setItem(row, 3, QTableWidgetItem(self._fmt_duration(stats["时长秒"])))
        self._summary_table.resizeColumnsToContents()

    def _fill_detail(self, filtered: list[dict]) -> None:
        headers = ["任务开始时间", "录入人", "用途", "录入量", "时长"]
        self._detail_table.setColumnCount(len(headers))
        self._detail_table.setHorizontalHeaderLabels(headers)
        sorted_tasks = sorted(filtered, key=lambda x: x["开始时间"], reverse=True)
        self._detail_table.setRowCount(len(sorted_tasks))
        for row, t in enumerate(sorted_tasks):
            self._detail_table.setItem(row, 0, QTableWidgetItem(t["开始时间"]))
            self._detail_table.setItem(row, 1, QTableWidgetItem(t["人员"]))
            self._detail_table.setItem(row, 2, QTableWidgetItem(t["用途"]))
            self._detail_table.setItem(row, 3, QTableWidgetItem(str(t["录入量"])))
            self._detail_table.setItem(row, 4, QTableWidgetItem(self._fmt_duration(t["时长秒"])))
        self._detail_table.resizeColumnsToContents()

    # ── 导出 ──────────────────────────────────────────────────────────────────

    def _export_excel(self) -> None:
        from datetime import datetime as _dt
        path, _ = QFileDialog.getSaveFileName(
            self, "保存工作量报告",
            f"录入工作量报告_{_dt.now().strftime('%Y%m%d')}.xlsx",
            "Excel 文件 (*.xlsx)",
        )
        if not path:
            return
        tasks, _ = self._parse_log()
        filtered = self._apply_filters(tasks)

        from collections import defaultdict
        from openpyxl import Workbook as _WB
        summary: dict[str, dict] = defaultdict(lambda: {"任务次数": 0, "录入标本数": 0, "时长秒": 0})
        for t in filtered:
            s = summary[t["人员"]]
            s["任务次数"] += 1
            s["录入标本数"] += t["录入量"]
            s["时长秒"] += t["时长秒"]

        wb = _WB()
        ws1 = wb.active
        ws1.title = "汇总"
        ws1.append(["录入人", "任务次数", "录入标本数", "累计时长"])
        for person_name, stats in sorted(summary.items()):
            ws1.append([person_name, stats["任务次数"], stats["录入标本数"],
                        self._fmt_duration(stats["时长秒"])])

        ws2 = wb.create_sheet("明细")
        ws2.append(["任务开始时间", "录入人", "用途", "录入量", "时长"])
        for t in sorted(filtered, key=lambda x: x["开始时间"], reverse=True):
            ws2.append([t["开始时间"], t["人员"], t["用途"], t["录入量"],
                        self._fmt_duration(t["时长秒"])])

        wb.save(path)
        QMessageBox.information(self, "完成", f"已导出：{path}")
