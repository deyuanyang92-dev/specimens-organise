from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


@dataclass(frozen=True)
class SpeciesMatch:
    chinese_name: str
    latin_name: str
    family_name: str
    family_latin: str

    @property
    def genus_name(self) -> str:
        return self.latin_name.split(" ", 1)[0].strip()


@dataclass(frozen=True)
class FamilyMatch:
    family_name: str
    family_latin: str


class SpeciesMatcher:
    """Live matcher backed by the preset Excel file.

    The file is reloaded when its mtime changes so users can update the preset
    workbook while the application is running.
    """

    def __init__(self, preset_path: Path):
        self.preset_path = preset_path
        self._mtime: float | None = None
        self._rows: list[SpeciesMatch] = []

    def matches(self, query: str, limit: int = 20) -> list[SpeciesMatch]:
        return self.species_matches(query, limit=limit)

    def species_matches(self, query: str, limit: int = 50) -> list[SpeciesMatch]:
        self._reload_if_needed()
        text = query.strip().lower()
        if not text:
            return []

        rows = _dedupe_species(self._rows)
        matched = [row for row in rows if _rank_species(row, text)[0] < 99]
        return sorted(matched, key=lambda row: _rank_species(row, text))[:limit]

    def family_matches(self, query: str, limit: int = 50) -> list[FamilyMatch]:
        self._reload_if_needed()
        text = query.strip().lower()
        if not text:
            return []

        rows = _dedupe_families(self._rows)
        matched = [row for row in rows if _rank_family(row, text)[0] < 99]
        return sorted(matched, key=lambda row: _rank_family(row, text))[:limit]

    def resolve_unique_species(self, query: str) -> SpeciesMatch | None:
        self._reload_if_needed()
        text = query.strip().lower()
        if not text:
            return None
        rows = _dedupe_species(self._rows)

        exact = [row for row in rows if row.chinese_name.lower() == text or row.latin_name.lower() == text]
        if len(exact) == 1:
            return exact[0]

        prefix = [
            row for row in rows
            if row.chinese_name.lower().startswith(text) or row.latin_name.lower().startswith(text)
        ]
        if len(prefix) == 1:
            return prefix[0]
        return None

    def resolve_unique_family(self, query: str) -> FamilyMatch | None:
        self._reload_if_needed()
        text = query.strip().lower()
        if not text:
            return None
        rows = _dedupe_families(self._rows)

        exact = [row for row in rows if row.family_name.lower() == text or row.family_latin.lower() == text]
        if len(exact) == 1:
            return exact[0]

        prefix = [
            row for row in rows
            if row.family_name.lower().startswith(text) or row.family_latin.lower().startswith(text)
        ]
        if len(prefix) == 1:
            return prefix[0]
        return None

    def all_rows(self) -> Iterable[SpeciesMatch]:
        self._reload_if_needed()
        return list(self._rows)

    def _reload_if_needed(self) -> None:
        if not self.preset_path.exists():
            self._rows = []
            self._mtime = None
            return
        mtime = self.preset_path.stat().st_mtime
        if self._mtime == mtime:
            return
        self._mtime = mtime
        self._rows = self._load_rows()

    def _load_rows(self) -> list[SpeciesMatch]:
        wb = load_workbook(self.preset_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            header = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
            mapping = {name: idx for idx, name in enumerate(header)}
            required = ["物种中文名", "物种拉丁名", "科中文名", "科拉丁名"]
            if any(name not in mapping for name in required):
                return []
            rows: list[SpeciesMatch] = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                chinese = _clean(row[mapping["物种中文名"]])
                if not chinese:
                    continue
                rows.append(
                    SpeciesMatch(
                        chinese_name=chinese,
                        latin_name=_clean(row[mapping["物种拉丁名"]]),
                        family_name=_clean(row[mapping["科中文名"]]),
                        family_latin=_clean(row[mapping["科拉丁名"]]),
                    )
                )
            return rows
        finally:
            wb.close()


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _dedupe_species(rows: Iterable[SpeciesMatch]) -> list[SpeciesMatch]:
    seen: set[tuple[str, str, str, str]] = set()
    unique: list[SpeciesMatch] = []
    for row in rows:
        key = (
            row.chinese_name.lower(),
            row.latin_name.lower(),
            row.family_name.lower(),
            row.family_latin.lower(),
        )
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    return unique


def _dedupe_families(rows: Iterable[SpeciesMatch]) -> list[FamilyMatch]:
    seen: set[tuple[str, str]] = set()
    unique: list[FamilyMatch] = []
    for row in rows:
        family_name = row.family_name.strip()
        family_latin = row.family_latin.strip()
        if not family_name and not family_latin:
            continue
        key = (family_name.lower(), family_latin.lower())
        if key in seen:
            continue
        seen.add(key)
        unique.append(FamilyMatch(family_name=family_name, family_latin=family_latin))
    return unique


def _rank_species(item: SpeciesMatch, text: str) -> tuple[int, int, str]:
    chinese = item.chinese_name.lower()
    latin = item.latin_name.lower()
    if chinese == text:
        return (0, len(chinese), chinese)
    if latin == text:
        return (1, len(latin), chinese)
    if chinese.startswith(text):
        return (2, len(chinese), chinese)
    if latin.startswith(text):
        return (3, len(latin), chinese)
    if text in chinese:
        return (4, len(chinese), chinese)
    if text in latin:
        return (5, len(latin), chinese)
    return (99, len(chinese), chinese)


def _rank_family(item: FamilyMatch, text: str) -> tuple[int, int, str]:
    family = item.family_name.lower()
    latin = item.family_latin.lower()
    if family == text:
        return (0, len(family), family)
    if latin == text:
        return (1, len(latin), family)
    if family.startswith(text):
        return (2, len(family), family)
    if latin.startswith(text):
        return (3, len(latin), family)
    if text in family:
        return (4, len(family), family)
    if text in latin:
        return (5, len(latin), family)
    return (99, len(family), family)
