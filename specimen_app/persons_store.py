"""入库人员管理 — 团队成员持久化 + 双向同步 (规范化软件设计 2026-05 新增)。

数据流:
    settings.json: ``team_members`` (全局,本地缓存)
                ↕ 双向同步
    工作区: ``数据/入库人员.xlsx`` (真权威,跟工作区拷贝走)

API:
- TeamMember dataclass + 字段
- load_members() -> list[TeamMember] (合并 settings + 工作区 xlsx,工作区优先)
- save_members(members, workspace) (写双方)
- sync_on_workspace_open(workspace) (新打开工作区时合并)
- color_for(name) (MD5 → HSL → #RRGGBB)
- avatar_text(name) (姓首字 中文 1 字 / 英文 2 字母)
"""

from __future__ import annotations

import colorsys
import hashlib
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Optional


# 中文姓名首字 / 英文姓名首 2 字 — 头像显示
def avatar_text(name: str) -> str:
    if not name:
        return "?"
    name = name.strip()
    if not name:
        return "?"
    # 中文 (UTF-8 编码下首字)
    first = name[0]
    if "一" <= first <= "鿿":  # CJK 统一汉字范围
        return first
    # 英文取前 2 字符大写
    return name[:2].upper()


def color_for(name: str) -> str:
    """根据姓名 hash 算固定头像色。Material Design 风:HSL 中等饱和度+亮度。"""
    if not name:
        return "#7c8a99"
    h_int = int(hashlib.md5(name.encode("utf-8")).hexdigest()[:8], 16) % 360
    r, g, b = colorsys.hls_to_rgb(h_int / 360.0, 0.50, 0.55)
    return "#{:02x}{:02x}{:02x}".format(int(r * 255), int(g * 255), int(b * 255))


# ---- TeamMember dataclass ----
ROLE_OPTIONS: dict[str, str] = {
    "recorder": "录入员",
    "checker": "核对员",
    "admin": "管理员",
}

PURPOSE_OPTIONS: list[str] = ["入库", "整理", "核查", "其他"]


@dataclass(slots=True)
class TeamMember:
    name: str
    pinyin: str = ""
    role: str = "recorder"
    starred: int = 0  # 0-5
    pinned: bool = False
    default_purpose: str = ""
    note: str = ""
    created_at: str = ""
    last_used_at: str = ""
    color_hint: str = ""

    def to_dict(self) -> dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, data: dict) -> "TeamMember":
        # 字段校验 + 默认值兜底
        name = str(data.get("name", "")).strip()
        if not name:
            raise ValueError("TeamMember requires non-empty name")
        starred_raw = data.get("starred", 0)
        try:
            starred = max(0, min(5, int(starred_raw)))
        except (TypeError, ValueError):
            starred = 0
        role = str(data.get("role", "recorder"))
        if role not in ROLE_OPTIONS:
            role = "recorder"
        return cls(
            name=name,
            pinyin=str(data.get("pinyin", "")),
            role=role,
            starred=starred,
            pinned=bool(data.get("pinned", False)),
            default_purpose=str(data.get("default_purpose", "")),
            note=str(data.get("note", "")),
            created_at=str(data.get("created_at", "")),
            last_used_at=str(data.get("last_used_at", "")),
            color_hint=str(data.get("color_hint", "")) or color_for(name),
        )

    def ensure_color(self) -> None:
        if not self.color_hint:
            self.color_hint = color_for(self.name)


def sort_key(m: TeamMember) -> tuple:
    """钉位 → 星标 desc → 最近用 desc → 字母。"""
    return (
        0 if m.pinned else 1,
        -m.starred,
        -_to_timestamp(m.last_used_at),
        m.name,
    )


def _to_timestamp(iso: str) -> float:
    if not iso:
        return 0.0
    try:
        return datetime.fromisoformat(iso).timestamp()
    except Exception:
        return 0.0


# ---- 工作区 xlsx 文件名 ----
TEAM_MEMBER_XLSX = "入库人员.xlsx"
TEAM_MEMBER_HEADERS = [
    "姓名", "拼音", "角色", "星标", "钉位",
    "默认用途", "备注", "创建时间", "最近使用", "颜色",
]


def _xlsx_path(workspace: Optional[Path]) -> Optional[Path]:
    if workspace is None:
        return None
    return Path(workspace) / "数据" / TEAM_MEMBER_XLSX


def _to_xlsx_row(m: TeamMember) -> list:
    return [
        m.name, m.pinyin, m.role, m.starred, "是" if m.pinned else "",
        m.default_purpose, m.note, m.created_at, m.last_used_at, m.color_hint,
    ]


def _from_xlsx_row(row: list) -> Optional[TeamMember]:
    if not row or not str(row[0]).strip():
        return None
    def _s(v): return str(v).strip() if v is not None else ""
    try:
        starred = int(row[3]) if len(row) > 3 and row[3] not in (None, "") else 0
    except Exception:
        starred = 0
    pinned = (len(row) > 4 and str(row[4]).strip() == "是")
    return TeamMember(
        name=_s(row[0]),
        pinyin=_s(row[1]) if len(row) > 1 else "",
        role=_s(row[2]) if len(row) > 2 else "recorder",
        starred=max(0, min(5, starred)),
        pinned=pinned,
        default_purpose=_s(row[5]) if len(row) > 5 else "",
        note=_s(row[6]) if len(row) > 6 else "",
        created_at=_s(row[7]) if len(row) > 7 else "",
        last_used_at=_s(row[8]) if len(row) > 8 else "",
        color_hint=_s(row[9]) if len(row) > 9 else "",
    )


def _read_xlsx(path: Path) -> list[TeamMember]:
    """读工作区 入库人员.xlsx;失败返空 list。"""
    if not path.exists():
        return []
    try:
        from openpyxl import load_workbook  # lazy
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if len(rows) < 2:
        return []
    result: list[TeamMember] = []
    for row in rows[1:]:
        m = _from_xlsx_row(list(row) if row else [])
        if m:
            m.ensure_color()
            result.append(m)
    return result


def _write_xlsx(path: Path, members: list[TeamMember]) -> None:
    """原子写工作区 入库人员.xlsx。"""
    import os
    path.parent.mkdir(parents=True, exist_ok=True)
    from openpyxl import Workbook  # lazy
    wb = Workbook()
    try:
        ws = wb.active
        ws.title = "入库人员"
        ws.append(TEAM_MEMBER_HEADERS)
        for m in members:
            m.ensure_color()
            ws.append(_to_xlsx_row(m))
        tmp = path.with_suffix(f".{os.getpid()}.tmp")
        wb.save(tmp)
        tmp.replace(path)
    finally:
        try:
            wb.close()
        except Exception:
            pass


# ---- 公共 API ----
def load_members(workspace: Optional[Path] = None) -> list[TeamMember]:
    """合并 settings.json + 工作区 xlsx 为单一团队列表。

    优先级: 工作区 xlsx 优先 (它跟工作区走,是权威);settings 中 xlsx 没有的成员补进来。
    返回已按 sort_key 排序的列表。
    """
    settings_members: dict[str, TeamMember] = {}
    try:
        from .app_settings import load_settings
        for d in load_settings().team_members or []:
            try:
                m = TeamMember.from_dict(d)
                settings_members[m.name] = m
            except Exception:
                continue
    except Exception:
        pass

    xlsx_members: dict[str, TeamMember] = {}
    xlsx_path = _xlsx_path(workspace) if workspace else None
    if xlsx_path is not None:
        for m in _read_xlsx(xlsx_path):
            xlsx_members[m.name] = m

    # 合并: xlsx 优先,settings 补漏
    merged: dict[str, TeamMember] = {}
    merged.update(settings_members)
    merged.update(xlsx_members)  # xlsx 覆盖同名 settings

    result = list(merged.values())
    for m in result:
        m.ensure_color()
    result.sort(key=sort_key)
    return result


def save_members(members: list[TeamMember], workspace: Optional[Path] = None) -> None:
    """同时写 settings.json + 工作区 xlsx。"""
    for m in members:
        m.ensure_color()
    # settings
    try:
        from .app_settings import load_settings, save_settings
        s = load_settings()
        s.team_members = [m.to_dict() for m in members]
        save_settings(s)
    except Exception:
        pass
    # workspace xlsx
    xlsx_path = _xlsx_path(workspace) if workspace else None
    if xlsx_path is not None:
        try:
            _write_xlsx(xlsx_path, members)
        except Exception:
            pass


def sync_on_workspace_open(workspace: Path) -> list[TeamMember]:
    """工作区打开时合并 settings ↔ xlsx,写回双方,返回最终列表。"""
    merged = load_members(workspace)
    save_members(merged, workspace)
    return merged


def update_last_used(name: str, workspace: Optional[Path] = None) -> None:
    """用户选用某人 → 更新 last_used_at。无对应 member 静默忽略。"""
    if not name:
        return
    members = load_members(workspace)
    now = datetime.now().isoformat(sep=" ", timespec="seconds")
    changed = False
    for m in members:
        if m.name == name:
            m.last_used_at = now
            changed = True
            break
    if changed:
        save_members(members, workspace)


def find_member(name: str, members: Optional[list[TeamMember]] = None) -> Optional[TeamMember]:
    if not name:
        return None
    if members is None:
        members = load_members()
    for m in members:
        if m.name == name:
            return m
    return None
