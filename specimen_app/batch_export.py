"""批量导出功能：类似 NCBI Batch Entrez，粘贴入库编号后一键导出标本数据与照片。

使用方式：
- 从凭证列表多选后右键 →「批量导出选中」
- 或点击工具栏「批量导出」按钮，在对话框中手动粘贴编号

修改记录（# 注释保留原有逻辑，注明变更原因与兼容性）。
"""

from __future__ import annotations

import os
import re
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

# 规范化软件设计 2026-05 P1 优化:openpyxl 改函数内 lazy import,启动期不触发加载。
# 模块常量 _HEADER_FONT 等改为 lazy property(_get_header_style).
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPlainTextEdit,
    QPushButton,
    QSlider,
    QSpinBox,
    QVBoxLayout,
)

if TYPE_CHECKING:
    from .excel_store import ExcelStore
    from openpyxl import Workbook  # 仅 type hint 用,运行时 from __future__ annotations 不解析

# Excel 表头样式 — lazy(模块加载时不触发 openpyxl import)
_HEADER_STYLE_CACHE: tuple | None = None

def _get_header_style() -> tuple:
    """首次调用时加载 openpyxl 并构造 (Font, PatternFill, Alignment)。"""
    global _HEADER_STYLE_CACHE
    if _HEADER_STYLE_CACHE is None:
        from openpyxl.styles import Font, PatternFill, Alignment
        _HEADER_STYLE_CACHE = (
            Font(bold=True, size=11),
            PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
            Alignment(horizontal="center"),
        )
    return _HEADER_STYLE_CACHE

# 照片导出格式：「保持原格式」= 原样 shutil.copy2（旧行为，兼容）；其余用 Pillow 重新编码。
PHOTO_EXPORT_FORMATS = ("保持原格式", "JPG", "PNG", "TIFF")
_FORMAT_EXT = {"JPG": ".jpg", "PNG": ".png", "TIFF": ".tif"}

_UNSAFE_CHARS = re.compile(r'[/\\:*?"<>|]')


def _safe_dirname(name: str) -> str:
    """Return a filesystem-safe directory name from *name*.

    Replaces Windows-illegal chars with '_', strips leading/trailing dots,
    and caps at 60 characters (prevents MAX_PATH overrun with 科/属/种 nesting).
    """
    cleaned = _UNSAFE_CHARS.sub("_", name).strip(". ")
    return cleaned[:60] or "_"


def _species_dirname(cls: dict) -> str:
    """Build species folder name from classification dict (中文种名_拉丁种名 or fallback)."""
    cn  = (cls.get("种名*") or "").strip()
    lat = (cls.get("种拉丁") or "").strip()
    if cn and lat:
        return _safe_dirname(f"{cn}_{lat}")
    return _safe_dirname(cn or lat or "未知种")


# Each resolver: (spec_dict, cls_dict) → folder name string (already safe).
# Order of keys = default display order in the UI list.
_LEVEL_RESOLVERS = {
    "采集地点缩写": lambda spec, cls: _safe_dirname(spec.get("采集地点缩写*") or "未知地点"),
    "目":          lambda spec, cls: _safe_dirname(cls.get("目") or "未知目"),
    "科":          lambda spec, cls: _safe_dirname(cls.get("科*") or "未分类"),
    "属":          lambda spec, cls: _safe_dirname(cls.get("属名") or "未知属"),
    "种名":        lambda spec, cls: _species_dirname(cls),
}
_FOLDER_LEVEL_DEFAULTS = {"科", "属", "种名"}  # checked by default (= original 按科/属/种名 behavior)


def _parse_voucher_numbers(text: str) -> list[str]:
    """从文本中解析入库编号列表。

    支持换行、逗号、空格、分号、中文逗号作为分隔符。
    返回去重并保持顺序的编号列表。
    """
    # 统一替换各种分隔符为换行
    cleaned = text.replace(",", "\n").replace("，", "\n")
    cleaned = cleaned.replace(";", "\n").replace("；", "\n")
    cleaned = cleaned.replace(" ", "\n")
    parts = [p.strip() for p in cleaned.split("\n")]
    seen: set[str] = set()
    result: list[str] = []
    for part in parts:
        if part and part not in seen:
            seen.add(part)
            result.append(part)
    return result


def _write_header_row(ws, headers: list[str]) -> None:
    """写入带样式的表头行。"""
    header_font, header_fill, header_alignment = _get_header_style()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


def _auto_width(ws, min_width: int = 8, max_width: int = 60) -> None:
    """自动调整列宽（根据内容）。"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col_cells:
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                # 中文字符按 2 个字符宽度计算
                text = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in text)
                max_len = max(max_len, length)
        if col_letter:
            ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


class BatchExportDialog(QDialog):
    """批量导出对话框。

    支持从凭证列表多选带入编号，也可手动粘贴。
    可选择导出标本信息、分类信息、照片路径、照片文件，
    支持打包为 ZIP 方便分发。
    """

    def __init__(
        self,
        store: "ExcelStore",
        preselected: list[str] | None = None,
        parent=None,
        photo_focus: bool = False,
    ):
        # photo_focus=True：从「入库汇总 → 导出选中照片」入口进来，只关心照片文件，
        #   默认只勾「照片文件」、三个 Excel sheet 默认不勾，标题改「导出照片」。
        # photo_focus=False：工具栏/凭证列表右键的旧入口，保持原默认（旧行为不变）。
        super().__init__(parent)
        self.store = store
        self._photo_focus = photo_focus
        self.setWindowTitle("导出照片" if photo_focus else "批量导出")
        self.setMinimumSize(520, 480)

        layout = QVBoxLayout(self)

        # ---- 入库编号输入 ----
        layout.addWidget(QLabel("入库编号（支持换行 / 逗号 / 空格 / 分号分隔）："))
        self._text_edit = QPlainTextEdit()
        self._text_edit.setPlaceholderText("粘贴入库编号，每行一个...\n例如：\nYZZ000001\nYZZ000042")
        self._text_edit.setMaximumHeight(120)
        if preselected:
            self._text_edit.setPlainText("\n".join(preselected))
        self._text_edit.textChanged.connect(self._on_text_changed)
        layout.addWidget(self._text_edit)

        self._count_label = QLabel("已识别 0 个有效编号")
        self._count_label.setStyleSheet("color: #59666b; font-size: 11px;")
        layout.addWidget(self._count_label)

        # ---- 导出内容选项 ----
        options_group = QGroupBox("导出内容")
        options_layout = QVBoxLayout(options_group)

        # 旧逻辑：标本/分类/照片路径默认勾选、照片文件默认不勾。
        # photo_focus 入口只导照片：三个非照片选项隐藏 + 取消勾选（控件仍创建，
        # _do_export 仍按 .isChecked() 引用、判断逻辑不动；隐藏行不占布局空间）。
        self._chk_specimen = QCheckBox("标本信息")
        self._chk_specimen.setChecked(not photo_focus)
        options_layout.addWidget(self._chk_specimen)

        self._chk_classification = QCheckBox("分类信息")
        self._chk_classification.setChecked(not photo_focus)
        options_layout.addWidget(self._chk_classification)

        self._chk_photo_paths = QCheckBox("照片路径清单")
        self._chk_photo_paths.setChecked(not photo_focus)
        options_layout.addWidget(self._chk_photo_paths)

        for _chk in (self._chk_specimen, self._chk_classification, self._chk_photo_paths):
            _chk.setVisible(not photo_focus)

        self._chk_photo_files = QCheckBox("照片文件（复制到导出目录）")
        self._chk_photo_files.setChecked(photo_focus)
        options_layout.addWidget(self._chk_photo_files)

        # ZIP 打包选项（仅当勾选照片文件时可用）
        zip_row = QHBoxLayout()
        zip_row.setContentsMargins(24, 0, 0, 0)
        self._chk_zip = QCheckBox("打包为 ZIP")
        self._chk_zip.setToolTip("将所有导出内容打包为一个 ZIP 文件，方便分享")
        self._chk_photo_files.toggled.connect(
            lambda checked: self._chk_zip.setEnabled(checked)
        )
        self._chk_zip.setEnabled(self._chk_photo_files.isChecked())
        zip_row.addWidget(self._chk_zip)
        zip_row.addStretch()
        options_layout.addLayout(zip_row)

        layout.addWidget(options_group)

        # ---- 照片导出格式 / 压缩 ----
        # 原代码照片只能原样复制；这里新增格式转换 + 压缩，整组仅在勾选「照片文件」时启用。
        fmt_group = QGroupBox("照片导出格式 / 压缩")
        fmt_layout = QVBoxLayout(fmt_group)

        fmt_row = QHBoxLayout()
        fmt_row.addWidget(QLabel("导出格式："))
        self._photo_format = QComboBox()
        self._photo_format.addItems(PHOTO_EXPORT_FORMATS)
        self._photo_format.setToolTip("「保持原格式」= 原样复制；其余用 Pillow 重新编码导出")
        self._photo_format.currentIndexChanged.connect(self._sync_photo_format_enabled)
        fmt_row.addWidget(self._photo_format)
        fmt_row.addStretch()
        fmt_layout.addLayout(fmt_row)

        quality_row = QHBoxLayout()
        self._quality_label = QLabel("JPEG 质量：")
        quality_row.addWidget(self._quality_label)
        self._quality_slider = QSlider(Qt.Horizontal)
        self._quality_slider.setRange(1, 100)
        self._quality_slider.setValue(90)
        self._quality_value = QLabel("90")
        self._quality_value.setFixedWidth(28)
        self._quality_slider.valueChanged.connect(
            lambda v: self._quality_value.setText(str(v))
        )
        quality_row.addWidget(self._quality_slider, stretch=1)
        quality_row.addWidget(self._quality_value)
        fmt_layout.addLayout(quality_row)

        resize_row = QHBoxLayout()
        self._chk_resize = QCheckBox("限制最大边长")
        self._chk_resize.setToolTip("超过该像素的照片等比缩小（不放大）；TIFF 原图始终不变")
        self._chk_resize.toggled.connect(self._sync_photo_format_enabled)
        resize_row.addWidget(self._chk_resize)
        self._resize_spin = QSpinBox()
        self._resize_spin.setRange(100, 100000)
        self._resize_spin.setValue(4000)
        self._resize_spin.setSuffix(" px")
        resize_row.addWidget(self._resize_spin)
        resize_row.addStretch()
        fmt_layout.addLayout(resize_row)

        # 文件夹层级：勾选 + 拖拽排序（原「按分类建子目录」复选框 → 原下拉预设 → 现自由组合列表）
        # 默认勾选「科/属/种名」= 原有默认行为；全不勾 = 平铺。
        fmt_layout.addWidget(QLabel("文件夹层级（勾选 + 拖拽排序）："))
        self._folder_levels = QListWidget()
        self._folder_levels.setDragDropMode(QAbstractItemView.InternalMove)
        self._folder_levels.setFixedHeight(130)
        self._folder_levels.setToolTip(
            "勾选的层级按列表顺序拼成导出路径（可上下拖动调整顺序）。\n"
            "全不勾选 = 平铺（所有照片在「照片/」根下）。\n"
            "无论如何组合，文件名末尾均追加 _入库编号。"
        )
        for _lvl_label in _LEVEL_RESOLVERS:
            _item = QListWidgetItem(_lvl_label)
            _item.setFlags(_item.flags() | Qt.ItemIsUserCheckable)
            _item.setCheckState(Qt.Checked if _lvl_label in _FOLDER_LEVEL_DEFAULTS else Qt.Unchecked)
            self._folder_levels.addItem(_item)
        fmt_layout.addWidget(self._folder_levels)

        layout.addWidget(fmt_group)

        # 「照片文件」勾选状态联动整组启用（与 ZIP 选项同一个信号源）。
        self._chk_photo_files.toggled.connect(self._sync_photo_format_enabled)
        self._sync_photo_format_enabled()

        # ---- 输出目录 ----
        output_row = QHBoxLayout()
        output_row.addWidget(QLabel("输出目录："))
        self._output_path = QLabel("（请选择目录）")
        self._output_path.setStyleSheet("color: #59666b;")
        output_row.addWidget(self._output_path, stretch=1)
        browse_btn = QPushButton("浏览...")
        browse_btn.clicked.connect(self._browse_output_dir)
        output_row.addWidget(browse_btn)
        layout.addLayout(output_row)

        # ---- 按钮 ----
        buttons = QDialogButtonBox()
        export_btn = buttons.addButton("导出", QDialogButtonBox.AcceptRole)
        export_btn.setStyleSheet("QPushButton { font-weight: bold; padding: 4px 16px; }")
        buttons.addButton("取消", QDialogButtonBox.RejectRole)
        buttons.accepted.connect(self._do_export)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self._on_text_changed()

    # ---- 事件处理 ----

    def _on_text_changed(self) -> None:
        """更新已识别编号计数。"""
        vouchers = _parse_voucher_numbers(self._text_edit.toPlainText())
        valid = [v for v in vouchers if self.store.get_specimen(v)]
        self._count_label.setText(
            f"已识别 {len(vouchers)} 个编号（{len(valid)} 个有效，"
            f"{len(vouchers) - len(valid)} 个不存在）"
        )

    def _browse_output_dir(self) -> None:
        directory = QFileDialog.getExistingDirectory(self, "选择导出目录")
        if directory:
            self._output_path.setText(directory)
            self._output_path.setStyleSheet("color: #000;")

    def _sync_photo_format_enabled(self) -> None:
        """照片格式/压缩组的联动启用：仅勾「照片文件」时可用；质量条仅 JPG 可用；
        最大边长仅在格式≠保持原格式时可用、数值框再随勾选启用。"""
        photo_on = self._chk_photo_files.isChecked()
        fmt = self._photo_format.currentText()
        self._photo_format.setEnabled(photo_on)
        is_jpg = photo_on and fmt == "JPG"
        self._quality_label.setEnabled(is_jpg)
        self._quality_slider.setEnabled(is_jpg)
        self._quality_value.setEnabled(is_jpg)
        can_resize = photo_on and fmt != "保持原格式"
        self._chk_resize.setEnabled(can_resize)
        self._resize_spin.setEnabled(can_resize and self._chk_resize.isChecked())
        self._folder_levels.setEnabled(photo_on)

    # ---- 导出逻辑 ----

    def _do_export(self) -> None:
        """执行批量导出。

        流程：解析编号 → 验证勾选 → 创建输出目录 → 逐项导出 → 可选打包 ZIP。
        所有 `#` 注释说明每一步的设计意图与兼容性考虑。
        """
        # 1. 解析入库编号
        vouchers = _parse_voucher_numbers(self._text_edit.toPlainText())
        if not vouchers:
            QMessageBox.warning(self, "无编号", "请粘贴或输入至少一个入库编号。")
            return

        # 过滤不存在的编号
        valid_vouchers: list[str] = []
        missing: list[str] = []
        for v in vouchers:
            if self.store.get_specimen(v):
                valid_vouchers.append(v)
            else:
                missing.append(v)

        if not valid_vouchers:
            QMessageBox.warning(self, "无有效编号", "输入的所有编号均不存在于当前工作区。")
            return

        # 2. 验证至少勾选一个导出项
        if not any([
            self._chk_specimen.isChecked(),
            self._chk_classification.isChecked(),
            self._chk_photo_paths.isChecked(),
            self._chk_photo_files.isChecked(),
        ]):
            QMessageBox.warning(self, "未选择内容", "请至少勾选一项要导出的内容。")
            return

        # 3. 确定输出目录
        out_dir_text = self._output_path.text().strip()
        if out_dir_text == "（请选择目录）":
            QMessageBox.warning(self, "未选择目录", "请选择导出目录。")
            return

        out_dir = Path(out_dir_text)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_dir = out_dir / f"导出_{timestamp}"
        try:
            export_dir.mkdir(parents=True, exist_ok=True)
        except OSError as e:
            QMessageBox.critical(self, "创建目录失败", f"无法创建导出目录：{e}")
            return

        # 4. 逐项导出
        errors: list[str] = []
        photo_count = 0

        try:
            # 写入合并 Excel 文件
            from openpyxl import Workbook  # lazy, P1 优化
            wb = Workbook()
            # 规范化软件设计 2026-05 P1 审查修复:用 try/finally 确保 Workbook 在任意路径都 close,
            # 防 save 抛异常时文件句柄泄漏。
            try:
                # 删除默认空白 sheet（openpyxl 新工作簿自带 "Sheet"）
                wb.remove(wb.active)

                if self._chk_specimen.isChecked():
                    self._export_specimen_sheet(wb, valid_vouchers, errors)

                if self._chk_classification.isChecked():
                    self._export_classification_sheet(wb, valid_vouchers, errors)

                if self._chk_photo_paths.isChecked():
                    self._export_photo_paths_sheet(wb, valid_vouchers, errors)

                if wb.sheetnames:
                    xlsx_path = export_dir / "导出汇总.xlsx"
                    wb.save(xlsx_path)
            finally:
                try:
                    wb.close()
                except Exception:
                    pass

            # 复制照片文件
            if self._chk_photo_files.isChecked():
                photo_count = self._export_photo_files(valid_vouchers, export_dir, errors)

        except Exception as exc:
            QMessageBox.critical(self, "导出失败", f"导出过程中出现异常：{exc}")
            return

        # 5. 可选 ZIP 打包
        # 规范化软件设计 2026-05 P1 审查修复:
        # 旧:make_archive 成功后 rmtree 若失败(权限/锁),已生成 zip 但源目录残留。
        # 现:先 make_archive,成功后单独 try rmtree;rmtree 失败不影响 final_path(已是 zip),
        #     仅追加警告让用户知道残留目录。
        final_path = export_dir
        if self._chk_photo_files.isChecked() and self._chk_zip.isChecked():
            zip_base = str(export_dir)
            try:
                final_zip = shutil.make_archive(zip_base, "zip", export_dir.parent, export_dir.name)
                final_path = Path(final_zip)
                # 打包成功 → 尝试清理源目录(失败不阻断)
                try:
                    shutil.rmtree(export_dir)
                except Exception as rm_exc:
                    errors.append(f"ZIP 已生成,但清理临时目录失败:{rm_exc}(可手动删 {export_dir})")
            except Exception as exc:
                errors.append(f"ZIP 打包失败:{exc}")

        # 6. 结果反馈
        missing_warning = ""
        if missing:
            missing_warning = f"\n\n跳过的无效编号：{', '.join(missing[:10])}"
            if len(missing) > 10:
                missing_warning += f" ...等共 {len(missing)} 个"

        error_warning = ""
        if errors:
            error_warning = f"\n\n警告：\n" + "\n".join(errors[:5])
            if len(errors) > 5:
                error_warning += f"\n...等共 {len(errors)} 个问题"

        QMessageBox.information(
            self,
            "导出完成",
            f"已成功导出 {len(valid_vouchers)} 个标本。\n"
            f"输出位置：{final_path}\n"
            f"导出照片：{photo_count} 张"
            + missing_warning
            + error_warning,
        )
        self.accept()

    def _export_specimen_sheet(self, wb: Workbook, vouchers: list[str], errors: list[str]) -> None:
        """导出标本信息 sheet。"""
        from .models import SPECIMEN_HEADERS

        ws = wb.create_sheet("标本信息")
        _write_header_row(ws, SPECIMEN_HEADERS)
        for row_idx, voucher in enumerate(vouchers, 2):
            specimen = self.store.get_specimen(voucher) or {}
            for col_idx, header in enumerate(SPECIMEN_HEADERS, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(specimen.get(header, "")))
        _auto_width(ws)

    def _export_classification_sheet(self, wb: Workbook, vouchers: list[str], errors: list[str]) -> None:
        """导出分类信息 sheet。"""
        from .models import CLASSIFICATION_HEADERS

        ws = wb.create_sheet("分类信息")
        _write_header_row(ws, CLASSIFICATION_HEADERS)
        for row_idx, voucher in enumerate(vouchers, 2):
            classification = self.store.get_classification(voucher) or {}
            for col_idx, header in enumerate(CLASSIFICATION_HEADERS, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(classification.get(header, "")))
        _auto_width(ws)

    def _export_photo_paths_sheet(self, wb: Workbook, vouchers: list[str], errors: list[str]) -> None:
        """导出照片路径清单 sheet。

        每行列出：入库编号、文件名、原始文件名、相对路径、绝对路径（解析后）。
        原代码无此功能；这里新增为独立 sheet，供用户核对照片位置。
        """
        headers = ["入库编号", "文件名", "原始文件名", "相对路径", "绝对路径", "文件存在"]
        ws = wb.create_sheet("照片路径")
        _write_header_row(ws, headers)
        row_idx = 2
        for voucher in vouchers:
            photos = self.store.get_photos(voucher)
            for photo in photos:
                resolved = self.store.resolve_photo_path(photo)
                ws.cell(row=row_idx, column=1, value=voucher)
                ws.cell(row=row_idx, column=2, value=str(photo.get("文件名", "")))
                ws.cell(row=row_idx, column=3, value=str(photo.get("原始文件名", "")))
                ws.cell(row=row_idx, column=4, value=str(photo.get("相对路径", "")))
                ws.cell(row=row_idx, column=5, value=str(resolved))
                ws.cell(row=row_idx, column=6, value="是" if resolved.exists() else "否")
                row_idx += 1
        _auto_width(ws)

    def _export_photo_files(self, vouchers: list[str], export_dir: Path, errors: list[str]) -> int:
        """导出照片文件到导出目录的 /照片 子文件夹。

        旧逻辑：一律 shutil.copy2 原样复制。保留为「保持原格式」分支（兼容）。
        新增：
        - 选 JPG/PNG/TIFF 时用 Pillow 重新编码（可调质量 + 可选等比缩放）。
        - 所有照片文件名末尾追加 _入库编号（stem_voucher.ext）。
        - 勾选「按分类建子目录」时，按「科/属/种名」建多级子目录（fallback：未分类/未知属/未知种）。
        处理同名冲突：使用 _2, _3 后缀（基于目标目录+文件名，避免跨目录误计）。
        返回成功导出的照片数量。
        """
        base_photo_dir = export_dir / "照片"
        base_photo_dir.mkdir(exist_ok=True)
        count = 0
        seen_names: dict[tuple, int] = {}  # (str(photo_dir), filename) → collision count

        fmt = self._photo_format.currentText()
        quality = self._quality_slider.value()
        max_edge = self._resize_spin.value() if self._chk_resize.isChecked() else 0

        # Collect active (checked) levels in current display order — supports drag-reorder.
        active_levels = [
            self._folder_levels.item(i).text()
            for i in range(self._folder_levels.count())
            if self._folder_levels.item(i).checkState() == Qt.Checked
        ]

        for voucher in vouchers:
            photos = self.store.get_photos(voucher)

            # --- 确定目标子目录（每个标本固定一个，该标本所有照片归入同一目录）---
            # 原逻辑：只有「按科/属/种名」和「平铺」两种。现改为自由勾选+排序（_folder_levels）。
            if active_levels:
                spec = self.store.get_specimen(voucher) or {}
                cls  = self.store.get_classification(voucher) or {}
                photo_dir = base_photo_dir
                for lvl in active_levels:
                    photo_dir = photo_dir / _LEVEL_RESOLVERS[lvl](spec, cls)
            else:
                photo_dir = base_photo_dir  # 平铺：全不勾选时
            photo_dir.mkdir(parents=True, exist_ok=True)

            for photo in photos:
                resolved = self.store.resolve_photo_path(photo)
                if not resolved.exists():
                    continue

                # --- 文件名：先追加 _入库编号，再转格式（保留旧行为结构，插入后缀）---
                base_name = str(photo.get("文件名", "")) or resolved.name
                stem, ext = os.path.splitext(base_name)
                filename = f"{stem}_{voucher}{ext}"  # 追加入库编号后缀

                # 转格式时把扩展名换成目标格式（「保持原格式」不动，等同旧行为）。
                if fmt != "保持原格式":
                    stem2, _ = os.path.splitext(filename)
                    filename = f"{stem2}{_FORMAT_EXT[fmt]}"

                # 处理同名冲突（按目标目录单独计数，不同目录间互不干扰）。
                dir_key = (str(photo_dir), filename)
                if dir_key in seen_names:
                    seen_names[dir_key] += 1
                    stem3, ext3 = os.path.splitext(filename)
                    dest_name = f"{stem3}_{seen_names[dir_key]}{ext3}"
                else:
                    seen_names[dir_key] = 1
                    dest_name = filename

                dest = photo_dir / dest_name
                try:
                    if fmt == "保持原格式":
                        shutil.copy2(resolved, dest)  # 原分支：原样复制
                    else:
                        self._reencode_photo(resolved, dest, fmt, quality, max_edge)
                    count += 1
                except Exception as exc:  # Pillow 解码/编码可能抛多种异常，逐张兜底不中断
                    errors.append(f"导出照片失败 [{filename}]：{exc}")

        return count

    @staticmethod
    def _reencode_photo(src: Path, dest: Path, fmt: str, quality: int, max_edge: int) -> None:
        """用 Pillow 把 src 重新编码为目标格式写到 dest。

        max_edge>0 时按最大边长等比缩小（不放大）。TIFF 等多页/特殊图先靠 Pillow
        原生解码；解不开会抛异常,由调用方逐张兜底记入 errors。

        规范化软件设计 2026-05 P1 审查修复:
        旧:`img = ImageOps.exif_transpose(opened)` 后 `img` 可能是新对象,
            JPG 路径再 `img.convert("RGB")` 又创新对象,中间 Image 没显式关。
        现:用嵌套 try/finally 确保所有中间 Image 对象的 close。
        """
        from PIL import Image, ImageOps

        with Image.open(src) as opened:
            img = ImageOps.exif_transpose(opened)
            try:
                if max_edge > 0:
                    img.thumbnail((max_edge, max_edge), Image.LANCZOS)
                if fmt == "JPG":
                    rgb = img.convert("RGB")
                    try:
                        rgb.save(dest, "JPEG", quality=quality, optimize=True)
                    finally:
                        if rgb is not img:
                            try: rgb.close()
                            except Exception: pass
                elif fmt == "PNG":
                    img.save(dest, "PNG")
                else:  # TIFF
                    img.save(dest, "TIFF")
            finally:
                # exif_transpose 在大多数情况返回新对象,显式关确保不依赖 GC。
                if img is not opened:
                    try: img.close()
                    except Exception: pass
